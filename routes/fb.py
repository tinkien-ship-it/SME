"""Routes F&B (menu, bàn, order) — tách từ app.py."""
import json
import logging
import os
import re
import sqlite3
import traceback
from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO

import pandas as pd
import requests
from flask import (
    Response,
    abort,
    flash,
    jsonify,
    make_response,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from werkzeug.utils import secure_filename

from Services.invoice_buyer import DEFAULT_RETAIL_BUYER_NAME, normalize_retail_buyer_name
from Services.pos_offline_schema import ensure_pos_offline_schema, find_sale_by_client_uuid

logger = logging.getLogger(__name__)



from Services.fb_product_codes import (
    assign_raw_material_product_codes,
    assign_ready_made_product_codes,
    get_next_menu_code,
    get_next_raw_material_code,
)
from Services.hkd_sector import resolve_item_hkd_sector
from Services.sale_helpers import insert_sale_item_with_sector
from Services.inventory_stock_helpers import sync_inventory_quantity_from_moves
from Services.sme.sale_journal import sync_sale_journals
from Services.accounting_queue import enqueue_accounting_job
from Services.tenant_profile import get_current_tenant_profile
from Services.schema_compat import (
    ensure_sale_items_canonical,
    expand_use_sale_unit_values,
    normalize_use_sale_unit,
    sale_item_pk_column,
    sale_item_pk_expr,
    use_sale_unit_expr,
    use_sale_unit_insert_columns,
    use_sale_unit_where_clause,
)
from db_utils import (
    _is_locked_error,
    begin_immediate,
    get_db_connection,
    locked_user_message,
    rollback_quietly,
    sqlite_commit,
    sqlite_write_retry,
)


def _fb_table_cols(cursor, table: str) -> set[str]:
    from Services.schema_compat import table_cols_lower
    return table_cols_lower(cursor, table)


def _fb_use_sale_unit_expr(cursor, alias: str = 'si') -> str:
    return use_sale_unit_expr(cursor, alias)


def _fb_normalize_use_sale_unit(raw) -> int:
    return normalize_use_sale_unit(raw)


def _fb_load_sale_items(cursor, sale_id):
    use_expr = use_sale_unit_expr(cursor, 'si')
    pk_expr = sale_item_pk_expr(cursor, 'si')
    return cursor.execute(f"""
        SELECT 
            {pk_expr} AS sale_item_id,
            si.menu_id,
            si.quantity,
            si.price,
            si.line_total,
            si.unit AS sale_item_unit,
            {use_expr} AS db_use_sale_unit,
            m.product_type, 
            m.name as item_name, 
            m.product_id as origin_product_id,
            p.unit_ratio,
            p.unit as product_stock_unit,
            i.avg_cost, 
            i.quantity as stock_qty
        FROM sale_items si
        LEFT JOIN menu m ON si.menu_id = m.id
        LEFT JOIN products p ON m.product_id = p.id
        LEFT JOIN inventory i ON m.product_id = i.product_id
        WHERE si.sale_id = ?
    """, (sale_id,)).fetchall()


def _fb_fetch_recipes(cursor, menu_id):
    """Định mức NVL của món — dùng cho kiểm tra/trừ kho lúc checkout."""
    return cursor.execute("""
        SELECT mr.product_id, mr.quantity AS recipe_qty, p.name AS material_name, p.unit,
               COALESCE(i.quantity, 0) AS stock_qty, COALESCE(i.avg_cost, 0) AS avg_cost
        FROM menu_recipes mr
        JOIN products p ON mr.product_id = p.id
        LEFT JOIN inventory i ON mr.product_id = i.product_id
        WHERE mr.menu_id = ?
    """, (menu_id,)).fetchall()


def _fb_validate_stock(cursor, items):
    for item in items:
        menu_id = item['menu_id']
        order_qty = float(item['quantity'] or 0)
        if order_qty <= 0:
            continue
        product_type = item['product_type']
        raw_use_unit = item['db_use_sale_unit']
        use_sale_unit = 1 if raw_use_unit in [1, '1', True, 'true'] else 0

        if product_type == 'ready_made':
            if not item['origin_product_id']:
                continue
            ratio = float(item['unit_ratio'] or 1) or 1.0
            required_stock_qty = order_qty * ratio if use_sale_unit == 1 else order_qty
            current_stock = float(item['stock_qty'] or 0)
            stock_unit = item['product_stock_unit'] or item['sale_item_unit'] or 'Cái'
            if current_stock < required_stock_qty:
                return False, f"Kho không đủ hàng! '{item['item_name']}' cần {required_stock_qty} {stock_unit} nhưng hiện tại chỉ còn {current_stock} {stock_unit}."

        # Món có định mức NVL → kiểm tra tồn lúc checkout (ready_made đã xử lý ở trên)
        if menu_id:
            recipes = _fb_fetch_recipes(cursor, menu_id)
            for ingredient in recipes:
                required_qty = float(ingredient['recipe_qty'] or 0) * order_qty
                current_stock = float(ingredient['stock_qty'] or 0)
                if current_stock < required_qty:
                    return False, (
                        f"Không đủ nguyên liệu chế biến món '{item['item_name']}'! "
                        f"Nguyên liệu '{ingredient['material_name']}' cần {required_qty} {ingredient['unit']} "
                        f"nhưng kho chỉ còn {current_stock} {ingredient['unit']}."
                    )
        # Món chế biến không có định mức → không kiểm tra tại checkout (chốt cuối ngày qua draft_inventory)
    return True, None


def _ensure_draft_inventory_table(cursor):
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS draft_inventory (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            quantity REAL NOT NULL,
            note TEXT,
            is_processed INTEGER DEFAULT 0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)


def _fb_import_upsert_product(c, item):
    """Tạo/cập nhật sản phẩm nhập kho F&B; sinh mã Mxxx/NVLxxx + barcode."""
    pid = item.get('product_id')
    item_name = (item.get('name') or item.get('invoice_name') or '').strip()
    current_product_type = (
        item.get('type') or item.get('invoice_product_type') or 'ready_made'
    )

    fe_base_sale_price = Decimal(str(item.get('base_sale_price') or 0))
    fe_sale_price = Decimal(str(item.get('sale_price') or 0))
    fe_wholesale_unit = str(item.get('wholesale_unit') or '').strip()
    fe_ratio = Decimal(str(item.get('ratio') if item.get('ratio') is not None else 1))
    retail_unit = str(item.get('base_unit') or item.get('unit') or 'Cái').strip()

    if not pid:
        if not item_name:
            return None
        c.execute("""
            INSERT INTO products (
                name, unit, unit1, unit_ratio, base_price, price,
                product_type, business_line
            ) VALUES (?, ?, ?, ?, ?, ?, ?, 'F&B')
        """, (
            item_name,
            retail_unit,
            fe_wholesale_unit or None,
            float(fe_ratio),
            float(fe_base_sale_price),
            float(fe_sale_price),
            current_product_type,
        ))
        pid = c.lastrowid
        c.execute(
            "INSERT OR IGNORE INTO inventory (product_id, quantity, avg_cost) VALUES (?, 0, 0)",
            (pid,),
        )

    c.execute("""
        SELECT name, unit, unit1, unit_ratio, price as sale_price, base_price, product_type 
        FROM products WHERE id = ?
    """, (pid,))
    p_info = c.fetchone()
    if not p_info:
        return None

    if not fe_wholesale_unit:
        fe_wholesale_unit = str(p_info['unit1'] or '').strip()
    if fe_base_sale_price == 0:
        fe_base_sale_price = Decimal(str(p_info['base_price'] or 0))
    if fe_sale_price == 0:
        fe_sale_price = Decimal(str(p_info['sale_price'] or 0))
    if fe_ratio == 0:
        fe_ratio = Decimal(str(p_info['unit_ratio'] or 1))

    c.execute("""
        UPDATE products 
        SET name = COALESCE(NULLIF(?, ''), name),
            unit = ?, unit1 = ?, unit_ratio = ?, base_price = ?, price = ?,
            product_type = ?, business_line = 'F&B', updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
    """, (
        item_name, retail_unit, fe_wholesale_unit or None,
        float(fe_ratio), float(fe_base_sale_price), float(fe_sale_price),
        current_product_type, pid,
    ))

    if current_product_type == 'ready_made':
        has_wholesale = bool(fe_wholesale_unit)
        item_code, barcode, barcode1 = assign_ready_made_product_codes(
            c, pid, has_wholesale
        )

        c.execute("SELECT item_code FROM menu WHERE product_id = ?", (pid,))
        menu_entry = c.fetchone()

        if menu_entry:
            c.execute("""
                UPDATE menu SET name = ?, unit = ?, unit1 = ?, base_price = ?, price = ?,
                is_active = 1, product_type = 'ready_made', item_code = ?
                WHERE product_id = ?
            """, (
                item_name or p_info['name'], retail_unit,
                fe_wholesale_unit or None, float(fe_base_sale_price),
                float(fe_sale_price), item_code, pid,
            ))
        else:
            c.execute("""
                INSERT INTO menu (
                    item_code, name, unit, unit1, base_price, price,
                    is_active, product_type, product_id
                ) VALUES (?, ?, ?, ?, ?, ?, 1, 'ready_made', ?)
            """, (
                item_code, item_name or p_info['name'], retail_unit,
                fe_wholesale_unit or None, float(fe_base_sale_price),
                float(fe_sale_price), pid,
            ))

    elif current_product_type == 'raw_materials':
        has_wholesale = bool(fe_wholesale_unit)
        assign_raw_material_product_codes(c, pid, has_wholesale)

    return pid


def _fb_finalize_checkout(cursor, sale_id, table_id, customer_name, payment_method, sale_date, is_einvoice, items):
    from Services.stock_move_write import insert_stock_move, resolve_posting_warehouse_code

    sale_no = f"ĐH{str(sale_id).zfill(6)}"
    px_items = []
    wh = resolve_posting_warehouse_code(cursor.connection)

    for item in items:
        order_qty = float(item['quantity'] or 0)
        product_type = item['product_type']
        price = float(item['price'] or 0)
        line_total = float(item['line_total'] or 0)
        raw_use_unit = item['db_use_sale_unit']
        use_sale_unit = 1 if raw_use_unit in [1, '1', True, 'true'] else 0

        if product_type == 'ready_made' and item['origin_product_id']:
            origin_pid = item['origin_product_id']
            avg_cost = float(item['avg_cost'] or 0)
            stock_unit = item['product_stock_unit'] or item['sale_item_unit'] or 'Cái'
            ratio = float(item['unit_ratio'] or 1) or 1.0
            if use_sale_unit == 1:
                deduct_qty = order_qty * ratio
                display_price = price / ratio
            else:
                deduct_qty = order_qty
                display_price = price

            insert_stock_move(cursor, {
                'product_id': origin_pid,
                'date': sale_date,
                'type': 'SALE',
                'ref_id': sale_id,
                'quantity': -deduct_qty,
                'cost_price': avg_cost,
                'ref_document': sale_no,
                'ref_type': 'export',
                'type1': 'Bán',
                'note': f"Bán tại bàn {table_id}",
                'warehouse_code': wh,
            })
            sync_inventory_quantity_from_moves(cursor, origin_pid)

            px_items.append({
                "product_id": origin_pid,
                "product_name": item['item_name'],
                "unit": stock_unit,
                "quantity": deduct_qty,
                "price": display_price,
                "amount": line_total
            })

        # Món có định mức NVL → trừ kho ngay khi checkout
        if item['menu_id']:
            recipes = _fb_fetch_recipes(cursor, item['menu_id'])
            for ingredient in recipes:
                mat_id = ingredient['product_id']
                total_mat_qty = float(ingredient['recipe_qty'] or 0) * order_qty
                mat_cost = float(ingredient['avg_cost'] or 0)
                insert_stock_move(cursor, {
                    'product_id': mat_id,
                    'date': sale_date,
                    'type': 'SALE_RECIPE',
                    'ref_id': sale_id,
                    'quantity': -total_mat_qty,
                    'cost_price': mat_cost,
                    'ref_document': sale_no,
                    'ref_type': 'export',
                    'type1': 'Xuất nguyên liệu',
                    'note': f"Hao phí chế biến món '{item['item_name']}' tại bàn {table_id}",
                    'warehouse_code': wh,
                })
                sync_inventory_quantity_from_moves(cursor, mat_id)

                px_items.append({
                    "product_id": mat_id,
                    "product_name": ingredient['material_name'],
                    "unit": ingredient['unit'],
                    "quantity": total_mat_qty,
                    "price": 0,
                    "amount": 0
                })
        # Món processed không có định mức: không trừ kho tại checkout — kiểm kê cuối ngày (draft_inventory)

    final_total = sum(float(item['line_total'] or 0) for item in items)

    if px_items:
        last_px = cursor.execute(
            "SELECT voucher_no FROM phieu_xuat_kho WHERE voucher_no LIKE 'PX%' ORDER BY id DESC LIMIT 1").fetchone()
        px_num = (int(last_px['voucher_no'][2:]) + 1) if last_px else 1
        px_vno = f"PX{px_num:06d}"
        cursor.execute("""
            INSERT INTO phieu_xuat_kho (voucher_no, date, customer_name, items_json, total_amount, sale_id) 
            VALUES (?, ?, ?, ?, ?, ?)
        """, (px_vno, sale_date, customer_name, json.dumps(px_items, ensure_ascii=False), final_total, sale_id))

    user_note = ''
    sale_row = cursor.execute("SELECT note FROM sale WHERE id = ?", (sale_id,)).fetchone()
    if sale_row and sale_row['note']:
        try:
            meta = json.loads(sale_row['note'])
            user_note = meta.get('user_note', '')
        except (json.JSONDecodeError, TypeError):
            user_note = sale_row['note']

    cursor.execute("""
        UPDATE sale SET status = 'completed', total_amount = ?, payment_method = ?, customer_name = ?, 
        business_line = 'fb_service', date = ?, sale_no = ?, note = ? WHERE id = ?
    """, (final_total, payment_method, customer_name, sale_date, sale_no, user_note, sale_id))

    if payment_method in ["111", "112"]:
        last_pt = cursor.execute(
            "SELECT voucher_no FROM Phieu_thu WHERE voucher_no LIKE 'PT%' ORDER BY id DESC LIMIT 1").fetchone()
        pt_num = (int(last_pt['voucher_no'][2:]) + 1) if last_pt else 1
        pt_vno = f"PT{pt_num:06d}"
        cursor.execute("""
            INSERT INTO Phieu_thu (voucher_no, payer_name, amount, debit_account, credit_account, reason, reference_document, sale_id, date)
            VALUES (?, ?, ?, ?, '511', ?, ?, ?, ?)
        """, (pt_vno, customer_name, final_total, payment_method, f"Thanh toán F&B {sale_no}", sale_no, sale_id, sale_date))

    elif payment_method == "131":
        cursor.execute("""
            INSERT INTO cong_no (customer_name, debit_account, credit_account, date_of_debt, unpaid_amount, sale_id, sale_no)
            VALUES (?, '131', '511', ?, ?, ?, ?)
        """, (customer_name, sale_date, final_total, sale_id, sale_no))

    cursor.execute(
        "UPDATE tables SET current_sale_id = NULL, status = 'Available' WHERE id = ?", (table_id,))

    try:
        from Services.accounting_queue import ensure_sale_accounting_posted
        ensure_sale_accounting_posted(
            cursor.connection,
            sale_id,
            accounting_regime=get_current_tenant_profile().get('accounting_regime'),
            features=get_current_tenant_profile().get('features'),
            created_by=session.get('user_name') or (session.get('user') or {}).get('username'),
            sync_now=True,
        )
    except Exception as exc:
        logger.warning('ensure accounting F&B sale %s: %s', sale_id, exc)
    return final_total, sale_no


def complete_fb_bank_payment(sale_id):
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    def _run():
        begin_immediate(conn, label='fb_bank_payment')
        sale = cursor.execute("SELECT * FROM sale WHERE id = ?", (sale_id,)).fetchone()
        if not sale:
            return {"success": False, "error": "Không tìm thấy hóa đơn"}
        if sale['status'] == 'completed':
            sqlite_commit(conn, label='fb_write')
            return {"success": True, "already_completed": True}

        meta = {}
        try:
            meta = json.loads(sale['note'] or '{}')
        except (json.JSONDecodeError, TypeError):
            conn.rollback()
            return {"success": False, "error": "Thiếu metadata bàn cho đơn F&B pending"}

        table_id = meta.get('table_id')
        is_einvoice = meta.get('is_einvoice', False)
        if not table_id:
            conn.rollback()
            return {"success": False, "error": "Không xác định được bàn thanh toán"}

        table = cursor.execute(
            "SELECT current_sale_id FROM tables WHERE id = ?", (table_id,)).fetchone()
        if not table or table['current_sale_id'] != sale_id:
            conn.rollback()
            return {"success": False, "error": "Bàn không còn giữ đơn hàng này"}

        items = _fb_load_sale_items(cursor, sale_id)
        if not items:
            conn.rollback()
            return {"success": False, "error": "Đơn hàng rỗng"}

        ok, err = _fb_validate_stock(cursor, items)
        if not ok:
            conn.rollback()
            return {"success": False, "error": err}

        sale_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        customer_name = sale['customer_name'] or DEFAULT_RETAIL_BUYER_NAME
        payment_method = sale['payment_method'] or '112'

        _fb_finalize_checkout(
            cursor, sale_id, table_id, customer_name, payment_method, sale_date, is_einvoice, items
        )
        sqlite_commit(conn, label='fb_write')
        return {"success": True, "sale_id": sale_id}

    try:
        return sqlite_write_retry(_run, label='fb_bank_payment')
    except sqlite3.OperationalError as e:
        rollback_quietly(conn)
        if _is_locked_error(e):
            return {"success": False, "error": locked_user_message()}
        raise
    except Exception as e:
        rollback_quietly(conn)
        return {"success": False, "error": str(e)}
    finally:
        conn.close()


def register_fb_routes(app):
    """Đăng ký route F&B (giữ nguyên URL/endpoint)."""
    from helpers import allowed_file, thuần_thục_tên_file
    from app import UPLOAD_FOLDER, login_required

    @app.route('/F&B_service')
    @login_required
    def F_and_B_service():
        return render_template('F&B_service.html')

    @app.route('/api/fb/tables', methods=['GET'])
    @login_required
    def get_fb_tables():
        from Services.fb_schema import ensure_fb_schema

        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        try:
            ensure_fb_schema(conn)
            query = """
            SELECT t.id, t.name, a.name as area_name, t.current_sale_id,
            CASE WHEN t.current_sale_id IS NOT NULL THEN 'Busy' ELSE 'Available' END as status
            FROM tables t
            LEFT JOIN areas a ON t.area_id = a.id
            ORDER BY a.name, t.name
            """
            cursor.execute(query)
            tables = [dict(row) for row in cursor.fetchall()]
            return jsonify(tables)
        except Exception as e:
            logger.error('get_fb_tables: %s', e)
            return jsonify([])
        finally:
            conn.close()

    @app.route('/api/menu', methods=['GET'])
    @login_required
    def get_menu():
        is_active_only = request.args.get('active', default=1, type=int)
        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        try:
            query = """
                SELECT 
                    m.*, 
                    i.quantity AS stock_qty,
                    p.product_code AS ref_product_code
                FROM menu m
                LEFT JOIN products p ON m.product_id = p.id
                LEFT JOIN inventory i ON m.product_id = i.product_id
                WHERE 1=1
            """
            if is_active_only == 1:
                query += " AND m.is_active = 1"
            query += " ORDER BY m.category ASC, m.name ASC"
            cursor.execute(query)
            rows = cursor.fetchall()
            menu_items = []
            for row in rows:
                item = dict(row)
                item['base_price'] = item.get('base_price') or 0
                item['price'] = item.get('price') or 0
                if item.get('product_type') != 'ready_made':
                    item['stock_qty'] = None
                else:
                    item['stock_qty'] = item.get('stock_qty') or 0
                if not item.get('image_path'):
                    item['image_path'] = '/static/img/default-food.png'
                menu_items.append(item)
            print(f"--- Menu Debug: Đã tải {len(menu_items)} món từ bảng Menu ---")
            return jsonify(menu_items)
        except Exception as e:
            print(f"--- Lỗi SQL api/menu: {str(e)} ---")
            return jsonify({"message": f"Lỗi truy vấn thực đơn: {str(e)}"}), 500
        finally:
            conn.close()

    @app.route('/api/fb/get-next-sale-no', methods=['GET'])
    @login_required
    def get_next_sale_no():
        db = get_db_connection()
        cursor = db.cursor()
        try:
            cursor.execute("SELECT MAX(id) FROM sale")
            last_id = cursor.fetchone()[0] or 0
            next_no = f"HĐ{str(last_id + 1).zfill(6)}"
            return jsonify({"success": True, "sale_no": next_no})
        finally:
            db.close()

    @app.route('/api/fb/active-orders', methods=['GET'])
    @login_required
    def api_active_orders():
        from Services.fb_schema import ensure_fb_schema

        db = get_db_connection()
        db.row_factory = sqlite3.Row
        cursor = db.cursor()
        try:
            ensure_fb_schema(db)
            sale_cols = {
                (r[1] or '').lower()
                for r in db.execute('PRAGMA table_info(sale)')
            }
            sale_created = 's.created_at' if 'created_at' in sale_cols else 'NULL'
            if 'date' in sale_cols:
                sale_created = f'COALESCE({sale_created}, s.date)'
            cursor.execute(f"""
                SELECT 
                    s.id AS sale_id,
                    s.sale_no,
                    s.table_id,
                    t.name AS table_name,
                    {sale_created} AS created_at,
                    si.menu_id,
                    si.UseSaleUnit,
                    si.unit,
                    COALESCE(NULLIF(si.product_name, ''), si.item_name) AS product_name,
                    si.quantity,
                    COALESCE(si.quantity_served, 0) AS quantity_served,
                    si.price AS unit_price,
                    si.line_total,
                    COALESCE(si.created_at, {sale_created}) AS item_created_at,
                    si.served_at,
                    m.item_code,
                    m.unit1
                FROM sale s
                LEFT JOIN tables t ON t.id = s.table_id
                LEFT JOIN sale_items si ON si.sale_id = s.id
                LEFT JOIN menu m ON m.id = si.menu_id
                WHERE LOWER(COALESCE(s.status, '')) = 'draft'
                  AND s.table_id IS NOT NULL
                ORDER BY t.name, si.created_at ASC
            """)
            rows = cursor.fetchall()
            orders = {}
            now_time = datetime.now()
            for row in rows:
                sale_id = row['sale_id']
                if sale_id not in orders:
                    orders[sale_id] = {
                        'sale_id': sale_id,
                        'sale_no': row['sale_no'],
                        'table_id': row['table_id'],
                        'table_name': row['table_name'] or f"Bàn {row['table_id']}",
                        'created_at': row['created_at'],
                        'items': []
                    }
                if row['menu_id']:
                    minutes_waiting = 0
                    if row['item_created_at']:
                        try:
                            time_str = str(row['item_created_at'])
                            if 'Z' in time_str or '+' in time_str:
                                created = datetime.fromisoformat(
                                    time_str.replace('Z', '+00:00')).replace(tzinfo=None)
                            else:
                                created = datetime.strptime(time_str, '%Y-%m-%d %H:%M:%S')
                            minutes_waiting = max(
                                0, int((now_time - created).total_seconds() / 60))
                        except Exception:
                            pass
                    try:
                        use_sale_unit = int(
                            row['UseSaleUnit'] if row['UseSaleUnit'] is not None else 0)
                    except (ValueError, TypeError):
                        use_sale_unit = 0
                    display_unit = str(row['unit']).strip() if row['unit'] is not None else ''
                    orders[sale_id]['items'].append({
                        'menu_id': row['menu_id'],
                        'UseSaleUnit': use_sale_unit,
                        'unit': display_unit,
                        'product_name': row['product_name'],
                        'quantity': float(row['quantity'] or 0),
                        'qty_served': float(row['quantity_served'] or 0),
                        'unit_price': float(row['unit_price'] or 0),
                        'line_total': float(row['line_total'] or 0),
                        'item_code': row['item_code'],
                        'unit1': row['unit1'],
                        'created_at': row['item_created_at'],
                        'served_at': row['served_at'],
                        'minutes_waiting': minutes_waiting
                    })
            return jsonify({"success": True, "orders": list(orders.values())})
        except sqlite3.Error as e:
            logger.exception("ERROR active-orders: %s", e)
            return jsonify({"success": True, "orders": [], "message": str(e)})
        except Exception as e:
            logger.exception("ERROR active-orders: %s", e)
            return jsonify({"success": True, "orders": [], "message": str(e)})
        finally:
            db.close()

    @app.route('/api/fb/mark-served/<int:sale_id>/<int:menu_id>', methods=['POST'])
    @login_required
    def api_mark_item_served(sale_id, menu_id):
        from Services.fb_schema import ensure_fb_schema

        db = get_db_connection()
        db.row_factory = sqlite3.Row
        cursor = db.cursor()

        try:
            ensure_fb_schema(db)
            payload = request.get_json(silent=True) or {}
            use_sale_unit = _fb_normalize_use_sale_unit(payload.get('use_sale_unit', 0))
            use_expr = _fb_use_sale_unit_expr(cursor, 'si')
            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            cursor.execute(f"""
            UPDATE sale_items 
            SET quantity_served = quantity,
                served_at = ? 
            WHERE sale_id = ? AND menu_id = ? AND ({use_expr}) = ?
            """, (now, sale_id, menu_id, use_sale_unit))

            if cursor.rowcount == 0:
                return jsonify({"success": False, "message": "Không tìm thấy món trong đơn hàng"}), 404

            sqlite_commit(db, label='fb_write')

            return jsonify({
                "success": True,
                "message": "Đã xác nhận món đã phục vụ hoàn thành"
            })

        except Exception as e:
            db.rollback()
            print("ERROR mark-item-served:", str(e))
            return jsonify({"success": False, "message": str(e)}), 500
        finally:
            db.close()

    @app.route('/api/fb/get-sale/<int:table_id>')
    @login_required
    def get_sale_by_table(table_id):
        from Services.fb_schema import ensure_fb_schema

        db = get_db_connection()
        db.row_factory = sqlite3.Row
        cursor = db.cursor()

        try:
            ensure_fb_schema(db)
            use_expr = use_sale_unit_expr(cursor, 'si')
            pk_expr = sale_item_pk_expr(cursor, 'si')
            table_info = cursor.execute(
                "SELECT current_sale_id FROM tables WHERE id = ?",
                (table_id,)
            ).fetchone()

            if not table_info or not table_info['current_sale_id']:
                return jsonify({
                    "success": True,
                    "items": [],
                    "total_amount": 0,
                    "message": "Bàn trống"
                })

            sale_id = table_info['current_sale_id']

            sale_data = cursor.execute(
                "SELECT total_amount, status FROM sale WHERE id = ?",
                (sale_id,)
            ).fetchone()

            items_query = f"""
            SELECT 
                {pk_expr} AS sale_item_id,
                si.menu_id,
                {use_expr} AS UseSaleUnit,
                si.unit,
                m.item_code,
                COALESCE(NULLIF(m.name, ''), si.product_name, si.item_name) AS product_name,
                m.image_path,
                si.quantity,
                si.price AS unit_price,
                si.line_total
            FROM sale_items si
            LEFT JOIN menu m ON si.menu_id = m.id
            WHERE si.sale_id = ?
            """
            rows = cursor.execute(items_query, (sale_id,)).fetchall()

            items = []
            for row in rows:
                item = dict(row)
                try:
                    item['UseSaleUnit'] = int(item.get('UseSaleUnit', 0))
                except (ValueError, TypeError):
                    item['UseSaleUnit'] = 0
                if item.get('unit') is not None:
                    item['unit'] = str(item['unit']).strip()
                else:
                    item['unit'] = ''
                if not item.get('image_path'):
                    item['image_path'] = '/static/img/default-food.png'
                items.append(item)

            return jsonify({
                "success": True,
                "sale_id": sale_id,
                "status": sale_data['status'] if sale_data else 'Draft',
                "total_amount": sale_data['total_amount'] if sale_data else 0,
                "items": items
            })

        except Exception as e:
            print(f"--- Lỗi nghiêm trọng tại API get-sale: {str(e)} ---")
            return jsonify({"success": False, "message": f"Lỗi hệ thống: {str(e)}"}), 500
        finally:
            db.close()

    @app.route('/api/fb/checkout', methods=['POST'])
    @login_required
    def checkout_table():
        data = request.json or {}
        table_id = data.get('table_id')
        client_uuid = (data.get('client_uuid') or '').strip()

        customer_name = normalize_retail_buyer_name(data.get('customer_name'))
        payment_method = data.get('payment_method', '111')
        is_einvoice = data.get('is_einvoice', False)

        if not table_id:
            return jsonify({"success": False, "message": "Thiếu thông tin bàn cần thanh toán"}), 400

        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        try:
            from Services.fb_schema import ensure_fb_schema
            ensure_fb_schema(conn, commit=False)
            ensure_pos_offline_schema(conn, commit=False)
            if client_uuid:
                existing = find_sale_by_client_uuid(conn, client_uuid)
                if existing and str(existing.get('status') or '').lower() == 'completed':
                    try:
                        from Services.accounting_queue import ensure_sale_accounting_posted
                        profile = get_current_tenant_profile()
                        ensure_sale_accounting_posted(
                            conn,
                            int(existing['id']),
                            accounting_regime=profile.get('accounting_regime'),
                            features=profile.get('features'),
                            created_by=session.get('user_name') or (session.get('user') or {}).get('username'),
                            sync_now=True,
                        )
                    except Exception as acct_exc:
                        logger.warning(
                            'ensure_sale_accounting_posted F&B dedupe sale %s: %s',
                            existing.get('id'), acct_exc,
                        )
                    return jsonify({
                        "success": True,
                        "sale_id": existing['id'],
                        "sale_no": existing.get('sale_no') or f"ĐH{str(existing['id']).zfill(6)}",
                        "deduped": True,
                        "message": "Đơn đã được đồng bộ trước đó",
                    })

            def _checkout():
                begin_immediate(conn, label='fb_checkout')

                table = cursor.execute(
                    "SELECT current_sale_id FROM tables WHERE id = ?", (table_id,)).fetchone()
                if not table or not table['current_sale_id']:
                    conn.rollback()
                    return {"success": False, "message": "Bàn không có đơn hàng"}, 404

                sale_id = table['current_sale_id']
                now_dt = datetime.now()
                sale_date = now_dt.strftime('%Y-%m-%d %H:%M:%S')
                sale_no = f"ĐH{str(sale_id).zfill(6)}"

                row_sale = cursor.execute(
                    "SELECT status FROM sale WHERE id = ?", (sale_id,),
                ).fetchone()
                if row_sale and str(row_sale['status'] or '').lower() == 'completed':
                    sqlite_commit(conn, label='fb_write')
                    return {
                        "success": True,
                        "sale_id": sale_id,
                        "sale_no": sale_no,
                        "deduped": True,
                        "message": "Bàn đã thanh toán",
                    }, 200

                items = _fb_load_sale_items(cursor, sale_id)
                if not items:
                    conn.rollback()
                    return {"success": False, "message": "Đơn hàng rỗng, không thể thanh toán"}, 400

                orphan = [it for it in items if not it['menu_id']]
                if orphan:
                    conn.rollback()
                    return {
                        "success": False,
                        "message": "Đơn có dòng không hợp lệ (thiếu mã món). Vui lòng xóa và gọi lại món.",
                    }, 400

                ok, err = _fb_validate_stock(cursor, items)
                if not ok:
                    conn.rollback()
                    return {"success": False, "message": err}, 400

                final_total = sum(float(item['line_total'] or 0) for item in items)

                if payment_method == '112':
                    note_payload = json.dumps({
                        'table_id': table_id,
                        'is_einvoice': bool(is_einvoice),
                        'user_note': data.get('note', '')
                    }, ensure_ascii=False)
                    cursor.execute("""
                        UPDATE sale SET status = 'pending', total_amount = ?, payment_method = ?, customer_name = ?,
                        business_line = 'fb_service', date = ?, sale_no = ?, note = ? WHERE id = ?
                    """, (final_total, payment_method, customer_name, sale_date, sale_no, note_payload, sale_id))
                    sqlite_commit(conn, label='fb_write')
                    return {
                        "success": True,
                        "sale_id": sale_id,
                        "sale_no": sale_no,
                        "pending_qr": True,
                        "total_amount": final_total,
                        "request_einvoice": is_einvoice,
                        "message": "Chờ khách chuyển khoản qua QR"
                    }, 200

                _fb_finalize_checkout(
                    cursor, sale_id, table_id, customer_name, payment_method, sale_date, is_einvoice, items
                )
                if client_uuid:
                    try:
                        cursor.execute(
                            "UPDATE sale SET client_uuid = ? WHERE id = ?",
                            (client_uuid, sale_id),
                        )
                    except sqlite3.OperationalError:
                        pass

                sqlite_commit(conn, label='fb_write')
                return {
                    "success": True,
                    "sale_id": sale_id,
                    "sale_no": sale_no,
                    "request_einvoice": is_einvoice,
                    "message": "Thanh toán và trừ kho sỉ/lẻ thành công!"
                }, 200

            payload, status = sqlite_write_retry(_checkout, label='fb_checkout')
            return jsonify(payload), status

        except sqlite3.OperationalError as e:
            rollback_quietly(conn)
            if _is_locked_error(e):
                return jsonify({"success": False, "message": locked_user_message()}), 503
            raise
        except Exception as e:
            rollback_quietly(conn)
            logger.error('Lỗi Checkout F&B: %s', e)
            return jsonify({"success": False, "message": f"Lỗi hệ thống: {str(e)}"}), 500
        finally:
            conn.close()

    @app.route('/api/fb/update-status', methods=['POST'])
    @login_required
    def update_product_status():
        data = request.json
        menu_id = data.get('product_id')

        if not menu_id:
            return jsonify({"success": False, "message": "Thiếu ID món ăn"}), 400

        db = get_db_connection()
        db.row_factory = sqlite3.Row
        cursor = db.cursor()

        try:
            item = cursor.execute(
                "SELECT is_active FROM menu WHERE id = ?", (menu_id,)).fetchone()

            if not item:
                return jsonify({"success": False, "message": "Món ăn không tồn tại trong thực đơn"}), 404

            new_status = 0 if item['is_active'] == 1 else 1

            cursor.execute(
                "UPDATE menu SET is_active = ? WHERE id = ?", (new_status, menu_id))
            sqlite_commit(db, label='fb_write')

            status_text = "Đang kinh doanh" if new_status == 1 else "Tạm ngừng bán"

            print(f"--- Menu Update: ID {menu_id} chuyển sang {status_text} ---")

            return jsonify({
                "success": True,
                "new_status": new_status,
                "message": status_text
            })

        except Exception as e:
            db.rollback()
            return jsonify({"success": False, "message": f"Lỗi hệ thống: {str(e)}"}), 500
        finally:
            db.close()

    @app.route('/api/fb/add-item', methods=['POST'])
    @login_required
    def api_add_item_to_table():
        from Services.fb_schema import ensure_fb_schema

        data = request.json or {}
        table_id = data.get('table_id')
        menu_id = data.get('menu_id')
        use_sale_unit = normalize_use_sale_unit(data.get('use_sale_unit', 0))

        try:
            quantity = float(data.get('quantity', 1))
        except (ValueError, TypeError):
            quantity = 1.0

        if not table_id or menu_id is None or menu_id == '':
            return jsonify({"success": False, "message": "Thiếu bàn hoặc mã món"}), 400

        db = get_db_connection()
        db.row_factory = sqlite3.Row
        cursor = db.cursor()

        try:
            ensure_fb_schema(db, commit=False)

            def _add():
                begin_immediate(db, label='fb_add_item')

                table = cursor.execute(
                    "SELECT current_sale_id FROM tables WHERE id = ?", (table_id,)).fetchone()
                if not table:
                    db.rollback()
                    return {"success": False, "message": "Bàn không tồn tại"}, 404

                sale_id = table['current_sale_id']

                if not sale_id:
                    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    cursor.execute(
                        "INSERT INTO sale (status, table_id, created_at, total_amount) VALUES ('Draft', ?, ?, 0)",
                        (table_id, now_str)
                    )
                    sale_id = cursor.lastrowid
                    cursor.execute(
                        "UPDATE tables SET current_sale_id = ?, status = 'Busy' WHERE id = ?",
                        (sale_id, table_id),
                    )

                item_menu = cursor.execute(
                    "SELECT name, base_price, unit, price, unit1, product_id, product_type FROM menu WHERE id = ?",
                    (menu_id,)
                ).fetchone()

                if not item_menu:
                    db.rollback()
                    return {"success": False, "message": "Món ăn không tồn tại trong thực đơn"}, 404

                item_name = item_menu['name']
                product_id = item_menu['product_id']

                if use_sale_unit == 1:
                    price_to_sell = item_menu['price'] if item_menu['price'] is not None else 0
                    unit_to_save = item_menu['unit1']
                else:
                    price_to_sell = item_menu['base_price'] if item_menu['base_price'] is not None else 0
                    unit_to_save = item_menu['unit']

                where_unit = use_sale_unit_where_clause(cursor, 'si')
                existing_item = cursor.execute(f"""
                    SELECT quantity FROM sale_items si
                    WHERE si.sale_id = ? AND si.menu_id = ? AND {where_unit}
                """, (sale_id, menu_id, use_sale_unit)).fetchone()

                if existing_item:
                    new_qty = existing_item['quantity'] + quantity
                    new_line_total = new_qty * price_to_sell
                    unit_cols = use_sale_unit_insert_columns(cursor)
                    set_parts = ['quantity = ?', 'line_total = ?']
                    params = [new_qty, new_line_total]
                    for col in unit_cols:
                        set_parts.append(f'{col} = ?')
                        params.append(use_sale_unit)
                    params.extend([sale_id, menu_id, use_sale_unit])
                    where_plain = use_sale_unit_where_clause(cursor, alias=None)
                    cursor.execute(f"""
                        UPDATE sale_items SET {', '.join(set_parts)}
                        WHERE sale_id = ? AND menu_id = ? AND {where_plain}
                    """, params)
                else:
                    line_total = quantity * price_to_sell
                    hkd_sector = resolve_item_hkd_sector(
                        business_line='fb_service',
                        menu_product_type=item_menu['product_type'] if 'product_type' in item_menu.keys() else None,
                    )
                    base_cols = [
                        'sale_id', 'menu_id', 'product_id', 'product_name', 'quantity',
                        'unit', 'price', 'line_total', 'created_at',
                    ]
                    base_vals = [
                        sale_id, menu_id, product_id, item_name, quantity,
                        unit_to_save, price_to_sell, line_total,
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    ]
                    unit_cols = use_sale_unit_insert_columns(cursor)
                    base_cols.extend(unit_cols)
                    base_vals.extend(expand_use_sale_unit_values(cursor, use_sale_unit))
                    insert_sale_item_with_sector(
                        cursor, base_cols, base_vals, hkd_sector_code=hkd_sector,
                    )

                cursor.execute("""
                    UPDATE sale 
                    SET total_amount = (SELECT SUM(line_total) FROM sale_items WHERE sale_id = ?) 
                    WHERE id = ?
                """, (sale_id, sale_id))

                sqlite_commit(db, label='fb_write')
                return {
                    "success": True,
                    "sale_id": sale_id,
                    "message": f"Đã thêm {item_name} ({unit_to_save})",
                }, 200

            payload, status = sqlite_write_retry(_add, label='fb_add_item')
            return jsonify(payload), status

        except sqlite3.OperationalError as e:
            rollback_quietly(db)
            if _is_locked_error(e):
                return jsonify({"success": False, "message": locked_user_message()}), 503
            raise
        except Exception as e:
            rollback_quietly(db)
            return jsonify({"success": False, "message": f"Lỗi hệ thống: {str(e)}"}), 500
        finally:
            db.close()

    @app.route('/api/fb/update-item-quantity', methods=['POST'])
    @login_required
    def update_item_quantity():
        from Services.fb_schema import ensure_fb_schema

        data = request.json or {}
        table_id = data.get('table_id')
        menu_id = data.get('menu_id')
        sale_item_id = data.get('sale_item_id')

        try:
            change = float(data.get('change', 0))
        except (ValueError, TypeError):
            change = 0.0

        use_sale_unit = normalize_use_sale_unit(data.get('use_sale_unit', 0))

        db = get_db_connection()
        db.row_factory = sqlite3.Row
        cursor = db.cursor()
        pk_col = 'rowid'

        try:
            ensure_fb_schema(db, commit=False)
            pk_col = sale_item_pk_column(cursor)

            def _update_qty():
                begin_immediate(db, label='fb_update_item_qty')

                table = cursor.execute(
                    "SELECT current_sale_id FROM tables WHERE id = ?", (table_id,)).fetchone()
                sale_id = table['current_sale_id'] if table else None

                if not sale_id:
                    db.rollback()
                    return {"success": False, "message": "Không tìm thấy đơn hàng cho bàn này"}, 404

                use_expr = use_sale_unit_expr(cursor, 'si')

                if sale_item_id:
                    current_item = cursor.execute(f"""
                        SELECT {pk_col} AS id, menu_id, quantity,
                               COALESCE(quantity_served, 0) AS quantity_served
                        FROM sale_items
                        WHERE {pk_col} = ? AND sale_id = ?
                    """, (sale_item_id, sale_id)).fetchone()
                else:
                    if menu_id is None or menu_id == '':
                        db.rollback()
                        return {
                            "success": False,
                            "message": "Thiếu mã món — tải lại trang và thử lại",
                        }, 400
                    current_item = cursor.execute(f"""
                        SELECT {pk_col} AS id, menu_id, quantity,
                               COALESCE(quantity_served, 0) AS quantity_served
                        FROM sale_items si
                        WHERE si.sale_id = ?
                          AND si.menu_id = ?
                          AND ({use_expr}) = ?
                    """, (sale_id, menu_id, use_sale_unit)).fetchone()

                if not current_item:
                    db.rollback()
                    return {"success": False, "message": "Món ăn không tồn tại trong giỏ hàng"}, 404

                row_id = current_item['id']

                current_qty = float(current_item['quantity'] if current_item['quantity'] is not None else 0)
                qty_served = float(current_item['quantity_served'] if current_item['quantity_served'] is not None else 0)

                if change == 0:
                    if qty_served > 0:
                        db.rollback()
                        return {
                            "success": False,
                            "message": (
                                f"Món này đã phục vụ {int(qty_served)} phần, không thể xóa! "
                                "Vui lòng chỉ giảm số lượng món chưa lên bàn."
                            ),
                        }, 400
                else:
                    new_qty = current_qty + change
                    if new_qty < qty_served:
                        db.rollback()
                        return {
                            "success": False,
                            "message": (
                                f"Không thể giảm! Món này đã phục vụ {int(qty_served)} phần lên bàn cho khách."
                            ),
                        }, 400

                if change == 0:
                    cursor.execute(f"DELETE FROM sale_items WHERE {pk_col} = ?", (row_id,))
                else:
                    cursor.execute(f"""
                        UPDATE sale_items 
                        SET quantity = quantity + ?, 
                            line_total = (quantity + ?) * price 
                        WHERE {pk_col} = ?
                    """, (change, change, row_id))

                cursor.execute(f"""
                    DELETE FROM sale_items 
                    WHERE {pk_col} = ? AND quantity <= 0 AND COALESCE(quantity_served, 0) <= 0
                """, (row_id,))

                check_items = cursor.execute(
                    "SELECT COUNT(*) as cnt FROM sale_items WHERE sale_id = ?", (sale_id,)).fetchone()

                if check_items['cnt'] == 0:
                    cursor.execute(
                        "UPDATE tables SET current_sale_id = NULL, status = 'Available' WHERE id = ?",
                        (table_id,),
                    )
                    cursor.execute("DELETE FROM sale WHERE id = ?", (sale_id,))
                    max_id_row = cursor.execute("SELECT MAX(id) FROM sale").fetchone()
                    new_max_id = max_id_row[0] if max_id_row and max_id_row[0] is not None else 0
                    cursor.execute(
                        "UPDATE sqlite_sequence SET seq = ? WHERE name = 'sale'", (new_max_id,))
                    message = "Đơn hàng đã được dọn dẹp, bàn đã trống"
                    is_empty = True
                else:
                    cursor.execute("""
                        UPDATE sale 
                        SET total_amount = (SELECT SUM(line_total) FROM sale_items WHERE sale_id = ?) 
                        WHERE id = ?
                    """, (sale_id, sale_id))
                    message = "Cập nhật số lượng thành công"
                    is_empty = False

                sqlite_commit(db, label='fb_write')
                return {"success": True, "message": message, "is_empty": is_empty}, 200

            payload, status = sqlite_write_retry(_update_qty, label='fb_update_item_qty')
            return jsonify(payload), status

        except sqlite3.OperationalError as e:
            rollback_quietly(db)
            if _is_locked_error(e):
                return jsonify({"success": False, "message": locked_user_message()}), 503
            raise
        except Exception as e:
            rollback_quietly(db)
            logger.error('Lỗi update-item-quantity: %s', e)
            return jsonify({"success": False, "message": f"Lỗi hệ thống: {str(e)}"}), 500
        finally:
            db.close()

    # ==============================================================================
    # 1. API: LƯU MÓN ĂN THỦ CÔNG (CÓ XỬ LÝ UPLOAD HÌNH ẢNH)
    # ==============================================================================
    from werkzeug.utils import secure_filename

    @app.route('/api/fb/add-menu-manual', methods=['POST'])
    @login_required
    def add_menu_manual():
        item_code = request.form.get('item_code', '').strip()
        name = request.form.get('name', '').strip()
        unit = request.form.get('unit', '').strip()
        unit1 = request.form.get('unit1', '').strip()
        category = request.form.get('product_category')
        product_type = request.form.get('product_type')

        try:
            base_price = float(request.form.get('base_price', 0))
        except (ValueError, TypeError):
            base_price = 0.0

        try:
            price = float(request.form.get('price', 0))
        except (ValueError, TypeError):
            price = 0.0

        if not name or not unit:
            return jsonify({"success": False, "message": "Vui lòng nhập đầy đủ thông tin bắt buộc: Tên món & Đơn vị tính (*)"}), 400

        image_path = '/static/img/default-food.png'

        file_key = None
        if 'image' in request.files and request.files['image'].filename != '':
            file_key = 'image'
        elif 'm_image' in request.files and request.files['m_image'].filename != '':
            file_key = 'm_image'

        if file_key:
            file = request.files[file_key]
            if allowed_file(file.filename):
                unique_filename = thuần_thục_tên_file(file.filename)
                full_path = os.path.join(UPLOAD_FOLDER, unique_filename)

                if not os.path.exists(UPLOAD_FOLDER):
                    os.makedirs(UPLOAD_FOLDER, mode=0o755, exist_ok=True)

                try:
                    file.save(full_path)
                    print(f"--- [Thành công] Đã lưu thêm tệp tin món mới tại: {full_path} ---")
                    image_path = f"/static/img/{unique_filename}"
                except Exception as file_error:
                    print(f"--- [LỖI GHI ĐĨA VẬT LÝ]: {str(file_error)} ---")
                    return jsonify({"success": False, "message": f"Hệ thống không có quyền ghi file ảnh: {str(file_error)}"}), 500

        conn = get_db_connection()
        cursor = conn.cursor()

        try:
            if not item_code or item_code.upper() == 'AUTO':
                item_code = get_next_menu_code(cursor)
                print(f"--- [MÃ TỰ ĐỘNG SINH]: {item_code} cho món '{name}' ---")
            else:
                exist_check = cursor.execute(
                    "SELECT id FROM menu WHERE item_code = ?", (item_code,)).fetchone()
                if exist_check:
                    return jsonify({"success": False, "message": f"Mã món '{item_code}' đã tồn tại trong hệ thống thực đơn!"}), 400

            cursor.execute("""
            INSERT INTO menu (item_code, name, unit, unit1, base_price, price, category, product_type, image_path, is_active)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
            """, (item_code, name, unit, unit1, base_price, price, category, product_type, image_path))

            sqlite_commit(conn, label='fb_write')
            return jsonify({"success": True, "message": f"Đã thêm món '{name}' thành công với Mã món: {item_code}."}), 200

        except Exception as e:
            conn.rollback()
            print(f"--- Lỗi nghiêm trọng tại add_menu_manual: {str(e)} ---")
            return jsonify({"success": False, "message": f"Lỗi xử lý cơ sở dữ liệu: {str(e)}"}), 500
        finally:
            if conn:
                conn.close()

    @app.route('/api/fb/get-next-code', methods=['GET'])
    @login_required
    def get_next_code_api():
        """API lấy trước mã tiếp theo (Mxxx hoặc NVLxxx) cho preview trên form nhập kho."""
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            product_type = (request.args.get('type') or 'ready_made').strip()
            if product_type == 'raw_materials':
                next_code = get_next_raw_material_code(cursor)
            else:
                next_code = get_next_menu_code(cursor)
            return jsonify({
                "success": True,
                "next_code": next_code,
                "barcode": f"{next_code}01",
                "barcode1": f"{next_code}02",
            }), 200
        except Exception as e:
            print(f"--- Lỗi lấy mã gợi ý: {str(e)} ---")
            fallback = "NVL001" if request.args.get('type') == 'raw_materials' else "M001"
            return jsonify({"success": False, "next_code": fallback}), 500
        finally:
            if conn:
                conn.close()

    @app.route('/api/fb/edit-menu-manual', methods=['POST'])
    @login_required
    def edit_menu_manual():
        item_code = request.form.get('item_code')
        name = request.form.get('name')
        unit = request.form.get('unit')
        unit1 = request.form.get('unit1')
        category = request.form.get('product_category')
        product_type = request.form.get('product_type')

        try:
            base_price = float(request.form.get('base_price', 0))
        except (ValueError, TypeError):
            base_price = 0.0

        try:
            price = float(request.form.get('price', 0))
        except (ValueError, TypeError):
            price = 0.0

        if not item_code:
            return jsonify({"success": False, "message": "Thiếu mã định danh món ăn (item_code)!"}), 400

        if not name or not unit:
            return jsonify({"success": False, "message": "Vui lòng nhập đầy đủ thông tin bắt buộc (*)"}), 400

        conn = get_db_connection()
        conn.row_factory = sqlite3.Row if hasattr(sqlite3, 'Row') else None
        cursor = conn.cursor()

        try:
            current_item = cursor.execute(
                "SELECT image_path, name FROM menu WHERE item_code = ?", (item_code,)).fetchone()
            if not current_item:
                return jsonify({"success": False, "message": f"Không tìm thấy món ăn có mã '{item_code}' để chỉnh sửa!"}), 404

            image_path = current_item['image_path']

            if 'image' in request.files:
                file = request.files['image']
                if file and file.filename != '' and allowed_file(file.filename):
                    unique_filename = thuần_thục_tên_file(file.filename)
                    full_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)

                    if not os.path.exists(app.config['UPLOAD_FOLDER']):
                        os.makedirs(app.config['UPLOAD_FOLDER'], mode=0o755, exist_ok=True)

                    file.save(full_path)
                    print(f"--- [Đã lưu vật lý thành công tại]: {full_path} ---")
                    image_path = f"/static/img/{unique_filename}"

            cursor.execute("""
            UPDATE menu 
            SET name = ?,
                category = ?, 
                unit = ?, 
                unit1 = ?, 
                base_price = ?, 
                price = ?, 
                product_type = ?, 
                image_path = ?
            WHERE item_code = ?
            """, (name, category, unit, unit1, base_price, price, product_type, image_path, item_code))

            sqlite_commit(conn, label='fb_write')
            return jsonify({"success": True, "message": f"Đã cập nhật thông tin món '{name}' thành công."}), 200

        except sqlite3.Error as e:
            conn.rollback()
            print(f"--- Lỗi SQLite: {str(e)} ---")
            return jsonify({"success": False, "message": f"Lỗi cơ sở dữ liệu: {str(e)}"}), 500
        except Exception as e:
            conn.rollback()
            print(f"--- Lỗi hệ thống: {str(e)} ---")
            return jsonify({"success": False, "message": f"Lỗi máy chủ hệ thống: {str(e)}"}), 500
        finally:
            if conn:
                conn.close()

    # ==============================================================================
    # 2. API: LẤY CHI TIẾT ĐỊNH MỨC CỦA MỘT MÓN ĂN (SỬ DỤNG product_id)
    # ==============================================================================
    @app.route('/api/fb/recipe/<string:menu_id>', methods=['GET'])
    @login_required
    def get_menu_recipe(menu_id):
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            recipe = cursor.execute("""
            SELECT r.product_id, r.quantity, p.name as product_name, p.unit
            FROM menu_recipes r
            JOIN products p ON r.product_id = p.id
            WHERE r.menu_id = ?
            """, (menu_id,)).fetchall()

            recipe_list = [dict(row) for row in recipe]
            return jsonify({"success": True, "recipe": recipe_list})
        except Exception as e:
            print(f"--- Lỗi get_menu_recipe: {str(e)} ---")
            return jsonify({"success": False, "message": str(e)}), 500
        finally:
            conn.close()

    # ==============================================================================
    # 3. API: LƯU / THÊM MỚI NGUYÊN VẬT LIỆU VÀO ĐỊNH MỨC (SỬ DỤNG product_id)
    # ==============================================================================

    @app.route('/api/fb/recipe/save', methods=['POST'])
    @login_required
    def save_menu_recipe():
        data = request.json
        menu_id = data.get('menu_id')
        product_id = data.get('product_id')
        quantity = float(data.get('quantity', 0))

        if not menu_id or not product_id or quantity <= 0:
            return jsonify({"success": False, "message": "Dữ liệu cấu hình định mức không hợp lệ"}), 400

        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("""
                INSERT OR REPLACE INTO menu_recipes (menu_id, product_id, quantity)
                VALUES (?, ?, ?)
            """, (menu_id, product_id, quantity))
            sqlite_commit(conn, label='fb_write')
            return jsonify({"success": True, "message": "Đã cập nhật định mức nguyên vật liệu"})
        except Exception as e:
            conn.rollback()
            print(f"--- Lỗi save_menu_recipe: {str(e)} ---")
            return jsonify({"success": False, "message": str(e)}), 500
        finally:
            conn.close()

    # ==============================================================================
    # 4. API: XÓA THÀNH PHẦN KHỎI ĐỊNH MỨC MÓN ĂN (SỬ DỤNG product_id)
    # ==============================================================================

    @app.route('/api/fb/recipe/delete/<string:menu_id>/<int:product_id>', methods=['DELETE'])
    @login_required
    def delete_menu_recipe_item(menu_id, product_id):
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute(
                "DELETE FROM menu_recipes WHERE menu_id = ? AND product_id = ?", (menu_id, product_id))
            sqlite_commit(conn, label='fb_write')
            return jsonify({"success": True, "message": "Đã xóa thành phần định mức thành công"})
        except Exception as e:
            conn.rollback()
            print(f"--- Lỗi delete_menu_recipe_item: {str(e)} ---")
            return jsonify({"success": False, "message": str(e)}), 500
        finally:
            conn.close()

    @app.route('/api/fb/import-menu', methods=['POST'])
    @login_required
    def import_menu():
        if 'file' not in request.files:
            return jsonify({"success": False, "message": "Không tìm thấy file tải lên"}), 400

        file = request.files['file']
        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        try:
            df = pd.read_excel(file)
            df = df.fillna('')

            for _, row in df.iterrows():
                name = str(row.get('Tên Món (name)', '')).strip()

                if not name:
                    continue

                raw_item_code = str(row.get('Mã món (item_code)', '')).strip()

                cursor.execute(
                    "SELECT item_code FROM menu WHERE name = ? LIMIT 1", (name,))
                existing_item = cursor.fetchone()

                if existing_item:
                    item_code = existing_item['item_code']
                elif raw_item_code:
                    item_code = raw_item_code
                else:
                    cursor.execute(
                        "SELECT item_code FROM menu WHERE item_code LIKE 'M%' ORDER BY item_code DESC LIMIT 1")
                    last_m = cursor.fetchone()
                    if last_m:
                        try:
                            last_num = int(last_m['item_code'][1:])
                            item_code = f"M{str(last_num + 1).zfill(3)}"
                        except Exception:
                            item_code = "M001"
                    else:
                        item_code = "M001"

                category = str(row.get('Danh Mục (category)', '')).strip() or 'Chưa phân loại'
                unit = str(row.get('ĐVT lẻ (unit)', '')).strip() or 'Phần'
                unit1 = str(row.get('ĐVT sỉ (unit1)', '')).strip()
                product_type = str(row.get('Loại (processed/ready_made)', '')).strip() or 'processed'
                img = str(row.get('Đường dẫn ảnh (image_path)', '')).strip()

                try:
                    base_price = float(row.get('Giá Bán Lẻ (base_price)', 0))
                except Exception:
                    base_price = 0.0

                try:
                    price = float(row.get('Giá Bán Sỉ (price)', 0))
                except Exception:
                    price = 0.0

                product_id = row.get('Mã ID sản phẩm (product_id)', '')
                p_id = int(product_id) if product_id != '' else None

                query = """
                INSERT INTO menu (
                    item_code, name, category, unit, unit1, 
                    base_price, price, product_type, product_id, image_path, is_active
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
                ON CONFLICT(item_code) DO UPDATE SET
                    name = excluded.name,
                    category = excluded.category,
                    unit = excluded.unit,
                    unit1 = excluded.unit1,
                    base_price = excluded.base_price,
                    price = excluded.price,
                    product_type = excluded.product_type,
                    product_id = excluded.product_id,
                    image_path = excluded.image_path
                """
                cursor.execute(query, (
                    item_code, name, category, unit, unit1,
                    base_price, price, product_type, p_id, img
                ))

            sqlite_commit(conn, label='fb_write')
            return jsonify({"success": True, "message": "Import và cập nhật danh sách thực đơn thành công!"}), 200

        except Exception as e:
            if conn:
                conn.rollback()
            print(f"--- Lỗi import_menu: {str(e)} ---")
            return jsonify({"success": False, "message": f"Lỗi xử lý cấu trúc file: {str(e)}"}), 500
        finally:
            if conn:
                conn.close()

    @app.route('/api/fb/download-sample-menu')
    @login_required
    def download_sample_menu():
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        menu_data = cursor.execute("""
        SELECT item_code, name, category, unit, unit1, 
               base_price, price, product_type, product_id, image_path 
        FROM menu 
        WHERE is_active = 1
        ORDER BY item_code ASC
        """).fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Danh_Sach_Thuc_Don"

        headers = [
            "STT", "Mã món (item_code)", "Tên Món (name)", "Danh Mục (category)",
            "ĐVT lẻ (unit)", "ĐVT sỉ (unit1)", "Giá Bán Lẻ (base_price)",
            "Giá Bán Sỉ (price)", "Loại (processed/ready_made)",
            "Mã ID sản phẩm (product_id)", "Đường dẫn ảnh (image_path)"
        ]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            ws.column_dimensions[get_column_letter(col)].width = 20

        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['K'].width = 35

        if menu_data:
            for idx, row in enumerate(menu_data, 1):
                ws.append([
                    idx, row['item_code'], row['name'], row['category'],
                    row['unit'], row['unit1'], row['base_price'],
                    row['price'], row['product_type'], row['product_id'],
                    row['image_path']
                ])
        else:
            samples = [
                [1, "M001", "Cà phê đá", "Đồ uống", "Ly", "", 25000,
                 0, "processed", "", "/static/img/cafe-da.jpg"],
                [2, "M002", "Sting dâu", "Đồ uống", "Lon", "Thùng", 15000,
                 320000, "ready_made", 101, "/static/img/sting.jpg"],
                [3, "M003", "Bún bò Huế", "Món ăn", "Tô", "", 45000,
                 0, "processed", "", "/static/img/bun-bo.jpg"]
            ]
            for row_data in samples:
                ws.append(row_data)

            guide_font = Font(italic=True, color="6B7280")
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.font = guide_font

        ws.append([])
        note_text = (
            "HƯỚNG DẪN: \n"
            "1. Bạn có thể chỉnh sửa trực tiếp dữ liệu hiện có và upload lại để cập nhật hàng loạt.\n"
            "2. Để trống 'Mã món (item_code)' nếu muốn hệ thống tự sinh mã Mxxx mới theo thứ tự.\n"
            "3. 'Mã ID sản phẩm (product_id)' là mã liên kết với kho hàng (bắt buộc cho hàng Dùng Ngay (ready_made), không được xóa mã này khi sửa Thực Đơn.\n"
            "4. Nếu thêm món mới, hãy tiếp tục STT và nhập đầy đủ thông tin.\n"
            "5. Nếu là Món Chế Biến ghi 'processed', Món dùng ngay như Bia, Nước Ngot,... ghi 'ready_made' ở cột Loại."
        )
        ws.append([note_text])

        note_row = ws.max_row
        ws.merge_cells(start_row=note_row, start_column=1,
                       end_row=note_row + 5, end_column=len(headers))
        note_cell = ws.cell(row=note_row, column=1)
        note_cell.alignment = Alignment(wrap_text=True, vertical='top')
        note_cell.font = Font(color="FF0000", bold=True)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        conn.close()

        filename = "Mau_Nhap_Thuc_Don.xlsx"
        return send_file(
            output,
            download_name=filename,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    @app.route('/api/fb/table-setup/template')
    @login_required
    def download_table_setup_template():
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")

        ws_area = wb.active
        ws_area.title = "1. Danh muc Khu vuc"
        headers_area = ["STT", "Ten Khu Vuc"]
        ws_area.append(headers_area)

        ws_table = wb.create_sheet("2. Danh muc Ban")
        headers_table = ["STT", "Ten Ban", "Thuoc Khu Vuc"]
        ws_table.append(headers_table)

        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        try:
            db_areas = cursor.execute(
                "SELECT name FROM areas ORDER BY id ASC").fetchall()

            db_tables = cursor.execute("""
                SELECT t.name AS table_name, a.name AS area_name 
                FROM tables t
                LEFT JOIN areas a ON t.area_id = a.id
                ORDER BY a.id ASC, t.id ASC
            """).fetchall()

        except sqlite3.Error:
            db_areas = []
            db_tables = []
        finally:
            conn.close()

        if db_areas:
            for idx, area in enumerate(db_areas, start=1):
                ws_area.append([idx, area['name']])
        else:
            ws_area.append([1, "Tang Tret"])

        if db_tables:
            for idx, table in enumerate(db_tables, start=1):
                ws_table.append([idx, table['table_name'],
                                table['area_name'] or "Chưa phân khu vực"])
        else:
            ws_table.append([1, "Ban 01", "Tang Tret"])

        for ws in [ws_area, ws_table]:
            for col in range(1, len(ws[1]) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                ws.column_dimensions[get_column_letter(col)].width = 25

            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='Mau_Nhap_Ban_Khu_Vuc.xlsx', as_attachment=True)

    @app.route('/api/fb/import-table', methods=['POST'])
    @login_required
    def import_table_setup():
        if 'file' not in request.files:
            return jsonify({"success": False, "message": "Không tìm thấy file tải lên hệ thống."}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({"success": False, "message": "Tên file Excel không hợp lệ."}), 400

        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        try:
            begin_immediate(conn, label='fb_import_areas')
            sheet_names = excel_file.sheet_names

            area_sheet = next(
                (name for name in sheet_names if "Khu vuc" in name or "Khu Vực" in name), 0)
            table_sheet = next(
                (name for name in sheet_names if "Ban" in name or "Bàn" in name), 1)

            df_areas = excel_file.parse(area_sheet).fillna('')
            df_tables = excel_file.parse(table_sheet).fillna('') if table_sheet in sheet_names or isinstance(table_sheet, int) else pd.DataFrame()

            areas_count = 0
            tables_count = 0

            for _, row in df_areas.iterrows():
                area_name = str(row.get('Ten Khu Vuc', row.get(
                    'Tên Khu Vực', row.get('Khu Vực', '')))).strip()
                if not area_name:
                    continue

                raw_area_id = row.get('STT', row.get('stt', row.get('ID', row.get('id', ''))))
                try:
                    area_id = int(float(raw_area_id)) if str(raw_area_id).strip() != '' else None
                except (ValueError, TypeError):
                    area_id = None

                if area_id:
                    existing_area = cursor.execute(
                        "SELECT id FROM areas WHERE id = ?", (area_id,)).fetchone()
                    if existing_area:
                        cursor.execute(
                            "UPDATE areas SET name = ? WHERE id = ?", (area_name, area_id))
                    else:
                        cursor.execute(
                            "INSERT INTO areas (id, name) VALUES (?, ?)", (area_id, area_name))
                    areas_count += 1
                else:
                    cursor.execute(
                        "INSERT INTO areas (name) VALUES (?)", (area_name,))
                    areas_count += 1

            if not df_tables.empty:
                db_areas_snapshot = cursor.execute(
                    "SELECT id, name FROM areas").fetchall()
                area_map = {row_a['name']: row_a['id'] for row_a in db_areas_snapshot}

                for _, row in df_tables.iterrows():
                    table_name = str(row.get('Ten Ban', row.get(
                        'Tên Bàn', row.get('Bàn', '')))).strip()
                    area_ref = str(
                        row.get('Thuoc Khu Vuc', row.get('Thuộc Khu Vực', ''))).strip()

                    if not table_name:
                        continue

                    area_id = area_map.get(area_ref, None)

                    raw_table_id = row.get('STT', row.get('stt', row.get('ID', row.get('id', ''))))
                    try:
                        table_id = int(float(raw_table_id)) if str(raw_table_id).strip() != '' else None
                    except (ValueError, TypeError):
                        table_id = None

                    if table_id:
                        existing_table = cursor.execute(
                            "SELECT id FROM tables WHERE id = ?", (table_id,)).fetchone()
                        if existing_table:
                            cursor.execute("""
                                UPDATE tables 
                                SET name = ?, area_id = ? 
                                WHERE id = ?
                            """, (table_name, area_id, table_id))
                        else:
                            cursor.execute("""
                                INSERT INTO tables (id, name, area_id, status) 
                                VALUES (?, ?, ?, 'Available')
                            """, (table_id, table_name, area_id))
                        tables_count += 1
                    else:
                        cursor.execute("""
                            INSERT INTO tables (name, area_id, status) 
                            VALUES (?, ?, 'Available')
                        """, (table_name, area_id))
                        tables_count += 1

            sqlite_commit(conn, label='fb_write')
            return jsonify({
                "success": True,
                "message": f"Đồng bộ thành công {areas_count} khu vực và {tables_count} bàn vào hệ thống!",
                "areas_count": areas_count,
                "tables_count": tables_count
            }), 200

        except Exception as e:
            if conn:
                conn.rollback()
            print(f"--- [LỖI IMPORT THEO STT/ID]: {str(e)} ---")
            return jsonify({"success": False, "message": f"Máy chủ gặp sự cố xử lý dữ liệu: {str(e)}"}), 500
        finally:
            if conn:
                conn.close()

    # API Lấy danh sách Khu vực
    @app.route('/api/fb/areas', methods=['GET'])
    @login_required
    def get_fb_areas():
        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT id, name FROM areas ORDER BY name")
        areas = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return jsonify(areas)

    @app.route('/api/fb/save-area', methods=['POST'])
    @login_required
    def save_fb_area():
        data = request.json or {}
        name = (data.get('name') or '').strip()
        if not name:
            return jsonify({"success": False, "message": "Vui lòng nhập tên khu vực"}), 400

        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            dup = cursor.execute("SELECT id FROM areas WHERE LOWER(name) = LOWER(?)", (name,)).fetchone()
            if dup:
                return jsonify({"success": False, "message": f"Khu vực '{name}' đã tồn tại"}), 400
            cursor.execute("INSERT INTO areas (name) VALUES (?)", (name,))
            sqlite_commit(conn, label='fb_write')
            return jsonify({"success": True, "status": "success", "message": f"Đã thêm khu vực '{name}'"})
        except Exception as e:
            conn.rollback()
            return jsonify({"success": False, "message": str(e)}), 500
        finally:
            conn.close()

    @app.route('/api/fb/save-table', methods=['POST'])
    @login_required
    def save_fb_table():
        data = request.json or {}
        area_id = data.get('area_id')
        name = (data.get('name') or '').strip()

        if not area_id:
            return jsonify({"success": False, "message": "Vui lòng chọn khu vực"}), 400
        if not name:
            return jsonify({"success": False, "message": "Vui lòng nhập tên bàn"}), 400

        try:
            area_id = int(area_id)
        except (TypeError, ValueError):
            return jsonify({"success": False, "message": "Khu vực không hợp lệ"}), 400

        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        try:
            area = cursor.execute("SELECT id FROM areas WHERE id = ?", (area_id,)).fetchone()
            if not area:
                return jsonify({"success": False, "message": "Khu vực không tồn tại"}), 404

            dup = cursor.execute(
                "SELECT id FROM tables WHERE area_id = ? AND LOWER(name) = LOWER(?)",
                (area_id, name),
            ).fetchone()
            if dup:
                return jsonify({"success": False, "message": f"Bàn '{name}' đã có trong khu vực này"}), 400

            cursor.execute(
                "INSERT INTO tables (name, area_id, status) VALUES (?, ?, 'Available')",
                (name, area_id),
            )
            sqlite_commit(conn, label='fb_write')
            return jsonify({
                "success": True,
                "status": "success",
                "message": f"Đã thêm bàn '{name}'",
                "table_id": cursor.lastrowid,
            })
        except Exception as e:
            conn.rollback()
            return jsonify({"success": False, "message": str(e)}), 500
        finally:
            conn.close()

    # === Khôi phục từ SME1/app.py ===
    @app.route('/importFB')
    @login_required
    def importFB_stock():
        """Phiếu nhập kho cho ngành ăn uống — dùng chung template với /import."""
        from routes.inventory import peek_next_import_no
        from Services.import_line_helpers import list_active_warehouses

        conn = get_db_connection()
        try:
            warehouses = list_active_warehouses(conn)
            sqlite_commit(conn, label='fb_write')
        finally:
            conn.close()

        return render_template(
            'import.html',
            today=datetime.now().strftime('%Y-%m-%d'),
            import_mode='fb',
            next_import_no=peek_next_import_no('stock'),
            warehouses=warehouses,
        )


    @app.route('/api/fb_import', methods=['POST'])
    @login_required
    def api_fbimport_post():
        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
    
        try:
            from Services.import_line_helpers import (
                ensure_warehouse_schema,
                insert_import_detail_row,
            )
            ensure_warehouse_schema(conn)
            sqlite_commit(conn, label='fb_write')

            data = request.get_json()
            if not data:
                return jsonify({"error": "Dữ liệu không hợp lệ"}), 400

            # --- 1. LẤY DỮ LIỆU ĐẦU VÀO ---
            items = data.get('items', [])
            supplier_id = data.get('supplier_id')
            import_date = data.get('date')
            bill_date = data.get('bill_date')
            import_no = data.get('import_no')
            bill_no = data.get('bill_no')
            note = data.get('note')
            default_warehouse = (data.get('warehouse_code') or 'KHO_001').strip()
        
            extra_cost = Decimal(str(data.get('extra_cost', 0) or 0))
            payment_status_input = data.get('payment_status', 'Chưa thanh toán')
            payment_method = data.get('payment_method', 'cash')

            # Lấy thông tin nhà cung cấp (ưu tiên địa chỉ/MST nhập tay từ form)
            c.execute("SELECT name, address, tax_code FROM suppliers WHERE id = ?", (supplier_id,))
            sup_row = c.fetchone()
            supplier_name = sup_row['name'] if sup_row else f"NCC ID {supplier_id}"
            form_address = (data.get('address') or '').strip()
            form_tax = (data.get('tax_code') or '').strip()
            supplier_address = form_address or (sup_row['address'] if sup_row and sup_row['address'] else "")
            tax_code = form_tax or (sup_row['tax_code'] if sup_row else None)
            if supplier_id and (form_address or form_tax):
                c.execute(
                    """UPDATE suppliers
                       SET address = COALESCE(NULLIF(?, ''), address),
                           tax_code = COALESCE(NULLIF(?, ''), tax_code)
                       WHERE id = ?""",
                    (form_address, form_tax, supplier_id),
                )

            # --- 2. TÍNH TỔNG GIÁ TRỊ SAU CHIẾT KHẤU ĐỂ PHÂN BỔ EXTRA_COST ---
            total_base_for_allocation = Decimal('0')
            for i in items:
                qty_val = Decimal(str(i.get('qty', 0) or 0))
                price_val = Decimal(str(i.get('buyprice', 0) or 0))
                disc_pct = Decimal(str(i.get('discount_pct') if i.get('discount_pct') is not None else i.get('discountPct', 0)))
                if qty_val > 0:
                    total_base_for_allocation += (qty_val * price_val) * (Decimal('1') - disc_pct / Decimal('100'))

            total_base_safe = total_base_for_allocation if total_base_for_allocation > 0 else Decimal('0.0001')

            # --- 3. TẠO PHIẾU NHẬP HEADER ---
            c.execute('PRAGMA table_info(import)')
            import_cols = {col[1] for col in c.fetchall()}
            import_fields = [
                'date', 'supplier_id', 'import_no', 'bill_no', 'bill_date',
                'note', 'payment_status', 'extra_cost', 'total_value', 'paid_amount',
            ]
            import_values = [
                import_date, supplier_id, import_no, bill_no, bill_date,
                note, payment_status_input, float(extra_cost), 0, 0,
            ]
            if 'warehouse_code' in import_cols:
                import_fields.append('warehouse_code')
                import_values.append(default_warehouse)

            c.execute(
                f'INSERT INTO import ({", ".join(import_fields)}) '
                f'VALUES ({", ".join(["?"] * len(import_fields))})',
                import_values,
            )
            import_id = c.lastrowid

            c.execute('PRAGMA table_info(stock_moves)')
            sm_has_wh = 'warehouse_code' in {col[1] for col in c.fetchall()}

            # --- 4. XỬ LÝ TỪNG SẢN PHẨM HỖN HỢP (READY_MADE & RAW_MATERIALS) ---
            total_invoice_value = Decimal('0')
            items_for_json = []
            now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            for item in items:
                pid = _fb_import_upsert_product(c, item)
                if not pid:
                    continue

                warehouse_code = (item.get('warehouse_code') or default_warehouse or 'KHO_001').strip()
                item_name = (item.get('name') or item.get('invoice_name') or '').strip()
                fe_wholesale_unit = str(item.get('wholesale_unit') or '').strip()
                fe_ratio = Decimal(str(item.get('ratio') if item.get('ratio') is not None else 1))
                unit_in = str(item.get('unit', '') or item.get('base_unit', '')).strip().lower()
                retail_unit = str(item.get('base_unit') or item.get('unit') or 'Cái').strip()

                c.execute("""
                    SELECT name, unit, unit1, unit_ratio, price as sale_price, base_price, product_type 
                    FROM products WHERE id = ?
                """, (pid,))
                p_info = c.fetchone()
                if not p_info:
                    continue

                if not fe_wholesale_unit:
                    fe_wholesale_unit = str(p_info['unit1'] or '').strip()
                if fe_ratio == 0:
                    fe_ratio = Decimal(str(p_info['unit_ratio'] or 1))

                qty_in = Decimal(str(item.get('qty', 0) or 0))
                if qty_in <= 0:
                    continue

                is_wholesale = bool(fe_wholesale_unit) and unit_in == fe_wholesale_unit.lower()
                qty_retail = qty_in * fe_ratio if is_wholesale else qty_in

                price_in = Decimal(str(item.get('buyprice', 0) or 0))
                tax_p = Decimal(str(item.get('tax_pct') if item.get('tax_pct') is not None else item.get('taxPct', 0)))
                disc_p = Decimal(str(item.get('discount_pct') if item.get('discount_pct') is not None else item.get('discountPct', 0)))

                line_subtotal = qty_in * price_in
                line_disc = line_subtotal * (disc_p / 100)
                line_after_disc = line_subtotal - line_disc
                line_tax = line_after_disc * (tax_p / 100)
            
                # Phân bổ chi phí mua hàng dựa trên giá sau chiết khấu
                line_extra = (line_after_disc / total_base_safe) * extra_cost
            
                # Tổng trị giá dòng hàng bao gồm đầy đủ Thuế và Chi phí phát sinh
                line_total = line_after_disc + line_tax + line_extra
                total_invoice_value += line_total

                # Giá vốn tính trên 1 đơn vị form nhập hạch toán
                cost_price_invoice = line_total / qty_in if qty_in > 0 else Decimal('0')
            
                # GIÁ VỐN QUY ĐỔI VỀ ĐƠN VỊ LẺ (Giá mua + Thuế + Chi phí mua)
                cost_price_base = cost_price_invoice / fe_ratio if is_wholesale else cost_price_invoice

                # Lưu thông tin chi tiết hóa đơn nhập (theo đơn vị hạch toán gốc)
                insert_import_detail_row(c, import_id, {
                    'import_id': import_id,
                    'product_id': pid,
                    'qty': float(qty_in),
                    'buyprice': float(price_in),
                    'subtotal': float(line_subtotal),
                    'discount': float(line_disc),
                    'tax': float(line_tax),
                    'cost_price': float(cost_price_invoice),
                    'unit_type': 1 if is_wholesale else 0,
                    'tax_pct': float(tax_p),
                    'discount_pct': float(disc_p),
                    'product_name': item_name or p_info['name'],
                    'unit': (item.get('unit') or item.get('base_unit') or '').strip(),
                    'line_type': (item.get('type') or p_info['product_type'] or 'ready_made'),
                    'warehouse_code': warehouse_code,
                })

                # Cập nhật tồn kho và tính giá vốn bình quân gia quyền theo giá trị lẻ chuẩn
                c.execute("SELECT quantity, avg_cost FROM inventory WHERE product_id = ?", (pid,))
                inv = c.fetchone()
                old_q = Decimal(str(inv['quantity'] if inv else 0))
                old_c = Decimal(str(inv['avg_cost'] if inv else 0))

                new_q = old_q + qty_retail
                new_avg = ((old_q * old_c) + (qty_retail * cost_price_base)) / new_q if new_q > 0 else cost_price_base

                c.execute("""
                    INSERT INTO inventory (product_id, quantity, avg_cost) VALUES (?, ?, ?)
                    ON CONFLICT(product_id) DO UPDATE SET quantity = excluded.quantity, avg_cost = excluded.avg_cost
                """, (pid, float(new_q), float(new_avg)))

                # Ghi nhận stock_moves theo giá lẻ quy đổi
                retail_unit = str(p_info['unit'] or "Cái").strip()
                wholesale_unit_str = str(p_info['unit1'] or "").strip()
                sm_cols = "product_id, date, type, ref_id, quantity, cost_price, note, ref_document, ref_type, type1, unit, unit1, unit_ratio"
                sm_vals = [
                    pid, import_date, import_id, float(qty_retail), float(cost_price_base),
                    f"Nhập hàng từ {supplier_name}", import_no, retail_unit,
                    wholesale_unit_str, float(fe_ratio),
                ]
                sm_placeholders = "?, ?, 'import', ?, ?, ?, ?, ?, 'import', 'Nhập', ?, ?, ?"
                if sm_has_wh:
                    sm_cols += ", warehouse_code"
                    sm_placeholders += ", ?"
                    sm_vals.append(warehouse_code)
                c.execute(
                    f"INSERT INTO stock_moves ({sm_cols}) VALUES ({sm_placeholders})",
                    sm_vals,
                )

                # Ghi nhận inventory_transactions theo giá lẻ quy đổi
                c.execute("""
                    INSERT INTO inventory_transactions (product_id, type, type1, quantity, cost_price, reference_id, reference_type, note, created_at)
                    VALUES (?, 'import', 'Nhập', ?, ?, ?, 'import', ?, ?)
                """, (pid, float(qty_retail), float(cost_price_base), import_id, f"Nhập kho - PN#{import_no}", now_str))

                items_for_json.append({
                    "product_id": pid, "product_name": p_info['name'], "qty": float(qty_in), "unit": unit_in,
                    "buyprice": float(price_in), "line_total": float(line_total), "warehouse_code": warehouse_code
                })

            # --- 5. CẬP NHẬT TỔNG KẾT HEADER VÀ PHIẾU CHI CHI TIẾT ---
            total_final = float(total_invoice_value)
            final_paid = total_final if payment_status_input in ['Đã thanh toán', 'Thanh toán một phần'] else 0.0

            c.execute("UPDATE import SET total_value = ?, paid_amount = ? WHERE id = ?", (total_final, final_paid, import_id))

            c.execute("""
                INSERT INTO phieu_nhap_kho (import_no, date, supplier_name, items_json, total_amount, import_id)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (import_no, import_date, supplier_name, json.dumps(items_for_json, ensure_ascii=False), total_final, import_id))

            if payment_status_input in ['Đã thanh toán', 'Thanh toán một phần']:
                c.execute("SELECT MAX(CAST(SUBSTR(voucher_no, 3) AS INTEGER)) as max_num FROM phieu_chi WHERE voucher_no LIKE 'PC%' AND LENGTH(voucher_no) > 2")
                row = c.fetchone()
                max_num = row['max_num'] if (row and row['max_num'] is not None) else 0
                new_pc_num = max_num + 1
            
                while True:
                    res_pc_vouch = f"PC{new_pc_num:06d}"
                    c.execute("SELECT id FROM phieu_chi WHERE voucher_no = ?", (res_pc_vouch,))
                    if not c.fetchone():
                        break
                    new_pc_num += 1
            
                credit_acc = '111' if payment_method == 'cash' else '112'
                c.execute("""
                    INSERT INTO phieu_chi (voucher_no, receiver_name, address, amount, credit_account, debit_account, reason, source_type, reference_document, source_id, date)
                    VALUES (?, ?, ?, ?, ?, '331', ?, 'import', ?, ?, ?)
                """, (res_pc_vouch, supplier_name, supplier_address, final_paid, credit_acc, f"Thanh toán nhập hàng {import_no}", bill_no, import_id, import_date))

            # Đồng bộ trạng thái hóa đơn điện tử gốc từ nhà cung cấp
            bill_no_clean = str(bill_no).strip() if bill_no else ""
            if bill_no_clean and bill_no_clean.lower() not in ['none', 'nan']:
                c.execute("UPDATE supplier_invoice SET status = 'imported' WHERE invoice_no = ? AND seller_tax_code = ? AND status != 'imported'", (bill_no_clean, tax_code))

            sqlite_commit(conn, label='fb_write')
            return jsonify({"success": True, "import_id": import_id, "voucher_no": import_no})
        except Exception as e:
            conn.rollback()
            traceback.print_exc()
            return jsonify({"error": f"Lỗi hệ thống: {str(e)}"}), 500
        finally:
            conn.close()

    @app.route('/api/fb_import/update/<int:import_id>', methods=['PUT'])
    @login_required
    def api_fbimport_update(import_id):
        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        try:
            data = request.get_json()
            if not data:
                return jsonify({"error": "Dữ liệu không hợp lệ"}), 400

            # --- 1. LẤY THÔNG TIN PHIẾU CŨ ---
            c.execute("SELECT * FROM import WHERE id = ?", (import_id,))
            imp = c.fetchone()
            if not imp:
                return jsonify({"error": "Không tìm thấy phiếu nhập"}), 404
        
            imp_keys = imp.keys()
            import_date_old = imp['date']
            import_no = imp['import_no']

            # --- 2. XỬ LÝ THÔNG TIN NCC ---
            supplier_name = (data.get('supplier_name') or '').strip()
            if not supplier_name:
                return jsonify({"error": "Tên nhà cung cấp không được để trống"}), 400
            tax_code = (data.get('tax_code') or '').strip() or None
            address = (data.get('address') or '').strip() or None
            supplier_id = imp['supplier_id'] if 'supplier_id' in imp_keys else None

            if supplier_id:
                c.execute("UPDATE suppliers SET name=?, tax_code=?, address=? WHERE id=?", (supplier_name, tax_code, address, supplier_id))
            else:
                c.execute("SELECT id FROM suppliers WHERE LOWER(name)=LOWER(?)", (supplier_name,))
                row = c.fetchone()
                if row:
                    supplier_id = row['id']
                    c.execute("UPDATE suppliers SET tax_code=?, address=? WHERE id=?", (tax_code, address, supplier_id))
                else:
                    c.execute("INSERT INTO suppliers (name, tax_code, address) VALUES (?, ?, ?)", (supplier_name, tax_code, address))
                    supplier_id = c.lastrowid

            # --- 3. ĐỌC THÔNG TIN CHUNG ---
            import_date = data.get('date', imp['date'])
            bill_date = data.get('bill_date') or (imp['bill_date'] if 'bill_date' in imp_keys else None)
            bill_no = data.get('bill_no') or (imp['bill_no'] if 'bill_no' in imp_keys else (imp['invoice_no'] if 'invoice_no' in imp_keys else None))
            payment_status = data.get('payment_status') or (imp['payment_status'] if 'payment_status' in imp_keys else 'Chưa thanh toán')
            payment_method_input = data.get('payment_method') or (imp['payment_method'] if 'payment_method' in imp_keys else 'cash')
            payment_method = 'cash' if '111' in str(payment_method_input) or payment_method_input in ['cash', 'Tiền mặt'] else 'bank'

            extra_cost = Decimal(str(data.get('extra_cost', 0) or (imp['extra_cost'] if 'extra_cost' in imp_keys else 0) or 0))
            note = data.get('note') or (imp['note'] if 'note' in imp_keys else '')
            now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            new_items = data.get('items', []) or []

            # Tính tổng trị giá cơ sở sau chiết khấu phân bổ
            total_base_for_allocation = Decimal('0')
            for it in new_items:
                qty_val = Decimal(str(it.get('qty', 0) or 0))
                price_val = Decimal(str(it.get('buyprice', 0) or 0))
                disc_pct = Decimal(str(it.get('discount_pct') if it.get('discount_pct') is not None else it.get('discountPct', 0)))
                if qty_val > 0:
                    total_base_for_allocation += (qty_val * price_val) * (Decimal('1') - disc_pct / Decimal('100'))

            if total_base_for_allocation <= 0:
                total_base_for_allocation = Decimal('0.0001')

            total_value = Decimal('0')

            for item in new_items:
                resolved_pid = _fb_import_upsert_product(c, item)
                if resolved_pid and not item.get('product_id'):
                    item['product_id'] = resolved_pid

            with conn:
                # Lấy chi tiết cũ để tiến hành hoàn nhập kho trước khi tính toán
                c.execute("SELECT product_id, qty, cost_price, unit_type FROM import_details WHERE import_id = ?", (import_id,))
                old_details = c.fetchall() or []

                product_units = {}
                if new_items:
                    product_ids = list(set(int(it.get('product_id', 0)) for it in new_items if it.get('product_id')))
                    if product_ids:
                        placeholders = ','.join('?' * len(product_ids))
                        c.execute(f"SELECT id, unit, unit1, unit_ratio FROM products WHERE id IN ({placeholders})", product_ids)
                        for p in c.fetchall():
                            product_units[p['id']] = {
                                'unit': p['unit'] or 'Cái', 'unit1': p['unit1'] or '', 'ratio': Decimal(str(p['unit_ratio'] or 1))
                            }

                # --- HOÀN KHO LŨY KẾ THEO PHƯƠNG PHÁP ĐẢO NGHỊCH BIẾN ĐỘNG ---
                for old in old_details:
                    old_cols = old.keys()
                    pid = old['product_id']
                    old_qty_input = Decimal(str(old['qty'] or 0))
                    old_cost = Decimal(str(old['cost_price'] or 0))
                    old_unit_type = old['unit_type'] if 'unit_type' in old_cols and old['unit_type'] is not None else 0

                    if old_qty_input <= 0: 
                        continue

                    ratio = product_units.get(pid, {}).get('ratio', Decimal('1'))
                    old_qty_base = old_qty_input * ratio if old_unit_type == 1 else old_qty_input

                    c.execute("SELECT COALESCE(SUM(ABS(quantity)), 0) FROM stock_moves WHERE product_id = ? AND date < ? AND type IN ('SALE', 'RETURN_IMPORT', 'export', 'ADJUSTMENT_OUT')", (pid, import_date_old))
                    total_out_before = Decimal(str(c.fetchone()[0] or 0))

                    remaining_old_base = max(old_qty_base - total_out_before, Decimal('0'))
                    remaining_value_old = remaining_old_base * old_cost

                    new_item = next((it for it in new_items if int(it.get('product_id', 0)) == pid), None)
                    if not new_item:
                        if remaining_old_base > 0:
                            c.execute("UPDATE inventory SET quantity = quantity - ? WHERE product_id = ?", (float(remaining_old_base), pid))
                        continue

                    qty_new_input = Decimal(str(new_item.get('qty', 0)))
                    if qty_new_input <= 0: 
                        continue

                    buyprice = Decimal(str(new_item.get('buyprice', 0)))
                    discountPct = Decimal(str(new_item.get('discount_pct') if new_item.get('discount_pct') is not None else new_item.get('discountPct', 0)))
                    taxPct = Decimal(str(new_item.get('tax_pct') if new_item.get('tax_pct') is not None else new_item.get('taxPct', 0)))

                    line_total = qty_new_input * buyprice
                    discount_amt = line_total * discountPct / Decimal('100')
                    after_discount = line_total - discount_amt
                    tax_amt = after_discount * taxPct / Decimal('100')
                
                    allocated_extra = extra_cost * (after_discount / total_base_for_allocation) if extra_cost > 0 else Decimal('0')
                    cost_value_full = after_discount + tax_amt + allocated_extra
                    cost_price_invoice = cost_value_full / qty_new_input

                    input_unit = str(new_item.get('unit') or '').strip()
                    p_unit = product_units.get(pid, {})
                    db_unit1 = str(p_unit.get('unit1') or '').strip()
                    new_unit_type = 1 if input_unit.lower() == db_unit1.lower() and db_unit1 != '' else 0
                
                    qty_new_base = qty_new_input * ratio if new_unit_type == 1 else qty_new_input
                
                    # GIÁ VỐN QUY ĐỔI VỀ ĐƠN VỊ LẺ (Giá mua + Thuế + Chi phí mua)
                    cost_price_base = cost_price_invoice / ratio if new_unit_type == 1 else cost_price_invoice
                
                    remaining_new_base = max(qty_new_base - total_out_before, Decimal('0'))
                    remaining_value_new = remaining_new_base * cost_price_base

                    c.execute("SELECT quantity, avg_cost FROM inventory WHERE product_id = ?", (pid,))
                    inv = c.fetchone()
                    cur_qty = Decimal(str(inv['quantity'] or 0)) if inv else Decimal('0')
                    cur_value = cur_qty * Decimal(str(inv['avg_cost'] or 0)) if inv else Decimal('0')

                    final_qty = cur_qty - remaining_old_base + remaining_new_base
                    final_value = cur_value - remaining_value_old + remaining_value_new
                    final_avg = final_value / final_qty if final_qty > 0 else Decimal('0')

                    if inv:
                        c.execute("UPDATE inventory SET quantity=?, avg_cost=? WHERE product_id=?", (float(final_qty), float(final_avg), pid))
                    else:
                        c.execute("INSERT INTO inventory (product_id, quantity, avg_cost) VALUES (?, ?, ?)", (pid, float(final_qty), float(cost_price_base)))

                    total_value += cost_value_full

                # --- DỌN SẠCH VÀ LƯU CHI TIẾT MỚI ---
                c.execute("DELETE FROM import_details WHERE import_id = ?", (import_id,))
                c.execute("DELETE FROM stock_moves WHERE ref_id = ? AND type IN ('import', 'RETURN_IMPORT')", (import_id,))
                c.execute("DELETE FROM inventory_transactions WHERE reference_id = ? AND reference_type = 'import'", (import_id,))

                for item in new_items:
                    qty = Decimal(str(item.get('qty', 0) or 0))
                    if qty <= 0: continue
                    pid = int(item.get('product_id', 0))
                    if not pid: continue

                    buyprice = Decimal(str(item.get('buyprice', 0)))
                    discountPct = Decimal(str(item.get('discount_pct') if item.get('discount_pct') is not None else item.get('discountPct', 0)))
                    taxPct = Decimal(str(item.get('tax_pct') if item.get('tax_pct') is not None else item.get('taxPct', 0)))
                
                    line_total = qty * buyprice
                    discount_amt = Decimal(str(item.get('discount_amount'))) if item.get('discount_amount') is not None else (line_total * discountPct / Decimal('100'))
                    after_discount = line_total - discount_amt
                    tax_amt = Decimal(str(item.get('tax_amount'))) if item.get('tax_amount') is not None else (after_discount * taxPct / Decimal('100'))
                
                    allocated_extra = extra_cost * (after_discount / total_base_for_allocation) if extra_cost > 0 else Decimal('0')
                    cost_value_full = after_discount + tax_amt + allocated_extra
                    cost_price_invoice = cost_value_full / qty

                    input_unit = str(item.get('unit') or '').strip()
                    p_unit = product_units.get(pid, {})
                    db_unit = str(p_unit.get('unit') or 'Cái').strip()
                    db_unit1 = str(p_unit.get('unit1') or '').strip()
                    ratio = p_unit.get('ratio', Decimal('1'))
                
                    unit_type = 1 if input_unit.lower() == db_unit1.lower() and db_unit1 != '' else 0
                    qty_base_final = qty * ratio if unit_type == 1 else qty
                
                    # GIÁ VỐN LẺ CHUẨN ĐÃ BAO GỒM THUẾ VÀ CHI PHÍ GỒM BỔ
                    cost_price_base = cost_price_invoice / ratio if unit_type == 1 else cost_price_invoice

                    c.execute("""INSERT INTO import_details (import_id, product_id, qty, buyprice, discount, tax, tax_pct, discount_pct, cost_price, unit_type)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                              (import_id, pid, float(qty), float(buyprice), float(discount_amt), float(tax_amt), float(taxPct), float(discountPct), float(cost_price_invoice), unit_type))

                    c.execute("""INSERT INTO stock_moves (product_id, date, type, ref_id, quantity, cost_price, note, ref_document, ref_type, type1, unit, unit1, unit_ratio)
                                VALUES (?, ?, 'import', ?, ?, ?, ?, ?, 'import', 'Nhập', ?, ?, ?)""",
                              (pid, import_date, import_id, float(qty_base_final), float(cost_price_base), f"Nhập kho – PN#{import_no}", import_no, db_unit, db_unit1, float(ratio)))

                    c.execute("""INSERT INTO inventory_transactions (product_id, type, type1, quantity, cost_price, reference_id, reference_type, note, created_at)
                                VALUES (?, 'import', 'Nhập', ?, ?, ?, 'import', ?, ?)""",
                              (pid, float(qty_base_final), float(cost_price_base), import_id, f"Nhập kho - PN#{import_no}", now_str))

                # --- CÂN ĐỐI PHIẾU CHI VÀ TÀI CHÍNH SỔ SÁCH ---
                c.execute("DELETE FROM Phieu_chi WHERE source_type = 'import' AND source_id = ?", (import_id,))
                if payment_status in ['Đã thanh toán', 'Thanh toán một phần']:
                    final_paid = float(total_value)
                    credit_acc = '111' if payment_method == 'cash' else '112'
                
                    c.execute("SELECT MAX(CAST(SUBSTR(voucher_no, 3) AS INTEGER)) as max_num FROM Phieu_chi WHERE voucher_no LIKE 'PC%' AND LENGTH(voucher_no) > 2")
                    max_num = (c.fetchone()['max_num'] or 0)
                    new_pc_num = max_num + 1
                
                    while True:
                        res_pc_vouch = f"PC{new_pc_num:06d}"
                        c.execute("SELECT id FROM Phieu_chi WHERE voucher_no = ?", (res_pc_vouch,))
                        if not c.fetchone(): break
                        new_pc_num += 1
                
                    c.execute("""INSERT INTO Phieu_chi (voucher_no, receiver_name, address, amount, credit_account, debit_account, reason, source_type, reference_document, source_id, preparer, date)
                                VALUES (?, ?, ?, ?, ?, '331', ?, 'import', ?, ?, ?, ?)""",
                              (res_pc_vouch, supplier_name, address, final_paid, credit_acc, f"Thanh toán tiền mua hàng số {import_no}", bill_no, import_id, session.get('user_name', 'Admin'), import_date))

                # Cập nhật thông tin Header phiếu nhập tổng hợp
                update_fields = ["supplier_id=?", "date=?"]
                update_values = [supplier_id, import_date]
                for field in ["bill_date", "bill_no", "payment_status", "payment_method", "extra_cost", "note", "total_value"]:
                    if field in imp_keys:
                        update_fields.append(f"{field}=?")
                        val = float(total_value) if field == "total_value" else (float(extra_cost) if field == "extra_cost" else locals()[field])
                        update_values.append(val)
                update_values.append(import_id)
                c.execute(f"UPDATE import SET {', '.join(update_fields)} WHERE id=?", tuple(update_values))

            sqlite_commit(conn, label='fb_write')
            return jsonify({"success": True, "message": "Cập nhật phiếu nhập và cân đối dòng tiền thành công!"})
        except Exception as e:
            conn.rollback()
            traceback.print_exc()
            return jsonify({"error": str(e)}), 500
        finally:
            conn.close()
    @app.route('/api/inventory/ingredients', methods=['GET'])
    @login_required
    def get_inventory_ingredients():
        from db.init import ensure_products_schema

        conn = None
        try:
            conn = get_db_connection()
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            ensure_products_schema(conn)

            query = """
                SELECT 
                    p.id,
                    p.product_code,
                    p.name,
                    p.unit,
                    p.product_type,
                    p.business_line,
                    COALESCE(i.quantity, 0) AS current_stock
                FROM products p
                LEFT JOIN inventory i ON p.id = i.product_id
                WHERE p.business_line = 'F&B'
                  AND p.product_type = 'raw_materials'
                ORDER BY p.name ASC
            """
        
            cursor.execute(query)
            rows = cursor.fetchall()
        
            ingredients = []
            for row in rows:
                ingredients.append({
                    "id": row["id"],
                    "product_code": row["product_code"] if row["product_code"] else f"PROD-{row['id']}",
                    "name": row["name"] if row["name"] else "Chưa đặt tên",
                    "unit": row["unit"] if row["unit"] else "ĐVT",
                    "product_type": row["product_type"],
                    "business_line": row["business_line"],
                    "current_stock": round(float(row["current_stock"] or 0), 4)
                })
            
            return jsonify(ingredients), 200

        except sqlite3.Error as e:
            logger.error("Lỗi ingredients API: %s", e)
            return jsonify([]), 200
        except Exception as e:
            logger.error("Lỗi ingredients API: %s", e)
            return jsonify([]), 200
        finally:
            if conn:
                conn.close()

    @app.route('/api/fb/stocktake-sheet', methods=['GET'])
    @login_required
    def fb_stocktake_sheet():
        """Danh sách NVL + tồn sổ để kiểm kê cuối ngày."""
        from Services.fb_schema import ensure_fb_schema
        from db.init import ensure_products_schema

        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        try:
            ensure_fb_schema(conn, commit=False)
            ensure_products_schema(conn)
            _ensure_draft_inventory_table(cursor)
            sqlite_commit(conn, label='fb_write')

            rows = cursor.execute("""
                SELECT p.id AS product_id, p.product_code, p.name, p.unit,
                       COALESCE(i.quantity, 0) AS book_stock,
                       COALESCE(i.avg_cost, 0) AS avg_cost
                FROM products p
                LEFT JOIN inventory i ON p.id = i.product_id
                WHERE p.business_line = 'F&B' AND p.product_type = 'raw_materials'
                ORDER BY p.name ASC
            """).fetchall()

            pending = cursor.execute("""
                SELECT d.product_id,
                       SUM(CASE WHEN d.quantity < 0 THEN ABS(d.quantity) ELSE 0 END) AS pending_used
                FROM draft_inventory d
                WHERE d.is_processed = 0
                GROUP BY d.product_id
            """).fetchall()
            pending_map = {r['product_id']: float(r['pending_used'] or 0) for r in pending}

            items = []
            for row in rows:
                book = round(float(row['book_stock'] or 0), 4)
                pending_used = round(pending_map.get(row['product_id'], 0), 4)
                items.append({
                    "product_id": row['product_id'],
                    "product_code": row['product_code'] or f"PROD-{row['product_id']}",
                    "name": row['name'],
                    "unit": row['unit'] or 'ĐVT',
                    "book_stock": book,
                    "avg_cost": float(row['avg_cost'] or 0),
                    "pending_used": pending_used,
                    "suggested_count": round(max(0, book - pending_used), 4) if pending_used > 0 else book,
                })

            return jsonify({"success": True, "items": items, "date": datetime.now().strftime('%Y-%m-%d')})
        except Exception as e:
            logger.error("stocktake-sheet: %s", e)
            return jsonify({"success": False, "message": str(e)}), 500
        finally:
            conn.close()

    @app.route('/api/fb/stocktake-batch', methods=['POST'])
    @login_required
    def fb_stocktake_batch():
        """
        Kiểm kê cuối ngày: nhập tồn thực → hệ thống tính lượng đã dùng = tồn sổ - tồn thực.
        Ghi nháp xuất kho (chưa trừ kho chính) — chờ chốt cuối ngày lập 1 phiếu xuất.
        """
        data = request.json or {}
        entries = data.get('items', [])
        if not entries:
            return jsonify({"success": False, "message": "Chưa có dữ liệu kiểm kê"}), 400

        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        try:
            begin_immediate(conn, label='fb_inventory_count')
            _ensure_draft_inventory_table(cursor)
            skipped = []
            now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            for entry in entries:
                try:
                    product_id = int(entry.get('product_id'))
                    counted_qty = float(entry.get('counted_qty'))
                except (TypeError, ValueError):
                    continue

                if counted_qty < 0:
                    skipped.append({"product_id": product_id, "reason": "Tồn thực không hợp lệ"})
                    continue

                row = cursor.execute("""
                    SELECT p.name, p.unit, COALESCE(i.quantity, 0) AS book_stock
                    FROM products p
                    LEFT JOIN inventory i ON p.id = i.product_id
                    WHERE p.id = ? AND p.product_type = 'raw_materials'
                """, (product_id,)).fetchone()

                if not row:
                    skipped.append({"product_id": product_id, "reason": "Không tìm thấy NVL"})
                    continue

                book_stock = float(row['book_stock'] or 0)
                used_qty = round(book_stock - counted_qty, 4)

                cursor.execute("""
                    DELETE FROM draft_inventory
                    WHERE product_id = ? AND is_processed = 0
                      AND COALESCE(note, '') LIKE 'Kiểm kê cuối ngày%'
                """, (product_id,))

                if used_qty <= 0:
                    if used_qty < 0:
                        skipped.append({
                            "product_id": product_id,
                            "name": row['name'],
                            "reason": f"Tồn thực ({counted_qty}) lớn hơn tồn sổ ({book_stock})"
                        })
                    continue

                note = (
                    f"Kiểm kê cuối ngày | Tồn sổ {book_stock} {row['unit']} | "
                    f"Tồn thực {counted_qty} {row['unit']} | Đã dùng {used_qty} {row['unit']}"
                )
                cursor.execute("""
                    INSERT INTO draft_inventory (product_id, quantity, note, is_processed, created_at)
                    VALUES (?, ?, ?, 0, ?)
                """, (product_id, -used_qty, note, now_str))

                saved.append({
                    "product_id": product_id,
                    "name": row['name'],
                    "unit": row['unit'],
                    "book_stock": book_stock,
                    "counted_qty": counted_qty,
                    "used_qty": used_qty,
                })

            if not saved:
                conn.rollback()
                return jsonify({
                    "success": False,
                    "message": "Không có NVL nào cần xuất. Kiểm tra lại tồn thực đã nhập.",
                    "skipped": skipped,
                }), 400

            sqlite_commit(conn, label='fb_write')
            return jsonify({
                "success": True,
                "message": f"Đã ghi nhận kiểm kê {len(saved)} NVL. Cuối ngày bấm Chốt để lập 1 phiếu xuất kho.",
                "saved": saved,
                "skipped": skipped,
            })
        except Exception as e:
            conn.rollback()
            logger.error("stocktake-batch: %s", e)
            return jsonify({"success": False, "message": str(e)}), 500
        finally:
            conn.close()

    @app.route('/api/fb/add-draft-inventory', methods=['POST'])
    @login_required
    def add_draft_item():
        data = request.json or {}
        product_id = data.get('product_id')
        qty = float(data.get('quantity', 0))
        note = (data.get('note') or 'Xuất kho thủ công').strip()

        if qty > 0:
            qty = -qty

        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        try:
            _ensure_draft_inventory_table(cursor)
            abs_qty = abs(qty)
            if abs_qty <= 0:
                return jsonify({"success": False, "message": "Số lượng phải lớn hơn 0"}), 400

            row = cursor.execute("""
                SELECT p.name, COALESCE(i.quantity, 0) AS stock_qty, p.unit
                FROM products p
                LEFT JOIN inventory i ON p.id = i.product_id
                WHERE p.id = ?
            """, (product_id,)).fetchone()

            if not row:
                return jsonify({"success": False, "message": "Không tìm thấy nguyên liệu"}), 404

            stock_qty = float(row['stock_qty'] or 0)
            if stock_qty < abs_qty:
                return jsonify({
                    "success": False,
                    "message": f"Không đủ hàng! '{row['name']}' chỉ còn {stock_qty} {row['unit']}."
                }), 400

            cursor.execute("""
                INSERT INTO draft_inventory (product_id, quantity, note, is_processed)
                VALUES (?, ?, ?, 0)
            """, (product_id, qty, note))

            sqlite_commit(conn, label='fb_write')
            return jsonify({"success": True, "status": "success", "message": "Ghi nhận xuất kho nháp thành công"})

        except Exception as e:
            conn.rollback()
            print(f"--- Lỗi add-draft-item: {str(e)} ---")
            return jsonify({"status": "error", "message": f"Lỗi hệ thống: {str(e)}"}), 500
        finally:
            conn.close()

    @app.route('/api/fb/get-draft-history', methods=['GET'])
    # @login_required # Mở ghi chú này nếu bạn muốn giới hạn quyền đăng nhập như hàm add_draft_item
    def get_draft_history():
        conn = None
        try:
            conn = get_db_connection()
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            _ensure_draft_inventory_table(cursor)
            sqlite_commit(conn, label='fb_write')
        
            query = """
                SELECT 
                    d.id,
                    COALESCE(d.created_at, datetime('now', 'localtime')) AS created_at,
                    COALESCE(p.name, 'Hàng hóa/Nguyên liệu đã xóa') AS ingredient_name,
                    d.quantity,
                    COALESCE(p.unit, 'ĐVT') AS unit,
                    COALESCE(d.note, '') AS note,
                    d.is_processed
                FROM draft_inventory d
                LEFT JOIN products p ON d.product_id = p.id
                WHERE d.is_processed = 0
                ORDER BY d.id DESC
                LIMIT 100
            """
        
            cursor.execute(query)
            rows = cursor.fetchall()
        
            # Chuyển đổi dữ liệu SQLite sang danh sách JSON gửi về Frontend
            draft_history = []
            for row in rows:
                qty = float(row["quantity"] or 0)
                draft_history.append({
                    "id": row["id"],
                    "created_at": row["created_at"],
                    "ingredient_name": row["ingredient_name"],
                    "quantity": round(abs(qty), 4),
                    "qty_signed": round(qty, 4),
                    "unit": row["unit"],
                    "note": row["note"],
                    "is_export": qty < 0,
                })
            
            return jsonify(draft_history), 200

        except sqlite3.Error as e:
            print(f"--- Lỗi SQLite get-draft-history: {str(e)} ---")
            # Trả về mảng rỗng kèm mã 200 để Frontend không bị lỗi cú pháp nhận diện JSON
            return jsonify([]), 200
        
        except Exception as e:
            print(f"--- Lỗi hệ thống get-draft-history: {str(e)} ---")
            return jsonify([]), 200
        
        finally:
            if conn:
                conn.close()

    # API Xóa lịch sử kho nháp
    @app.route('/api/fb/delete-draft/<int:draft_id>', methods=['DELETE'])
    @login_required
    def delete_draft_item(draft_id):
        conn = get_db_connection()
        cursor = conn.cursor()
    
        try:
            # Thực hiện xóa bản ghi nháp theo ID
            cursor.execute("DELETE FROM draft_inventory WHERE id = ?", (draft_id,))
        
            # Kiểm tra xem có bản ghi nào thực sự bị xóa không
            if cursor.rowcount == 0:
                return jsonify({
                    "status": "error", 
                    "message": f"Không tìm thấy phiếu nháp có ID {draft_id} hoặc phiếu đã bị xóa trước đó."
                }), 404
            
            sqlite_commit(conn, label='fb_write')
            return jsonify({
                "status": "success", 
                "message": "Xóa mục nháp thành công"
            })
        
        except Exception as e:
            conn.rollback()
            print(f"--- Lỗi delete_draft_item: {str(e)} ---")
            return jsonify({
                "status": "error", 
                "message": f"Lỗi hệ thống khi xóa: {str(e)}"
            }), 500
        
        finally:
            conn.close()

    @app.route('/api/fb/finalize-inventory', methods=['POST'])
    @login_required
    def finalize_inventory():
        data = request.json or {}
        customer_name = data.get('customer_name') or 'Xuất NVL chế biến cuối ngày'
        sale_id = data.get('sale_id', None) # Có thể truyền hoặc không tùy nghiệp vụ

        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        try:
            begin_immediate(conn, label='fb_end_of_day_export')
            _ensure_draft_inventory_table(cursor)
            drafts = cursor.execute("""
                SELECT d.id, d.product_id, d.quantity as qty_change, d.note as draft_note,
                       p.name, p.unit, COALESCE(i.avg_cost, 0) AS avg_cost,
                       COALESCE(i.quantity, 0) AS stock_qty
                FROM draft_inventory d
                JOIN products p ON d.product_id = p.id
                LEFT JOIN inventory i ON p.id = i.product_id
                WHERE d.is_processed = 0
            """).fetchall()

            if not drafts:
                return jsonify({"success": False, "message": "Không có dữ liệu nháp để thực hiện chốt"}), 400

            # --- KIỂM TRA LỖI XUẤT ÂM TRƯỚC KHI GHI NHẬN CHÍNH THỨC ---
            for row in drafts:
                qty_change = row['qty_change']
                stock_qty = row['stock_qty'] if row['stock_qty'] is not None else 0

                if qty_change < 0 and (stock_qty + qty_change) < 0:
                    conn.rollback()
                    return jsonify({
                        "success": False,
                        "message": f"Không thể chốt! Yêu cầu xuất '{row['name']}' ({abs(qty_change)} {row['unit']}) vượt quá số lượng trong kho ({stock_qty} {row['unit']})."
                    }), 400

            sale_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            px_items = []      #-- Danh sách lưu chi tiết mặt hàng xuất để đẩy vào phiếu xuất kho
            final_px_total = 0 #-- Tổng giá trị phiếu xuất kho (Tích lũy theo GIÁ VỐN)

            # 2. Cập nhật kho chính thức, ghi thẻ kho (moves/transactions) và chuẩn bị danh sách phiếu xuất
            for row in drafts:
                pid = row['product_id']
                qty_change = row['qty_change']
                avg_cost = row['avg_cost'] or 0
            
                # Xác định động loại giao dịch dựa vào dấu âm hay dương của qty_change
                move_type = 'import' if qty_change > 0 else 'export'
                type1_label = 'nhập kho' if qty_change > 0 else 'chế biến'
                note_content = row['draft_note'] if row['draft_note'] else f"Chốt kho {type1_label}: {row['name']}"

                # A. Cập nhật kho lớn (Cộng trực tiếp đại số: số dương tăng kho, số âm giảm kho)
                cursor.execute("""
                    UPDATE inventory
                    SET quantity = COALESCE(quantity, 0) + ?
                    WHERE product_id = ?
                """, (qty_change, pid))

                if cursor.rowcount == 0:
                    cursor.execute(
                        "INSERT INTO inventory (product_id, quantity, avg_cost) VALUES (?, ?, ?)",
                        (pid, qty_change, avg_cost),
                    )
            
                # B. Ghi log lịch sử vào bảng stock_moves
                cursor.execute("""
                    INSERT INTO stock_moves (product_id, date, type, quantity, cost_price, ref_type, type1, note)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (pid, sale_date, move_type, qty_change, avg_cost, move_type, type1_label, note_content))

                # C. Ghi log vào bảng inventory_transactions
                cursor.execute("""
                    INSERT INTO inventory_transactions (product_id, type1, type, quantity, cost_price, reference_type, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (pid, type1_label, move_type, qty_change, avg_cost, move_type, sale_date))

                # D. Nếu dòng nháp này là XUẤT KHO (qty_change < 0), gom dữ liệu để lập phiếu xuất kho theo giá vốn
                if qty_change < 0:
                    qty_positive = abs(qty_change)
                    line_amount = qty_positive * avg_cost # Thành tiền xuất kho = Số lượng * Giá vốn bình quân
                    final_px_total += line_amount
                
                    px_items.append({
                        "product_id": pid, 
                        "product_name": row['name'], 
                        "unit": row['unit'],
                        "quantity": qty_positive, 
                        "price": avg_cost,   # Đảm bảo dùng GIÁ VỐN avg_cost thay vì giá bán thực đơn
                        "amount": line_amount
                    })

            # 3. Tạo Phiếu Xuất Kho (Phieu_xuat_kho) chính thức nếu có sản phẩm bị trừ kho
            px_vno = None
            if px_items:
                last_px = cursor.execute("SELECT voucher_no FROM phieu_xuat_kho WHERE voucher_no LIKE 'PX%' ORDER BY id DESC LIMIT 1").fetchone()
                px_num = (int(last_px['voucher_no'][2:]) + 1) if last_px else 1
                px_vno = f"PX{px_num:06d}"

                cursor.execute("""
                    INSERT INTO phieu_xuat_kho (voucher_no, date, customer_name, items_json, total_amount, sale_id) 
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (px_vno, sale_date, customer_name, json.dumps(px_items, ensure_ascii=False), final_px_total, sale_id))

            cursor.execute("UPDATE draft_inventory SET is_processed = 1 WHERE is_processed = 0")

            sqlite_commit(conn, label='fb_write')
            msg = f"Đã chốt kho và lập phiếu xuất {px_vno} ({len(px_items)} NVL)" if px_vno else "Đã chốt kho (không có xuất NVL)"
            return jsonify({
                "success": True,
                "message": msg,
                "px_voucher_no": px_vno,
                "export_lines": len(px_items),
            })

        except Exception as e:
            conn.rollback()
            print(f"--- Lỗi finalize_inventory: {str(e)} ---")
            return jsonify({"success": False, "message": f"Lỗi hệ thống: {str(e)}"}), 500
        finally:
            conn.close()

