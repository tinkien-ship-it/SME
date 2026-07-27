"""Routes nhập kho, tồn kho, kiểm kê — tách từ app.py."""
import json
import logging
import re
import sqlite3
import traceback
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO

import pandas as pd
import requests
from flask import (
    Response,
    abort,
    flash,
    jsonify,
    make_response,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from flask_login import login_required

from db_utils import get_db_connection
from helpers import format_price
from Services.inventory_stock_helpers import (
    apply_wac_inbound,
    apply_wac_outbound,
    import_base_qty,
    import_cost_to_base,
    ledger_quantity,
    rebuild_all_wac_from_moves,
    reconcile_all_inventory,
    reverse_import_moves_wac,
    sync_inventory_quantity_from_moves,
    sync_inventory_quantities,
)
from Services.return_import_checkout import process_return_import_checkout

logger = logging.getLogger(__name__)


def _format_import_date_display(raw_date):
    if not raw_date:
        return ""
    text = str(raw_date)
    try:
        return datetime.strptime(text[:19], '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y %H:%M')
    except ValueError:
        try:
            return datetime.strptime(text[:10], '%Y-%m-%d').strftime('%d/%m/%Y')
        except ValueError:
            return text[:10]


def _calc_import_line_amounts(quantity, unit_price, discount_amount, tax_amount, subtotal,
                              discount_pct=0, tax_pct=0):
    """Tính chiết khấu, thuế, tổng thanh toán cho một dòng import_details."""
    from Services.import_line_helpers import calc_import_detail_line_amounts
    return calc_import_detail_line_amounts(
        quantity, unit_price, discount_amount, tax_amount, subtotal,
        discount_pct, tax_pct,
    )


def fetch_import_items_detail_report(cursor, start_date, end_date, search_query=None):
    """Lấy chi tiết hàng mua từ import_details trong khoảng ngày."""
    cursor.execute("PRAGMA table_info(import_details)")
    detail_cols = {col[1] for col in cursor.fetchall()}

    unit_type_sel = "COALESCE(ii.unit_type, 0) AS unit_type" if 'unit_type' in detail_cols else "0 AS unit_type"
    disc_pct_sel = "COALESCE(ii.discount_pct, 0) AS discount_pct" if 'discount_pct' in detail_cols else "0 AS discount_pct"
    tax_pct_sel = "COALESCE(ii.tax_pct, 0) AS tax_pct" if 'tax_pct' in detail_cols else "0 AS tax_pct"
    detail_name_sel = "COALESCE(ii.product_name, p.name, '—') AS product_name" if 'product_name' in detail_cols else "COALESCE(p.name, '—') AS product_name"
    detail_unit_col = "ii.unit" if 'unit' in detail_cols else None
    line_type_sel = "COALESCE(ii.line_type, 'goods') AS line_type" if 'line_type' in detail_cols else "'goods' AS line_type"
    if detail_unit_col:
        unit_base = f"COALESCE({detail_unit_col}, p.unit, 'Cái')"
    else:
        unit_base = "COALESCE(p.unit, 'Cái')"

    cursor.execute("PRAGMA table_info(products)")
    product_cols = {col[1] for col in cursor.fetchall()}
    if 'product_code' in product_cols and 'barcode' in product_cols:
        product_code_sel = "COALESCE(p.product_code, p.barcode, '') AS product_code"
        product_code_search = "COALESCE(p.product_code, p.barcode, '') LIKE ?"
    elif 'product_code' in product_cols:
        product_code_sel = "COALESCE(p.product_code, '') AS product_code"
        product_code_search = "COALESCE(p.product_code, '') LIKE ?"
    elif 'barcode' in product_cols:
        product_code_sel = "COALESCE(p.barcode, '') AS product_code"
        product_code_search = "COALESCE(p.barcode, '') LIKE ?"
    else:
        product_code_sel = "CAST(ii.product_id AS TEXT) AS product_code"
        product_code_search = "CAST(ii.product_id AS TEXT) LIKE ?"

    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='supplier_invoice'")
    has_supplier_invoice = cursor.fetchone() is not None
    if has_supplier_invoice:
        pdf_url_sel = """
            (
                SELECT '/api/invoices/inward/' || si.id || '/pdf'
                FROM supplier_invoice si
                WHERE TRIM(COALESCE(si.invoice_no, '')) = TRIM(COALESCE(i.bill_no, ''))
                  AND TRIM(COALESCE(i.bill_no, '')) != ''
                  AND (
                      COALESCE(s.tax_code, '') = ''
                      OR COALESCE(si.seller_tax_code, '') = ''
                      OR TRIM(si.seller_tax_code) = TRIM(COALESCE(s.tax_code, ''))
                  )
                ORDER BY si.id DESC
                LIMIT 1
            ) AS invoice_pdf_url
        """
    else:
        pdf_url_sel = "NULL AS invoice_pdf_url"

    sql = f"""
        SELECT
            i.id AS import_id,
            i.date AS import_date,
            COALESCE(i.import_no, 'PN' || printf('%06d', i.id)) AS import_no,
            COALESCE(i.bill_no, '') AS bill_no,
            COALESCE(s.tax_code, '') AS supplier_tax_code,
            {pdf_url_sel},
            COALESCE(s.name, '') AS supplier_name,
            {product_code_sel},
            {detail_name_sel},
            {unit_base} AS base_unit,
            p.unit1 AS wholesale_unit,
            COALESCE(p.unit_ratio, 1) AS unit_ratio,
            {line_type_sel},
            {unit_type_sel},
            ii.qty AS quantity,
            ii.buyprice AS unit_price,
            COALESCE(ii.discount, 0) AS discount_amount,
            COALESCE(ii.tax, 0) AS tax_amount,
            COALESCE(ii.subtotal, 0) AS subtotal,
            {disc_pct_sel},
            {tax_pct_sel}
        FROM import_details ii
        INNER JOIN import i ON i.id = ii.import_id
        LEFT JOIN suppliers s ON s.id = i.supplier_id
        LEFT JOIN products p ON p.id = ii.product_id
        WHERE ii.qty > 0
          AND date(i.date) >= date(?)
          AND date(i.date) <= date(?)
    """
    params = [start_date, end_date]
    if search_query:
        sql += f"""
          AND (
                {product_code_search}
             OR COALESCE(ii.product_name, p.name, '') LIKE ?
             OR COALESCE(p.name, '') LIKE ?
             OR COALESCE(i.import_no, 'PN' || printf('%06d', i.id), '') LIKE ?
             OR COALESCE(i.bill_no, '') LIKE ?
             OR COALESCE(s.name, '') LIKE ?
          )
        """
        like = f"%{search_query}%"
        params.extend([like, like, like, like, like, like])

    sql += " ORDER BY i.date DESC, i.id DESC, ii.rowid"
    cursor.execute(sql, params)

    rows = []
    summary = {
        "total_quantity": 0.0,
        "total_discount": 0.0,
        "total_tax": 0.0,
        "total_payment": 0.0,
        "line_count": 0,
    }

    for r in cursor.fetchall():
        row = dict(r)
        unit_type = int(row.get('unit_type') or 0)
        if unit_type == 1 and row.get('wholesale_unit'):
            unit = (row.get('wholesale_unit') or '').strip() or row.get('base_unit') or 'Cái'
        else:
            unit = (row.get('base_unit') or '').strip() or 'Cái'

        amounts = _calc_import_line_amounts(
            row['quantity'],
            row['unit_price'],
            row['discount_amount'],
            row['tax_amount'],
            row['subtotal'],
            row.get('discount_pct'),
            row.get('tax_pct'),
        )
        qty = float(row['quantity'] or 0)
        item = {
            "import_id": row['import_id'],
            "import_date": (row['import_date'] or '')[:10],
            "import_date_display": _format_import_date_display(row['import_date']),
            "import_no": row['import_no'],
            "bill_no": row['bill_no'] or "",
            "supplier_tax_code": row.get('supplier_tax_code') or "",
            "invoice_pdf_url": row.get('invoice_pdf_url') or "",
            "supplier_name": row['supplier_name'] or "",
            "product_code": row['product_code'] or "",
            "product_name": row['product_name'],
            "unit": unit,
            "quantity": qty,
            "unit_price": float(row['unit_price'] or 0),
            "discount_pct": float(row.get('discount_pct') or 0),
            "tax_pct": float(row.get('tax_pct') or 0),
            "discount_amount": amounts["discount_amount"],
            "tax_amount": amounts["tax_amount"],
            "line_total": amounts["line_total"],
        }
        rows.append(item)
        summary["total_quantity"] += qty
        summary["total_discount"] += amounts["discount_amount"]
        summary["total_tax"] += amounts["tax_amount"]
        summary["total_payment"] += amounts["line_total"]
        summary["line_count"] += 1

    return rows, summary


def clean_xml_value(val):
    if not val: return 0
    # Thay dấu phẩy thành dấu chấm và chuyển sang float
    return float(val.replace(',', '.'))
def xml_to_invoice_json(xml_content: str):
    try:
        import xml.etree.ElementTree as ET
        xml_content = xml_content.strip().encode('utf-8').decode('utf-8-sig', errors='replace')
        root = ET.fromstring(xml_content)

        # Tìm NDHDon hoặc DLHDon
        ndhdon = root.find('.//NDHDon') or root.find('.//DLHDon')
        if ndhdon is None:
            raise ValueError("Không tìm thấy NDHDon hoặc DLHDon")

        ttchung = ndhdon.find('.//TTChung')
        nban   = ndhdon.find('.//NBan')

        data = {
            "SHDon": (ttchung.findtext('SHDon') or '').strip(),
            "NLap":  (ttchung.findtext('NLap')  or '').strip(),
            "NBanTen":  (nban.findtext('Ten')   or '').strip(),
            "NBanMST":  (nban.findtext('MST')   or '').strip(),
            "NBanDChi": (nban.findtext('DChi')  or '').strip(),
            "DSHHDVu": []
        }

        for hh in ndhdon.findall('.//HHDVu'):
            try:
                # Xử lý Thuế suất
                vat_str = (hh.findtext('TSuat') or '0').replace('%','').strip()
                vat_rate = float(vat_str) if vat_str.replace('.','').isdigit() else 0
                
                # --- BỔ SUNG CHIẾT KHẤU ---
                # Lấy % chiết khấu (TLCKhau)
                tl_ck_str = (hh.findtext('TLCKhau') or '0').replace('%','').strip()
                tl_ck = float(tl_ck_str) if tl_ck_str.replace('.','').isdigit() else 0
                
                # Lấy số tiền chiết khấu (STCKhau hoặc TienCKhau)
                st_ck = float(hh.findtext('STCKhau', hh.findtext('TienCKhau', '0')) or 0)

                data["DSHHDVu"].append({
                    "THHDVu": (hh.findtext('THHDVu') or '').strip(),
                    "DVTinh": (hh.findtext('DVTinh') or '').strip(),
                    "SLuong": float(hh.findtext('SLuong') or 0),
                    "DGia":   float(hh.findtext('DGia')   or 0),
                    "TSuat":  f"{int(vat_rate)}",
                    "TyLeCK": tl_ck,      # Đồng bộ key với JS
                    "STCKhau": st_ck      # Số tiền chiết khấu
                })
            except:
                continue

        return data
    except Exception as e:
        raise ValueError(f"Lỗi parse XML: {str(e)}")


def _next_import_no_from_db(c, mode='stock'):
    """Sinh số phiếu kế tiếp: PNxxxxxx (nhập kho) hoặc HTxxxxxx (hạch toán dịch vụ)."""
    mode = (mode or 'stock').strip().lower()
    prefix = 'HT' if mode == 'service' else 'PN'

    if mode == 'service':
        c.execute("""
            SELECT MAX(CAST(SUBSTR(import_no, 3) AS INTEGER))
            FROM import
            WHERE import_no LIKE 'HT%'
        """)
    else:
        c.execute("""
            SELECT MAX(CAST(SUBSTR(ref_document, 3) AS INTEGER))
            FROM stock_moves
            WHERE ref_document LIKE 'PN%'
              AND ref_type = 'import'
        """)

    row = c.fetchone()
    max_num = row[0] if row and row[0] is not None else 0
    next_num = max_num + 1
    next_no = f"{prefix}{next_num:06d}"

    while True:
        if mode == 'service':
            c.execute("SELECT 1 FROM import WHERE import_no = ?", (next_no,))
        else:
            c.execute("SELECT 1 FROM stock_moves WHERE ref_document = ?", (next_no,))
        if not c.fetchone():
            break
        next_num += 1
        next_no = f"{prefix}{next_num:06d}"

    return next_no


def peek_next_import_no(mode='stock'):
    """Lấy số phiếu kế tiếp — dùng render trang /import."""
    conn = get_db_connection()
    try:
        c = conn.cursor()
        return _next_import_no_from_db(c, mode)
    finally:
        conn.close()


def register_inventory_routes(app):
    """Đăng ký route kho / nhập hàng (giữ nguyên URL/endpoint)."""
    from auth import admin_or_master_required

    @app.route('/api/parse-xml-invoice', methods=['POST'])
    def parse_xml_invoice():
        try:

            xml_content = None
            data = request.get_json(silent=True) or {}

            # ==================== LẤY XML ====================

            xml_url = data.get('xml_url')

            if xml_url:

                # Nếu là link tương đối
                if xml_url.startswith('/'):
                    base_url = request.host_url.rstrip('/')
                    xml_url = base_url + xml_url

                try:
                    r = requests.get(xml_url, timeout=15)
                    r.raise_for_status()
                    xml_content = r.text
                except requests.exceptions.RequestException as url_err:
                    return jsonify({
                        "success": False,
                        "error": f"Không tải được XML từ link: {str(url_err)}"
                    }), 400

            # upload file
            elif 'file' in request.files:
                file = request.files['file']
                xml_content = file.read().decode('utf-8')

            # xml string
            elif data.get('xml'):
                xml_content = data.get('xml')

            if not xml_content:
                return jsonify({
                    "success": False,
                    "error": "Không có dữ liệu XML"
                }), 400

            # ==================== PARSE XML ====================

            root = ET.fromstring(xml_content)

            ndhdon = root.find('.//NDHDon') or root.find('NDHDon')

            if ndhdon is None:
                return jsonify({
                    "success": False,
                    "error": "Không tìm thấy NDHDon"
                }), 400

            # ==================== NHÀ BÁN ====================

            nban = ndhdon.find('.//NBan')

            supplier = {
                "name": "",
                "tax_code": "",
                "address": ""
            }

            if nban is not None:
                supplier["name"] = nban.findtext('Ten', '').strip()
                supplier["tax_code"] = nban.findtext('MST', '').strip()
                supplier["address"] = nban.findtext('DChi', '').strip()

            # ==================== NGƯỜI MUA ====================

            nmua = ndhdon.find('.//NMua')

            buyer = {
                "name": "",
                "tax_code": "",
                "address": "",
                "contact_name": "",
                "payment_method": ""
            }

            if nmua is not None:

                buyer["name"] = nmua.findtext('Ten', '').strip()
                buyer["tax_code"] = nmua.findtext('MST', '').strip()
                buyer["address"] = nmua.findtext('DChi', '').strip()
                buyer["contact_name"] = nmua.findtext('HVTNMHang', '').strip()

                # Hình thức thanh toán
                for ttin in nmua.findall('.//TTin'):

                    ttruong = ttin.findtext('TTruong', '').lower()

                    if "thanh toán" in ttruong or "payment" in ttruong:
                        buyer["payment_method"] = ttin.findtext('DLieu', '').strip()

            # ==================== THÔNG TIN HÓA ĐƠN ====================

            ttchung = root.find('.//TTChung')

            invoice = {
                "invoice_no": "",
                "invoice_date": ""
            }

            if ttchung is not None:

                raw_date = ttchung.findtext('NLap', '').strip()

                if 'T' in raw_date:
                    raw_date = raw_date.split('T')[0]

                invoice["invoice_no"] = ttchung.findtext('SHDon', '').strip()
                invoice["invoice_date"] = raw_date

    # ==================== DANH SÁCH HÀNG ====================

            items = []

            for hh in ndhdon.findall('.//HHDVu'):

                try:
                    name = hh.findtext('THHDVu', '').strip()
                    unit = hh.findtext('DVTinh', 'Cái').strip()

                    qty = float(hh.findtext('SLuong', '0') or 0)
                    price = float(hh.findtext('DGia', '0') or 0)

                    # --- VAT ---
                    vat_text = hh.findtext('TSuat', '0').replace('%', '').strip()
                    vat_rate = 0
                    if vat_text.replace('.', '').isdigit():
                        vat_rate = float(vat_text)

                    # --- CHIẾT KHẤU (MỚI BỔ SUNG) ---
                    # 1. Lấy tỷ lệ chiết khấu (%)
                    discount_pct_text = hh.findtext('TLCKhau', '0').replace('%', '').strip()
                    discount_pct = 0.0
                    try:
                        if discount_pct_text:
                            discount_pct = float(discount_pct_text)
                    except:
                        discount_pct = 0.0

                    # 2. Lấy số tiền chiết khấu (STCKhau hoặc TienCKhau)
                    discount_amount = float(hh.findtext('STCKhau', hh.findtext('TienCKhau', '0')) or 0)

                    items.append({
                        "name": name,
                        "unit": unit,
                        "quantity": qty,
                        "price": price,
                        "vat_rate": vat_rate,
                        "discount_percent": discount_pct,   # Tỷ lệ %
                        "discount_amount": discount_amount  # Số tiền
                    })

                except Exception:
                    continue

            # ==================== RESPONSE ====================

            return jsonify({
                "success": True,
                "supplier": supplier,
                "buyer": buyer,
                "invoice": invoice,
                "items": items
            })

        except Exception as e:

            import traceback
            traceback.print_exc()

            return jsonify({
                "success": False,
                "error": str(e)
            }), 500

    import requests
    from flask import Response

    #===API HỖ TRỢ TẢI XML ĐIỀN FORM XUẤT HÓA ĐƠN THAY THẾ===#
    @app.route("/api/proxy/xml")
    def proxy_xml():
        url = request.args.get("url")

        if not url:
            return {"success": False, "error": "Missing URL"}, 400

        try:
            r = requests.get(url, timeout=20)

            return Response(
                r.content,
                content_type="text/xml"
            )

        except Exception as e:
            return {"success": False, "error": str(e)}, 500

    #===API TẢI JSON TẠO PHIẾU NHẬP KHO TỰ ĐỘNG TỪ TRANG QUẢN LÝ HÓA ĐƠN ĐẦU VÀO===#
    def xml_to_invoice_json(xml_content: str):
        try:
            import xml.etree.ElementTree as ET
            xml_content = xml_content.strip().encode('utf-8').decode('utf-8-sig', errors='replace')
            root = ET.fromstring(xml_content)

            # Tìm NDHDon hoặc DLHDon
            ndhdon = root.find('.//NDHDon') or root.find('.//DLHDon')
            if ndhdon is None:
                raise ValueError("Không tìm thấy NDHDon hoặc DLHDon")

            ttchung = ndhdon.find('.//TTChung')
            nban   = ndhdon.find('.//NBan')

            data = {
                "SHDon": (ttchung.findtext('SHDon') or '').strip(),
                "NLap":  (ttchung.findtext('NLap')  or '').strip(),
                "NBanTen":  (nban.findtext('Ten')   or '').strip(),
                "NBanMST":  (nban.findtext('MST')   or '').strip(),
                "NBanDChi": (nban.findtext('DChi')  or '').strip(),
                "DSHHDVu": []
            }

            for hh in ndhdon.findall('.//HHDVu'):
                try:
                    # Xử lý Thuế suất
                    vat_str = (hh.findtext('TSuat') or '0').replace('%','').strip()
                    vat_rate = float(vat_str) if vat_str.replace('.','').isdigit() else 0
                
                    # --- BỔ SUNG CHIẾT KHẤU ---
                    # Lấy % chiết khấu (TLCKhau)
                    tl_ck_str = (hh.findtext('TLCKhau') or '0').replace('%','').strip()
                    tl_ck = float(tl_ck_str) if tl_ck_str.replace('.','').isdigit() else 0
                
                    # Lấy số tiền chiết khấu (STCKhau hoặc TienCKhau)
                    st_ck = float(hh.findtext('STCKhau', hh.findtext('TienCKhau', '0')) or 0)

                    data["DSHHDVu"].append({
                        "THHDVu": (hh.findtext('THHDVu') or '').strip(),
                        "DVTinh": (hh.findtext('DVTinh') or '').strip(),
                        "SLuong": float(hh.findtext('SLuong') or 0),
                        "DGia":   float(hh.findtext('DGia')   or 0),
                        "TSuat":  f"{int(vat_rate)}",
                        "TyLeCK": tl_ck,      # Đồng bộ key với JS
                        "STCKhau": st_ck      # Số tiền chiết khấu
                    })
                except:
                    continue

            return data
        except Exception as e:
            raise ValueError(f"Lỗi parse XML: {str(e)}")

    import json
    import sqlite3
    from flask import jsonify

    @app.route('/api/invoice/<int:invoice_id>/xml')
    def get_invoice_xml(invoice_id):
        conn = None
        try:
            # ===== CONNECT DB =====
            conn = get_db_connection()
            conn.row_factory = sqlite3.Row
            c = conn.cursor()

            # ===== QUERY =====
            c.execute("""
                SELECT * 
                FROM supplier_invoice 
                WHERE id = ?
            """, (invoice_id,))
        
            row = c.fetchone()

            # ===== NOT FOUND =====
            if row is None:
                return jsonify({
                    "success": False,
                    "error": f"Không tìm thấy hóa đơn với id = {invoice_id}"
                }), 404

            # ===== CHECK FIELD =====
            raw_data = row['xml_data']
            if not raw_data:
                return jsonify({
                    "success": False,
                    "error": "xml_data rỗng"
                }), 400

            # ===== AUTO DETECT JSON / XML =====
            raw_data = raw_data.strip()

            # ===== SUCCESS =====
            from Services.inward_invoice_helpers import normalize_supplier_invoice_payload
            try:
                data = normalize_supplier_invoice_payload(raw_data)
            except ValueError as norm_err:
                if isinstance(raw_data, str) and raw_data.strip().startswith(('{', '[')):
                    data = json.loads(raw_data)
                else:
                    return jsonify({'success': False, 'error': str(norm_err)}), 400

            return jsonify({
                'success': True,
                'data': data
            })

        except Exception as e:
            import traceback
            print("🔥 ERROR /api/invoice/<id>/xml:")
            print(traceback.format_exc())

            return jsonify({
                "success": False,
                "error": str(e)
            }), 500

        finally:
            if conn:
                conn.close()

    @app.route('/api/import', methods=['POST'])
    @login_required
    def api_import_post():
        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
    
        try:
            from Services.inward_invoice_helpers import ensure_import_service_schema
            from Services.import_line_helpers import (
                ensure_warehouse_schema,
                insert_import_detail_row,
                tracks_retail_inventory,
            )
            from Services.fixed_assets_helpers import (
                ensure_fixed_assets_schema,
                register_fixed_asset_from_import,
                register_tool_from_import,
            )
            ensure_import_service_schema(conn)
            ensure_warehouse_schema(conn)
            ensure_fixed_assets_schema(conn)
            conn.commit()

            data = request.get_json()
            if not data:
                return jsonify({"error": "Dữ liệu không hợp lệ"}), 400

            # --- 1. LẤY DỮ LIỆU ĐẦU VÀO ---
            items = data.get('items', [])
            supplier_id = data.get('supplier_id')
            import_date = data.get('date')
            bill_date = data.get('bill_date')
            import_no = data.get('import_no')
            bill_no = data.get('bill_no')
            tax_code = data.get('tax_code')
            note = data.get('note')
            extra_cost = Decimal(str(data.get('extra_cost', 0) or 0))
            payment_status_input = data.get('payment_status', 'Chưa thanh toán')
            from_invoice_id = data.get('from_invoice_id')
            payment_method = data.get('payment_method', 'cash')
            default_warehouse = (data.get('warehouse_code') or 'KHO_001').strip()

            # Lấy thông tin nhà cung cấp
            c.execute("SELECT name, address FROM suppliers WHERE id = ?", (supplier_id,))
            sup_row = c.fetchone()
            supplier_name = sup_row['name'] if sup_row else f"NCC ID {supplier_id}"
            supplier_address = sup_row['address'] if sup_row and sup_row['address'] else ""

            # --- 2. TÍNH TỔNG GIÁ TRỊ GỐC ĐỂ PHÂN BỔ CHI PHÍ ---
            total_base = sum(
                Decimal(str(i.get('qty', 0) or 0)) * Decimal(str(i.get('buyprice', 0) or 0))
                for i in items
            )
            total_base_safe = total_base if total_base > 0 else Decimal('1')

            c.execute('PRAGMA table_info(import)')
            import_cols = {col[1] for col in c.fetchall()}
            import_fields = [
                'date', 'supplier_id', 'import_no', 'bill_no', 'bill_date',
                'note', 'payment_status', 'extra_cost', 'total_value', 'paid_amount',
            ]
            import_values = [
                import_date, supplier_id, import_no, bill_no, bill_date,
                note, payment_status_input, float(extra_cost), 0, 0,
            ]
            if 'warehouse_code' in import_cols:
                import_fields.append('warehouse_code')
                import_values.append(default_warehouse)
            if 'from_invoice_id' in import_cols and from_invoice_id:
                import_fields.append('from_invoice_id')
                import_values.append(int(from_invoice_id))

            placeholders = ', '.join(['?'] * len(import_fields))
            c.execute(
                f'INSERT INTO import ({", ".join(import_fields)}) VALUES ({placeholders})',
                import_values,
            )
            import_id = c.lastrowid

            c.execute('PRAGMA table_info(stock_moves)')
            sm_has_wh = 'warehouse_code' in {col[1] for col in c.fetchall()}

            # --- 4. TỐI ƯU: LẤY THÔNG TIN SẢN PHẨM TRƯỚC (BATCH SELECT) ---
            p_ids = [i.get('product_id') for i in items if i.get('product_id')]
            p_map = {}
            if p_ids:
                placeholders_p = ','.join(['?'] * len(p_ids))
                c.execute(
                    f"SELECT id, name, unit, unit1, unit_ratio, barcode, product_code FROM products WHERE id IN ({placeholders_p})",
                    p_ids,
                )
                p_map = {row['id']: row for row in c.fetchall()}

            # --- 5. XỬ LÝ CHI TIẾT TỪNG DÒNG ---
            total_invoice_value = Decimal('0')
            items_for_json = []

            for item in items:
                line_type = (item.get('line_type') or item.get('product_type') or 'goods').strip().lower()
                warehouse_code = (item.get('warehouse_code') or default_warehouse or 'KHO_001').strip()
                qty_in = Decimal(str(item.get('qty', 0) or 0))
                if qty_in <= 0:
                    continue

                price_in = Decimal(str(item.get('buyprice', 0) or 0))
                tax_p = Decimal(str(item.get('tax_pct', 0) or 0))
                disc_p = Decimal(str(item.get('discountPct', 0) or item.get('discount_pct', 0) or 0))
                unit_in = str(item.get('unit', '') or 'Cái').strip()
                inv_name = (item.get('invoice_name') or item.get('invoice_name_hidden') or item.get('name') or '').strip()

                line_subtotal = qty_in * price_in
                line_disc = line_subtotal * (disc_p / Decimal('100'))
                line_after_disc = line_subtotal - line_disc
                line_tax = line_after_disc * (tax_p / Decimal('100'))
                line_extra = (line_subtotal / total_base_safe) * extra_cost
                line_total = line_after_disc + line_tax + line_extra
                total_invoice_value += line_total

                if line_type == 'service':
                    insert_import_detail_row(c, import_id, {
                        'import_id': import_id,
                        'product_id': None,
                        'qty': float(qty_in),
                        'buyprice': float(price_in),
                        'subtotal': float(line_subtotal),
                        'discount': float(line_disc),
                        'tax': float(line_tax),
                        'cost_price': float(line_total / qty_in) if qty_in else 0,
                        'tax_pct': float(tax_p),
                        'discount_pct': float(disc_p),
                        'payment_amt': float(line_total),
                        'product_name': inv_name or item.get('name') or '',
                        'unit': unit_in,
                        'line_type': 'service',
                        'warehouse_code': warehouse_code,
                    })
                    items_for_json.append({
                        'product_name': inv_name or item.get('name') or '',
                        'unit': unit_in,
                        'qty': float(qty_in),
                        'buyprice': float(price_in),
                        'line_type': 'service',
                        'line_total': float(line_total),
                    })
                    continue

                pid = item.get('product_id')
                p_info = p_map.get(pid)
                if not p_info:
                    continue

                if inv_name:
                    c.execute("""
                        INSERT OR IGNORE INTO product_aliases (product_id, invoice_name, supplier_id)
                        VALUES (?, ?, ?)
                    """, (pid, inv_name.strip(), supplier_id))

                unit_in_lower = unit_in.lower()
                product_name = p_info['name']
                retail_unit = str(p_info['unit'] or 'Cái').strip()
                wholesale_unit = str(p_info['unit1'] or '').strip().lower()
                ratio = Decimal(str(p_info['unit_ratio'] or 1))

                is_wholesale = wholesale_unit and unit_in_lower == wholesale_unit
                qty_retail = qty_in * ratio if is_wholesale else qty_in
                cost_per_retail = line_total / qty_retail if qty_retail > 0 else Decimal('0')

                items_for_json.append({
                    'product_id': pid,
                    'product_name': product_name,
                    'barcode': p_info['barcode'] or '',
                    'unit': item.get('unit'),
                    'qty': float(qty_in),
                    'buyprice': float(price_in),
                    'discount_pct': float(disc_p),
                    'tax_pct': float(tax_p),
                    'line_type': line_type,
                    'warehouse_code': warehouse_code,
                    'line_total': float(line_total),
                })

                insert_import_detail_row(c, import_id, {
                    'import_id': import_id,
                    'product_id': pid,
                    'qty': float(qty_in),
                    'buyprice': float(price_in),
                    'subtotal': float(line_subtotal),
                    'discount': float(line_disc),
                    'tax': float(line_tax),
                    'cost_price': float(cost_per_retail),
                    'unit_type': 1 if is_wholesale else 0,
                    'tax_pct': float(tax_p),
                    'discount_pct': float(disc_p),
                    'payment_amt': float(line_total),
                    'product_name': product_name,
                    'unit': unit_in,
                    'line_type': line_type,
                    'warehouse_code': warehouse_code,
                })
                detail_id = c.lastrowid

                if line_type == 'fixed_asset':
                    register_fixed_asset_from_import(
                        c,
                        import_id=import_id,
                        import_detail_id=detail_id,
                        product_id=pid,
                        product_code=p_info['product_code'] or '',
                        product_name=product_name,
                        import_no=import_no,
                        import_date=import_date,
                        warehouse_code=warehouse_code,
                        qty=float(qty_in),
                        buyprice=float(price_in),
                        tax_amount=float(line_tax),
                        discount_amount=float(line_disc),
                        line_total=float(line_total),
                        subtotal=float(line_subtotal),
                    )
                elif line_type == 'tools':
                    register_tool_from_import(
                        c,
                        import_id=import_id,
                        import_detail_id=detail_id,
                        product_id=pid,
                        product_code=p_info['product_code'] or '',
                        product_name=product_name,
                        import_no=import_no,
                        import_date=import_date,
                        warehouse_code=warehouse_code,
                        qty=float(qty_in),
                        buyprice=float(price_in),
                        tax_amount=float(line_tax),
                        line_total=float(line_total),
                        subtotal=float(line_subtotal),
                        discount_amount=float(line_disc),
                    )

                params_detail = (
                    import_id, pid, float(qty_in), float(price_in), float(line_subtotal),
                    float(line_disc), float(line_tax), float(cost_per_retail),
                    1 if is_wholesale else 0, float(tax_p), float(disc_p),
                )

                if tracks_retail_inventory(line_type):
                    c.execute(
                        "INSERT INTO chi_tiet_phieu_nhap_kho (import_id, product_id, quantity, buyprice, subtotal, discount_amount, tax_amount, cost_price, unit_type, tax_pct, discount_pct) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                        params_detail,
                    )
                    apply_wac_inbound(c, pid, float(qty_retail), float(line_total))
                    move_note = f"Nhập từ {supplier_name} (Gốc: {qty_in} {unit_in}) — {warehouse_code}"
                    if sm_has_wh:
                        c.execute("""
                            INSERT INTO stock_moves (product_id, date, type, ref_id, quantity, cost_price, note, ref_document, ref_type, type1, unit, unit1, unit_ratio, warehouse_code)
                            VALUES (?, ?, 'import', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (pid, import_date, import_id, float(qty_retail), float(cost_per_retail), move_note, import_no, 'import', 'Nhập', retail_unit, wholesale_unit, float(ratio), warehouse_code))
                    else:
                        c.execute("""
                            INSERT INTO stock_moves (product_id, date, type, ref_id, quantity, cost_price, note, ref_document, ref_type, type1, unit, unit1, unit_ratio)
                            VALUES (?, ?, 'import', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (pid, import_date, import_id, float(qty_retail), float(cost_per_retail), move_note, import_no, 'import', 'Nhập', retail_unit, wholesale_unit, float(ratio)))
                    c.execute("""
                        INSERT INTO inventory_transactions
                        (product_id, type, type1, quantity, cost_price, reference_id, reference_type, note, created_at)
                        VALUES (?, 'import', 'Nhập', ?, ?, ?, 'import', ?, ?)
                    """, (pid, float(qty_retail), float(cost_per_retail), import_id, f"Nhập kho - PN#{import_no} ({warehouse_code})", import_date))
                    sync_inventory_quantity_from_moves(c, pid)

            # --- 6. CẬP NHẬT HEADER & LƯU PHIẾU IN ---
            total_final_float = float(total_invoice_value)
            final_paid = total_final_float if payment_status_input == 'Đã thanh toán' else 0.0
            c.execute("UPDATE import SET total_value = ?, paid_amount = ? WHERE id = ?", (total_final_float, final_paid, import_id))

            items_json_str = json.dumps(items_for_json, ensure_ascii=False)
            c.execute("""
                INSERT INTO phieu_nhap_kho (import_no, date, bill_no, bill_date, supplier_name, supplier_id, items_json, total_amount, import_id, note)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (import_no, import_date, bill_no, bill_date, supplier_name, supplier_id, items_json_str, total_final_float, import_id, note))

            # --- 7. MỘT PHIẾU CHI cho cả phiếu nhập (HH + VT + DV + TSCĐ + CCDC) ---
            res_pc_vouch = None
            if payment_status_input == 'Đã thanh toán':
                from Services.inward_invoice_helpers import next_pc_voucher_no
                res_pc_vouch = next_pc_voucher_no(c)
                credit_acc = '111' if payment_method == 'cash' else '112'
                bill_ref = bill_no or import_no
                c.execute("""
                    INSERT INTO phieu_chi (voucher_no, receiver_name, address, amount, credit_account, debit_account, reason, source_type, reference_document, source_id, preparer, date)
                    VALUES (?, ?, ?, ?, ?, '331', ?, 'import', ?, ?, ?, ?)
                """, (
                    res_pc_vouch, supplier_name, supplier_address, final_paid, credit_acc,
                    f'Thanh toán phiếu nhập {import_no}' + (f' (HĐ {bill_ref})' if bill_ref else ''),
                    bill_ref, import_id, session.get('user_name', 'Admin'), import_date,
                ))

            # --- 8. CẬP NHẬT TRẠNG THÁI HÓA ĐƠN NCC (NẾU CÓ) ---
            # Chuyển đổi thành chuỗi, loại bỏ khoảng trắng và viết thường để kiểm tra an toàn
            bill_no_clean = str(bill_no).strip() if bill_no else ""
        
            # Chỉ xử lý nếu bill_no hợp lệ (không rỗng, không phải chữ 'none' hoặc 'nan')
            if bill_no_clean and bill_no_clean.lower() not in ['none', 'nan']:
                c.execute("""
                    UPDATE supplier_invoice 
                    SET status = 'imported' 
                    WHERE invoice_no = ? AND seller_tax_code = ? AND status != 'imported'
                """, (bill_no_clean, tax_code))

            conn.commit()
            return jsonify({"success": True, "import_id": import_id, "voucher_no": import_no, "phieu_chi_voucher": res_pc_vouch, "paid_amount": final_paid})

        except Exception as e:
            conn.rollback()
            logging.error(f"LỖI NHẬP KHO: {str(e)}", exc_info=True)
            return jsonify({"error": f"Lỗi hệ thống: {str(e)}"}), 500
        finally:
            conn.close()

    @app.route('/api/import/service', methods=['POST'])
    @login_required
    def api_import_service_post():
        """Hạch toán mua dịch vụ HKD: import_details (không kho).
        Chưa thanh toán: không lập phiếu chi (công nợ N642/C331 theo import).
        Đã thanh toán: một phiếu chi N642/C111|112."""
        from Services.inward_invoice_helpers import (
            ensure_import_service_schema,
            import_details_allows_null_product_id,
            next_pc_voucher_no,
        )

        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        c = conn.cursor()

        try:
            ensure_import_service_schema(conn)
            conn.commit()

            if not import_details_allows_null_product_id(conn):
                return jsonify({
                    'success': False,
                    'error': 'Cần cập nhật cấu trúc DB: import_details.product_id phải cho phép NULL. '
                             'Vui lòng khởi động lại ứng dụng hoặc liên hệ quản trị.',
                }), 500

            data = request.get_json(silent=True) or {}
            items = data.get('items') or []
            supplier_id = data.get('supplier_id')
            import_date = data.get('date')
            bill_date = data.get('bill_date')
            import_no = (data.get('import_no') or '').strip()
            bill_no = (data.get('bill_no') or '').strip()
            tax_code = (data.get('tax_code') or '').strip()
            note = (data.get('note') or '').strip() or 'Hạch toán chi phí dịch vụ'
            extra_cost = Decimal(str(data.get('extra_cost', 0) or 0))
            payment_status_input = data.get('payment_status', 'Chưa thanh toán')
            from_invoice_id = data.get('from_invoice_id')
            payment_method = data.get('payment_method', 'cash')

            if not supplier_id:
                return jsonify({'success': False, 'error': 'Thiếu nhà cung cấp'}), 400
            if not import_date or not import_no:
                return jsonify({'success': False, 'error': 'Thiếu ngày hoặc số phiếu'}), 400
            if not items:
                return jsonify({'success': False, 'error': 'Vui lòng nhập ít nhất một dòng dịch vụ'}), 400

            c.execute('SELECT name, address FROM suppliers WHERE id = ?', (supplier_id,))
            sup_row = c.fetchone()
            supplier_name = sup_row['name'] if sup_row else f'NCC ID {supplier_id}'
            supplier_address = sup_row['address'] if sup_row and sup_row['address'] else ''

            c.execute('PRAGMA table_info(import)')
            import_cols = {col[1] for col in c.fetchall()}
            c.execute('PRAGMA table_info(import_details)')
            detail_cols = {col[1] for col in c.fetchall()}

            import_fields = [
                'date', 'supplier_id', 'import_no', 'bill_no', 'bill_date',
                'note', 'payment_status', 'extra_cost', 'total_value', 'paid_amount',
            ]
            import_values = [
                import_date, supplier_id, import_no, bill_no, bill_date,
                note, payment_status_input, float(extra_cost), 0, 0,
            ]
            if 'doc_type' in import_cols:
                import_fields.append('doc_type')
                import_values.append('service')
            if 'from_invoice_id' in import_cols and from_invoice_id:
                import_fields.append('from_invoice_id')
                import_values.append(int(from_invoice_id))

            placeholders = ', '.join(['?'] * len(import_fields))
            c.execute(
                f"INSERT INTO import ({', '.join(import_fields)}) VALUES ({placeholders})",
                import_values,
            )
            import_id = c.lastrowid

            total_invoice_value = Decimal('0')
            for item in items:
                name = (item.get('name') or item.get('invoice_name') or '').strip()
                qty_in = Decimal(str(item.get('qty', 0) or 0))
                if not name or qty_in <= 0:
                    continue

                price_in = Decimal(str(item.get('buyprice', 0) or 0))
                tax_p = Decimal(str(item.get('tax_pct', 0) or 0))
                disc_p = Decimal(str(item.get('discountPct', 0) or item.get('discount_pct', 0) or 0))
                unit_in = (item.get('unit') or 'Lần').strip()

                line_subtotal = qty_in * price_in
                line_disc = line_subtotal * (disc_p / Decimal('100'))
                line_after_disc = line_subtotal - line_disc
                line_tax = line_after_disc * (tax_p / Decimal('100'))
                line_total = line_after_disc + line_tax
                total_invoice_value += line_total

                per_unit_cost = line_total / qty_in if qty_in > 0 else Decimal('0')

                cols = ['import_id', 'qty', 'buyprice', 'subtotal', 'discount', 'tax', 'cost_price']
                vals = [
                    import_id, float(qty_in), float(price_in), float(line_subtotal),
                    float(line_disc), float(line_tax), float(per_unit_cost),
                ]
                if 'product_id' in detail_cols:
                    pid = item.get('product_id')
                    if pid is not None and str(pid).strip() not in ('', '0'):
                        cols.append('product_id')
                        vals.append(int(pid))
                    else:
                        cols.append('product_id')
                        vals.append(None)
                if 'unit_type' in detail_cols:
                    cols.append('unit_type')
                    vals.append(0)
                if 'tax_pct' in detail_cols:
                    cols.append('tax_pct')
                    vals.append(float(tax_p))
                if 'discount_pct' in detail_cols:
                    cols.append('discount_pct')
                    vals.append(float(disc_p))
                if 'payment_amt' in detail_cols:
                    cols.append('payment_amt')
                    vals.append(float(line_total))
                if 'product_name' in detail_cols:
                    cols.append('product_name')
                    vals.append(name)
                if 'unit' in detail_cols:
                    cols.append('unit')
                    vals.append(unit_in)
                if 'line_type' in detail_cols:
                    cols.append('line_type')
                    vals.append('service')

                ph = ', '.join(['?'] * len(vals))
                c.execute(
                    f"INSERT INTO import_details ({', '.join(cols)}) VALUES ({ph})",
                    vals,
                )

            if total_invoice_value <= 0:
                conn.rollback()
                return jsonify({'success': False, 'error': 'Tổng giá trị phiếu phải lớn hơn 0'}), 400

            total_final = total_invoice_value + extra_cost
            total_final_float = float(total_final)
            is_paid = payment_status_input == 'Đã thanh toán'
            final_paid = total_final_float if is_paid else 0.0
            c.execute(
                'UPDATE import SET total_value = ?, paid_amount = ? WHERE id = ?',
                (total_final_float, final_paid, import_id),
            )

            preparer = session.get('user_name', 'Admin')
            res_pc_vouch = None
            if is_paid:
                res_pc_vouch = next_pc_voucher_no(c)
                credit_acc = '111' if payment_method == 'cash' else '112'
                c.execute("""
                    INSERT INTO phieu_chi (
                        voucher_no, receiver_name, address, amount,
                        debit_account, credit_account,
                        expense_type, reason, source_type, reference_document, source_id, preparer, date
                    ) VALUES (?, ?, ?, ?, '642', ?, 'CP_DV', ?, 'import_service', ?, ?, ?, ?)
                """, (
                    res_pc_vouch, supplier_name, supplier_address, total_final_float,
                    credit_acc,
                    f'Thanh toán tiền mua dịch vụ số {import_no}',
                    bill_no or import_no, import_id, preparer, import_date,
                ))

            bill_no_clean = bill_no.strip() if bill_no else ''
            if bill_no_clean and bill_no_clean.lower() not in ('none', 'nan'):
                c.execute("""
                    UPDATE supplier_invoice
                    SET status = 'accounted'
                    WHERE invoice_no = ? AND seller_tax_code = ? AND status NOT IN ('imported', 'accounted')
                """, (bill_no_clean, tax_code))
                if from_invoice_id and c.rowcount == 0:
                    c.execute(
                        "UPDATE supplier_invoice SET status = 'accounted' WHERE id = ?",
                        (int(from_invoice_id),),
                    )

            conn.commit()
            return jsonify({
                'success': True,
                'import_id': import_id,
                'voucher_no': import_no,
                'phieu_chi_voucher': res_pc_vouch,
                'paid_amount': final_paid,
            })

        except Exception as e:
            conn.rollback()
            logging.error('LỖI HẠCH TOÁN DỊCH VỤ: %s', e, exc_info=True)
            return jsonify({'success': False, 'error': str(e)}), 500
        finally:
            conn.close()

    #=== API TÌM TÊN SẢN PHẨM LIÊN KẾT - HÀNG CÓ TÊN TRÊN HÓA ĐƠN KHÁC VỚI TÊN TRÊN POS ===#
    @app.route('/api/products/smart-link', methods=['POST'])
    def smart_link():
        """Tìm kiếm sản phẩm dựa trên tên hóa đơn lạ"""
        data = request.json
        inv_name = data.get('invoice_name', '').strip()
        supplier_id = data.get('supplier_id')
    
        conn = get_db_connection()
        # 1. Tìm chính xác trong bảng Alias
        alias = conn.execute(
            "SELECT product_id FROM product_aliases WHERE invoice_name = ? AND supplier_id = ?", 
            (inv_name, supplier_id)
        ).fetchone()
    
        if alias:
            return jsonify({'success': True, 'match_type': 'EXACT', 'product_id': alias['product_id']})

        # 2. Gợi ý dựa trên mã số kỹ thuật (ví dụ: 63174)
        codes = re.findall(r'\d{4,6}', inv_name)
        if codes:
            suggested = conn.execute(
                "SELECT id, name FROM products WHERE name LIKE ?", (f'%{codes[0]}%',)
            ).fetchone()
            if suggested:
                return jsonify({
                    'success': True, 
                    'match_type': 'SUGGEST', 
                    'product_id': suggested['id'], 
                    'product_name': suggested['name']
                })

        return jsonify({'success': False, 'match_type': 'NONE'})

    #===HÀM ĐƠN VỊ SỈ, LẺ===#
    def suggest_wholesale_unit(unit):
        unit = (unit or '').lower()
        mapping = {
            'Lon': ('Thùng', 24),
            'Chai': ('Thùng', 24),
            'M': ('Cuộn', 100),
            'hộp': ('Thùng', 12),
            'cái': ('Thùng', 10),
            'kg': ('Bao', 25),
            'Lít': ('Thùng', 20),
            'Viên': ('Lốc', 10),
            'M': ('Cây', 4)
        }
        for key, (w_unit, ratio) in mapping.items():
            if key in unit:
                return {"unit": w_unit, "ratio": ratio}
        return {"unit": '', "ratio": 1}

    # === VIEW IMPORT POS (chi tiết) ===
    @app.route('/import/view/<int:import_id>')
    # @login_required  # Bỏ comment nếu cần đăng nhập
    # @admin_or_master_required  # Nếu chỉ admin mới xem được
    def view_import(import_id):
        try:
            from Services.import_line_helpers import prepare_import_edit_json

            imp = get_import_detail(import_id)
            if not imp:
                abort(404, description="Phiếu nhập không tồn tại.")

            view_items = []
            for raw in imp.get('items') or []:
                item = dict(raw)
                item['display_unit'] = str(
                    item.get('unit') or item.get('invoice_unit') or item.get('base_unit') or 'Cái'
                ).strip() or 'Cái'
                item['discount_amount'] = float(item.get('discount') or 0)
                item['tax_amount'] = float(item.get('tax') or 0)
                qty = float(item.get('qty') or 0)
                price = float(item.get('buyprice') or 0)
                line_total = float(item.get('line_total') or (qty * price))
                subtotal = float(item.get('subtotal') or (line_total - item['discount_amount']))
                item['line_total'] = line_total
                item['subtotal'] = subtotal
                item['payment_amount'] = float(
                    item.get('payment_amount') or (subtotal + item['tax_amount'])
                )
                view_items.append(item)

            payload = prepare_import_edit_json({**imp, 'items': view_items})

            return render_template(
                'import_view.html',
                imp=payload,
                items=payload.get('items') or [],
                is_service_import=bool(imp.get('is_service_import')),
            )

        except Exception as e:
            import traceback
            traceback.print_exc()
            print(f"Lỗi xem phiếu nhập ID {import_id}: {str(e)}")
            abort(500, description="Lỗi xử lý dữ liệu phiếu nhập.")

    # ======= IN Phiếu Nhập Kho 03_VT ======#
    @app.route('/DanhSachPhieuNhapKho')
    def DanhSachPhieuNhapKho():
        return render_template('KeToanHKD/DanhSachPhieuNhapKho.html')

    #====PHIẾU NHẬP KHO THEO HÓA ĐƠN NHẬP KHO TRÊN POS====#
    @app.route('/print_import/<int:import_id>')
    def print_import(import_id):
        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        c = conn.cursor()

        try:
            # 1. Lấy thông tin hộ kinh doanh (MST, Tên, Địa chỉ)
            c.execute("SELECT * FROM business_info LIMIT 1")
            business_info = c.fetchone()

            # 2. Lấy thông tin phiếu nhập
            c.execute("""
                SELECT i.*, COALESCE(s.name, 'Nhà cung cấp') AS supplier_name
                FROM import i
                LEFT JOIN suppliers s ON s.id = i.supplier_id
                WHERE i.id = ?
            """, (import_id,))
            imp = c.fetchone()

            if not imp:
                return "Không tìm thấy phiếu", 404

            # 3. Lấy chi tiết hàng hóa - ĐÃ FIX ĐỂ HIỂN THỊ ĐÚNG ĐƠN VỊ GỐC
            # Kiểm tra cột tồn tại trong import_details
            c.execute("PRAGMA table_info(import_details)")
            details_cols = {col[1] for col in c.fetchall()}

            has_unit = 'unit' in details_cols
            has_unit_type = 'unit_type' in details_cols

            select_fields = [
                "id.*",
                "p.name AS product_name",
                "p.unit AS base_unit",
                "p.unit1 AS wholesale_unit",
                "p.barcode",
                "p.product_code"
            ]

            # Thêm cột unit và unit_type nếu tồn tại
            if has_unit:
                select_fields.append("id.unit AS import_unit")
            if has_unit_type:
                select_fields.append("id.unit_type")

            query = f"""
                SELECT {', '.join(select_fields)}
                FROM import_details id
                JOIN products p ON p.id = id.product_id
                WHERE id.import_id = ?
            """
            c.execute(query, (import_id,))
            raw_items = c.fetchall()

            # Xử lý đơn vị hiển thị cho từng sản phẩm
            items = []
            for row in raw_items:
                item = dict(row)

                # Xác định đơn vị in ra (đồng bộ với return_import)
                if has_unit and item.get('import_unit'):
                    item['display_unit'] = item['import_unit'].strip() or item['base_unit'] or '—'
                elif has_unit_type and item.get('unit_type') == 1 and item.get('wholesale_unit'):
                    item['display_unit'] = item['wholesale_unit'].strip() or '—'
                else:
                    item['display_unit'] = item.get('base_unit', '—')

                # Có thể thêm các trường khác nếu cần
                items.append(item)

            conn.close()

            return render_template(
                'KeToanHKD/PhieuNhapKho_printpos.html',
                imp=imp,
                items=items,
            )

        except Exception as e:
            if conn:
                conn.close()
            print(f"ERROR in print_import {import_id}: {str(e)}")
            return f"Lỗi in phiếu: {str(e)}", 500

    # === BỔ SUNG: RESET SỐ PHIẾU NHẬP KẾ TOÁN (ADMIN ONLY)  ===
    @app.route('/api/import/reset_sequence', methods=['POST'])
    @login_required
    def reset_import_sequence():
        conn = get_db_connection()
        c = conn.cursor()
        try:
            c.execute("UPDATE import_sequence SET current_seq = 0 WHERE id = 1")
            conn.commit()
            return jsonify({"success": True, "message": "Đã reset số phiếu nhập về NK000001"})
        except Exception as e:
            conn.rollback()
            return jsonify({"error": str(e)}), 500
        finally:
            conn.close()

    # --- API XỬ LÝ TRẢ HÀNG NHẬP (TẠO PHIẾU TRẢ) ---
    import json
    import traceback
    from flask import jsonify, request, session
    from datetime import datetime

    @app.route('/api/return/import', methods=['POST'])
    @login_required
    def api_return_import_post():
        """
        [API] Xử lý trả hàng nhập - Đồng bộ đơn vị, tồn kho
        Bổ sung: Tự động tạo Phiếu Thu (PT) và Phiếu Xuất Kho (PX)
        """
        data = request.get_json() or {}
        required_keys = ['import_id', 'product_id', 'quantity', 'reason', 'date']
    
        # Kiểm tra thiếu trường
        for k in required_keys:
            if k not in data:
                return jsonify({"error": f"Thiếu trường bắt buộc: {k}"}), 400

        try:
            import_id = int(data['import_id'])
            product_id = int(data['product_id'])
            return_qty_input = float(data['quantity'])
            reason = data['reason'].strip()
            return_date = data['date']
        except (ValueError, TypeError, KeyError):
            return jsonify({"error": "Dữ liệu đầu vào không hợp lệ"}), 400

        if return_qty_input <= 0:
            return jsonify({"error": "Số lượng trả phải lớn hơn 0"}), 400

        conn = get_db_connection()
        c = conn.cursor()
        try:
            # 1. Lấy thông tin sản phẩm và đơn vị quy đổi
            c.execute("""
                SELECT name, unit AS base_unit, unit1 AS wholesale_unit, unit_ratio, barcode
                FROM products WHERE id = ?
            """, (product_id,))
            prod = c.fetchone()
            if not prod:
                return jsonify({"error": "Sản phẩm không tồn tại"}), 404

            ratio = float(prod['unit_ratio'] or 1)
            product_name = prod['name']
            product_barcode = prod['barcode'] or str(product_id)

            # 2. Kiểm tra phiếu nhập và lấy thông tin nhà cung cấp
            c.execute("""
                SELECT i.supplier_id, s.name as supplier_name, s.address as supplier_address, s.tax_code
                FROM import i
                LEFT JOIN suppliers s ON i.supplier_id = s.id
                WHERE i.id = ?
            """, (import_id,))
            imp_sup = c.fetchone()
            if not imp_sup:
                return jsonify({"error": "Phiếu nhập không tồn tại"}), 404

            supplier_name = imp_sup['supplier_name'] or 'NCC Không xác định'
            supplier_address = imp_sup['supplier_address'] or ''
            tax_code = imp_sup['tax_code'] or ''

            # 3. Lấy thông tin dòng nhập gốc (chỉ đọc — không sửa chứng từ gốc)
            c.execute("""
                SELECT qty, unit_type, cost_price, buyprice,
                       COALESCE(subtotal, 0) AS subtotal,
                       COALESCE(tax, 0) AS tax_amount,
                       COALESCE(discount, 0) AS discount_amount
                FROM import_details
                WHERE import_id = ? AND product_id = ?
            """, (import_id, product_id))
            item = c.fetchone()
            if not item:
                return jsonify({"error": "Sản phẩm không có trong phiếu nhập"}), 404

            original_qty = float(item['qty'])
            is_wholesale_entry = bool(item['unit_type'])
            cost_price_base = import_cost_to_base(item['cost_price'], item['unit_type'], ratio)

            display_unit = prod['wholesale_unit'] if is_wholesale_entry else prod['base_unit']
            original_qty_base = import_base_qty(original_qty, item['unit_type'], ratio)

            # 4. Kiểm tra số lượng có thể trả (theo bảng return_import, không sửa import_details)
            c.execute("SELECT COALESCE(SUM(quantity), 0) FROM return_import WHERE import_id = ? AND product_id = ?", (import_id, product_id))
            already_returned = float(c.fetchone()[0])
            if return_qty_input > (original_qty - already_returned) + 0.0001:
                return jsonify({"error": f"Số lượng tối đa có thể trả: {original_qty - already_returned}"}), 400

            return_qty_base = import_base_qty(return_qty_input, item['unit_type'], ratio)

            # Giá trị thuế/ mua hàng (tầng HĐ — PT thu NCC hoàn), không dùng cho WAC kho
            line_purchase_value = (
                float(item['subtotal']) - float(item['discount_amount']) + float(item['tax_amount'])
            )
            if line_purchase_value <= 0 and original_qty_base > 0:
                line_purchase_value = cost_price_base * original_qty_base
            refund_amount = (
                line_purchase_value * (return_qty_base / original_qty_base)
                if original_qty_base > 0 else 0.0
            )

            # 5. Xuất kho theo cost_price gốc của phiếu nhập (WAC chính xác khi trả NCC)
            _, cost_used = apply_wac_outbound(c, product_id, return_qty_base, cost_price_base)
            cost_out = return_qty_base * cost_used

            # 6. Ghi lịch sử kho (Stock Moves)
            c.execute("""
                INSERT INTO stock_moves (product_id, date, type, ref_id, quantity, cost_price, note, ref_document, ref_type, type1)
                VALUES (?, ?, 'RETURN_IMPORT', ?, ?, ?, ?, ?, ?, ?)
            """, (product_id, return_date, import_id, -return_qty_base, cost_price_base,
                  f"Trả hàng nhập PN{str(import_id).zfill(6)} - {reason}", f"PN{str(import_id).zfill(6)}", 'export', 'Trả HN'))

            c.execute("""
                INSERT INTO inventory_transactions
                (product_id, type, type1, quantity, cost_price, reference_id, reference_type, note, created_at)
                VALUES (?, 'export', 'Trả HN', ?, ?, ?, ?, ?, ?)
            """, (product_id, -return_qty_base, cost_price_base, import_id, f"PN{str(import_id).zfill(6)}", f"Trả hàng nhập PN{str(import_id).zfill(6)} - {reason}", return_date))

            sync_inventory_quantity_from_moves(c, product_id)

            # 7. Phiếu xuất kho — giá vốn theo cost_price PN gốc
            c.execute("SELECT voucher_no FROM phieu_xuat_kho WHERE voucher_no LIKE 'PX%' ORDER BY id DESC LIMIT 1")
            last_px = c.fetchone()
            if last_px and last_px['voucher_no'] and len(last_px['voucher_no']) > 2:
                try:
                    px_num = int(last_px['voucher_no'][2:]) + 1
                except: px_num = 1
            else: px_num = 1
            px_voucher_no = f"PX{px_num:06d}"

            px_unit_price = cost_price_base * (ratio if is_wholesale_entry else 1)
            px_items = [{
                "product_id": product_id,
                "product_name": product_name,
                "product_code": product_barcode,
                "unit": display_unit,
                "quantity": return_qty_input,
                "price": px_unit_price,
                "amount": cost_out
            }]

            c.execute("""
                INSERT INTO phieu_xuat_kho (voucher_no, date, customer_name, items_json, total_amount, note)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (px_voucher_no, return_date, supplier_name, json.dumps(px_items, ensure_ascii=False), cost_out, f"Trả hàng theo PN{str(import_id).zfill(6)}"))

            # 8. Ghi nhận phiếu trả NCC (chứng từ mới — không sửa import_details gốc)
            c.execute("""
                INSERT INTO return_import (import_id, product_id, quantity, cost_price, date, reason)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (import_id, product_id, return_qty_input, cost_price_base, return_date, reason))

            # 9. Phiếu thu — theo giá trị mua/thuế (tiền NCC hoàn)
            voucher_thu = None
            if refund_amount > 0:
                c.execute("SELECT voucher_no FROM phieu_thu WHERE voucher_no LIKE 'PT%' ORDER BY id DESC LIMIT 1")
                last_pt = c.fetchone()
                pt_num = int(last_pt['voucher_no'][2:]) + 1 if last_pt else 1
                voucher_thu = f"PT{pt_num:06d}"

                c.execute("""
                    INSERT INTO phieu_thu (voucher_no, payer_name, address, tax_code, amount, debit_account, credit_account, reason, reference_document, date)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (voucher_thu, supplier_name, supplier_address, tax_code, refund_amount, '111', '156', f"Thu tiền trả hàng PN{str(import_id).zfill(6)}", f"PN{str(import_id).zfill(6)}", return_date))

            conn.commit()
            return jsonify({
                "success": True, 
                "px_voucher": px_voucher_no,
                "pt_voucher": voucher_thu,
                "refund": round(refund_amount, 2),
                "cost_out": round(cost_out, 2),
                "import_cost_base": round(cost_price_base, 4),
            }), 200

        except Exception as e:
            conn.rollback()
            traceback.print_exc() # In lỗi ra Console để kiểm tra dòng bị lỗi
            return jsonify({"error": "Lỗi hệ thống", "detail": str(e)}), 500
        finally:
            conn.close()

    @app.route('/api/return/import/checkout', methods=['POST'])
    @login_required
    def api_return_import_checkout():
        """Trả NCC nhiều dòng: sale + kho + PX + PT/CN + (tuỳ chọn) HĐĐT."""
        from flask import current_app

        data = request.get_json() or {}
        if not data.get('import_id'):
            return jsonify({'success': False, 'error': 'Thiếu import_id'}), 400

        conn = get_db_connection()
        c = conn.cursor()
        try:
            c.execute('BEGIN IMMEDIATE')
            result = process_return_import_checkout(c, data)
            conn.commit()

            invoice_result = None
            if data.get('issue_invoice'):
                issue_fn = current_app.config.get('issue_invoice_for_sale')
                if issue_fn:
                    loai = data.get('loai_hdon', 1)
                    try:
                        loai = int(loai)
                    except (TypeError, ValueError):
                        loai = 1
                    invoice_result = issue_fn(result['sale_id'], loai_hdon=loai)
                else:
                    invoice_result = {'success': False, 'error': 'Chưa cấu hình xuất HĐĐT'}

            return jsonify({
                'success': True,
                **result,
                'invoice': invoice_result,
            })
        except ValueError as e:
            conn.rollback()
            return jsonify({'success': False, 'error': str(e)}), 400
        except Exception as e:
            conn.rollback()
            traceback.print_exc()
            return jsonify({'success': False, 'error': str(e)}), 500
        finally:
            conn.close()

    # === API TẢI DANH SÁCH PHIẾU NHẬP POS (/api/import GET) IMPORT_LIST ===
    @app.route('/api/import', methods=['GET'])
    def api_import_get():
        q = request.args.get('q', '').strip()
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        filter_today = request.args.get('filter_today', '').lower() == 'true'

        conn = get_db_connection()
        c = conn.cursor()

        try:
            # Lấy columns từ bảng import
            c.execute("PRAGMA table_info(import)")
            columns = [col[1] for col in c.fetchall()]

            base_cols = ['id', 'import_no', 'date', 'supplier_id', 'bill_no']
            select_parts = [f"i.{col}" for col in base_cols if col in columns]

            total_col = next((col for col in ['total_payment', 'total', 'total_value'] if col in columns), None)
            select_parts.append(f"COALESCE(i.{total_col},0) as total_goods_amount_db" if total_col else "0 as total_goods_amount_db")

            extra_col = next((col for col in ['extra_cost','extraCost'] if col in columns), None)
            select_parts.append(f"COALESCE(i.{extra_col},0) as extraCost" if extra_col else "0 as extraCost")

            status_col = next((col for col in ['status','payment_status'] if col in columns), None)
            # SỬ DỤNG ALIAS 'status_raw' để lấy giá trị thô từ DB
            select_parts.append(f"i.{status_col} as status_raw" if status_col else "'' as status_raw")

            select_parts.append("s.name as supplier_name")

            sql = f"SELECT {', '.join(select_parts)} FROM import i LEFT JOIN suppliers s ON i.supplier_id = s.id"

            where_conditions = []
            params = []

            final_start_datetime = final_end_datetime = None
            if filter_today:
                today = date.today().strftime('%Y-%m-%d')
                final_start_datetime = f"{today} 00:00:00"
                final_end_datetime = f"{today} 23:59:59"
            elif start_date_str and end_date_str:
                try:
                    datetime.strptime(start_date_str,'%Y-%m-%d')
                    datetime.strptime(end_date_str,'%Y-%m-%d')
                    final_start_datetime = f"{start_date_str} 00:00:00"
                    final_end_datetime = f"{end_date_str} 23:59:59"
                except ValueError:
                    pass

            if final_start_datetime and final_end_datetime:
                where_conditions.append("i.date BETWEEN ? AND ?")
                params.extend([final_start_datetime, final_end_datetime])

            if q:
                like = f"%{q}%"
                search_conditions = []
                if 'import_no' in columns: search_conditions.append("i.import_no LIKE ?")
                if 'bill_no' in columns: search_conditions.append("i.bill_no LIKE ?")
                search_conditions.append("s.name LIKE ?")
                where_conditions.append("(" + " OR ".join(search_conditions) + ")")
                params.extend([like]*len(search_conditions))

            if where_conditions:
                sql += " WHERE " + " AND ".join(where_conditions)

            sql += " ORDER BY i.date DESC, i.id DESC LIMIT 100"
        
            c.execute(sql, params)
            rows = c.fetchall()

            result = []
            for row in rows:
                r = dict(row) 
                total_goods_amount = float(r.get('total_goods_amount_db') or 0)
                extra_cost = float(r.get('extraCost') or 0)
                status_raw = r.get('status_raw') or ''

                total_value = total_goods_amount + extra_cost
                # Trả về trạng thái thanh toán thô (raw status)
                payment_status = status_raw 

                date_str = r.get('date') or ''
                try:
                    date_obj = datetime.strptime(date_str,'%Y-%m-%d %H:%M:%S')
                    r['date'] = date_obj.strftime('%Y-%m-%d') 
                except:
                    r['date'] = date_str 

                r.update({
                    'payment_amt': total_goods_amount,
                    'extraCost': extra_cost,
                    'total_value': total_value,
                    # Đảm bảo trường này được gửi đi với giá trị thô
                    'payment_status': payment_status
                })

                r.pop('status_raw',None)
                r.pop('total_goods_amount_db',None)
                result.append(r)

            return jsonify(result), 200

        except Exception as e:
            print("ERROR /api/import:", e)
            import traceback
            traceback.print_exc()
            return jsonify({"error":"Lỗi server","detail":str(e)}), 500
        finally:
            conn.close()

    #==== API LẤY CHI TIẾT PHIẾU NHẬP ====#
    from flask import jsonify
    from datetime import datetime
    import traceback

    @app.route('/api/import/<int:import_id>', methods=['GET'])
    def api_import_detail(import_id):
        conn = get_db_connection()
        c = conn.cursor()
        try:
            # Lấy tất cả cột của bảng import để kiểm tra an toàn
            c.execute("PRAGMA table_info(import)")
            import_cols = {col[1] for col in c.fetchall()}

            # Các cột cơ bản (luôn chọn nếu tồn tại)
            select_header = []
            if 'id' in import_cols: select_header.append("i.id")
            if 'import_no' in import_cols: select_header.append("i.import_no")
            if 'date' in import_cols: select_header.append("i.date")
            if 'supplier_id' in import_cols: select_header.append("i.supplier_id")
            if 'bill_no' in import_cols: select_header.append("i.bill_no")

            # Total amount - fallback nhiều cột
            total_col = next((c for c in ['total_value', 'total_payment', 'total'] if c in import_cols), None)
            select_header.append(f"COALESCE(i.{total_col}, 0) AS total_goods_amount" if total_col else "0 AS total_goods_amount")

            # Extra cost
            extra_col = next((c for c in ['extra_cost', 'extraCost'] if c in import_cols), None)
            select_header.append(f"COALESCE(i.{extra_col}, 0) AS extra_cost" if extra_col else "0 AS extra_cost")

            # Payment status
            status_col = next((c for c in ['payment_status', 'status'] if c in import_cols), None)
            select_header.append(f"COALESCE(i.{status_col}, '') AS payment_status" if status_col else "'' AS payment_status")

            select_header.append("s.name AS supplier_name")
            if import_cols:
                pass
            header_sql = f"""
                SELECT {', '.join(select_header)},
                       s.address AS supplier_address,
                       COALESCE(s.tax_code, '') AS supplier_tax_code,
                       COALESCE(s.phone, '') AS supplier_phone,
                       COALESCE(s.email, '') AS supplier_email
                FROM import i
                LEFT JOIN suppliers s ON i.supplier_id = s.id
                WHERE i.id = ?
            """
            c.execute(header_sql, (import_id,))
            header_row = c.fetchone()
            if not header_row:
                return jsonify({"error": "Không tìm thấy phiếu nhập"}), 404

            import_data = dict(header_row)

            # Format ngày
            if import_data.get('date'):
                try:
                    date_obj = datetime.strptime(import_data['date'][:10], '%Y-%m-%d')  # Chỉ lấy ngày nếu có giờ
                    import_data['date'] = date_obj.strftime('%d/%m/%Y')
                except Exception:
                    pass

            import_data['total_value'] = float(import_data.get('total_goods_amount', 0)) + float(import_data.get('extra_cost', 0))
            import_data['payment_status_display'] = import_data.get('payment_status', 'Chưa thanh toán')

            # === CHI TIẾT SẢN PHẨM - RẤT AN TOÀN ===
            c.execute("PRAGMA table_info(import_details)")
            details_cols = {col[1] for col in c.fetchall()}

            select_items = [
                "ii.id",
                "ii.product_id",
                "COALESCE(NULLIF(TRIM(ii.product_name), ''), p.name, '') AS product_name",
                "p.unit AS base_unit",
                "p.unit1 AS wholesale_unit",
                "COALESCE(p.unit_ratio, 1) AS unit_ratio",
                "ii.qty",
                "ii.buyprice",
                "COALESCE(ii.discount, 0) AS discount",
                "COALESCE(ii.tax, 0) AS tax",
                "COALESCE(ii.discount_pct, 0) AS discount_pct",
                "COALESCE(ii.tax_pct, 0) AS tax_pct",
            ]
            if 'line_type' in details_cols:
                select_items.append("COALESCE(ii.line_type, 'goods') AS line_type")

            # Chỉ thêm nếu cột tồn tại
            if 'unit' in details_cols:
                select_items.append("COALESCE(NULLIF(TRIM(ii.unit), ''), p.unit, 'Cái') AS import_unit")
            if 'unit_type' in details_cols:
                select_items.append("ii.unit_type")

            items_sql = f"""
                SELECT {', '.join(select_items)}
                FROM import_details ii
                LEFT JOIN products p ON ii.product_id = p.id
                WHERE ii.import_id = ?
            """
            c.execute(items_sql, (import_id,))

            items = []
            for row in c.fetchall():
                it = dict(row)

                # FIX: Xác định đơn vị hiển thị đúng (ưu tiên từ import_details)
                if 'import_unit' in it and it['import_unit']:
                    it['unit'] = it['import_unit'].strip() or it.get('base_unit', '—')
                elif 'unit_type' in it and it['unit_type'] == 1 and it.get('wholesale_unit'):
                    it['unit'] = it['wholesale_unit'].strip() or '—'
                else:
                    it['unit'] = it.get('base_unit', '—')

                # Tính remaining_qty theo đơn vị gốc
                c.execute("""
                    SELECT COALESCE(SUM(quantity), 0) AS returned
                    FROM return_import
                    WHERE import_id = ? AND product_id = ?
                """, (import_id, it['product_id']))
                returned = float(c.fetchone()['returned'] or 0)

                it['quantity'] = float(it['qty'])
                it['remaining_qty'] = it['quantity'] - returned
                it['unit_type'] = int(it.get('unit_type') or 0)
                it['unit_ratio'] = float(it.get('unit_ratio') or 1)

                for k in ['base_unit', 'wholesale_unit', 'import_unit', 'qty']:
                    it.pop(k, None)

                items.append(it)

            import_data['supplier'] = {
                'name': import_data.get('supplier_name') or '',
                'address': import_data.get('supplier_address') or '',
                'tax_code': import_data.get('supplier_tax_code') or '',
                'phone': import_data.get('supplier_phone') or '',
                'email': import_data.get('supplier_email') or '',
            }
            import_data['items'] = items
            return jsonify(import_data), 200

        except Exception as e:
            traceback.print_exc()  # In lỗi ra console để debug
            return jsonify({
                "error": "Lỗi tải chi tiết phiếu",
                "detail": str(e)
            }), 500

        finally:
            conn.close()

    # === EDITTING IMPORT === @app.route('/import/edit/<int:import_id>')
    # HÀM LẤY CHI TIẾT PHIẾU NHẬP KHO ĐỂ CHỈNH SỬA CHO ROUTE – @app.route('/import/edit/<int:import_id>')
    # =====================================================
    from flask import render_template, flash, redirect, url_for
    import sqlite3

    def _normalize_db_date(value):
        """Chuẩn hóa ngày từ SQLite/datetime → chuỗi YYYY-MM-DD cho JSON/JS."""
        if value is None:
            return None
        if hasattr(value, 'strftime'):
            return value.strftime('%Y-%m-%d')
        text = str(value).strip()
        if not text:
            return None
        if 'T' in text:
            text = text.split('T', 1)[0]
        elif ' ' in text:
            text = text.split(' ', 1)[0]
        return text[:10] if len(text) >= 10 else text

    def get_import_detail(import_id, db=None):
        """
        Trả về dict chi tiết phiếu nhập + danh sách sản phẩm
        DÙNG CHO: import_edit, import_view, print_import...
        Sử dụng trường unit_type từ bảng import_details để quyết định invoice_unit
        """
        close_db = False
        if db is None:
            db = get_db_connection()  # Đảm bảo get_db_connection() có row_factory = sqlite3.Row
            close_db = True
        try:
            c = db.cursor()

            # ----------------------------
            # 1. Lấy thông tin tổng hợp phiếu nhập
            # ----------------------------
            c.execute("SELECT * FROM import WHERE id = ?", (import_id,))
            row = c.fetchone()
            if not row:
                return None
            
            imp = dict(row)
            imp_keys = row.keys()

            # Lấy thông tin nhà cung cấp liên kết an toàn
            if imp.get('supplier_id'):
                c.execute("SELECT name, tax_code, address FROM suppliers WHERE id = ?", (imp['supplier_id'],))
                sup_row = c.fetchone()
                if sup_row:
                    imp['supplier_name'] = sup_row['name']
                    imp['tax_code'] = sup_row['tax_code']
                    imp['address'] = sup_row['address']

            # Fallback đồng bộ tên trường dữ liệu header cũ/mới
            if 'invoice_no' in imp_keys and not imp.get('bill_no'):
                imp['bill_no'] = imp['invoice_no']
            if 'bill_date' not in imp_keys:
                imp['bill_date'] = imp.get('date')
            if 'payment_method' not in imp_keys:
                imp['payment_method'] = 'cash'

            imp['date'] = _normalize_db_date(imp.get('date'))
            imp['bill_date'] = _normalize_db_date(imp.get('bill_date'))

            calculated_total = 0
            items = []

            from Services.import_line_helpers import (
                detect_service_import,
                enrich_stock_detail_for_edit,
                fetch_import_details_raw,
                is_service_detail_row,
                map_service_detail_for_edit,
            )

            raw_rows = fetch_import_details_raw(c, import_id)
            is_service_import = detect_service_import(imp, raw_rows)

            for row in raw_rows:
                if is_service_detail_row(row):
                    item = map_service_detail_for_edit(row)
                else:
                    item = enrich_stock_detail_for_edit(c, row)
                calculated_total += float(item.get('payment_amount') or 0)
                items.append(item)

            imp['is_service_import'] = is_service_import

            # ----------------------------
            # 3. Tổng hợp thông tin thanh toán phiếu nhập
            # ----------------------------
            extra_cost = float(imp.get('extra_cost') or 0)
            imp['total_payment'] = calculated_total
            imp['total_value'] = calculated_total + extra_cost
            imp['items'] = items

            return imp

        except Exception:
            logging.exception("get_import_detail(%s)", import_id)
            raise

        finally:
            if close_db and db:
                db.close()

    @app.route('/api/import/<int:import_id>/edit', methods=['GET'])
    @login_required
    def api_import_edit_detail(import_id):
        """API tải dữ liệu phiếu nhập cho trang sửa."""
        from Services.import_line_helpers import prepare_import_edit_json

        try:
            imp = get_import_detail(import_id)
            if not imp:
                return jsonify({"success": False, "error": "Không tìm thấy phiếu nhập"}), 404
            return jsonify({"success": True, "data": prepare_import_edit_json(imp)})
        except Exception as exc:
            logging.exception("api_import_edit_detail(%s)", import_id)
            return jsonify({
                "success": False,
                "error": f"Lỗi tải chi tiết phiếu nhập: {exc}",
            }), 500

    @app.route('/import/edit/<int:import_id>')
    @login_required
    def import_edit(import_id):
        conn = get_db_connection()
        try:
            row = conn.execute(
                "SELECT id, import_no FROM import WHERE id = ?",
                (import_id,),
            ).fetchone()
        finally:
            conn.close()

        if not row:
            flash('Không tìm thấy phiếu nhập!', 'danger')
            return redirect(url_for('import_list'))

        return render_template(
            'Import_edit.html',
            import_id=import_id,
            import_no=row['import_no'] or import_id,
        )

    # === API UPDATE IMPORT SAU KHI SỬA PHIẾU NHẬP KHO ===
    @app.route('/api/import/update/<int:import_id>', methods=['PUT'])
    @login_required
    def api_import_update(import_id):
        conn = get_db_connection()
        conn.row_factory = sqlite3.Row  # Định dạng kết quả trả về kiểu Row
        c = conn.cursor()
        try:
            data = request.get_json()
            if not data:
                return jsonify({"error": "Dữ liệu không hợp lệ"}), 400

            # --- 1. LẤY PHIẾU NHẬP CŨ ---
            c.execute("SELECT * FROM import WHERE id = ?", (import_id,))
            imp = c.fetchone()
            if not imp:
                return jsonify({"error": "Không tìm thấy phiếu nhập"}), 404
        
            imp_keys = imp.keys()  # Lấy danh sách cột thực tế của bảng import để kiểm tra động
            import_date_old = imp['date']
            import_no = imp['import_no']

            # ================== XỬ LÝ NHÀ CUNG CẤP ==================
            supplier_name = (data.get('supplier_name') or '').strip()
            if not supplier_name:
                return jsonify({"error": "Tên nhà cung cấp không được để trống"}), 400
            tax_code = (data.get('tax_code') or '').strip() or None
            address = (data.get('address') or '').strip() or None
            supplier_id = imp['supplier_id'] if 'supplier_id' in imp_keys else None

            if supplier_id:
                c.execute("UPDATE suppliers SET name=?, tax_code=?, address=? WHERE id=?",
                          (supplier_name, tax_code, address, supplier_id))
            else:
                c.execute("SELECT id FROM suppliers WHERE LOWER(name)=LOWER(?)", (supplier_name,))
                row = c.fetchone()
                if row:
                    supplier_id = row['id']
                    c.execute("UPDATE suppliers SET tax_code=?, address=? WHERE id=?", (tax_code, address, supplier_id))
                else:
                    c.execute("INSERT INTO suppliers (name, tax_code, address) VALUES (?, ?, ?)",
                              (supplier_name, tax_code, address))
                    supplier_id = c.lastrowid

            # ================== THÔNG TIN CHUNG ==================
            import_date = data.get('date', imp['date'])
        
            bill_date_old = imp['bill_date'] if 'bill_date' in imp_keys else None
            bill_date = data.get('bill_date') or bill_date_old

            bill_no_old = imp['bill_no'] if 'bill_no' in imp_keys else (imp['invoice_no'] if 'invoice_no' in imp_keys else None)
            bill_no = data.get('bill_no') or bill_no_old

            payment_status_old = imp['payment_status'] if 'payment_status' in imp_keys else 'Chưa thanh toán'
            payment_status = data.get('payment_status') or payment_status_old or 'Chưa thanh toán'

            payment_method_old = imp['payment_method'] if 'payment_method' in imp_keys else 'cash'
            payment_method_input = data.get('payment_method') or payment_method_old or 'cash'
        
            if payment_method_input and '111' in str(payment_method_input):
                payment_method = 'cash'
            elif payment_method_input and '112' in str(payment_method_input):
                payment_method = 'bank'
            else:
                payment_method = 'cash' if payment_method_input in ['cash', 'Tiền mặt'] else 'bank'

            extra_cost_old = imp['extra_cost'] if 'extra_cost' in imp_keys else 0
            extra_cost = Decimal(str(data.get('extra_cost', 0) or extra_cost_old or 0))
        
            note_old = imp['note'] if 'note' in imp_keys else ''
            note = data.get('note') or note_old or ''
        
            now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            new_items = data.get('items', []) or []

            # Tính tổng giá trị cơ sở để phân bổ chi phí phát sinh
            total_base_for_allocation = Decimal('0')
            for it in new_items:
                qty_val = Decimal(str(it.get('qty', 0) or 0))
                price_val = Decimal(str(it.get('buyprice', 0) or 0))
                disc_pct = Decimal(str(it.get('discount_pct') if it.get('discount_pct') is not None else it.get('discountPct', 0)))
                if qty_val > 0:
                    total_base_for_allocation += (qty_val * price_val) * (Decimal('1') - disc_pct / Decimal('100'))

            if total_base_for_allocation <= 0:
                total_base_for_allocation = Decimal('0.0001')

            total_value = Decimal('0')

            from Services.import_line_helpers import (
                build_service_line_insert_fields,
                fetch_import_details_raw,
                insert_import_detail_row,
                is_service_line_payload,
            )
            from Services.inward_invoice_helpers import next_pc_voucher_no

            old_raw_rows = fetch_import_details_raw(c, import_id)
            doc_type = (imp['doc_type'] if 'doc_type' in imp_keys else '').strip().lower()
            is_service_doc = doc_type == 'service'

            with conn:
                # ================== LẤY CHI TIẾT CŨ TỪ BẢNG import_details ==================
                c.execute(
                    "SELECT product_id, qty, cost_price, unit_type, COALESCE(line_type, 'goods') AS line_type "
                    "FROM import_details WHERE import_id = ?",
                    (import_id,),
                )
                old_details = c.fetchall() or []
                old_stock_details = [
                    od for od in old_details
                    if (od['line_type'] or '').strip().lower() != 'service' and od['product_id']
                ]

                # Tải cấu trúc quy đổi đơn vị của mặt hàng mới (chỉ dòng có product_id)
                product_units = {}
                if new_items:
                    product_ids = list({
                        int(it.get('product_id', 0))
                        for it in new_items
                        if it.get('product_id') and not is_service_line_payload(it)
                    })
                    if product_ids:
                        placeholders = ','.join('?' * len(product_ids))
                        c.execute(
                            f"SELECT id, unit, unit1, unit_ratio FROM products WHERE id IN ({placeholders})",
                            product_ids,
                        )
                        for p in c.fetchall():
                            product_units[p['id']] = {
                                'unit': p['unit'] or 'Cái',
                                'unit1': p['unit1'] or '',
                                'ratio': Decimal(str(p['unit_ratio'] or 1)),
                            }

                # ================== HOÀN WAC PHIẾU NHẬP CŨ (chỉ dòng hàng hóa có kho) ==================
                sync_pids = set()
                if old_stock_details:
                    wac_pids = reverse_import_moves_wac(c, import_id)
                    sync_pids.update(wac_pids)

                # ================== DỌN SẠCH NHẬT KÝ CHI TIẾT CŨ & LƯU MỚI ==================
                c.execute("DELETE FROM import_details WHERE import_id = ?", (import_id,))
                c.execute(
                    "DELETE FROM stock_moves WHERE ref_id = ? AND type IN ('import', 'RETURN_IMPORT')",
                    (import_id,),
                )
                c.execute(
                    "DELETE FROM inventory_transactions WHERE reference_id = ? AND reference_type = 'import'",
                    (import_id,),
                )

                # Ghi lại lịch sử chi tiết phiếu nhập mới
                for item in new_items:
                    qty = Decimal(str(item.get('qty', 0) or 0))
                    if qty <= 0:
                        continue

                    if is_service_line_payload(item):
                        svc_fields, line_total = build_service_line_insert_fields(
                            import_id, item, extra_cost, total_base_for_allocation,
                        )
                        if not svc_fields:
                            continue
                        insert_import_detail_row(c, import_id, svc_fields)
                        total_value += line_total
                        continue

                    pid = int(item.get('product_id', 0) or 0)
                    if not pid:
                        continue

                    buyprice = Decimal(str(item.get('buyprice', 0)))
                    discountPct = Decimal(str(item.get('discount_pct') if item.get('discount_pct') is not None else item.get('discountPct', 0)))
                    taxPct = Decimal(str(item.get('tax_pct') if item.get('tax_pct') is not None else item.get('taxPct', 0)))
                
                    line_total = qty * buyprice
                
                    discount_amt = Decimal(str(item.get('discount_amount'))) if item.get('discount_amount') is not None else (line_total * discountPct / Decimal('100'))
                    after_discount = line_total - discount_amt
                    tax_amt = Decimal(str(item.get('tax_amount'))) if item.get('tax_amount') is not None else (after_discount * taxPct / Decimal('100'))
                
                    allocated_extra = extra_cost * (after_discount / total_base_for_allocation) if extra_cost > 0 else Decimal('0')
                    cost_value_full = after_discount + tax_amt + allocated_extra
                
                    input_unit = str(item.get('unit') or '').strip()
                    p_unit = product_units.get(pid, {})
                    db_unit = str(p_unit.get('unit') or 'Cái').strip()
                    db_unit1 = str(p_unit.get('unit1') or '').strip()
                    ratio = p_unit.get('ratio', Decimal('1'))
                
                    unit_type = 1 if input_unit.lower() == db_unit1.lower() and db_unit1 != '' else 0
                    qty_base_final = qty * ratio if unit_type == 1 else qty
                    cost_price_base = cost_value_full / qty_base_final if qty_base_final > 0 else Decimal('0')

                    total_value += cost_value_full

                    apply_wac_inbound(c, pid, float(qty_base_final), float(cost_value_full))

                    c.execute("""INSERT INTO import_details
                                (import_id, product_id, qty, buyprice, discount, tax, tax_pct, discount_pct, cost_price, unit_type)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                              (import_id, pid, float(qty), float(buyprice), float(discount_amt), float(tax_amt), float(taxPct), float(discountPct), float(cost_price_base), unit_type))

                    c.execute("""INSERT INTO stock_moves
                                (product_id, date, type, ref_id, quantity, cost_price, note, ref_document, ref_type, type1, unit, unit1, unit_ratio)
                                VALUES (?, ?, 'import', ?, ?, ?, ?, ?, 'import', 'Nhập', ?, ?, ?)""",
                              (pid, import_date, import_id, float(qty_base_final), float(cost_price_base),
                               f"Nhập kho – PN#{import_no}", import_no, db_unit, db_unit1, float(ratio)))

                    c.execute("""INSERT INTO inventory_transactions
                                (product_id, type, type1, quantity, cost_price, reference_id, reference_type, note, created_at)
                                VALUES (?, 'import', 'Nhập', ?, ?, ?, 'import', ?, ?)""",
                              (pid, float(qty_base_final), float(cost_price_base), import_id, f"Nhập kho - PN#{import_no}", now_str))

                    sync_pids.add(pid)

                for old in old_stock_details:
                    sync_pids.add(old['product_id'])
                if sync_pids:
                    sync_inventory_quantities(c, [p for p in sync_pids if p])

                # ================== XỬ LÝ QUỸ SỔ SÁCH DÒNG TIỀN (PHIẾU CHI) ==================
                c.execute(
                    "DELETE FROM phieu_chi WHERE source_type IN ('import', 'import_service') AND source_id = ?",
                    (import_id,),
                )
                c.execute(
                    "DELETE FROM Phieu_chi WHERE source_type IN ('import', 'import_service') AND source_id = ?",
                    (import_id,),
                )

                if payment_status in ['Đã thanh toán', 'Thanh toán một phần']:
                    final_paid = float(total_value + extra_cost)
                    credit_acc = '111' if payment_method == 'cash' else '112'

                    if is_service_doc:
                        res_pc_vouch = next_pc_voucher_no(c)
                        c.execute("""
                            INSERT INTO phieu_chi (
                                voucher_no, receiver_name, address, amount,
                                debit_account, credit_account,
                                expense_type, reason, source_type, reference_document, source_id, preparer, date
                            ) VALUES (?, ?, ?, ?, '642', ?, 'CP_DV', ?, 'import_service', ?, ?, ?, ?)
                        """, (
                            res_pc_vouch, supplier_name, address, final_paid, credit_acc,
                            f'Thanh toán tiền mua dịch vụ số {import_no}',
                            bill_no, import_id, session.get('user_name', 'Admin'), import_date,
                        ))
                    else:
                        c.execute("""
                            SELECT MAX(CAST(SUBSTR(voucher_no, 3) AS INTEGER)) as max_num 
                            FROM Phieu_chi 
                            WHERE voucher_no LIKE 'PC%' AND LENGTH(voucher_no) > 2
                        """)
                        row = c.fetchone()
                    
                        max_num = row['max_num'] if (row and row['max_num'] is not None) else 0
                        new_pc_num = max_num + 1
                    
                        while True:
                            res_pc_vouch = f"PC{new_pc_num:06d}"
                            c.execute("SELECT id FROM Phieu_chi WHERE voucher_no = ?", (res_pc_vouch,))
                            if not c.fetchone():
                                break
                            new_pc_num += 1
                    
                        c.execute("""
                            INSERT INTO Phieu_chi (
                                voucher_no, receiver_name, address, amount, credit_account, 
                                debit_account, reason, source_type, reference_document, source_id, preparer, date
                            )
                            VALUES (?, ?, ?, ?, ?, '331', ?, 'import', ?, ?, ?, ?)
                        """, (res_pc_vouch, supplier_name, address, final_paid, credit_acc, 
                              f"Thanh toán tiền mua hàng số {import_no}", bill_no, import_id, 
                              session.get('user_name', 'Admin'), import_date))

                # ================== ĐỒNG BỘ TRẠNG THÁI HÓA ĐƠN NCC GỐC ==================
                bill_no_clean = str(bill_no).strip() if bill_no else ""
                if bill_no_clean and bill_no_clean.lower() not in ['none', 'nan']:
                    c.execute("""
                        UPDATE supplier_invoice 
                        SET status = 'imported' 
                        WHERE invoice_no = ? AND seller_tax_code = ? AND status != 'imported'
                    """, (bill_no_clean, tax_code))

                # ================== CẬP NHẬT HEADER PHIẾU NHẬP TỔNG HỢP ==================
                update_fields = ["supplier_id=?", "date=?"]
                update_values = [supplier_id, import_date]

                if "bill_date" in imp_keys:
                    update_fields.append("bill_date=?")
                    update_values.append(bill_date)
                if "bill_no" in imp_keys:
                    update_fields.append("bill_no=?")
                    update_values.append(bill_no)
                elif "invoice_no" in imp_keys:
                    update_fields.append("invoice_no=?")
                    update_values.append(bill_no)

                if "payment_status" in imp_keys:
                    update_fields.append("payment_status=?")
                    update_values.append(payment_status)
                if "payment_method" in imp_keys:
                    update_fields.append("payment_method=?")
                    update_values.append(payment_method)
                if "extra_cost" in imp_keys:
                    update_fields.append("extra_cost=?")
                    update_values.append(float(extra_cost))
                if "note" in imp_keys:
                    update_fields.append("note=?")
                    update_values.append(note)
                if "total_value" in imp_keys:
                    update_fields.append("total_value=?")
                    update_values.append(float(total_value))

                update_values.append(import_id)
                query_update = f"UPDATE import SET {', '.join(update_fields)} WHERE id=?"
                c.execute(query_update, tuple(update_values))

            conn.commit()
            return jsonify({"success": True, "message": "Cập nhật phiếu nhập và cân đối dòng tiền thành công!"})

        except Exception as e:
            conn.rollback()
            traceback.print_exc()
            return jsonify({"error": str(e)}), 500
        finally:
            conn.close()

    #==== API XÓA PHIẾU NHẬP ====#
    @app.route('/api/import/delete/<int:import_id>', methods=['DELETE'])
    @login_required
    @admin_or_master_required
    def delete_import(import_id):
        conn = None
        try:
            conn = get_db_connection()
            conn.row_factory = sqlite3.Row
            c = conn.cursor()

            # 1. Lấy thông tin phiếu nhập và NCC
            c.execute("""
                SELECT i.id, i.import_no, i.date AS import_date, i.bill_no, i.supplier_id,
                       i.from_invoice_id,
                       s.tax_code AS supplier_tax_code
                FROM "import" i
                LEFT JOIN suppliers s ON s.id = i.supplier_id
                WHERE i.id = ?
            """, (import_id,))
            import_row = c.fetchone()
            if not import_row:
                return jsonify({"success": False, "error": "Không tìm thấy phiếu nhập"}), 404
        
            import_no = import_row['import_no']
            import_date = import_row['import_date']
            bill_no = import_row['bill_no']
            supplier_id = import_row['supplier_id']
            tax_code = import_row['supplier_tax_code']
            from_invoice_id = import_row['from_invoice_id'] if 'from_invoice_id' in import_row.keys() else None

            bill_no_clean = str(bill_no).strip() if bill_no else ""

            c.execute("SELECT COUNT(*) AS cnt FROM return_import WHERE import_id = ?", (import_id,))
            if int(c.fetchone()['cnt'] or 0) > 0:
                return jsonify({
                    "success": False,
                    "error": "Phiếu nhập đã phát sinh trả hàng NCC, không thể xóa.",
                }), 403

            # 2. Kiểm tra TSCĐ/CCDC đã kích hoạt
            from Services.fixed_assets_helpers import (
                count_active_assets_by_import_id,
                delete_assets_by_import_id,
            )
            fa_act, tools_act = count_active_assets_by_import_id(c, import_id)
            if fa_act or tools_act:
                return jsonify({
                    "success": False,
                    "error": "Không thể xóa: TSCĐ/CCDC từ phiếu nhập đã đưa vào sử dụng.",
                }), 403

            # 3. Lấy chi tiết hàng hóa
            c.execute("""
                SELECT d.product_id, d.qty, d.unit_type,
                       COALESCE(d.line_type, p.product_type, 'goods') AS line_type,
                       COALESCE(p.unit_ratio, 1) AS unit_ratio, p.name
                FROM import_details d
                JOIN products p ON p.id = d.product_id
                WHERE d.import_id = ?
            """, (import_id,))
            items_in_import = c.fetchall()

            sync_pids = set()
            outbound_types = (
                'export', 'sale', 'SALE', 'adjustment_out', 'ADJUSTMENT_OUT',
                'RETURN_IMPORT', 'RETURN_SALE',
            )
            placeholders = ','.join('?' * len(outbound_types))

            for item in items_in_import:
                line_type = (item["line_type"] or 'goods').strip().lower()
                if line_type in ('fixed_asset', 'tools', 'service'):
                    continue

                p_id = item["product_id"]
                p_name = item["name"]
                imported_qty_retail = import_base_qty(item["qty"], item["unit_type"], item["unit_ratio"])

                c.execute(f"""
                    SELECT COUNT(*) AS count FROM stock_moves
                    WHERE product_id = ?
                      AND date >= ?
                      AND type IN ({placeholders})
                """, (p_id, import_date, *outbound_types))
                has_moved = c.fetchone()['count']

                if has_moved > 0:
                    return jsonify({
                        "success": False,
                        "error": f"Không thể xóa vì sản phẩm '{p_name}' đã phát sinh xuất/trả kho sau ngày nhập.",
                    }), 403

                curr_qty = ledger_quantity(c, p_id)
                if imported_qty_retail > curr_qty + 0.0001:
                    return jsonify({
                        "success": False,
                        "error": f"Sản phẩm '{p_name}' tồn kho ({curr_qty}) nhỏ hơn số lượng nhập ({imported_qty_retail}).",
                    }), 403
                sync_pids.add(p_id)

            # 3. Hoàn WAC theo stock_moves import gốc rồi xóa chứng từ
            wac_pids = reverse_import_moves_wac(c, import_id)
            sync_pids.update(wac_pids)

            # 4. Xóa dữ liệu liên quan
            delete_assets_by_import_id(c, import_id)
            c.execute("DELETE FROM import_details WHERE import_id = ?", (import_id,))
            c.execute("DELETE FROM chi_tiet_phieu_nhap_kho WHERE import_id = ?", (import_id,))
            c.execute("DELETE FROM phieu_nhap_kho WHERE import_id = ?", (import_id,))
            c.execute(
                "DELETE FROM stock_moves WHERE ref_id = ? AND type IN ('import', 'RETURN_IMPORT')",
                (import_id,),
            )
            c.execute(
                "DELETE FROM inventory_transactions WHERE reference_id = ? AND reference_type = 'import'",
                (import_id,),
            )
            c.execute('DELETE FROM "import" WHERE id = ?', (import_id,))
            c.execute("DELETE FROM phieu_chi WHERE source_id = ? AND source_type = 'import'", (import_id,))
            c.execute("DELETE FROM Phieu_chi WHERE source_id = ? AND source_type = 'import'", (import_id,))

            from Services.inward_invoice_helpers import reset_supplier_invoice_after_import_removed
            reset_supplier_invoice_after_import_removed(
                c,
                from_invoice_id=from_invoice_id,
                bill_no=bill_no_clean,
                tax_code=tax_code,
            )

            sync_inventory_quantities(c, list(sync_pids))

            # 5. Đồng bộ sequence SQLite
            tables_to_reset = ['import', 'import_details', 'phieu_nhap_kho', 'chi_tiet_phieu_nhap_kho', 'stock_moves', 'inventory_transactions', 'phieu_chi']
            for table in tables_to_reset:
                formatted = f'"{table}"' if table == "import" else table
                c.execute(f"""
                    UPDATE sqlite_sequence
                    SET seq = (SELECT COALESCE(MAX(id), 0) FROM {formatted})
                    WHERE name = ?
                """, (table,))

            conn.commit()
            return jsonify({"success": True, "message": "Đã hủy phiếu nhập, hoàn kho và cập nhật giá vốn thành công!"})

        except Exception as e:
            if conn:
                conn.rollback()
            import traceback
            traceback.print_exc()
            return jsonify({"success": False, "error": f"Lỗi hệ thống: {str(e)}"}), 500
        finally:
            if conn:
                conn.close()

    # API: KIỂM TRA TRÙNG SỐ PHIẾU
    @app.route('/api/import/check_duplicate', methods=['POST'])
    @login_required
    def api_check_import_duplicate():
        data = request.json
        import_no = data.get('import_no', '').strip()
        if not import_no:
            return jsonify({"exists": False})

        conn = get_db_connection()
        c = conn.cursor()
        try:
            c.execute("SELECT 1 FROM import WHERE import_no = ?", (import_no,))
            exists = c.fetchone() is not None
            return jsonify({"exists": exists})
        finally:
            conn.close()

    @app.route('/api/import/next_sequence', methods=['POST'])
    @login_required
    def api_import_next_sequence():
        payload = request.get_json(silent=True) or {}
        mode = (payload.get('mode') or request.args.get('mode') or 'stock').strip().lower()

        conn = get_db_connection()
        conn.execute("BEGIN IMMEDIATE")
        c = conn.cursor()

        try:
            next_no = _next_import_no_from_db(c, mode)
            conn.commit()
            return jsonify({"success": True, "next_no": next_no})
        except Exception as e:
            conn.rollback()
            print(f"Error generating import sequence ({mode}): {str(e)}")
            fallback = "HT000001" if mode == 'service' else "PN000001"
            return jsonify({
                "success": False,
                "error": str(e),
                "next_no": fallback,
            }), 500
        finally:
            conn.close()

    #============= API IN TEM BARCODE GIẤY A4, LOẠI 40 TEM TOMMY TỪ PHIẾU NHẬP===============#

    def _barcode_print_item_from_row(row):
        item = dict(row)
        item['price_formatted'] = format_price(item.get('price'))
        item['base_price_formatted'] = format_price(item.get('base_price'))
        return item

    def _apply_stock_barcode_defaults(item):
        """Mặc định số tem in = tồn kho (trang products)."""
        import math
        stock_qty = float(item.get('quantity') or 0)
        retail_qty = max(0, int(math.floor(stock_qty + 1e-9)))
        item['quantity'] = stock_qty
        item['quantity_display'] = (
            str(retail_qty) if abs(stock_qty - retail_qty) < 1e-9
            else f'{stock_qty:,.3f}'.replace(',', '.')
        )
        item['default_qty_retail'] = retail_qty
        item['default_qty_wholesale'] = 0
        return item

    def _append_barcodes_for_item(barcodes, item, qty_retail, qty_wholesale):
        qty_retail = int(qty_retail or 0)
        qty_wholesale = int(qty_wholesale or 0)
        for _ in range(qty_retail):
            if item.get('barcode'):
                barcodes.append({
                    'code': item['barcode'],
                    'name': item['name'],
                    'unit': item.get('unit') or 'Cái',
                    'price': item.get('base_price', item.get('price', 0)),
                })
        for _ in range(qty_wholesale):
            if item.get('barcode1'):
                barcodes.append({
                    'code': item['barcode1'],
                    'name': item['name'],
                    'unit': item.get('unit1') or 'Thùng',
                    'price': item.get('price', 0),
                })

    def _collect_extra_barcodes_from_form(request, cursor):
        barcodes = []
        i = 1
        while True:
            pid_str = request.form.get(f'extra_product_id_{i}')
            if not pid_str:
                break
            try:
                pid = int(pid_str)
            except (TypeError, ValueError):
                i += 1
                continue
            cursor.execute(
                """
                SELECT id AS product_id, name, unit, unit1, price, base_price, barcode, barcode1
                FROM products WHERE id = ?
                """,
                (pid,),
            )
            p = cursor.fetchone()
            if p:
                item = dict(p)
                _append_barcodes_for_item(
                    barcodes,
                    item,
                    request.form.get(f'extra_qty_retail_{i}', 0),
                    request.form.get(f'extra_qty_wholesale_{i}', 0),
                )
            i += 1
        return barcodes

    def _render_barcode_stamp_or_select(barcodes):
        if not barcodes:
            barcodes = [{'code': '', 'name': 'Chưa chọn tem', 'unit': '', 'price': 0}]
        return render_template('print_barcode_40stamp.html', barcodes=barcodes)

    @app.route('/print_barcode_import_select/<int:import_id>', methods=['GET', 'POST'])
    def print_barcode_import_select(import_id):
        conn = get_db_connection()
        c = conn.cursor()
        try:
            c.execute("SELECT import_no, date FROM import WHERE id = ?", (import_id,))
            imp_row = c.fetchone()
            if not imp_row:
                return "Không tìm thấy phiếu nhập", 404
            imp = dict(import_no=imp_row[0], date=imp_row[1])

            c.execute("""
                SELECT
                    ii.product_id,
                    ii.qty AS quantity,
                    p.name,
                    p.unit,
                    p.unit1,
                    COALESCE(p.unit_ratio, 1) AS unit_ratio,
                    COALESCE(p.price, 0) AS price,
                    COALESCE(p.base_price, 0) AS base_price,
                    COALESCE(p.barcode, '') AS barcode,
                    COALESCE(p.barcode1, '') AS barcode1
                FROM import_details ii
                JOIN products p ON p.id = ii.product_id
                WHERE ii.import_id = ?
                ORDER BY p.name
            """, (import_id,))
            items = [_barcode_print_item_from_row(row) for row in c.fetchall()]

            if request.method == 'POST':
                barcodes = []
                for item in items:
                    pid = item['product_id']
                    _append_barcodes_for_item(
                        barcodes,
                        item,
                        request.form.get(f'qty_retail_{pid}', 0),
                        request.form.get(f'qty_wholesale_{pid}', 0),
                    )
                barcodes.extend(_collect_extra_barcodes_from_form(request, c))
                return _render_barcode_stamp_or_select(barcodes)

            return render_template('print_barcode_import.html', imp=imp, items=items)
        finally:
            conn.close()

    @app.route('/print_barcode_product_select/<int:product_id>', methods=['GET', 'POST'])
    def print_barcode_product_select(product_id):
        conn = get_db_connection()
        c = conn.cursor()
        try:
            c.execute("""
                SELECT
                    p.id AS product_id,
                    COALESCE(i.quantity, 0) AS quantity,
                    p.name,
                    p.unit,
                    p.unit1,
                    COALESCE(p.unit_ratio, 1) AS unit_ratio,
                    COALESCE(p.price, 0) AS price,
                    COALESCE(p.base_price, 0) AS base_price,
                    COALESCE(p.barcode, '') AS barcode,
                    COALESCE(p.barcode1, '') AS barcode1
                FROM products p
                LEFT JOIN inventory i ON i.product_id = p.id
                WHERE p.id = ?
            """, (product_id,))
            row = c.fetchone()
            if not row:
                return "Không tìm thấy sản phẩm", 404

            item = _apply_stock_barcode_defaults(_barcode_print_item_from_row(row))
            items = [item]
            imp = dict(import_no=item['name'], date='')

            if request.method == 'POST':
                barcodes = []
                pid = item['product_id']
                _append_barcodes_for_item(
                    barcodes,
                    item,
                    request.form.get(f'qty_retail_{pid}', 0),
                    request.form.get(f'qty_wholesale_{pid}', 0),
                )
                barcodes.extend(_collect_extra_barcodes_from_form(request, c))
                return _render_barcode_stamp_or_select(barcodes)

            return render_template(
                'print_barcode_import.html',
                imp=imp,
                items=items,
                product_only=True,
            )
        finally:
            conn.close()

    #====== ROUTE IN BARCODE TỪ MÁY POS===========#
    @app.route('/pos/barcode/print')
    @login_required
    def pos_print_barcode():
        return render_template('pos_barcode_print.html')

    @app.route('/api/warehouses', methods=['GET'])
    @login_required
    def api_list_warehouses():
        conn = get_db_connection()
        try:
            from Services.import_line_helpers import list_active_warehouses
            return jsonify({'success': True, 'data': list_active_warehouses(conn)})
        finally:
            conn.close()

    @app.route('/import')
    @login_required
    def import_stock():
        mode = (request.args.get('mode') or 'stock').strip().lower()
        warehouses = []
        conn = get_db_connection()
        try:
            from Services.import_line_helpers import list_active_warehouses
            warehouses = list_active_warehouses(conn)
        finally:
            conn.close()
        return render_template(
            'import.html',
            today=datetime.now().strftime('%Y-%m-%d'),
            import_mode=mode,
            next_import_no=peek_next_import_no(mode),
            warehouses=warehouses,
        )

    @app.route('/import_list')
    @login_required
    def import_list():
        return render_template('import_list.html')

    # =====Route NHẬN DỮ LIỆU HÀNG TỒN KHO BAN ĐẦU — ghi đủ inventory + stock_moves =====
    @app.route('/api/inventory/import', methods=['POST'])
    @login_required
    @admin_or_master_required
    def api_import_stock():
        conn = get_db_connection()
        c = conn.cursor()

        data = request.get_json(silent=True) or {}
        items = data.get('items', []) or data.get('products', []) or []

        if not items:
            return jsonify({"success": False, "error": "Không có dữ liệu"}), 400

        try:
            inserted = 0
            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            for item in items:
                pid = item.get('product_id') or item.get('id') or item.get('productId')
                qty = float(item.get('quantity') or item.get('qty') or 0)
                note = str(item.get('note') or item.get('notes') or '').strip()
                if not pid or qty <= 0:
                    continue

                pid = int(pid)
                c.execute("SELECT avg_cost FROM inventory WHERE product_id = ?", (pid,))
                inv_row = c.fetchone()
                avg_cost = float(inv_row[0] or 0) if inv_row else float(item.get('cost_price') or item.get('avg_cost') or 0)

                c.execute("""
                    INSERT INTO stock_moves
                    (product_id, date, type, ref_id, quantity, cost_price, note, ref_document, ref_type, type1)
                    VALUES (?, ?, 'import', NULL, ?, ?, ?, 'initial', 'import', 'Nhập')
                """, (pid, now, qty, avg_cost, note or 'Nhập kho từ phiếu', 'PN-INIT'))

                c.execute("""
                    INSERT INTO inventory_transactions
                    (product_id, type, type1, quantity, cost_price, note, created_at)
                    VALUES (?, 'import', 'Nhập', ?, ?, ?, ?)
                """, (pid, qty, avg_cost, note or 'Nhập kho từ phiếu', now))

                if inv_row:
                    c.execute("UPDATE inventory SET avg_cost = ? WHERE product_id = ?", (avg_cost, pid))
                else:
                    c.execute(
                        "INSERT INTO inventory (product_id, quantity, avg_cost) VALUES (?, 0, ?)",
                        (pid, avg_cost),
                    )
                sync_inventory_quantity_from_moves(c, pid)
                inserted += 1

            conn.commit()
            return jsonify({
                "success": True,
                "message": f"Nhập kho thành công {inserted} sản phẩm!",
                "inserted": inserted,
            })

        except Exception as e:
            conn.rollback()
            print("LỖI NHẬP KHO TỪ IMPORTFORM:", e)
            return jsonify({"success": False, "error": "Lỗi server: " + str(e)}), 500

        finally:
            conn.close()

    #=== Kiểm Kê Kho ===#
    @app.route('/inventory/check', methods=['GET', 'POST'])
    @login_required
    def inventory_check():
        conn = get_db_connection()
        conn.row_factory = sqlite3.Row 
        c = conn.cursor()

        if request.method == 'POST':
            data = request.get_json()
            check_date = data.get('check_date', date.today().isoformat())
            # Thêm giờ phút giây để tránh trùng ID khi query MAX
            now_full = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            items = data.get('items', [])

            try:
                for item in items:
                    pid = item['product_id']
                    checked_qty = Decimal(str(item.get('checked_qty', 0)))

                    c.execute("SELECT avg_cost FROM inventory WHERE product_id = ?", (pid,))
                    row = c.fetchone()
                    ledger_qty = Decimal(str(ledger_quantity(c, pid)))
                    current_avg_cost = Decimal(str(row['avg_cost'] if (row and row['avg_cost']) else 0))

                    diff = checked_qty - ledger_qty
                    if diff == 0:
                        continue

                    # 2. Xác định Loại, Số phiếu và Ghi chú
                    if diff > 0:  # KIỂM KÊ DƯ -> TẠO PHIẾU NHẬP (PN)
                        move_type = 'import'
                        # Đếm số phiếu nhập để tạo ref_id
                        c.execute("""SELECT COUNT(DISTINCT COALESCE(ref_document, ref_id)) FROM stock_moves WHERE type IN ('import', 'RETURN_SALE')""")
                        new_seq = (c.fetchone()[0] or 0) + 1
                        ref_doc = f"PN{str(new_seq).zfill(6)}"
                        final_note = item.get('note') or "Nhập hàng dư theo số kiểm kê"
                    else:  # KIỂM KÊ THIẾU -> TẠO PHIẾU XUẤT (PX)
                        move_type = 'export'
                        # Đếm số phiếu xuất để tạo ref_id (bao gồm cả SALE, export...)
                        c.execute("""
                            SELECT COUNT(DISTINCT COALESCE(ref_document, ref_id)) 
                            FROM stock_moves 
                            WHERE type IN ('SALE', 'RETURN_IMPORT', 'DELETE_IMPORT', 'export')
                        """)
                        new_seq = (c.fetchone()[0] or 0) + 1
                        ref_doc = f"PX{str(new_seq).zfill(6)}"
                        final_note = item.get('note') or "Xuất điều chỉnh hàng thiếu theo kiểm kê"

                    # 3. Ghi stock_moves rồi sync inventory từ sổ cái
                    c.execute("""
                        INSERT INTO stock_moves 
                        (product_id, date, type, ref_id, ref_document, ref_type, quantity, note, type1, cost_price)
                        VALUES (?, ?, ?, ?, ?, 'inventory_check', ?, ?, 'Kiểm Kê', ?)
                    """, (
                        pid,
                        now_full,
                        move_type,
                        new_seq,
                        ref_doc,
                        float(diff),
                        final_note,
                        float(current_avg_cost),
                    ))
                    if row:
                        c.execute(
                            "UPDATE inventory SET avg_cost = ? WHERE product_id = ?",
                            (float(current_avg_cost), pid),
                        )
                    else:
                        c.execute(
                            "INSERT INTO inventory (product_id, quantity, avg_cost) VALUES (?, 0, ?)",
                            (pid, float(current_avg_cost)),
                        )
                    sync_inventory_quantity_from_moves(c, pid)

                conn.commit()
                return jsonify({"success": True, "message": "Kiểm kê thành công", "redirect": "/inventory/moves"})
            except Exception as e:
                conn.rollback()
                return jsonify({"success": False, "error": str(e)}), 500
            finally:
                conn.close()

        # GET: Load danh sách sản phẩm (giữ nguyên)
        c.execute("""
            SELECT p.id, p.name, p.unit, COALESCE(i.quantity, 0) as quantity
            FROM products p
            LEFT JOIN inventory i ON p.id = i.product_id
            ORDER BY p.id
        """)
        products = c.fetchall()
        conn.close()
        return render_template('inventory_check.html', products=products)

    #=== TẢI FILE EXCEL ĐỂ LẬP DANH SÁCH HÀNG NHẬP KHO===#
    from io import BytesIO
    from flask import send_file
    import openpyxl
    from openpyxl.comments import Comment
    from openpyxl.styles import Font, PatternFill, Alignment

    @app.route('/initial_stock/template')
    @login_required
    def initial_stock_template():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tồn kho ban đầu"

        # ==================== HEADER (dòng 1) ====================
        headers = [
            "Tên sản phẩm (*)", 
            "Đơn vị cơ bản (*)", 
            "Đơn vị sỉ/quy đổi", 
            "Tỷ lệ quy đổi (nếu có unit1)", 
            "Giá vốn nhập (buyprice)", 
            "Giá bán lẻ (base_price)", 
            "Giá bán sỉ (price)", 
            "Số lượng tồn đầu kỳ (*)", 
            "Ghi chú"
        ]
        ws.append(headers)

        # Style header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18 if col in [1,8] else 14

        # ==================== Dòng hướng dẫn / chú thích (dòng 2) ====================
        ws.append([
            "Ví dụ: Ống nước Bình Minh 21",
            "Ví dụ: Mét",
            "Ví dụ: Cây (để trống nếu không có)",
            "Ví dụ: 4 (1 Cây = 4 mét), mặc định 1",
            "Ví dụ: 10.000 đ/m (giá nhập)",
            "Ví dụ: 15.000 đ/m (giá bán lẻ)",
            "Ví dụ: 50.000 đ/cây (giá bán sỉ nếu có unit1)",
            "Ví dụ: 15.5 / 120 (lưu ý nhập số lượng lẻ/đơn vị nhỏ nhất)",
            "Ví dụ: Tồn đầu kỳ 07/02/2026"
        ])

        # Style dòng hướng dẫn
        guide_font = Font(italic=True, color="4B5563")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=2, column=col)
            cell.font = guide_font
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Thêm comment vào các cột bắt buộc
        ws['A1'].comment = Comment("Bắt buộc điền. Dùng để nhận diện sản phẩm.", "Hệ thống")
        ws['B1'].comment = Comment("Bắt buộc. Ví dụ: Cái, Mét, Bộ, Kg...", "Hệ thống")
        ws['H1'].comment = Comment("Bắt buộc. Số lượng tồn kho thực tế ban đầu.", "Hệ thống")

        # ==================== Tạo file và trả về ====================
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            download_name='Mau_Nhap_Ton_Kho_Ban_Dau.xlsx',
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    #=== API NHẬP DANH SÁCH VÀ SỐ LƯỢNG HÀNG TỒN KHO===#
    @app.route('/api/initial_stock/import', methods=['POST'])
    @login_required
    def api_initial_stock_import():
        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
    
        try:
            file = request.files.get('file')
            if not file:
                return jsonify({"success": False, "error": "Không tìm thấy file"}), 400

            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active

            imported_count = 0
            total_value_all = Decimal('0')
            errors = []
            current_now = datetime.now()
            import_date = current_now.strftime('%Y-%m-%d %H:%M:%S')
        
            # --- 1. LOGIC ĐÁNH SỐ PHIẾU PNxxxxxx ---
            # Đếm số lượng ref_id duy nhất của các loại nhập kho trong stock_moves
            c.execute("""
                SELECT COUNT(DISTINCT ref_id) 
                FROM stock_moves 
                WHERE LOWER(type) IN ('import', 'return_sale', 'delete_sale')
            """)
            current_pn_count = c.fetchone()[0] or 0
            # Tạo số phiếu mới định dạng PN + 6 chữ số
            new_voucher_no = f"PN{str(current_pn_count + 1).zfill(6)}"

            # Ghi chú chung
            master_note = "Nhập tồn kho ban đầu từ Excel"

            # --- 2. TẠO HEADER PHIẾU NHẬP KHO ---
            # Sử dụng new_voucher_no cho cột import_no (số chứng từ chính)
            c.execute("""
                INSERT INTO phieu_nhap_kho (
                    import_no, date, bill_no, supplier_name, supplier_id, 
                    total_amount, note
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (new_voucher_no, import_date, "Tồn kho cũ", "Nhập tồn kho cũ", 0, 0, master_note))
        
            import_id = c.lastrowid

            # Duyệt từ dòng 3
            for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
                if not row or not row[0]:
                    continue

                try:
                    name        = str(row[0]).strip()
                    unit        = str(row[1]).strip() if row[1] else "Cái"
                    unit1       = str(row[2]).strip() if row[2] else None
                    unit_ratio  = Decimal(str(row[3] or 1))
                    buyprice    = Decimal(str(row[4] or 0))
                    base_price  = Decimal(str(row[5] or 0))
                    price       = Decimal(str(row[6] or 0))
                    quantity    = Decimal(str(row[7] or 0))
                    note_row    = str(row[8]).strip() if row[8] else "Nhập tồn kho cũ"

                    if quantity <= 0: continue

                    # --- 3. XỬ LÝ SẢN PHẨM ---
                    c.execute("SELECT id FROM products WHERE name = ? LIMIT 1", (name,))
                    product_row = c.fetchone()

                    if product_row:
                        pid = product_row['id']
                        c.execute("""
                            UPDATE products SET 
                                unit=?, unit1=?, unit_ratio=?, buyprice=?, base_price=?, price=?, updated_at=?
                            WHERE id=?
                        """, (unit, unit1, float(unit_ratio), float(buyprice), float(base_price), float(price), import_date, pid))
                    else:
                        c.execute("SELECT MAX(id) as max_id FROM products")
                        max_id_row = c.fetchone()
                        next_id = (max_id_row['max_id'] or 0) + 1
                    
                        p_code = f"SP{str(next_id).zfill(4)}" 
                        b_code = f"{p_code}01"
                        b_code1 = f"{p_code}02" if unit1 else None

                        c.execute("""
                            INSERT INTO products (
                                product_code, barcode, barcode1, name, 
                                unit, unit1, unit_ratio, buyprice, base_price, price, updated_at
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (p_code, b_code, b_code1, name, unit, unit1, float(unit_ratio), float(buyprice), float(base_price), float(price), import_date))
                        pid = c.lastrowid

                    # --- 4. CHI TIẾT PHIẾU NHẬP ---
                    line_total = quantity * buyprice
                    total_value_all += line_total
                
                    c.execute("""
                        INSERT INTO chi_tiet_phieu_nhap_kho (
                            import_id, product_id, quantity, buyprice, 
                            subtotal, discount_amount, tax_amount, cost_price, unit_type, note, date
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?)
                    """, (import_id, pid, float(quantity), float(buyprice), float(line_total), 0, 0, float(buyprice), 'Nhập tồn kho cũ', import_date))

                    # --- 5. STOCK_MOVES (Lịch sử biến động) ---
                    # Quan trọng: ref_document lưu số phiếu PNxxxxxx
                    c.execute("""
                        INSERT INTO stock_moves (
                            product_id, date, type, ref_id, ref_type, 
                            quantity, cost_price, note, type1, ref_document
                        ) VALUES (?, ?, 'import', ?, 'initial_import', ?, ?, ?, 'Nhập tồn cũ', ?)
                    """, (pid, import_date, import_id, float(quantity), float(buyprice), note_row, new_voucher_no))

                    # --- 6. CẬP NHẬT GIÁ VỐN + SYNC TỒN TỪ SỔ CÁI ---
                    c.execute("SELECT avg_cost FROM inventory WHERE product_id = ?", (pid,))
                    inv_row = c.fetchone()
                    old_c = float(inv_row[0] or 0) if inv_row else float(buyprice)
                    ledger_before = ledger_quantity(c, pid)
                    new_avg = (
                        ((ledger_before * old_c) + (float(quantity) * float(buyprice)))
                        / (ledger_before + float(quantity))
                    ) if (ledger_before + float(quantity)) > 0 else float(buyprice)
                    if inv_row:
                        c.execute(
                            "UPDATE inventory SET avg_cost = ? WHERE product_id = ?",
                            (new_avg, pid),
                        )
                    else:
                        c.execute(
                            "INSERT INTO inventory (product_id, quantity, avg_cost) VALUES (?, 0, ?)",
                            (pid, new_avg),
                        )
                    sync_inventory_quantity_from_moves(c, pid)

                    # --- 7. THẺ KHO ---
                    c.execute("""
                        INSERT INTO inventory_transactions (
                            product_id, type, type1, quantity, cost_price, 
                            reference_id, reference_type, note, created_at
                        ) VALUES (?, 'import', 'Nhập tồn cũ', ?, ?, ?, 'initial_import', ?, ?)
                    """, (pid, float(quantity), float(buyprice), import_id, note_row, import_date))

                    imported_count += 1

                except Exception as e:
                    errors.append(f"Dòng {row_idx}: {str(e)}")
                    continue

            # Cập nhật tổng tiền
            c.execute("UPDATE phieu_nhap_kho SET total_amount = ? WHERE id = ?", (float(total_value_all), import_id))

            conn.commit()
            return jsonify({
                "success": True, 
                "message": f"Đã nhập thành công {imported_count} mặt hàng.",
                "voucher_no": new_voucher_no,
                "import_id": import_id
            })

        except Exception as e:
            if conn: conn.rollback()
            return jsonify({"success": False, "error": str(e)}), 500
        finally:
            if conn: conn.close()


    #=============================================================== Start of Inventory, Stock_moves & Inventory_transactions' APIs and Route================================================#
    # === API: BÁO CÁO TỒN KHO CHI TIẾT (inventory_detail.html) ===#
    @app.route('/api/inventory/detail', methods=['POST'])
    @login_required
    def api_inventory_detail():
        data = request.get_json()
        start_date_str = data.get('start_date')  # YYYY-MM-DD
        end_date_str = data.get('end_date')      # YYYY-MM-DD

        if not start_date_str or not end_date_str:
            return jsonify({"success": False, "error": "Thiếu tham số ngày"}), 400

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        except ValueError:
            return jsonify({"success": False, "error": "Định dạng ngày không hợp lệ"}), 400

        period_start = start_date.replace(hour=0, minute=0, second=0)
        period_end = end_date.replace(hour=23, minute=59, second=59)

        try:
            conn = get_db_connection()
            conn.row_factory = sqlite3.Row
            c = conn.cursor()

            # SIÊU TRUY VẤN: Khớp hoàn hảo với dữ liệu xuất kho lưu số ÂM
            c.execute("""
                SELECT 
                    p.id AS product_id,
                    p.name AS product_name,
                    p.product_code AS product_code,
                    COALESCE(p.unit, 'Cái') AS unit_name,
                
                    -- 1. Tính số lượng tồn đầu kỳ (Chỉ cần SUM hết vì xuất đã tự mang dấu âm)
                    COALESCE(SUM(CASE 
                        WHEN sm.date < ? AND sm.type IN ('import', 'RETURN_SALE', 'DELETE_SALE', 'SALE', 'export', 'RETURN_IMPORT', 'DELETE_IMPORT', 'export_for_use', 'export_material', 'adjustment') THEN sm.quantity
                        ELSE 0 
                    END), 0) AS beginning_quantity,

                    -- 2. Tính giá trị tồn đầu kỳ (SUM hết vì số lượng xuất âm nhân cost_price dương sẽ ra giá trị âm)
                    COALESCE(SUM(CASE 
                        WHEN sm.date < ? AND sm.type IN ('import', 'RETURN_SALE', 'DELETE_SALE', 'SALE', 'export', 'RETURN_IMPORT', 'DELETE_IMPORT', 'export_for_use', 'export_material', 'adjustment') THEN sm.quantity * sm.cost_price
                        ELSE 0 
                    END), 0) AS beginning_value,

                    -- 3. Phát sinh Nhập trong kỳ (Các loại chứng từ làm tăng kho, quantity dương)
                    COALESCE(SUM(CASE 
                        WHEN sm.date >= ? AND sm.date <= ? AND sm.type IN ('import', 'RETURN_SALE', 'DELETE_SALE') THEN sm.quantity
                        WHEN sm.date >= ? AND sm.date <= ? AND sm.type = 'adjustment' AND sm.quantity > 0 THEN sm.quantity
                        ELSE 0 
                    END), 0) AS import_quantity,
                
                    COALESCE(SUM(CASE 
                        WHEN sm.date >= ? AND sm.date <= ? AND sm.type IN ('import', 'RETURN_SALE', 'DELETE_SALE') THEN sm.quantity * sm.cost_price
                        WHEN sm.date >= ? AND sm.date <= ? AND sm.type = 'adjustment' AND sm.quantity > 0 THEN sm.quantity * sm.cost_price
                        ELSE 0 
                    END), 0) AS import_value,

                    -- 4. Phát sinh Xuất trong kỳ (Dùng dấu trừ phía trước để chuyển số ÂM trong DB thành số DƯƠNG hiển thị báo cáo)
                    COALESCE(SUM(CASE 
                        WHEN sm.date >= ? AND sm.date <= ? AND sm.type IN ('SALE', 'export', 'RETURN_IMPORT', 'DELETE_IMPORT', 'export_for_use', 'export_material') THEN -sm.quantity
                        WHEN sm.date >= ? AND sm.date <= ? AND sm.type = 'adjustment' AND sm.quantity < 0 THEN -sm.quantity
                        ELSE 0 
                    END), 0) AS export_quantity,

                    COALESCE(SUM(CASE 
                        WHEN sm.date >= ? AND sm.date <= ? AND sm.type IN ('SALE', 'export', 'RETURN_IMPORT', 'DELETE_IMPORT', 'export_for_use', 'export_material') THEN -sm.quantity * sm.cost_price
                        WHEN sm.date >= ? AND sm.date <= ? AND sm.type = 'adjustment' AND sm.quantity < 0 THEN -sm.quantity * sm.cost_price
                        ELSE 0 
                    END), 0) AS export_value

                FROM products p
                LEFT JOIN stock_moves sm ON p.id = sm.product_id
                WHERE COALESCE(p.product_type, 'goods') != 'service'
                  AND UPPER(COALESCE(p.product_code, '')) NOT LIKE 'DV%'
                GROUP BY p.id
                ORDER BY p.name
            """, (
                # Tham số cho Tồn đầu kỳ
                period_start, period_start,
                # Tham số cho Nhập trong kỳ
                period_start, period_end, period_start, period_end,
                period_start, period_end, period_start, period_end,
                # Tham số cho Xuất trong kỳ
                period_start, period_end, period_start, period_end,
                period_start, period_end, period_start, period_end
            ))

            rows = c.fetchall()
            report = []

            for r in rows:
                beg_qty = float(r['beginning_quantity'])
                beg_val = float(r['beginning_value'])
                imp_qty = float(r['import_quantity'])
                imp_val = float(r['import_value'])
                exp_qty = float(r['export_quantity']) # Đây là số dương (đã đổi dấu từ SQL)
                exp_val = float(r['export_value'])    # Đây là số dương (đã đổi dấu từ SQL)

                # Vì exp_qty thu được từ SQL đã đổi sang số dương, nên công thức tính cuối kỳ sẽ là trừ đi
                end_qty = beg_qty + imp_qty - exp_qty
                end_val = beg_val + imp_val - exp_val

                report.append({
                    "product_id": r['product_id'],
                    "product_code": r['product_code'] or "",
                    "product_name": r['product_name'],
                    "unit_name": r['unit_name'],
                
                    "beginning_quantity": beg_qty,
                    "beginning_value": max(0.0, beg_val),
                
                    "import_quantity": imp_qty,
                    "import_value": imp_val,
                
                    "export_quantity": exp_qty,
                    "export_value": exp_val,
                
                    "ending_quantity": end_qty,
                    "ending_value": max(0.0, end_val)
                })

            return jsonify({
                "success": True,
                "period": f"{start_date_str} → {end_date_str}",
                "data": report
            })

        except Exception as e:
            print("LỖI BÁO CÁO TỒN KHO:", e)
            import traceback
            traceback.print_exc()
            return jsonify({"success": False, "error": "Lỗi hệ thống", "detail": str(e)}), 500
        finally:
            if 'conn' in locals() and conn:
                conn.close()

    @app.route('/api/inventory/reconcile', methods=['POST'])
    @login_required
    @admin_or_master_required
    def api_inventory_reconcile():
        """Đối soát inventory.quantity ← SUM(stock_moves) cho toàn bộ sản phẩm."""
        conn = None
        try:
            conn = get_db_connection()
            c = conn.cursor()
            fixes = reconcile_all_inventory(c)
            conn.commit()
            return jsonify({
                "success": True,
                "fixed_count": len(fixes),
                "fixes": fixes[:100],
                "message": f"Đã đối soát {len(fixes)} sản phẩm.",
            })
        except Exception as e:
            if conn:
                conn.rollback()
            return jsonify({"success": False, "error": str(e)}), 500
        finally:
            if conn:
                conn.close()

    @app.route('/api/inventory/reconcile-wac', methods=['POST'])
    @login_required
    @admin_or_master_required
    def api_inventory_reconcile_wac():
        """Tính lại avg_cost từ stock_moves + đối soát số lượng."""
        conn = None
        try:
            conn = get_db_connection()
            c = conn.cursor()
            qty_fixes = reconcile_all_inventory(c)
            wac_fixes = rebuild_all_wac_from_moves(c)
            conn.commit()
            return jsonify({
                'success': True,
                'qty_fixed_count': len(qty_fixes),
                'wac_fixed_count': len(wac_fixes),
                'wac_fixes': wac_fixes[:100],
                'message': f'Đối soát SL: {len(qty_fixes)} SP; rebuild WAC: {len(wac_fixes)} SP.',
            })
        except Exception as e:
            if conn:
                conn.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500
        finally:
            if conn:
                conn.close()

    @app.route('/api/inventory', methods=['GET'])
    @login_required
    def api_inventory():
        """
        API endpoint lấy dữ liệu Tồn kho hiện tại và Lịch sử nhập xuất theo WAC
        ĐÃ SỬA LỖI LỆCH NGÀY (bao gồm giao dịch đến hết ngày cuối cùng)
        """
        start_date_str = request.args.get('start') # Ví dụ: "2025-11-23"
        end_date_str = request.args.get('end')     # Ví dụ: "2025-11-23"
        query_search = f"%{request.args.get('q', '')}%"

        if not start_date_str or not end_date_str:
            return jsonify({"current_stock": [], "history": []})

        # --- SỬA LỖI LỆCH NGÀY QUAN TRỌNG NHẤT ---
        # 1. Ngày Bắt đầu: Bắt đầu từ 00:00:00 của ngày được chọn
        start_of_day = f"{start_date_str} 00:00:00"
    
        # 2. Ngày Kết thúc: Kết thúc vào 23:59:59 của ngày được chọn
        end_of_day = f"{end_date_str} 23:59:59"
        # ----------------------------------------
    
        try:
            conn = get_db_connection()

            # --- 1. Tồn kho: inventory + join products (mã/tên/ĐVT) ---
            # Chỉ SP đã có phát sinh stock_moves; loại dịch vụ; gồm cả tồn = 0
            current_stock_query = """
                SELECT
                    i.product_id,
                    p.product_code,
                    p.name,
                    p.barcode,
                    COALESCE(p.unit, 'Cái') AS unit,
                    COALESCE(
                        (SELECT SUM(sm.quantity) FROM stock_moves sm WHERE sm.product_id = i.product_id),
                        i.quantity,
                        0
                    ) AS quantity,
                    COALESCE(i.avg_cost, 0) AS avg_cost
                FROM inventory i
                INNER JOIN products p ON p.id = i.product_id
                WHERE COALESCE(p.product_type, 'goods') != 'service'
                  AND EXISTS (
                      SELECT 1 FROM stock_moves sm WHERE sm.product_id = i.product_id
                  )
                  AND (p.name LIKE ? OR p.barcode LIKE ? OR p.product_code LIKE ?)
                ORDER BY p.name
            """
            current_stock = conn.execute(
                current_stock_query,
                (query_search, query_search, query_search),
            ).fetchall()
            current_stock_list = []
            for row in current_stock:
                qty = row["quantity"]
                avg_cost = row["avg_cost"]
                current_stock_list.append({
                    "product_id": row["product_id"],
                    "product_code": row["product_code"] or "",
                    "name": row["name"],
                    "barcode": row["barcode"],
                    "unit": row["unit"],
                    "quantity": qty,
                    "avg_cost": avg_cost,
                    "total_value": qty * avg_cost,
                })

            # --- 2. Lịch sử nhập xuất: stock_moves + join products ---
            # Chỉ giao dịch kho thực tế; loại dịch vụ
            history_query = """
                SELECT
                    sm.date,
                    sm.product_id,
                    p.product_code,
                    p.name,
                    p.barcode,
                    COALESCE(p.unit, sm.unit, 'Cái') AS unit,
                    sm.type1,
                    sm.quantity,
                    sm.cost_price,
                    sm.ref_type AS reference_type,
                    sm.ref_id AS reference_id,
                    sm.note,
                    sm.ref_document
                FROM stock_moves sm
                INNER JOIN products p ON p.id = sm.product_id
                WHERE COALESCE(p.product_type, 'goods') != 'service'
                  AND sm.date BETWEEN ? AND ?
                  AND (p.name LIKE ? OR p.barcode LIKE ? OR p.product_code LIKE ?)
                ORDER BY sm.date DESC
            """
            history_data = conn.execute(
                history_query,
                (start_of_day, end_of_day, query_search, query_search, query_search),
            ).fetchall()
            history_list = []

            for row in history_data:
                qty = row["quantity"]
                cost = row["cost_price"] if row["cost_price"] is not None else 0
                if cost == 0 and row["type1"] in ["Bán", "Khách Trả"]:
                    c = conn.cursor()
                    c.execute("SELECT avg_cost FROM inventory WHERE product_id = ?", (row["product_id"],))
                    avg = c.fetchone()
                    cost = avg["avg_cost"] if avg else 0

                total_value = qty * cost

                # --- Chuẩn hóa nguồn tham chiếu ---
                source_ref = f"Ref: {row['reference_id']}"
                if row['type1'] == 'Nhập':
                    source_ref = f"PNK #{row['reference_id']}"
                elif row['type1'] == 'Bán':
                    source_ref = f"HĐ #{row['reference_id']}"
                elif row['type1'] == 'Điều chỉnh':
                    source_ref = f"Điều chỉnh #{row['reference_id']}"
                elif row['type1'] == 'Trả HN':
                    source_ref = f"Trả nhập #{row['reference_id']}"
                elif row['type1'] == 'Khách Trả':
                    source_ref = f"Trả bán #{row['reference_id']}"

                history_list.append({
                    "date": row["date"],
                    "product_id": row["product_id"],
                    "product_code": row["product_code"] or "",
                    "name": row["name"],
                    "barcode": row["barcode"],
                    "unit": row["unit"],
                    "type": row["type1"],
                    "quantity": qty,
                    "cost_price": cost,
                    "total_value": total_value,
                    "source": source_ref,
                    "note": row["note"],
                })

            # Đảm bảo đóng kết nối ngay cả khi có lỗi (nên dùng context manager)
            conn.close() 
        
            return jsonify({
                "current_stock": current_stock_list,
                "history": history_list
            })

        except Exception as e:
            print(f"Lỗi khi xử lý API inventory: {e}")
            # Đóng kết nối nếu nó vẫn mở
            try:
                conn.close()
            except:
                pass
            return jsonify({"success": False, "message": "Lỗi máy chủ khi tải dữ liệu tồn kho.", "error": str(e)}), 500

    # Routes & API Inventory # Nhân viên có quyền view_inventory mới thấy trang tồn kho
    @app.route('/inventory')
    @login_required
    def inventory():
        return render_template('inventory.html')

    @app.route('/inventory/detail')
    @login_required
    def inventory_detail():
        return render_template('inventory_detail.html') 
    @app.route('/return/import')
    @login_required
    def return_import_page():
        return render_template('return_import.html')

    @app.route('/import_details')
    @login_required
    def import_details_page():
        return render_template('import_details.html')

    @app.route('/api/import/details-report', methods=['POST'])
    @login_required
    def api_import_details_report():
        data = request.get_json(silent=True) or {}
        start_date = (data.get('start_date') or '').strip()
        end_date = (data.get('end_date') or '').strip()
        search_query = (data.get('q') or '').strip() or None

        if not start_date or not end_date:
            return jsonify({"success": False, "error": "Thiếu tham số ngày bắt đầu/kết thúc"}), 400

        try:
            datetime.strptime(start_date, '%Y-%m-%d')
            datetime.strptime(end_date, '%Y-%m-%d')
        except ValueError:
            return jsonify({"success": False, "error": "Định dạng ngày không hợp lệ (YYYY-MM-DD)"}), 400

        if start_date > end_date:
            return jsonify({"success": False, "error": "Ngày bắt đầu phải nhỏ hơn hoặc bằng ngày kết thúc"}), 400

        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        try:
            cursor = conn.cursor()
            rows, summary = fetch_import_items_detail_report(
                cursor, start_date, end_date, search_query
            )
            return jsonify({
                "success": True,
                "data": rows,
                "summary": summary,
                "start_date": start_date,
                "end_date": end_date,
            })
        except Exception as e:
            traceback.print_exc()
            return jsonify({"success": False, "error": str(e)}), 500
        finally:
            conn.close()

    @app.route('/api/import/invoice_pdf_url', methods=['GET'])
    @login_required
    def api_import_invoice_pdf_url():
        """Lấy link PDF hóa đơn mua từ supplier_invoice."""
        bill_no = (request.args.get('bill_no') or '').strip()
        tax_code = (request.args.get('tax_code') or '').strip()
        if not bill_no:
            return jsonify({"success": False, "error": "Thiếu số hóa đơn mua"}), 400

        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        try:
            c = conn.cursor()
            c.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name='supplier_invoice'"
            )
            if not c.fetchone():
                return jsonify({"success": False, "error": "Chưa có dữ liệu hóa đơn đầu vào"}), 404

            sql = """
                SELECT id, pdf_url FROM supplier_invoice
                WHERE TRIM(COALESCE(invoice_no, '')) = ?
                  AND pdf_url IS NOT NULL AND TRIM(pdf_url) != ''
            """
            params = [bill_no]
            if tax_code:
                sql += " AND TRIM(COALESCE(seller_tax_code, '')) = ?"
                params.append(tax_code)
            sql += " ORDER BY id DESC LIMIT 1"
            c.execute(sql, params)
            row = c.fetchone()
            if not row and tax_code:
                c.execute(
                    """
                    SELECT id, pdf_url FROM supplier_invoice
                    WHERE TRIM(COALESCE(invoice_no, '')) = ?
                      AND pdf_url IS NOT NULL AND TRIM(pdf_url) != ''
                    ORDER BY id DESC LIMIT 1
                    """,
                    (bill_no,),
                )
                row = c.fetchone()
            if not row or not row['id']:
                return jsonify({"success": False, "error": "Không tìm thấy link PDF hóa đơn mua"}), 404
            return jsonify({
                "success": True,
                "pdf_url": f"/api/invoices/inward/{row['id']}/pdf",
                "invoice_id": row['id'],
            })
        except Exception as e:
            traceback.print_exc()
            return jsonify({"success": False, "error": str(e)}), 500
        finally:
            conn.close()
