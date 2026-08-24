# app.py - POS SYSTEM - UPDATED (dotenv + stock_moves + sale complete)
import atexit
import calendar
import glob
import hashlib
import json
import logging
import os
import pyotp
import random
import re
import schedule
import secrets
import shutil
import smtplib
import sqlite3
import string
import tempfile
import threading
import time
import traceback
import uuid
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from email.message import EmailMessage
from functools import wraps
from io import BytesIO
from sqlite3 import Row
from typing import Optional
from unicodedata import normalize
from xml.dom import minidom
from zoneinfo import ZoneInfo  # Python 3.9+

import pandas as pd
import pdfkit  # IN PDF
import pymysql
import requests  # GỌI API HÓA ĐƠN ĐIỆN TỬ
from authlib.integrations.flask_client import OAuth
from dotenv import load_dotenv
from flask import (
    Flask,
    abort,
    current_app,
    flash,
    g,
    jsonify,
    make_response,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    session,
    url_for,
)
from flask_bcrypt import Bcrypt  # BẢO MẬT MẬT KHẨU
from flask_login import LoginManager, UserMixin, current_user, login_required, login_user, logout_user
from flask_mail import Mail, Message
from flask_sqlalchemy import SQLAlchemy
from flask_wtf import FlaskForm
from num2words import num2words
from requests.auth import HTTPBasicAuth  # Hóa đơn điện tử Viettel
from sqlalchemy import func
from urllib3.exceptions import InsecureRequestWarning
from wtforms import DateField, SelectField, StringField
from wtforms.validators import DataRequired, Length, Regexp
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

from db_utils import BASE_DIR, MAIN_DB_PATH, get_db_connection
from tenant_middleware import (
    add_user_to_mapping,
    get_tenant_by_username,
    init_tenant,
    init_tenant_middleware,
    update_user_email_in_mapping,
)

requests.packages.urllib3.disable_warnings(category=InsecureRequestWarning)

load_dotenv()
# ÉP FLASK TÌM ĐÚNG THƯ MỤC templates – FIX LỖI VĨNH VIỄN!
template_dir = os.path.abspath('templates')
app = Flask(__name__, template_folder=template_dir)

# ==============================================================================
# --- KHỐI CẤU HÌNH ĐƯỜNG DẪN ĐỒNG BỘ CHUẨN TUYỆT ĐỐI (CHỈ KHAI BÁO 1 LẦN) ---
# BASE_DIR, MAIN_DB_PATH, get_db_connection → db_utils.py
# ==============================================================================
STATIC_DIR    = os.path.join(BASE_DIR, 'static')
BACKUP_DIR    = os.path.join(BASE_DIR, 'backups')
UPLOAD_FOLDER = os.path.join(STATIC_DIR, 'img')      # Ép lưu chuẩn vào POS/static/img/

# Đăng ký thư mục upload chuẩn vào cấu hình của Flask
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

# ==============================================================================

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def thuần_thục_tên_file(filename):
    """
    Hàm dọn dẹp tên file chạy tốt trên cả Windows (Localhost) và Linux (VPS).
    Khử sạch dấu tiếng Việt, đổi khoảng trắng thành dấu gạch ngang.
    Giữ đúng tên gốc ban đầu, không thêm bất kỳ nội dung hay timestamp nào.
    """
    if '.' in filename:
        name_part, ext_part = filename.rsplit('.', 1)
        ext_part = ext_part.lower()
    else:
        name_part, ext_part = filename, 'jpg'

    # Chuyển thành chữ thường để chuẩn hóa
    name_part = name_part.lower()

    # Mảng chuyển đổi toàn bộ ký tự tiếng Việt có dấu thành không dấu chuẩn tuyệt đối
    co_dau = "áàảãạăắằẳẵặâấầẩẫậéèẻẽẹêếềểễệíìỉĩịóòỏõọôốồổỗộơớờởỡợúùủũụưứừửữựýỳỷỹỵđ"
    khong_dau = "aaaaaaaaaaaaaaaaaeeeeeeeeeeeiiiiiooooooooooooooooouuuuuuuuuuuyyyyyd"
    bang_dich = str.maketrans(co_dau, khong_dau)
    name_part = name_part.translate(bang_dich)

    # Thay thế các ký tự đặc biệt, khoảng trắng, gạch dưới thành dấu gạch ngang đơn (-)
    name_part = re.sub(r'[^\w\s-]', '', name_part)
    name_part = re.sub(r'[\s_]+', '-', name_part)
    name_part = name_part.strip('-')

    if not name_part:
        name_part = "uploaded_file"

    return f"{name_part}.{ext_part}"

# Khởi tạo OAuth
oauth = OAuth(app) 
google = oauth.register(
    name='google',
    client_id='YOUR_CLIENT_ID',
    client_secret='YOUR_CLIENT_SECRET',
    server_metadata_url='https://accounts.google.com/.well-known/openid-configuration',
    client_kwargs={'scope': 'openid email profile'}
)

# Import các hàm hỗ trợ sau khi đã khai báo app
init_tenant_middleware(app, get_db_connection)

# Khởi tạo Multi-Tenant
init_tenant(app)

# Tự động thêm tenant_id vào tất cả url_for
from flask import url_for as original_url_for

def url_for(endpoint, **values):
    tenant_id = getattr(g, 'tenant_id', None)
    if tenant_id and endpoint != 'static' and 'tenant_id' not in values:
        values.setdefault('tenant_id', tenant_id)
    return original_url_for(endpoint, **values)

app.add_template_global(url_for, 'url_for')



def so_thanh_chu(so):
    don_vi = ['', 'một', 'hai', 'ba', 'bốn', 'năm', 'sáu', 'bảy', 'tám', 'chín']
    hang_chuc = ['', 'mười', 'hai mươi', 'ba mươi', 'bốn mươi', 'năm mươi', 'sáu mươi', 'bảy mươi', 'tám mươi', 'chín mươi']
    nhom_lon = ['', 'nghìn', 'triệu', 'tỉ']
    
    def doc_ba_chu_so(n):
        if n == 0: return ''
        chuoi = ''
        hang_tram = n // 100
        n %= 100
        if hang_tram > 0:
            chuoi += don_vi[hang_tram] + ' trăm'
            if n > 0: chuoi += ' '
        if n > 0:
            if n < 10:
                chuoi += ('linh ' if hang_tram > 0 else '') + don_vi[n]
            elif n < 20:
                hang_dv = n % 10
                if hang_dv == 0:
                    chuoi += 'mười'
                elif hang_dv == 5:
                    chuoi += 'mười lăm'
                else:
                    chuoi += 'mười ' + don_vi[hang_dv]
            else:
                hang_ch = n // 10
                hang_dv = n % 10
                chuoi += hang_chuc[hang_ch]
                if hang_dv > 0:
                    chuoi += ' '
                    if hang_dv == 1: chuoi += 'mốt'
                    elif hang_dv == 5: chuoi += 'lăm'
                    else: chuoi += don_vi[hang_dv]
        return chuoi
    
    try:
        so = int(so)
    except:
        return "Không đồng chẵn"
    
    if so == 0:
        return "Không đồng chẵn"
    
    if so < 0:
        return "âm " + so_thanh_chu(-so)
    
    ket_qua = []
    i = 0
    while so > 0:
        nhom = so % 1000
        so //= 1000
        chuoi_nhom = doc_ba_chu_so(nhom)
        if chuoi_nhom:
            chuoi_nhom += (' ' + nhom_lon[i]) if i > 0 else ''
            ket_qua.append(chuoi_nhom.strip())
        i += 1
    
    ket_qua = ' '.join(reversed(ket_qua)).strip() + " đồng chẵn"

    # ← THÊM ĐOẠN NÀY ĐỂ VIẾT HOA CHỮ CÁI ĐẦU
    if ket_qua:
        ket_qua = ket_qua[0].upper() + ket_qua[1:]

    return ket_qua

# ĐĂNG KÝ FILTER – BẮT BUỘC PHẢI CÓ 2 DÒNG NÀY!
app.jinja_env.filters['so_thanh_chu'] = so_thanh_chu

logging.getLogger(__name__).info("Filter 'so_thanh_chu' đã được đăng ký.")

def format_price(price):
    if price is None:
        return "0"
    try:
        return f"{int(float(price)):,.0f}".replace(",", ".")
    except:
        return "0"

@app.template_filter('format_number')
def format_number(value):
    try:
        n = float(value)
        return f"{n:,.0f}".replace(",", ".")
    except:
        return "0"

def get_business_config():
    conn = get_db_connection()
    # Lấy thông tin hộ kinh doanh đang hoạt động mới nhất
    config = conn.execute('SELECT * FROM business_info WHERE is_active = 1 ORDER BY id DESC').fetchone()
    conn.close()
    return dict(config) if config else {}

@app.context_processor
def inject_business_info():
    """
    Cung cấp thông tin doanh nghiệp cho tất cả Template Frontend.
    Ưu tiên lấy dữ liệu từ bảng 'business_info' của database hiện tại (g.db_path).
    """
    # 1. Khởi tạo giá trị mặc định từ tenant_info (nếu có)
    # Điều này giúp hiển thị tên shop cơ bản ngay cả khi bảng business_info trống
    business_data = dict(g.tenant_info) if g.get('tenant_info') else {}

    try:
        # 2. Kết nối tới DB hiện tại (Hàm này đã được fix dùng g.db_path ở các bước trước)
        conn = get_db_connection()
        row = conn.execute("SELECT * FROM business_info LIMIT 1").fetchone()
        conn.close()

        if row:
            # Chuyển row thành dict và cập nhật/ghi đè vào business_data
            # Dữ liệu trong bảng business_info của Tenant sẽ có độ ưu tiên cao nhất
            business_data.update(dict(row))

        return {'info': business_data}

    except Exception as e:
        # Trong trường hợp bảng business_info chưa tồn tại ở một số tenant cũ
        print(f"Lỗi inject_business_info ({g.get('tenant_id', 'Main')}): {e}")
        return {'info': business_data}

@app.context_processor
def inject_now():
    return {'now': datetime.now()}

@app.context_processor
def inject_current_user():
    user = session.get('user', {})
    return dict(
        current_user = {
            'id': user.get('id'),
            'username': user.get('username'),
            'role': user.get('role', 'guest'),
            'full_name': user.get('full_name', 'Khách'),
            'permissions': user.get('permissions', '').split(',')
        }
    )

# ================== 2 FILTER BẮT BUỘC – DÁN NGAY SAU app = Flask(__name__) ==================
def format_date(value, fmt='%d/%m/%Y'):
    if not value:
        return '—'
    if isinstance(value, str):
        value = value.strip()
        for f in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y', '%d-%m-%Y'):
            try:
                return datetime.strptime(value, f).strftime(fmt)
            except:
                pass
        return value
    return value.strftime(fmt) if hasattr(value, 'strftime') else str(value)

def vnd(value):
    try:
        return "{:,.0f}".format(float(value or 0)).replace(",", ".") + " ₫"
    except:
        return "0 ₫"

# 2 DÒNG QUAN TRỌNG NHẤT – KHÔNG ĐƯỢC THIẾU!!!
app.jinja_env.filters['format_date'] = format_date
app.jinja_env.filters['vnd'] = vnd
# =========================================================================================

bcrypt = Bcrypt(app) # Khởi tạo bcrypt
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'pos_secret_key_2026_secure')

# --- Database Setup (SQLite) ---
DATABASE = 'database.db'

def parse_date(date_str):
    if not date_str:
        return None
    date_str = str(date_str).strip()
    # Định dạng chính xác trong DB của bạn
    formats = [
        '%Y-%m-%d %H:%M:%S',   # 2026-01-04 11:20:06 và 2026-01-05 07:25:21 ← HOÀN HẢO CHO BẠN
        '%Y-%m-%d %H:%M',
        '%Y-%m-%d',
        '%d/%m/%Y %H:%M:%S',
        '%d/%m/%Y',
        '%d-%m-%Y %H:%M:%S',
        '%d-%m-%Y',
        '%Y/%m/%d %H:%M:%S',
        '%Y/%m/%d',
    ]
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    # Nếu vẫn lỗi, in ra để bạn biết (có thể xóa dòng này sau)
    print(f"[ERROR PARSE DATE] Không parse được: '{date_str}'")
    return None

def get_product_list_with_stock(query=None):
    """
    Truy vấn sản phẩm và LEFT JOIN với tồn kho (inventory) để lấy số lượng.
    Bao gồm cột barcode1 và xử lý lỗi chi tiết.
    """
    conn = get_db_connection()
    if conn is None:
        return jsonify({"success": False, "error": "Không thể kết nối cơ sở dữ liệu."}), 500
        
    c = conn.cursor()
    
    sql = """
        SELECT
            p.id, p.name, p.barcode, p.base_price, p.unit, 
            p.unit1, p.unit_ratio, p.price as sale_price,
            p.barcode1,
            COALESCE(i.quantity, 0) AS quantity
        FROM products p
        LEFT JOIN inventory i ON p.id = i.product_id
        WHERE 1=1
    """
    params = []
    
    if query:
        # Tìm kiếm trong 3 cột: tên, mã vạch cơ bản, hoặc mã vạch đơn vị bán
        sql += " AND (p.name LIKE ? OR p.barcode LIKE ? OR p.barcode1 LIKE ?)"
        # Đảm bảo 3 tham số được truyền vào
        params.extend([f'%{query}%', f'%{query}%', f'%{query}%'])
    
    sql += " LIMIT 50"
        
    try:
        c.execute(sql, tuple(params))
        products = c.fetchall()
        
        result = [dict(row) for row in products]
        
        return jsonify(result), 200
        
    except sqlite3.OperationalError as e:
        # Xử lý lỗi SQL cụ thể, ví dụ: 'no such column' hoặc 'no such table'
        conn.rollback()
        print(f"LỖI VẬN HÀNH SQL (Kiểm tra tên cột/bảng): {e}")
        return jsonify({"success": False, "error": f"Lỗi truy vấn SQL: {e}. Vui lòng kiểm tra console server."}), 500
    except Exception as e:
        # Xử lý các lỗi khác
        conn.rollback()
        print(f"LỖI HỆ THỐNG KHÁC KHI TÌM KIẾM: {e}")
        return jsonify({"success": False, "error": f"Lỗi hệ thống không xác định: {e}"}), 500
    finally:
        conn.close()

# ==================== HÀM KẾT NỐI DATABASE THỐNG NHẤT (MULTI-TENANT) ====================

@app.teardown_appcontext
def close_db(error):
    """Đóng connection sau khi request kết thúc"""
    db = g.pop('db', None)
    if db is not None:
        db.close()

@app.teardown_appcontext
def close_db(error):
    if hasattr(g, 'db'):
        g.db.close()

#===== Kết thúc phần bổ sung cho mô hình Tenants ======#

# Converter an toàn cho DATE và TIMESTAMP
def convert_date(val):
    if val is None:
        return None
    if isinstance(val, bytes):
        val = val.decode('utf-8', errors='ignore').strip()
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        try:
            # Lấy 10 ký tự đầu (YYYY-MM-DD)
            return date.fromisoformat(val[:10])
        except:
            return None
    return None


def convert_datetime(val):
    if val is None:
        return None
    if isinstance(val, bytes):
        val = val.decode('utf-8', errors='ignore').strip()
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        try:
            return datetime.fromisoformat(val.replace(' ', 'T')[:19])
        except:
            return None
    return None


# Đăng ký converter
sqlite3.register_converter("DATE", convert_date)
sqlite3.register_converter("TIMESTAMP", convert_datetime)

def dict_factory(cursor, row):
    d = {}
    for idx, col in enumerate(cursor.description):
        d[col[0]] = row[idx]
    return d

def format_vn_date(date_obj):
    if not date_obj:
        return '—'
    if isinstance(date_obj, str):
        date_obj = parse_date(date_obj)
    return date_obj.strftime('%d/%m/%Y') if date_obj else '—'

def format_vn_number(num):
    try:
        return f"{int(float(num)):,}".replace(",", ".")
    except:
        return "0"

# THÊM FILTER ĐỊNH DẠNG TIỀN TỆ VIỆT NAM
def format_currency(value):
    if value is None:
        return "0"
    try:
        return "{:,.0f}".format(float(value)).replace(",", ".")
    except:
        return "0"

app.jinja_env.filters['format_currency'] = format_currency

app.jinja_env.filters['format_vn_date'] = format_vn_date
app.jinja_env.filters['format_vn_number'] = format_vn_number
app.config['parse_date'] = parse_date

# Kéo vào runtime khi cần: xmlsec, lxml, win32crypt — ký số USB token là môi trường đặc thù
try:
    import xmlsec
    from lxml import etree
    import win32crypt
    HAS_XMLSEC = True
except Exception:
    HAS_XMLSEC = False

# === CẤU HÌNH BAN ĐẦU ===
database = 'database.db'

# --- PHẦN 1: API QUÉT MÃ VẠCH (2 ĐƠN VỊ TÍNH) ---
@app.route('/api/scan', methods=['POST'])
def scan_barcode():
    barcode = request.json.get('barcode', '').strip()
    conn = get_db_connection()
    db_type = os.getenv('DB_TYPE', 'sqlite').lower()
    
    try:
        # Sử dụng placeholder phù hợp (symbol) cho từng DB 
        symbol = '?' if db_type == 'sqlite' else '%s'
        query = f"""
            SELECT id, barcode, barcode1, name, unit, unit1, base_price, price, unit_ratio 
            FROM products 
            WHERE barcode = {symbol} OR barcode1 = {symbol}
        """
        
        if db_type == 'sqlite':
            product = conn.execute(query, (barcode, barcode)).fetchone()
        else:
            with conn.cursor() as cursor:
                cursor.execute(query, (barcode, barcode))
                product = cursor.fetchone()

        if product:
            # Logic nhận diện đơn vị dựa trên cột barcode và barcode1
            is_unit1 = (barcode == product['barcode1'])
            return jsonify({
                "success": True,
                "data": {
                    "id": product['id'],
                    "name": product['name'],
                    "unit": product['unit1'] if is_unit1 else product['unit'],
                    "price": product['price'] if is_unit1 else product['base_price'],
                    "useUnit1": is_unit1,
                    "ratio": product['unit_ratio']
                }
            })
        return jsonify({"success": False, "message": "Không tìm thấy sản phẩm"}), 404
    finally:
        conn.close()

# === CẤU HÌNH TỪ THÔNG SỐ KĨ THUẬT VIETTEL ===
# Cấu hình database (thay đổi URI theo db của bạn, ví dụ PostgreSQL, MySQL, hoặc SQLite)
_sqlalchemy_db_url = os.getenv('DATABASE_URL', 'sqlite:///pos.db')
if _sqlalchemy_db_url.startswith('postgresql://'):
    _sqlalchemy_db_url = 'postgresql+psycopg://' + _sqlalchemy_db_url[len('postgresql://'):]
app.config['SQLALCHEMY_DATABASE_URI'] = _sqlalchemy_db_url  # Mặc định SQLite cho test
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# Model BusinessInfo (từ models.py, nhưng tích hợp vào đây cho full code)
class BusinessInfo(db.Model):
    __tablename__ = 'business_info'
    
    id = db.Column(db.Integer, primary_key=True)
    tax_code = db.Column(db.String(20), nullable=False, unique=True)      # MST
    business_name = db.Column(db.String(255))                             # Tên doanh nghiệp
    address = db.Column(db.String(500))                                   # Địa chỉ nếu có
    # Các cột khác nếu cần: phone, email, is_active, branch_code, v.v.
    
    def __repr__(self):
        return f"<BusinessInfo {self.tax_code} - {self.business_name}>"

# Tạo bảng nếu chưa tồn tại (chạy lần đầu)
with app.app_context():
    db.create_all()

# Config Viettel từ .env
VIETTEL_CONFIG = {
    "api_url": "https://demo-sinvoice.viettel.vn:8443/InvoiceAPI/InvoiceWS",
    "username": os.getenv("VIETTEL_USERNAME"),  # Ví dụ: "0100109106-215"
    "password": os.getenv("VIETTEL_PASSWORD"),
}


# === DATABASE ===

# ====================================================================
#  MOCK DATABASE SETUP (REQUIRED for 'db' and 'cursor' definitions)
# ====================================================================

class MockCursor:
    """Simulates a database cursor for demonstration."""
    def __init__(self):
        # Initial ID to mock a sequence
        self.lastrowid = 1000 
    def execute(self, query, params=None):
        # Mock execution and simulate ID generation
        if "INSERT INTO sale" in query:
             self.lastrowid += 1
        pass
    def fetchone(self):
        return None
    def close(self):
        pass

class MockDB:
    """Simulates a database connection."""
    def __init__(self):
        self.cursor_instance = MockCursor()

    def cursor(self):
        return self.cursor_instance

    def commit(self):
        # Database commit operation
        pass

    def rollback(self):
        # Database rollback operation
        pass

# Initialize the global DB and Cursor objects used by the API route
db = MockDB()
cursor = db.cursor()

def init_db_columns():
    conn = get_db_connection()
    c = conn.cursor()
    try:
        c.execute("PRAGMA table_info(products)")
        columns = [col[1] for col in c.fetchall()]
        
        if 'unit1' not in columns:
            c.execute("ALTER TABLE products ADD COLUMN unit1 TEXT")
        if 'unit_ratio' not in columns:
            c.execute("ALTER TABLE products ADD COLUMN unit_ratio INTEGER DEFAULT 1")
        if 'price' not in columns:
            c.execute("ALTER TABLE products ADD COLUMN price REAL")
        
        conn.commit()
    except Exception as e:
        print("Lỗi khởi tạo cột:", e)
    finally:
        conn.close()

def format_date_for_frontend(date_str):
    if not date_str:
        return ""
    
    # Danh sách các định dạng mà database của bạn có thể đang lưu
    formats = ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d']
    
    for fmt in formats:
        try:
            # Thử parse chuỗi ngày tháng
            dt = datetime.strptime(str(date_str), fmt)
            # Trả về định dạng chuẩn ISO mà Frontend luôn đọc được
            return dt.strftime('%Y-%m-%d')
        except ValueError:
            continue
            
    return date_str  # Trả về nguyên bản nếu không parse được

# Gọi khi app khởi động
with app.app_context():
    init_db_columns()
def init_db():
    conn = get_db_connection()
    c = conn.cursor()

# --- LƯU Ý: Cần có bảng 'products', 'customers', 'staff' để tạo khóa ngoại ---
# Dưới đây là các bảng giả định:
    c.execute("""CREATE TABLE IF NOT EXISTS customers (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS staff (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT)""")

    # Bảng products
    c.execute('''CREATE TABLE IF NOT EXISTS products (
            	 id INTEGER PRIMARY KEY AUTOINCREMENT,
            	 barcode TEXT UNIQUE,
		 barcode1 TEXT UNIQUE,
                 product_code TEXT UNIQUE,
            	 name TEXT NOT NULL,
            	 unit TEXT DEFAULT 'Cái',
            	 unit1 TEXT DEFAULT 'Thùng',
		 UseSaleUnit INTEGER DEFAULT 0,
            	 buyprice REAL DEFAULT 0,
            	 base_price REAL DEFAULT 0,
              	 unit_ratio REAL DEFAULT 1,
            	 price REAL DEFAULT 0,
		 FOREIGN KEY (sale_id) REFERENCES sale(id)
            )
        ''')
#Bảng Tôn Kho (Để tính giá vốn bình quân gia quyền)
    c.execute('''CREATE TABLE IF NOT EXISTS inventory (
    		 product_id INTEGER PRIMARY KEY,
    		 quantity REAL DEFAULT 0,
    		 avg_cost REAL DEFAULT 0,
    		 last_updated TEXT DEFAULT CURRENT_TIMESTAMP,
    		 FOREIGN KEY (product_id) REFERENCES products(id) ON DELETE CASCADE
	)''')
    # Bảng suppliers
    c.execute('''CREATE TABLE IF NOT EXISTS suppliers (
                 id INTEGER PRIMARY KEY AUTOINCREMENT, code TEXT UNIQUE, name TEXT NOT NULL,
                 phone TEXT, email TEXT, address TEXT, note TEXT, tax_code TEXT)''')

    c.execute('''
        CREATE TABLE IF NOT EXISTS Operating_Cost (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,                         -- Ngày ghi sổ (Dạng YYYY-MM-DD)
            note TEXT,                                  -- Diễn giải (Cột D)
            employee_salary REAL DEFAULT 0,             -- Chi phí nhân công (Cột 2)
            electric_cost REAL DEFAULT 0,               -- Chi phí điện (Cột 3)
            water_cost REAL DEFAULT 0,                  -- Chi phí nước (Cột 4)
            telecomunication_cost REAL DEFAULT 0,       -- Chi phí viễn thông (Cột 5)
            premise_warehouse_cost REAL DEFAULT 0,      -- Chi phí thuê mặt bằng (Cột 6)
            management_cost REAL DEFAULT 0,             -- Chi phí quản lý/VPP (Cột 7)
            other_cost REAL DEFAULT 0,                  -- Chi phí khác (Cột 8)
            total_cost REAL GENERATED ALWAYS AS (
            employee_salary + electric_cost + water_cost + 
            telecomunication_cost + premise_warehouse_cost + 
            management_cost + other_cost) VIRTUAL
        )
    ''')
    # Bảng import & import_details
    c.execute('''
        CREATE TABLE IF NOT EXISTS import (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            import_no TEXT UNIQUE NOT NULL,
            date TEXT,
            supplier_id INTEGER,
            bill_no TEXT,
            note TEXT,
            payment_status TEXT,
            extra_cost REAL,
            total_value REAL
        )
    ''')

    c.execute('''
        CREATE TABLE IF NOT EXISTS import_details (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            import_id INTEGER NOT NULL,
            product_id INTEGER NOT NULL,
            qty REAL NOT NULL,                    -- Số lượng nhập
            buyprice REAL DEFAULT 0,              -- Đơn giá mua chưa bao gồm thuế
            cost_price REAL DEFAULT 0,            -- Giá vốn bao gồm thuế (dùng để tính total_value trong stock_moves)
            discount REAL DEFAULT 0,              -- Chiết khấu dòng
            tax REAL DEFAULT 0,                   -- Thuế dòng
            subtotal REAL NOT NULL,               -- Tổng tiền hàng dòng (có thể là trước hoặc sau thuế tùy định nghĩa)
            payment_amt REAL NOT NULL,            -- Số tiền thực tế thanh toán cho dòng này
            FOREIGN KEY(import_id) REFERENCES import(id) ON DELETE CASCADE,
            FOREIGN KEY(product_id) REFERENCES products(id)
        )
    ''')
    # === BỔ SUNG: BẢNG QUẢN LÝ SỐ THỨ TỰ PHIẾU NHẬP ===
    c.execute('''CREATE TABLE IF NOT EXISTS import_sequence (
                 id INTEGER PRIMARY KEY CHECK (id = 1),
                 current_seq INTEGER DEFAULT 0
                 )''')
    c.execute("SELECT COUNT(*) FROM import_sequence")
    if c.fetchone()[0] == 0:
        c.execute("INSERT OR REPLACE INTO import_sequence (id, current_seq) VALUES (1, 0)")

    # Bảng sale & sale_items
    c.execute('''CREATE TABLE IF NOT EXISTS sale (
                 id INTEGER PRIMARY KEY AUTOINCREMENT, date TEXT, total_amount REAL, discount_pct REAL, tax_pct REAL, UseSaleUnit INTEGER DEFAULT 0,
                 payment_method TEXT, customer_name TEXT, customer_phone TEXT, status TEXT, discount_amount REAL DEFAULT 0, tax_amount REAL DEFAULT 0, 
                 invoice_number TEXT DEFAULT '', invoice_provider TEXT DEFAULT 'Tự Tạo', note TEXT DEFAULT '',
		 FOREIGN KEY (order_id) REFERENCES orders(id),
		 FOREIGN KEY (staff_id) REFERENCES staff(id),
                 FOREIGN KEY (customer_id) REFERENCES customers(id)
             )''')

    c.execute('''CREATE TABLE IF NOT EXISTS sale_items (
                 sale_id INTEGER, product_id INTEGER, quantity REAL, price REAL,
                 cost_price REAL DEFAULT 0, UseSaleUnit INTEGER DEFAULT 0, unit_ratio REAL DEFAULT 1,
                 FOREIGN KEY (sale_id) REFERENCES sale (id),
		 FOREIGN KEY (order_id) REFERENCES orders(id),
                 FOREIGN KEY (product_id) REFERENCES products (id))''')
    # Bảng trả hàng (IMPORT/SALE)
    c.execute('''CREATE TABLE IF NOT EXISTS return_import (
                 id INTEGER PRIMARY KEY AUTOINCREMENT, date TEXT, import_id INTEGER,
                 product_id INTEGER, quantity REAL, reason TEXT,
                 cost_price REAL DEFAULT 0, refund_amount REAL DEFAULT 0,
                 FOREIGN KEY (import_id) REFERENCES import (id),
                 FOREIGN KEY (product_id) REFERENCES products (id))''')
    c.execute('''CREATE TABLE IF NOT EXISTS return_sales (
                 id INTEGER PRIMARY KEY AUTOINCREMENT, date TEXT, sale_id INTEGER,
                 product_id INTEGER, quantity REAL, reason TEXT,
                 FOREIGN KEY (sale_id) REFERENCES sale (id),
                 FOREIGN KEY (product_id) REFERENCES products (id))''')
    # Bảng users & settings
    c.execute('''CREATE TABLE IF NOT EXISTS users (
                 id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT UNIQUE NOT NULL,
                 password TEXT NOT NULL, role TEXT DEFAULT 'user', full_name TEXT,
                 permissions TEXT DEFAULT '', created_at TEXT DEFAULT (datetime('now')))''')
    c.execute('''CREATE TABLE IF NOT EXISTS settings (
                 key TEXT PRIMARY KEY, value TEXT)''')

    # Bảng stock_moves để lưu lịch sử nhập/xuất
    c.execute("""
        CREATE TABLE IF NOT EXISTS stock_moves (
       	    id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            type TEXT NOT NULL, -- 'IMPORT', 'SALE', 'RETURN_SALE', 'RETURN_IMPORT', 'ADJUST'
            unit TEXT,
            ref_document TEXT NOT NULL,
            ref_id INTEGER,
            ref_no TEXT,
            in_quantity REAL DEFAULT 0,
            out_quantity REAL DEFAULT 0,
            quantity REAL NOT NULL,
            avg_cost REAL DEFAULT 0,
            cost_price REAL DEFAULT 0,
            total_value REAL NOT NULL, -- Đã sửa: NOT NUL -> NOT NULL
            note TEXT,
            FOREIGN KEY (product_id) REFERENCES products(id)
      )
    """)

    # --- INDEX TỐI ƯU ---
    stock_moves_index_sql = """
    CREATE INDEX IF NOT EXISTS idx_stock_moves_product_date ON stock_moves (product_id, date);
    """

    # === BỔ SUNG: BẢNG INVENTORY TRANSACTIONS MỚI ===
    c.execute('''CREATE TABLE IF NOT EXISTS inventory_transactions (
                 id INTEGER PRIMARY KEY AUTOINCREMENT,
                 product_id INTEGER NOT NULL,
                 type TEXT NOT NULL CHECK(type IN ('import', 'export', 'adjust')),
		 type1 TEXT,
                 quantity INTEGER NOT NULL,
                 reference_id INTEGER,
                 reference_type TEXT,
                 cost_price REAL NOT NULL DEFAULT 0,
                 total_value REAL NOT NULL DEFAULT 0,
                 import_id INTEGER,             -- <--- ĐÃ THÊM: Liên kết tới phiếu nhập
                 sale_id INTEGER,               -- <--- ĐÃ THÊM: Liên kết tới hóa đơn bán
                 return_sale_id INTEGER,        -- <--- ĐÃ THÊM: Liên kết tới phiếu trả hàng bán
                 note TEXT,
                 created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                 FOREIGN KEY (product_id) REFERENCES products (id),
                 FOREIGN KEY (import_id) REFERENCES import (id),
                 FOREIGN KEY (sale_id) REFERENCES sale (id),
                 FOREIGN KEY (return_sale_id) REFERENCES return_sales (id)
    )''')

    print("Khởi tạo database thành công! Đăng nhập: admin / admin123")
    
    # Index để query nhanh
    c.execute('''CREATE INDEX IF NOT EXISTS idx_inventory_product ON inventory_transactions(product_id);''')
    c.execute('''CREATE INDEX IF NOT EXISTS idx_inventory_date ON inventory_transactions(created_at);''')
    # =================================================

   # === THÊM BẢNG MỚI – KẾ TOÁN HKD ===
    c.execute('''CREATE TABLE IF NOT EXISTS voucher_seq (
                 type TEXT PRIMARY KEY, seq INTEGER DEFAULT 0)''')
    c.execute("INSERT OR IGNORE INTO voucher_seq (type, seq) VALUES ('PT', 0), ('PC', 0), ('PN', 0), ('PX', 0)")

    # 5 CHỨNG TỪ
    c.execute('''CREATE TABLE IF NOT EXISTS phieu_thu (
                 id INTEGER PRIMARY KEY AUTOINCREMENT,
            so_phieu TEXT NOT NULL UNIQUE,
            ngay_lap DATE NOT NULL,
            nguoi_nop TEXT NOT NULL,
            dia_chi TEXT,
            ly_do_nop TEXT NOT NULL,
            so_tien INTEGER NOT NULL,
            hinh_thuc TEXT DEFAULT 'Tiền mặt',
            kem_theo TEXT,
            nguoi_lap TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_code TEXT,
            customer_name TEXT,
            customer_phone TEXT,
	    payment_method TEXT,
	    discount_amount REAL DEFAULT 0,
	    tax_amount REAL DEFAULT 0,
            total_amount REAL,
            status TEXT DEFAULT 'pending',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
	    FOREIGN KEY (staff_id) REFERENCES staff(id)
            FOREIGN KEY (customer_id) REFERENCES customers(id)
        )
    ''')

    c.execute('''CREATE TABLE IF NOT EXISTS phieu_chi (
                 id INTEGER PRIMARY KEY AUTOINCREMENT,
                 voucher_no TEXT UNIQUE,
                 date TEXT,
                 receiver_name TEXT,
                 address TEXT,
                 reason TEXT,
		 reference_document,
		 preparer,
                 amount REAL,
                 source_id INTEGER,
                 source_type TEXT DEFAULT 'manual')''')

    c.execute('''CREATE TABLE IF NOT EXISTS phieu_nhap_kho (
                 id INTEGER PRIMARY KEY AUTOINCREMENT,
                 voucher_no TEXT UNIQUE,
                 date TEXT,
                 supplier_name TEXT,
                 items_json TEXT,
                 total_amount REAL,
                 import_id INTEGER)''')

    c.execute('''CREATE TABLE IF NOT EXISTS phieu_xuat_kho (
                 id INTEGER PRIMARY KEY AUTOINCREMENT,
                 voucher_no TEXT UNIQUE,
                 date TEXT,
                 customer_name TEXT,
                 items_json TEXT,
                 total_amount REAL,
                 sale_id INTEGER)''')
    c.execute('''CREATE TABLE IF NOT EXISTS bang_luong (
             	id INTEGER PRIMARY KEY AUTOINCREMENT,
             	period TEXT,
            	 employee_name TEXT,
             	gross_salary REAL,
             	bhxh REAL DEFAULT 0,
             	bhyt REAL DEFAULT 0,
             	bhtn REAL DEFAULT 0,
            	 other_deductions REAL DEFAULT 0,
             	total_deductions REAL,
             	net_pay REAL,
            	 paid_date TEXT)''')

    c.execute('''CREATE TABLE IF NOT EXISTS so_theo_doi_tien_luong (
                 id INTEGER PRIMARY KEY AUTOINCREMENT,
                 period TEXT,
                 employee_name TEXT,
                 gross_salary REAL,
                 deductions REAL,
                 net_pay REAL,
                 paid_date TEXT)''')

    # 7 SỔ KẾ TOÁN
    c.execute('''CREATE TABLE IF NOT EXISTS so_chi_tiet_doanh_thu (
                 id INTEGER PRIMARY KEY AUTOINCREMENT,
                 period TEXT,
                 date TEXT,
                 voucher_no TEXT,
                 description TEXT,
                 revenue REAL,
                 vat REAL,
                 pit REAL,
                 total_tax REAL)''')

    c.execute('''CREATE TABLE IF NOT EXISTS so_chi_tiet_hang_hoa (
                 id INTEGER PRIMARY KEY AUTOINCREMENT,
                 period TEXT,
                 product_name TEXT,
                 unit TEXT,
                 begin_qty REAL,
                 import_qty REAL,
                 export_qty REAL,
                 end_qty REAL,
                 begin_value REAL,
                 end_value REAL)''')

    c.execute('''CREATE TABLE IF NOT EXISTS so_quy_tien_mat (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  period TEXT,
                  date TEXT,
                  voucher_no TEXT,
                  type TEXT,
                  amount REAL,
                  balance REAL)''')

    # Sổ tiền gửi ngân hàng (Đã thêm voucher_no và type)
    c.execute('''CREATE TABLE IF NOT EXISTS so_tien_gui_ngan_hang (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  period TEXT,
                  date TEXT,
                  voucher_no TEXT,             
                  type TEXT,                   
                  description TEXT,
                  amount REAL,
                  bank_name TEXT,
                  balance REAL)''')

    # Admin mặc định
    c.execute("SELECT COUNT(*) FROM users WHERE username='admin'")
    if c.fetchone()[0] == 0:
        pwd = bcrypt.generate_password_hash('admin123').decode('utf-8')
        c.execute("INSERT INTO users (username, password, role, full_name) VALUES (?, ?, ?, ?)",
                  ('admin', pwd, 'admin', 'Quản trị viên'))
    conn.commit()
    conn.close()

    migrate_database()

def migrate_database():
    conn = get_db_connection()
    c = conn.cursor()
    """
    Thêm cột nếu cần. Tránh lỗi ALTER TABLE khi cột đã tồn tại.
    """
    # helper: check column exists
    def has_column(table, column):
        try:
            c.execute(f"PRAGMA table_info({table})")
            cols = [r[1] for r in c.fetchall()]
            return column in cols
        except Exception:
            return False
    # columns to ensure
    extras = [
        ('import', 'bill_no', "TEXT"),
        ('suppliers', 'tax_code', "TEXT"),
        ('sale', 'invoice_number', "TEXT"),
        ('sale', 'invoice_provider', "TEXT"),
        ('return_import', 'cost_price', "REAL"),
        ('return_import', 'refund_amount', "REAL"),
        ('users', 'permissions', "TEXT"),
    ]
    for table, col, col_type in extras:
        if not has_column(table, col):
            try:
                c.execute(f"ALTER TABLE {table} ADD COLUMN {col} {col_type}")
            except sqlite3.OperationalError as e:
                # có thể table chưa tồn tại hoặc cột đã tồn tại; bỏ qua khác
                print(f"[MIGRATE] Không thể thêm {col} vào {table}: {e}")
    conn.commit()
    conn.close()

# ================================================================================ HELPERS ========================================================================================

def get_setting(key, default=""):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT value FROM settings WHERE key=?", (key,))
    result = c.fetchone()
    conn.close()
    return result[0] if result else default

def get_next_voucher_no(prefix):
    with get_db_connection() as conn:
        c = conn.cursor()
        c.execute("""
            INSERT INTO voucher_seq (type, seq) VALUES (?, 0)
            ON CONFLICT(type) DO UPDATE SET seq = seq + 1
        """, (prefix,))
        c.execute("SELECT seq FROM voucher_seq WHERE type=?", (prefix,))
        seq = c.fetchone()[0]
        conn.commit()
        return f"{prefix}{seq:06d}"

def load_esign_config():
    global ESIGN_CONFIG
    ESIGN_CONFIG = {
        "provider": get_setting("esign_provider"),
        "api_url": get_setting("esign_api_url") or os.getenv("ESIGN_API_URL", ""),
        "client_id": get_setting("esign_client_id") or os.getenv("ESIGN_CLIENT_ID", ""),
        "client_secret": get_setting("esign_client_secret") or os.getenv("ESIGN_CLIENT_SECRET", "")
    }

def validate_json(required_fields, data):
    for f in required_fields:
        if f not in data or data[f] in [None, '', []]:
            return f"Thiếu hoặc không hợp lệ trường: {f}"
    return None

def get_product_stock(product_id):
    conn = get_db_connection()
    c = conn.cursor()
    # Lấy tồn kho và giá vốn từ bảng inventory
    c.execute("SELECT quantity, avg_cost FROM inventory WHERE product_id=?", (product_id,))
    row = c.fetchone()
    conn.close()
    if row:
        return {'quantity': row['quantity'] or 0, 'avg_cost': row['avg_cost'] or 0}
    return {'quantity': 0, 'avg_cost': 0}

@app.template_filter('vnd')
def vnd(value):
    try:
        value = float(value)
        return f"{value:,.0f}".replace(",", ".")
    except:
        return value

# =========================================================================== START OF SETTINGS, LOGIN & DECORATORS ======================================================================#
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            # Với API → trả JSON, không redirect
            if request.path.startswith('/api/'):
                return jsonify({"success": False, "error": "Unauthorized"}), 401
            # Với trang web → redirect login
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def master_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('role') != 'master':
            if request.path.startswith('/api/'):
                return jsonify({"success": False, "error": "Forbidden"}), 403
            return redirect(url_for('sale'))
        return f(*args, **kwargs)
    return decorated_function

def admin_or_master_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        role = session.get('role')
        # Cho phép nếu là admin HOẶC là master
        if role not in ['admin', 'admin*', 'adminFB', 'master']:
            if request.path.startswith('/api/'):
                return jsonify({"success": False, "error": "Forbidden"}), 403
            return redirect(url_for('sale'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('role') != 'admin':
            if request.path.startswith('/api/'):
                return jsonify({"success": False, "error": "Forbidden"}), 403
            return redirect(url_for('sale'))
        return f(*args, **kwargs)
    return decorated_function

from functools import wraps
from flask import session, request, jsonify, flash, redirect, url_for, current_app

from functools import wraps
from flask import session, request, jsonify, flash, redirect, url_for, current_app

def require_permission(target_perm):
    """
    Decorator kiểm tra quyền truy cập.
    - Master và Admin có toàn quyền (bypass tất cả permission)
    - Admin* có quyền theo permission và mặc định điều hướng về /rental
    """
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            # 1. Kiểm tra đã đăng nhập chưa
            if 'user' not in session:
                if request.is_json or request.path.startswith('/api/'):
                    return jsonify({"success": False, "error": "Unauthorized"}), 401
                flash("Vui lòng đăng nhập", "warning")
                return redirect(url_for('login'))

            user = session['user']
            role = user.get('role', 'guest')
            # Chuyển chuỗi permissions thành list để kiểm tra cho chính xác
            perms = [p.strip() for p in user.get('permissions', '').split(',') if p.strip()]

            # ==================== QUYỀN SIÊU CAO (BYPASS) ====================
            # Master và Admin (không dấu *) có toàn quyền
            if role in ['master', 'admin', 'adminFB']:
                return f(*args, **kwargs)

            # ==================== KIỂM TRA QUYỀN THỰC TẾ ====================
            # admin* hoặc các role khác phải có target_perm trong danh sách perms
            if target_perm in perms:
                return f(*args, **kwargs)

            # ==================== TỪ CHỐI TRUY CẬP ====================
            msg = f"Không có quyền thực hiện: {target_perm}"
            current_app.logger.warning(f"Access Denied: {user.get('username')} (Role: {role}) tried to access {target_perm}")

            # Trả về JSON nếu là API
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({"success": False, "error": msg}), 403

            # Điều hướng trang chủ mặc định tùy theo Role
            flash(msg, "danger")
            
            # Fix: Nếu user là admin* thì mặc định về trang rental
            if role == 'admin*':
                return redirect(url_for('rental')) 
            
            # Các trường hợp khác về trang sale
            return redirect(url_for('sale'))

        return decorated
    return decorator


# --- Cấu hình Flask-Login ---
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = "Vui lòng đăng nhập để tiếp tục."
login_manager.login_message_category = "info"


# ====================== USER CLASS ======================
class User(UserMixin):
    def __init__(self, id, username, role, db_path, tenant_id, full_name, permissions=""):
        self.id = id
        self.username = username
        self.role = role
        self.db_path = db_path
        self.tenant_id = tenant_id
        self.full_name = full_name
        self.permissions = permissions

    def get_id(self):
        return str(self.id)  # Flask-Login yêu cầu trả về string

@login_manager.user_loader
def load_user(user_id):
    try:
        # Lấy db_path trực tiếp từ session
        db_path_raw = session.get('db_path')
        
        if not db_path_raw:
            # Nếu mất session db_path, thử dùng DB tổng để tìm (áp dụng cho tài khoản master)
            db_path = os.path.join(BASE_DIR, 'database.db')
        else:
            # Chuyển đổi thành đường dẫn tuyệt đối chuẩn xác trên Ubuntu VPS
            if os.path.isabs(db_path_raw):
                db_path = os.path.abspath(db_path_raw)
            else:
                db_path = os.path.join(BASE_DIR, db_path_raw)

        # Kiểm tra file DB có thực sự tồn tại trên VPS không, tránh lỗi đứng luồng
        if not os.path.exists(db_path):
            current_app.logger.error(f"Database path không tồn tại: {db_path}")
            return None

        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        user_row = conn.execute("SELECT * FROM users WHERE id = ?", (int(user_id),)).fetchone()
        conn.close()

        if user_row:
            u = dict(user_row)
            # Trả về đối tượng User chuẩn cho Flask-Login quản lý
            return User(
                id=u['id'],
                username=u['username'],
                role=str(u.get('role', '')).strip(), # Xóa khoảng trắng thừa
                db_path=db_path,
                tenant_id=session.get('last_tenant_id'),
                full_name=u.get('full_name') or u['username'],
                permissions=u.get('permissions', '')
            )
    except Exception as e:
        current_app.logger.error(f"Lỗi nghiêm trọng tại load_user: {e}")
    return None

from werkzeug.middleware.proxy_fix import ProxyFix

# Thêm cấu hình này ngay sau dòng khai báo app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

# Cấu hình bảo mật Cookie Session
app.config.update(
    SESSION_COOKIE_SECURE=True,   # Chỉ gửi cookie qua HTTPS (Vì bạn đã bật SSL thành công)
    SESSION_COOKIE_HTTPONLY=True, # Ngăn chặn Javascript đọc trộm cookie
    SESSION_COOKIE_SAMESITE='Lax' # Chống tấn công CSRF, giúp lưu session ổn định
)


#================================================================================= END OF LOGIN & DECORATORS ==========================================================================#

# ================================================================================== CÁC API & ROUTES PRODUCTS ===========================================================================#
# === API: QUẢN LÝ SẢN PHẨM (products.html) ===
@app.route('/api/products/manage', methods=['GET', 'POST', 'PUT', 'DELETE'])
@login_required
def product_manage():
    conn = get_db_connection()
    c = conn.cursor()
    
    try:
        # ==================== GET: TRUY VẤN DANH SÁCH ====================
        if request.method == 'GET':
            q = request.args.get('q', '')
            c.execute("""
                SELECT 
                    p.*,
                    COALESCE(i.quantity, 0) AS quantity,
                    COALESCE(i.avg_cost, 0) AS avg_cost
                FROM products p
                LEFT JOIN inventory i ON p.id = i.product_id
                WHERE p.name LIKE ? OR p.barcode LIKE ? 
                ORDER BY p.id DESC
            """, (f'%{q}%', f'%{q}%'))
            products = c.fetchall()
            return jsonify([dict(row) for row in products])
            
        # Lấy dữ liệu JSON từ body
         # Lấy dữ liệu JSON từ body (dùng cho POST, PUT, DELETE)
        data = request.get_json(silent=True) or {}
        product_id = data.get('id')

        # ==================== POST: THÊM MỚI ====================
        if request.method == 'POST':
            name = data.get('name', '').strip()
            if not name:
                return jsonify({"success": False, "error": "Tên sản phẩm bắt buộc"}), 400

            unit1 = data.get('unit1') or None
            
            # 1. Thêm sản phẩm
            c.execute("INSERT INTO products (name, unit, base_price, price, unit1, unit_ratio) VALUES (?, ?, ?, ?, ?, ?)",
                      (name, data.get('unit', 'Cái'), data.get('base_price', 0), data.get('price', 0), unit1, data.get('unit_ratio', 1)))
            new_id = c.lastrowid
            
            # 2. Khởi tạo tồn kho
            c.execute("INSERT INTO inventory (product_id, quantity, avg_cost) VALUES (?, 0, 0)", (new_id,))
            
            # 3. Tạo Mã Sản Phẩm & Mã Vạch tự động
            code = f"SP{new_id:04d}" 
            barcode = code + "01"
            barcode1 = code + "02" if unit1 else None
            
            c.execute("UPDATE products SET product_code=?, barcode=?, barcode1=? WHERE id=?", 
                      (code, barcode, barcode1, new_id))
            
            conn.commit()
            return jsonify({"success": True, "id": new_id, "product_code": code, "barcode": barcode})

        # ==================== PUT: CẬP NHẬT ====================
        elif request.method == 'PUT':
            product_id = data.get('id')
            if not product_id:
                return jsonify({"success": False, "error": "Thiếu ID sản phẩm"}), 400
                
            c.execute("""UPDATE products SET 
                         name=?, unit=?, base_price=?, price=?, unit1=?, unit_ratio=?
                         WHERE id=?""",
                      (data['name'], data.get('unit', 'Cái'), data.get('base_price', 0), data.get('price', 0), 
                       data.get('unit1'), data.get('unit_ratio', 1), product_id))
            conn.commit()
            return jsonify({"success": True, "id": product_id})

        # ==================== DELETE: XÓA SẢN PHẨM ====================
        elif request.method == 'DELETE':
            # Ưu tiên lấy ID từ Query String (args) hoặc JSON data
            p_id = request.args.get('id') or data.get('id')
            if not p_id:
                return jsonify({"success": False, "error": "Thiếu ID sản phẩm"}), 400

            # --- Sửa lỗi max_id is not defined ---
            res_max = c.execute("SELECT MAX(id) FROM products").fetchone()
            max_id_val = res_max[0] if res_max[0] else 0

            # 1. Kiểm tra phát sinh giao dịch (Dùng bảng sale_items của bạn)
            c.execute("SELECT 1 FROM sale_items WHERE product_id = ? LIMIT 1", (p_id,))
            if c.fetchone():
                return jsonify({"success": False, "error": "Sản phẩm đã có trong hóa đơn bán hàng"}), 400

            # 2. Kiểm tra tồn kho (Bảng inventory và stock_moves)
            c.execute("SELECT 1 FROM stock_moves WHERE product_id = ? LIMIT 1", (p_id,))
            if c.fetchone():
                return jsonify({"success": False, "error": "Sản phẩm đã phát sinh lịch sử nhập/xuất kho"}), 400

            # 3. Thực hiện xóa liên hoàn (Cascade delete thủ công)
            c.execute("DELETE FROM inventory WHERE product_id = ?", (p_id,))
            c.execute("DELETE FROM products WHERE id = ?", (p_id,))

            # 4. Logic làm sạch Auto-increment Sequence nếu xóa sản phẩm cuối
            if int(p_id) == max_id_val:
                c.execute("SELECT MAX(id) FROM products")
                new_max = c.fetchone()[0] or 0
                c.execute("UPDATE sqlite_sequence SET seq = ? WHERE name = 'products'", (new_max,))
            
            conn.commit()
            return jsonify({"success": True, "message": "Đã xóa sản phẩm thành công"})

    except sqlite3.Error as e:
        if conn: conn.rollback()
        return jsonify({"success": False, "error": f"Lỗi DB: {str(e)}"}), 500
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({"success": False, "error": f"Lỗi hệ thống: {str(e)}"}), 500
    finally:
        if conn: conn.close()

# API BATCH CẬP NHẬT GIÁ
@app.route('/api/products/batch_update', methods=['POST'])
@login_required
def batch_update_products():
    updates = request.json.get('updates', [])
    if not updates: return jsonify({"success": True})
    conn = get_db_connection()
    c = conn.cursor()
    try:
        for upd in updates:
            pid = upd['product_id']
            base_price = float(upd.get('base_price', 0))
            unit_ratio = int(upd.get('unit_ratio', 1))
            unit1 = upd.get('unit1', '').strip()
            price = float(upd.get('price', 0))
            c.execute("""
                UPDATE products SET base_price=?, price=?, unit1=?, unit_ratio=?
                WHERE id=?
            """, (base_price, price, unit1, unit_ratio, pid))
        conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()

# === API: CẬP NHẬT GIÁ BÁN SẢN PHẨM ===
@app.route('/api/products/update_prices', methods=['PUT'])
# @login_required
# @admin_or_master_required
def update_product_prices():
    conn = get_db_connection()
    c = conn.cursor()
    try:
        price_updates = request.get_json()
        if not isinstance(price_updates, list):
            return jsonify({'success': False, 'error': 'Dữ liệu phải là một mảng'}), 400

        for update in price_updates:
            product_id = update.get('id')
            if not product_id:
                continue

            # Xây dựng câu lệnh UPDATE động chỉ cho các trường được cung cấp
            set_clauses = []
            params = []

            if update.get('base_price') is not None:
                set_clauses.append("base_price = ?")
                params.append(update['base_price'])
            
            if update.get('unit1') is not None:
                set_clauses.append("unit1 = ?")
                params.append(update['unit1'])

            if update.get('unit_ratio') is not None:
                set_clauses.append("unit_ratio = ?")
                params.append(update['unit_ratio'])

            if update.get('price') is not None:
                set_clauses.append("price = ?")
                params.append(update['price'])

            if not set_clauses:
                continue # Bỏ qua nếu không có trường nào để cập nhật

            sql = f"UPDATE products SET {', '.join(set_clauses)} WHERE id = ?"
            params.append(product_id)
            
            c.execute(sql, tuple(params))

        conn.commit()
        return jsonify({'success': True})


    except sqlite3.Error as e:
        conn.rollback()
        return jsonify({'success': False, 'error': f'Lỗi Database khi cập nhật giá: {e}'}), 500    
    finally:
        # close_db(conn)
        pass

@app.route('/api/products', methods=['GET'])
def api_products():
    query = request.args.get('q', '').strip()
    # Nếu có tham số exact=1, ta sẽ lọc chính xác tên
    exact = request.args.get('exact', '0') == '1'
    return get_product_list_with_stock(query=query)

@app.route('/api/products/barcode/<barcode>', methods=['GET'])
def api_get_product_by_barcode(barcode):
    #"""Route xử lý tìm kiếm sản phẩm bằng mã vạch (dùng Enter)"""
    # Lấy danh sách sản phẩm theo mã vạch
    response, status_code = get_product_list_with_stock(barcode=barcode)
    
    if status_code != 200:
        return response, status_code
    
    products = response.get_json()
    
    # Trả về sản phẩm đầu tiên tìm được (vì barcode là UNIQUE)
    if products and products[0]:
        return jsonify(products[0])
    else:
        # Trả về 404/None nếu không tìm thấy
        return jsonify(None), 404

@app.route('/api/products/upsert', methods=['POST'])
def api_upsert_product():
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "error": "Không nhận được dữ liệu JSON"}), 400

    product_id = data.get('id')
    name = (data.get('name') or '').strip()
    unit = (data.get('unit') or 'Cái').strip()
    base_price = float(data.get('base_price') or 0)
    buyprice = float(data.get('buyprice') or 0)
    import_id = float(data.get('import_id') or 0)

    
    # Thông tin sỉ
    unit1 = (data.get('unit1') or '').strip() or None
    unit_ratio = float(data.get('unit_ratio') or 1)
    price = float(data.get('price') or 0) # Giá bán sỉ

    if not name:
        return jsonify({"success": False, "error": "Thiếu tên sản phẩm"}), 400

    conn = get_db_connection()
    c = conn.cursor()
    try:
        if product_id:
            # Logic Update
            c.execute("""
                UPDATE products SET 
                name=?, unit=?, buyprice=?, base_price=?, unit1=?, unit_ratio=?, price=?, import_id=?, updated_at=CURRENT_TIMESTAMP
                WHERE id=?
            """, (name, unit, buyprice, base_price, unit1, unit_ratio, price, import_id, product_id))
        else:
            # Logic Insert mới
            c.execute("""
                INSERT INTO products (name, unit, buyprice, base_price, unit1, unit_ratio, price, import_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (name, unit, buyprice, base_price, unit1, unit_ratio, price, import_id))
            product_id = c.lastrowid
            
            # --- TỰ TẠO MÃ VẠCH NẾU LÀ SẢN PHẨM MỚI ---
            formatted_code = f"SP{product_id:04d}"
            barcode = data.get('barcode') or (formatted_code + "01")
            barcode1 = data.get('barcode1') or ((formatted_code + "02") if unit1 else None)
            
            c.execute("""
                UPDATE products SET product_code=?, barcode=?, barcode1=? WHERE id=?
            """, (formatted_code, barcode, barcode1, product_id))
            
            # Khởi tạo tồn kho
            c.execute("INSERT OR IGNORE INTO inventory (product_id, quantity, avg_cost) VALUES (?, 0, 0)", (product_id,))

        conn.commit()
        
        # Lấy lại dữ liệu sau khi update/insert để trả về client
        c.execute("SELECT * FROM products WHERE id = ?", (product_id,))
        p = c.fetchone()
        
        return jsonify({
            "success": True,
            "product": dict(p)
        })
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        conn.close()

# === NEW: COMPLETE SALE (ghi sale_items, trừ kho, tính tổng, ghi stock_moves) ===
@app.route('/api/products')
def api_search_products():
    query = request.args.get('q', '').strip().lower()
    if not query:
        return jsonify([])

    conn = get_db_connection()
    c = conn.cursor()
    like = f"%{query}%"
    c.execute("""
        SELECT id, barcode, name, unit, quantity, base_price, 
               unit1, unit_ratio, price
        FROM products 
        WHERE quantity > 0 
          AND (LOWER(name) LIKE ? OR barcode LIKE ?)
        ORDER BY name 
        LIMIT 15
    """, (like, like))
    rows = c.fetchall()
    conn.close()

    products = []
    for row in rows:
        p = dict(row)
        # Đảm bảo price luôn có
        if not p['price'] and p['unit1'] and p['unit_ratio']:
            p['price'] = p['base_price'] * p['unit_ratio']
        products.append(p)
    return jsonify(products)

@app.get("/api/products/barcode/{barcode}")
def get_by_barcode(barcode: str):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("""
        SELECT id, barcode, name, unit, quantity, base_price, 
               unit1, unit_ratio, price
        FROM products 
        WHERE barcode = ? AND quantity > 0
    """, (barcode.upper(),))
    row = c.fetchone()
    conn.close()

    if not row:
        return jsonify({})

    p = dict(row)
    if not p['price'] and p['unit1'] and p['unit_ratio']:
        p['price'] = p['base_price'] * p['unit_ratio']
    return jsonify(p)

@app.route('/products')
@login_required
def products():
    return render_template('products.html')

# ====================================================================================End of Product's API & Routes================================================================# 





# === SALE routes (POS / Bán hàng) → routes/sale.py ===
from routes.sale import register_sale_routes
register_sale_routes(app)

# === INVOICE routes (Hóa đơn điện tử) → routes/invoice.py ===
from routes.invoice import register_invoice_routes
register_invoice_routes(app)

# === INVENTORY routes (Nhập kho / Tồn kho) → routes/inventory.py ===
from routes.inventory import register_inventory_routes
register_inventory_routes(app)

# === INWARD routes (Hóa đơn đầu vào) → routes/inward.py ===
from routes.inward import register_inward_routes
register_inward_routes(app)

# === KẾ TOÁN HKD routes → routes/ketoan_hkd.py ===
from routes.ketoan_hkd import register_ketoan_hkd_routes
register_ketoan_hkd_routes(app)

# === RENTAL routes → routes/rental.py ===
from routes.rental import register_rental_routes
register_rental_routes(app)

# === SETTINGS / LOGIN routes → routes/settings.py ===
from routes.settings import register_settings_routes
register_settings_routes(app)

# === KẾ TOÁN SME routes → routes/ketoan_sme.py ===
from routes.ketoan_sme import register_ketoan_sme_routes
register_ketoan_sme_routes(app)


#================================================================================ End of Inventory ===============================================================================#

@app.route('/export/so-chi-tiet-doanh-thu/<period>')
@login_required
def export_so_chi_tiet_doanh_thu_excel(period):
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from io import BytesIO

    # Lấy dữ liệu như trên...
    # so_chi, tong_doanh_thu, ...

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"DoanhThu_{period}"

    # Header
    headers = ["Ngày", "Số phiếu", "Diễn giải", "Doanh thu", "Thuế GTGT", "Tổng thu", "Ghi chú"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Dữ liệu
    for row in so_chi:
        ws.append([
            row['date'], row['voucher_no'], row['description'],
            row['doanh_thu'], row['vat'], row['total'], row['note'] or ''
        ])

    # Cộng
    ws.append(["", "", "CỘNG", tong_doanh_thu, tong_vat, tong_thu, ""])
    ws[f"D{len(so_chi)+2}"].font = Font(bold=True)
    ws[f"E{len(so_chi)+2}"].font = Font(bold=True)
    ws[f"F{len(so_chi)+2}"].font = Font(bold=True)

    # Định dạng
    for row in ws.iter_rows(min_row=1, max_row=len(so_chi)+2):
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=SoChiDoanhThu_{period}.xlsx'
    return response

@app.route('/')
def index():
    return redirect(url_for('sale'))

# ==================== HƯỚNG DẪN SỬ DỤNG ====================
@app.route('/huong-dan-su-dung')
@login_required  # hoặc bỏ nếu muốn ai cũng xem được
def huong_dan_su_dung():
    return render_template('huongdansudung.html')


def get_products_data(query=None):
    """Lấy danh sách sản phẩm, bao gồm giá vốn từ inventory."""
    conn = get_db_connection()
    if conn is None:
        return jsonify({"success": False, "error": "Lỗi kết nối database."}), 500
        
    c = conn.cursor()
    c.row_factory = sqlite3.Row 
    
    # === TRUY VẤN ĐÚNG: JOIN products VÀ inventory ===
    sql = """
        SELECT
            p.id, p.name, p.barcode, 
            p.base_price,                      -- Giá bán cơ bản (từ products)
            p.unit, 
            p.unit1, p.unit_ratio, 
            p.price,                           -- Giá bán đơn vị chuyển đổi
	    p.barcode1
            i.avg_cost                         -- <<< ĐÃ SỬA: Lấy Giá Vốn (avg_cost) từ inventory
        FROM products p
        LEFT JOIN inventory i ON p.id = i.product_id -- Nối với bảng inventory
        WHERE 1=1
    """
    params = []
    
    if query:
        # Nếu có tìm kiếm
        sql += " AND (p.name LIKE ? OR p.barcode LIKE ?)"
        params.extend([f'%{query}%', f'%{query}%'])
    
    sql += " LIMIT 100"
        
    try:
        c.execute(sql, tuple(params))
        products = c.fetchall()
        result = [dict(row) for row in products]
        return jsonify(result), 200
        
    except sqlite3.OperationalError as e:
        conn.rollback()
        print(f"LỖI VẬN HÀNH SQL (CONSOLE): {e}") 
        return jsonify({"success": False, "error": f"Lỗi SQL: {e}. Vui lòng kiểm tra console server."}), 500
    except Exception as e:
        conn.rollback()
        print(f"LỖI HỆ THỐNG KHÁC (CONSOLE): {e}")
        return jsonify({"success": False, "error": f"Lỗi hệ thống không xác định: {e}"}), 500
    finally:
        # Quan trọng: Đảm bảo conn.close() được gọi nếu bạn không dùng app context teardown
        # Nếu bạn đang dùng g.db, việc đóng kết nối sẽ do @app.teardown_appcontext xử lý
        if 'g' not in globals() or not hasattr(g, '_database'):
             conn.close()



@app.route('/suppliers')
@login_required
def suppliers_page():
    return render_template('suppliers.html')

#=== Tự Tạo Nhà Cung Cấp Khi Lập Phiếu Nhập Kho nếu Database chưa có===#
import re

@app.route('/api/suppliers/by-tax-code/<tax_code>', methods=['GET'])
@login_required
def get_or_create_supplier_by_tax(tax_code):
    """
    API: Lấy hoặc tự động tạo nhà cung cấp theo mã số thuế
    - Nếu MST đã tồn tại → trả về thông tin hiện có
    - Nếu chưa tồn tại → tạo mới với tên mặc định và trả về
    """
    # Chuẩn hóa mã số thuế
    tax_code = (tax_code or "").strip().upper().replace(" ", "")
    
    if not tax_code:
        return jsonify({
            "success": False,
            "error": "Mã số thuế không được để trống"
        }), 400

    # Kiểm tra định dạng MST Việt Nam cơ bản (10 số hoặc 13 số, có thể có dấu gạch ngang)
    if not re.match(r'^\d{10}(-\d{3})?$|^\d{13}$', tax_code):
        return jsonify({
            "success": False,
            "error": "Mã số thuế không đúng định dạng (10 hoặc 13 số)"
        }), 400

    try:
        with get_db_connection() as conn:  # ← sử dụng hàm get_db_connection() có sẵn của bạn
            conn.row_factory = sqlite3.Row
            c = conn.cursor()

            # Tìm nhà cung cấp theo MST
            c.execute("""
                SELECT id, name, tax_code, address, phone, email, created_at
                FROM suppliers 
                WHERE tax_code = ?
            """, (tax_code,))
            
            supplier = c.fetchone()

            if supplier:
                return jsonify({
                    "success": True,
                    "supplier": dict(supplier),
                    "created": False,
                    "message": "Đã tìm thấy nhà cung cấp"
                })

            # Tạo mới nếu không tìm thấy
            default_name = f"NCC {tax_code}"

            c.execute("""
                INSERT INTO suppliers 
                (name, tax_code, created_at)
                VALUES (?, ?, datetime('now'))
            """, (default_name, tax_code))

            new_id = c.lastrowid
            conn.commit()

            # Lấy lại thông tin vừa tạo
            c.execute("""
                SELECT id, name, tax_code, created_at
                FROM suppliers 
                WHERE id = ?
            """, (new_id,))
            
            new_supplier = c.fetchone()

            return jsonify({
                "success": True,
                "supplier": dict(new_supplier),
                "created": True,
                "message": "Đã tự động tạo nhà cung cấp mới"
            })

    except sqlite3.IntegrityError:
        # Trường hợp mã số thuế đã tồn tại (race condition)
        return jsonify({
            "success": False,
            "error": "Mã số thuế này đã được sử dụng"
        }), 409

    except sqlite3.Error as e:
        return jsonify({
            "success": False,
            "error": f"Lỗi cơ sở dữ liệu: {str(e)}"
        }), 500

    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Lỗi hệ thống: {str(e)}"
        }), 500




@app.route('/profit')
@login_required
def profit():
    return render_template('profit.html')

@app.route('/reports')
@login_required
def reports():
    return render_template('reports.html')

@app.route('/order')
@login_required
def order():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, order_code, customer_name, customer_phone, total_amount
        FROM orders
        WHERE status = 'pending' OR status = 'partial_paid'
        ORDER BY created_at DESC
    """)
    orders = cursor.fetchall()
    conn.close()

    orders_list = [
        {
            'id': o[0],           # o (chữ thường)
            'order_code': o[1],   # o (chữ thường)
            'customer_name': o[2],
            'customer_phone': o[3],
            'total': o[4]
        } for o in orders          # o (chữ thường) – PHẢI GIỐNG NHAU!
    ]

    return render_template('order.html', orders=orders_list)




# === API SUPPLIERS ===
@app.route('/api/suppliers', methods=['GET', 'POST', 'PUT', 'DELETE'])
@login_required
def api_suppliers():
    conn = get_db_connection()
    c = conn.cursor()
    try:
        if request.method == 'GET':
            q = request.args.get('q', '')
            if q:
                like = f"%{q}%"
                c.execute("SELECT * FROM suppliers WHERE code LIKE ? OR name LIKE ? OR phone LIKE ? OR tax_code LIKE ?", (like,)*4)
            else:
                c.execute("SELECT * FROM suppliers ORDER BY name")
            return jsonify([dict(row) for row in c.fetchall()])
        data = request.get_json() or {}
        if request.method == 'POST':
            name = data.get('name', '').strip()
            if not name: return jsonify({"error": "Tên NCC trống"}), 400
            c.execute("SELECT COUNT(*) FROM suppliers")
            count = c.fetchone()[0]
            code = data.get('code', '').strip() or f"NCC{count + 1:06d}"
            c.execute("INSERT INTO suppliers (code, name, phone, email, address, note, tax_code) VALUES (?, ?, ?, ?, ?, ?, ?)",
                      (code, name, data.get('phone',''), data.get('email',''), data.get('address',''), data.get('note',''), data.get('tax_code','')))
            conn.commit()
            return jsonify({"success": True, "code": code})
        if request.method in ['PUT', 'DELETE']:
            id_ = data.get('id')
            if not id_: return jsonify({"error": "Thiếu ID"}), 400
            if request.method == 'PUT':
                c.execute("UPDATE suppliers SET code=?, name=?, phone=?, email=?, address=?, note=?, tax_code=? WHERE id=?",
                          (data.get('code',''), data.get('name',''), data.get('phone',''), data.get('email',''), data.get('address',''), data.get('note',''), data.get('tax_code',''), id_))
            else:
                c.execute("DELETE FROM suppliers WHERE id=?", (id_,))
            conn.commit()
            return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()

#====NHẬP THÔNG TIN NHÀ CUNG CẤP VÀO BẢNG SUPPLIERS TỪ PHIẾU NHẬP===#
@app.route('/api/suppliers/upsert', methods=['POST'])
@login_required
def api_suppliers_upsert():
    conn = get_db_connection()
    c = conn.cursor()
    try:
        data = request.get_json()
        tax_code = data.get('tax_code', '').strip()
        name = data.get('name', '').strip()
        address = data.get('address', '').strip()
        phone = data.get('phone', '').strip()

        # Tìm kiếm ưu tiên MST, sau đó đến Tên
        supplier_id = None
        if tax_code:
            c.execute("SELECT id FROM suppliers WHERE tax_code = ?", (tax_code,))
            res = c.fetchone()
            if res: supplier_id = res['id']

        if not supplier_id:
            c.execute("SELECT id FROM suppliers WHERE name = ?", (name,))
            res = c.fetchone()
            if res: supplier_id = res['id']

        if supplier_id:
            # Update thông tin mới nhất từ XML
            c.execute("""
                UPDATE suppliers 
                SET tax_code = COALESCE(NULLIF(tax_code, ''), ?), 
                    address = ?, phone = ? 
                WHERE id = ?
            """, (tax_code, address, phone, supplier_id))
        else:
            # Tạo mới nếu hoàn toàn chưa có
            c.execute("SELECT COUNT(*) FROM suppliers")
            count = c.fetchone()[0]
            code = f"NCC{count + 1:06d}"
            c.execute("""
                INSERT INTO suppliers (code, name, tax_code, address, phone) 
                VALUES (?, ?, ?, ?, ?)
            """, (code, name, tax_code, address, phone))
            supplier_id = c.lastrowid

        conn.commit()
        return jsonify({"success": True, "supplier_id": supplier_id})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        conn.close()

# === API SUPPLIERS IMPORT ===
@app.route('/api/suppliers/import', methods=['POST'])
@login_required
@admin_or_master_required
def api_suppliers_import():
    conn = get_db_connection()
    c = conn.cursor()
    data = request.get_json()
    imported_count = 0
    errors = []
    try:
        for item in data:
            name = item.get('name', '').strip()
            if not name:
                errors.append({"item": item, "error": "Tên NCC trống"})
                continue
            code = item.get('code', '').strip()
            if not code:
                c.execute("SELECT COUNT(*) FROM suppliers")
                count = c.fetchone()[0]
                code = f"NCC{count + 1:06d}"
            try:
                c.execute("INSERT INTO suppliers (code, name, phone, email, address, note, tax_code) VALUES (?, ?, ?, ?, ?, ?, ?)",
                          (code, name, item.get('phone',''), item.get('email',''), item.get('address',''), item.get('note',''), item.get('tax_code','')))
                imported_count += 1
            except sqlite3.IntegrityError:
                errors.append({"item": item, "error": "Mã NCC hoặc trường UNIQUE đã tồn tại"})
            except Exception as e:
                errors.append({"item": item, "error": str(e)})
        conn.commit()
        return jsonify({"success": True, "count": imported_count, "errors": errors})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": f"Lỗi server: {str(e)}"}), 500
    finally:
        conn.close()

# === API ĐƠN HÀNG ===
@app.route('/api/orders', methods=['GET', 'POST', 'PUT'])
# @login_required # Giữ nguyên nếu bạn đang dùng decorator này
def api_orders():
    conn = get_db_connection()
    
    # Quan trọng: Đảm bảo cursor trả về kết quả dưới dạng dictionary để jsonify hoạt động tốt
    conn.row_factory = sqlite3.Row 
    c = conn.cursor()

    # --- 1. POST: TẠO ĐƠN HÀNG MỚI (Nháp) ---
    if request.method == 'POST':
        data = request.get_json()
        
        # Lấy các trường dữ liệu, bao gồm các trường MỚI
        customer_name = data.get('customer_name', '')
        customer_phone = data.get('customer_phone', '')
        customer_taxcode = data.get('customer_taxcode', '')
        customer_address = data.get('customer_address', '')
        note = data.get('note', '')
        
        # Đặt giá trị mặc định cho đơn hàng mới
        total = 0 
        status = 'Hoàn Thành' 
        payment_method = data.get('payment_method', 'Tiền mặt')

        sql = """
            INSERT INTO sale (
                date, customer_name, customer_phone, customer_taxcode, customer_address, 
                total, payment_method, note, status
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        params = (
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 
            customer_name, customer_phone, customer_taxcode, customer_address, 
            total, payment_method, note, status
        )
        
        try:
            c.execute(sql, params)
            conn.commit()
            return jsonify({"success": True, "id": c.lastrowid}), 201
        except Exception as e:
            conn.rollback()
            return jsonify({"success": False, "error": f"Lỗi tạo đơn hàng: {e}"}), 500


    # --- 2. PUT: CẬP NHẬT CHI TIẾT ĐƠN HÀNG ---
    if request.method == 'PUT':
        data = request.get_json()
        order_id = data.get('id')
        if not order_id:
            return jsonify({"success": False, "error": "Thiếu ID đơn hàng"}), 400
            
        sql = """
            UPDATE sale SET 
                customer_name=?, customer_phone=?, customer_taxcode=?, customer_address=?, note=? 
            WHERE id=?
        """
        params = (
            data.get('customer_name',''), 
            data.get('customer_phone',''), 
            data.get('customer_taxcode',''),   # Cột mới
            data.get('customer_address',''),   # Cột mới
            data.get('note',''), 
            order_id
        )
        
        try:
            c.execute(sql, params)
            conn.commit()
            return jsonify({"success": True}), 200
        except Exception as e:
            conn.rollback()
            return jsonify({"success": False, "error": f"Lỗi cập nhật đơn hàng: {e}"}), 500


    # --- 3. GET: LẤY DANH SÁCH ĐƠN HÀNG (có tìm kiếm) ---
    query = request.args.get('q', '').strip()
    
    # Lấy TẤT CẢ các cột cần thiết cho bảng Orders
    sql = """
        SELECT 
            id, date, customer_name, customer_phone, total_amount as total, invoice_number, 
            customer_taxcode, customer_address, status, note, payment_method 
        FROM sale 
        WHERE 1=1
    """
    params = []

    if query:
        # Tìm kiếm theo tên, SĐT, hoặc số hóa đơn
        sql += " AND (customer_name LIKE ? OR customer_phone LIKE ? OR invoice_number LIKE ?)"
        params.extend([f'%{query}%', f'%{query}%', f'%{query}%'])

    sql += " ORDER BY id DESC LIMIT 100"
    
    try:
        c.execute(sql, tuple(params))
        # Trả về kết quả dưới dạng JSON (List of Dicts)
        return jsonify([dict(row) for row in c.fetchall()]), 200
    except Exception as e:
        return jsonify({"success": False, "error": f"Lỗi truy vấn danh sách đơn hàng: {e}"}), 500
    finally:
        # Lưu ý: Nếu bạn đang sử dụng g.db, việc đóng kết nối sẽ do @app.teardown_appcontext xử lý.
        # Nếu không, bạn cần đảm bảo conn.close() được gọi.
        if 'g' not in globals() or not hasattr(g, '_database'):
             conn.close()

@app.route('/api/orders/items', methods=['POST'])
# @login_required # Giữ nguyên nếu bạn đang dùng decorator này
def api_orders_items():
    """Lưu chi tiết (sale_items), tính tổng tiền và cập nhật total_amount cho đơn hàng."""
    conn = get_db_connection()
    c = conn.cursor()
    data = request.get_json()
    
    order_id = data.get('id')
    items = data.get('items', []) # Danh sách các mặt hàng (sản phẩm, số lượng, giá)

    if not order_id:
        return jsonify({"success": False, "error": "Thiếu ID đơn hàng"}), 400

    try:
        # 1. Xóa các chi tiết cũ của đơn hàng này
        c.execute("DELETE FROM sale_items WHERE sale_id=?", (order_id,))
        
        # 2. Thêm lại các chi tiết mới và tính tổng tiền
        grand_total = 0
        
        for item in items:
            # Chuyển đổi an toàn sang float
            try:
                quantity = float(item.get('quantity', 0))
                price = float(item.get('price', 0))
            except (TypeError, ValueError):
                raise ValueError("Số lượng hoặc Đơn giá không hợp lệ.")
                
            item_total = quantity * price
            grand_total += item_total
            
            # Kiểm tra dữ liệu sản phẩm cơ bản
            if item.get('product_id') is None or not item.get('product_name'):
                 raise ValueError("Thiếu thông tin sản phẩm.")

            c.execute("""
                INSERT INTO sale_items (sale_id, product_id, product_name, unit_name, quantity, price, total) 
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                order_id, 
                item.get('product_id'), 
                item.get('product_name'), 
                item.get('unit_name', 'Cái'), # Giá trị mặc định là 'Cái' nếu không có
                quantity, 
                price, 
                item_total
            ))
            
        # 3. Cập nhật cột total_amount (theo tên cột CSDL của bạn) vào bảng sale
        c.execute("UPDATE sale SET total_amount=? WHERE id=?", (grand_total, order_id))
        
        conn.commit()
        return jsonify({"success": True, "total_amount": grand_total}), 200

    except Exception as e:
        conn.rollback()
        error_message = str(e) if isinstance(e, ValueError) else f"Lỗi xử lý chi tiết đơn hàng: {e}"
        return jsonify({"success": False, "error": error_message}), 500
    finally:
        # Đảm bảo đóng kết nối
        conn.close()

@app.route('/api/orders', methods=['POST'])
def api_create_order():
    data = request.get_json()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO orders (customer_name, customer_phone, note, total, status)
        VALUES (?, ?, ?, 0, 'pending')
    """, (data['customer_name'], data.get('customer_phone'), data.get('note')))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/orders/list')
def api_orders_list():
    # Lấy tham số
    start = request.args.get('start') # Ví dụ: "2025-11-23"
    end = request.args.get('end')     # Ví dụ: "2025-11-23"
    q = request.args.get('q', '').strip()
    
    # --- SỬA LỖI LỆCH NGÀY QUAN TRỌNG NHẤT ---
    # Thêm 00:00:00 cho ngày bắt đầu và 23:59:59 cho ngày kết thúc
    start_of_day = f"{start} 00:00:00" if start else None
    end_of_day = f"{end} 23:59:59" if end else None
    # ----------------------------------------
    
    try:
        # 1. Sử dụng 'with' để quản lý kết nối an toàn
        with sqlite3.connect(database) as conn:
            # Thiết lập để trả về kết quả dưới dạng dictionary
            conn.row_factory = sqlite3.Row
            c = conn.cursor()

            sql = """
                SELECT id, customer_name, total_amount, date, invoice_number, status
                FROM sale
                WHERE 1=1
            """
            params = []

            # Áp dụng ngày bắt đầu
            if start_of_day:
                sql += " AND date >= ?"
                params.append(start_of_day)
                
            # Áp dụng ngày kết thúc (ĐÃ FIX)
            if end_of_day:
                sql += " AND date <= ?"
                params.append(end_of_day)
                
            if q:
                like = f"%{q}%"
                sql += " AND (customer_name LIKE ? OR invoice_number LIKE ? OR CAST(id AS TEXT) LIKE ?)"
                params += [like, like, like]

            # Giữ nguyên việc giới hạn kết quả (LIMIT 500) để ngăn chặn việc tải quá nhiều dữ liệu
            sql += " ORDER BY date DESC, id DESC LIMIT 500"

            rows = c.execute(sql, params).fetchall()
            
            # 2. Chuyển đổi kết quả (đã tối ưu: sử dụng list comprehension)
            data = [dict(r) for r in rows]
            return jsonify(data)

    except sqlite3.Error as e:
        print(f"Lỗi Database khi lấy danh sách đơn hàng: {e}")
        return jsonify({"success": False, "message": f"Lỗi cơ sở dữ liệu: {e}"}), 500
    except Exception as e:
        print(f"Lỗi không xác định khi lấy danh sách đơn hàng: {e}")
        return jsonify({"success": False, "message": "Lỗi máy chủ không xác định."}), 500


# 3. Trang XUẤT E-INVOICE (in trực tiếp)

@app.route("/api/orders/<int:id>", methods=["DELETE"])
def delete_order(id):
    try:
        # Sử dụng 'with' để đảm bảo kết nối được đóng (conn.close()) dù có lỗi hay không.
        with sqlite3.connect(database) as conn:
            # Tự động bật chế độ commit
            # (Trong khối 'with' mặc định, commit sẽ được gọi khi khối kết thúc thành công, 
            # nhưng tốt hơn là nên gọi tường minh hoặc kiểm soát)
            
            c = conn.cursor()
            
            # Kiểm tra xem có bản ghi nào bị xóa không
            c.execute("DELETE FROM sale WHERE id = ?", (id,))
            
            # Lưu các thay đổi (commit transaction)
            conn.commit()
            
            # Kiểm tra số lượng hàng bị ảnh hưởng
            if c.rowcount == 0:
                # Nếu không có hàng nào bị xóa (ID không tồn tại)
                return jsonify({"success": False, "message": f"Không tìm thấy đơn hàng có ID: {id}"}), 404
                
            return jsonify({"success": True, "message": f"Đã xóa đơn hàng ID: {id}"})

    except sqlite3.Error as e:
        # Bắt lỗi database (ví dụ: database bị khóa, lỗi I/O, v.v.)
        print(f"Lỗi Database khi xóa đơn hàng ID {id}: {e}")
        return jsonify({"success": False, "message": f"Lỗi cơ sở dữ liệu: {e}"}), 500
    except Exception as e:
        # Bắt các lỗi chung khác
        print(f"Lỗi không xác định khi xóa đơn hàng ID {id}: {e}")
        return jsonify({"success": False, "message": "Lỗi máy chủ không xác định."}), 500

@app.route("/api/orders/upsert", methods=["POST"])
def upsert_order():
    data = request.json
    
    # 1. Kiểm tra dữ liệu đầu vào (Validation)
    customer_name = data.get("customer_name", "").strip()
    if not customer_name:
        return jsonify({"success": False, "message": "Tên khách hàng là bắt buộc."}), 400 # Bad Request

    try:
        # 2. Sử dụng 'with' và quản lý giao dịch
        with sqlite3.connect(database) as conn:
            c = conn.cursor()
            
            # Ghi chú: Nếu invoice_number không phải là Autoincrement, 
            # bạn nên tạo logic đánh số hóa đơn ở đây thay vì dùng None.
            c.execute("""
                INSERT INTO sale (invoice_number, customer_name, date, total_amount, status)
                VALUES (?, ?, DATE('now'), 0, 'Nháp')
            """, (None, customer_name))
            
            conn.commit()
            new_id = c.lastrowid

            return jsonify({"success": True, "id": new_id})

    except sqlite3.Error as e:
        # Lỗi xảy ra, transaction sẽ tự động được rollback (trong hầu hết các trường hợp SQLite)
        print(f"Lỗi Database khi tạo đơn hàng mới: {e}")
        return jsonify({"success": False, "message": f"Lỗi cơ sở dữ liệu: {e}"}), 500
    except Exception as e:
        print(f"Lỗi không xác định khi tạo đơn hàng mới: {e}")
        return jsonify({"success": False, "message": "Lỗi máy chủ không xác định."}), 500

@app.route('/api/orders', methods=['PUT'])
def api_update_order():
    data = request.get_json()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE orders 
        SET customer_name=?, customer_phone=?, note=?
        WHERE id=?
    """, (data['customer_name'], data.get('customer_phone'), data.get('note'), data['id']))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/orders/invoice', methods=['POST'])
def api_save_invoice():
    data = request.get_json()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE orders SET invoice_number=? WHERE id=?", (data['invoice_number'], data['id']))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/orders/invoice', methods=['POST'])
# @login_required # Giữ nguyên nếu bạn đang dùng decorator này
def api_update_invoice():
    """Cập nhật Số hóa đơn và trạng thái cho đơn hàng."""
    conn = get_db_connection()
    c = conn.cursor()
    data = request.get_json()
    
    order_id = data.get('id')
    invoice_number = data.get('invoice_number', '').strip()

    if not order_id or not invoice_number:
        return jsonify({"success": False, "error": "Thiếu ID đơn hàng hoặc Số hóa đơn"}), 400

    # Thực hiện cập nhật Số Hóa Đơn và chuyển trạng thái sang "Hoàn thành"
    sql = """
        UPDATE sale SET 
            invoice_number=?, 
            status='Hoàn thành' 
        WHERE id=?
    """
    params = (invoice_number, order_id)
    
    try:
        c.execute(sql, params)
        if c.rowcount == 0:
             return jsonify({"success": False, "error": "Không tìm thấy đơn hàng để cập nhật"}), 404

        conn.commit()
        return jsonify({"success": True, "id": order_id}), 200
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": f"Lỗi cập nhật số hóa đơn: {e}"}), 500
    finally:
        conn.close()


#======================================================================= Start of Báo Cáo Thuế ===================================================================#

# === API BÁO CÁO THUẾ ===
#=== Lấy Thông Tin Hộ Kinh Doanh cho trang kê khai thuế===#
@app.route('/api/company_info', methods=['GET'])
@login_required
def api_company_info():
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # 1. Ưu tiên lấy từ tờ khai GẦN NHẤT trong tax_declarations
        cursor.execute("""
            SELECT
                mst, mst_cu, tenNNT, ct05, ct06, ct10, tuGio, denGio,
                dthoaiNNT, emailNNT,
                ct12b_soNha, ct12c_maPhuong, ct12c_tenPhuong, ct12d_maTinh, ct12d_tenTinh,
                ct13a_soNha, ct13b_maPhuong, ct13b_tenPhuong, ct13d_maTinh, ct13d_tenTinh, nguoiKy
            FROM tax_declarations
            ORDER BY created_at DESC
            LIMIT 1
        """)
        row = cursor.fetchone()
        if row:
            columns = [desc[0] for desc in cursor.description]
            info = dict(zip(columns, row))
            # Chuyển None thành chuỗi rỗng để frontend không bị lỗi
            info = {k: v if v is not None else '' for k, v in info.items()}

            # Lấy danh sách ngành nghề từ bảng hkd_nganh_nghe theo mst
            mst = info.get('mst', '')
            if mst:
                cursor.execute("""
                    SELECT ma_nganh, ten_nganh
                    FROM hkd_nganh_nghe
                    WHERE mst = ?
                    ORDER BY thu_tu ASC
                """, (mst,))
                nganh_rows = cursor.fetchall()
                info['nganh_nghe'] = [
                    {"ma": row[0], "ten": row[1]} for row in nganh_rows
                ]
            else:
                info['nganh_nghe'] = []

            return jsonify({
                "success": True,
                "data": info,
                "source": "tax_declarations_latest"
            })

        # 2. Nếu chưa có tờ khai nào → fallback sang business_info
        cursor.execute("""
            SELECT
                tax_code AS mst,
                representative_name AS tenNNT,
                business_name AS ct05,
                bank_account AS ct06,
                address AS ct12b_soNha,
                '' AS ct12c_maPhuong,
                '' AS ct12c_tenPhuong,
                '' AS ct12d_maTinh,
                '' AS ct12d_tenTinh,
                '' AS ct13a_soNha,
                '' AS ct13b_maPhuong,
                '' AS ct13b_tenPhuong,
                '' AS ct13d_maTinh,
                '' AS ct13d_tenTinh,
                '' AS dthoaiNNT,
                '' AS emailNNT
            FROM business_info
            LIMIT 1
        """)
        row = cursor.fetchone()
        if row:
            columns = [desc[0] for desc in cursor.description]
            info = dict(zip(columns, row))
            info = {k: v if v is not None else '' for k, v in info.items()}

            # Vẫn lấy ngành nghề từ hkd_nganh_nghe nếu có mst
            mst = info.get('mst', '')
            if mst:
                cursor.execute("""
                    SELECT ma_nganh, ten_nganh
                    FROM hkd_nganh_nghe
                    WHERE mst = ?
                    ORDER BY thu_tu ASC
                """, (mst,))
                nganh_rows = cursor.fetchall()
                info['nganh_nghe'] = [
                    {"ma": row[0], "ten": row[1]} for row in nganh_rows
                ]
            else:
                info['nganh_nghe'] = []

            return jsonify({
                "success": True,
                "data": info,
                "source": "business_info"
            })

        # 3. Nếu cả hai bảng đều không có dữ liệu
        return jsonify({
            "success": False,
            "message": "Chưa có dữ liệu thông tin hộ kinh doanh trong hệ thống"
        })

    except sqlite3.Error as db_err:
        print(f"Lỗi SQLite trong /api/company_info: {db_err}")
        return jsonify({
            "success": False,
            "error": f"Lỗi cơ sở dữ liệu: {str(db_err)}"
        }), 500

    except Exception as e:
        import traceback
        print("Lỗi tổng quát trong /api/company_info:")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

    finally:
        if conn:
            conn.close()

@app.route('/api/tax_data')
@login_required
def api_tax_data():
    ky = request.args.get('ky')
    if not ky or '/' not in ky:
        return jsonify({"success": False, "error": "Kỳ không hợp lệ"}), 400

    try:
        q, y = map(int, ky.split('/'))
        start_month = (q - 1) * 3 + 1
        end_month = q * 3
        _, last_day = calendar.monthrange(y, end_month)

        start_search = f"{y}-{start_month:02d}-01 00:00:00"
        end_search = f"{y}-{end_month:02d}-{last_day:02d} 23:59:59"

        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        c = conn.cursor()

        # DOANH THU THỰC TẾ - CHỈ TỪ BẢNG SALE
        c.execute("""
            SELECT COALESCE(SUM(total_amount), 0) AS doanhthu
            FROM sale
            WHERE status = 'completed'
            AND date >= ? AND date <= ?
        """, (start_search, end_search))
        row = c.fetchone()
        doanhthu = float(row['doanhthu'] if row else 0)

        # CHI PHÍ ct24 (lương nhân viên)
        c.execute("""
            SELECT COALESCE(SUM(salary_rate), 0) AS ct24
            FROM Salary_Detail
            WHERE month BETWEEN ? AND ? AND year = ?
        """, (start_month, end_month, y))
        row = c.fetchone()
        ct24 = float(row['ct24'] if row else 0)

        # Hàm tiện ích lấy tổng chi phí theo loại
        def get_sum_expense(types):
            if not types:
                return 0.0
            placeholders = ','.join(['?'] * len(types))
            sql = f"""
                SELECT COALESCE(SUM(amount), 0) AS total
                FROM phieu_chi
                WHERE date >= ? AND date <= ?
                AND expense_type IN ({placeholders})
            """
            params = [start_search, end_search] + types
            c.execute(sql, params)
            row = c.fetchone()
            return float(row['total'] if row else 0)

        ct25 = get_sum_expense(['CP_DIEN'])
        ct26 = get_sum_expense(['CP_NUOC'])
        ct27 = get_sum_expense(['CP_INTERNET', 'CP_DT', 'CP_VIENTHONG'])
        ct28 = get_sum_expense(['CP_MATBANG', 'CP_THUEKHO'])

        # ct29: Chi phí quản lý khác + khấu hao TSCĐ
        ct29_raw = get_sum_expense(['CP_VPP', 'CP_DUNG_CU', 'CP_QUANLY'])

        ct29_khauhao = 0.0
        c.execute("""
            SELECT nguyen_gia_tinh_khau_hao, so_thang_khau_hao, ngay_bat_dau_su_dung
            FROM tai_san_co_dinh
            WHERE tinh_trang = 'Active'
        """)
        for asset in c.fetchall():
            ng_gia = float(asset['nguyen_gia_tinh_khau_hao'] or 0)
            s_thang = int(asset['so_thang_khau_hao'] or 1)
            try:
                ngay_bd_str = asset['ngay_bat_dau_su_dung'].split(' ')[0]
                ngay_bd = datetime.strptime(ngay_bd_str, '%Y-%m-%d')
            except:
                continue
            for m in range(start_month, end_month + 1):
                curr_date = datetime(y, m, 1)
                diff_months = (curr_date.year - ngay_bd.year) * 12 + (curr_date.month - ngay_bd.month)
                if 0 <= diff_months < s_thang:
                    ct29_khauhao += (ng_gia / s_thang)

        ct29 = ct29_raw + ct29_khauhao

        # ct30: Chi phí khác (các mã không thuộc trên)
        ct30 = get_sum_expense(['CP_KHAC'])  # Hoặc mở rộng danh sách nếu cần

        # Bảng kê hàng hóa - CHỈ LIỆT KÊ, KHÔNG ẢNH HƯỞNG DOANH THU [28]
        c.execute("SELECT id, name, unit FROM products ORDER BY name")
        products = c.fetchall()
        report_items = []

        for prod in products:
            pid = prod['id']
            # Tồn đầu kỳ
            c.execute("""
                SELECT COALESCE(SUM(quantity), 0) AS beg_qty
                FROM stock_moves
                WHERE product_id = ? AND date < ?
            """, (pid, start_search))
            row = c.fetchone()
            beg_qty = float(row['beg_qty'] if row else 0)

            # Giá vốn trung bình hiện tại
            c.execute("SELECT avg_cost FROM inventory WHERE product_id = ?", (pid,))
            row = c.fetchone()
            wac = float(row['avg_cost'] if row else 0)

            # Biến động trong kỳ
            c.execute("""
                SELECT
                    COALESCE(SUM(CASE WHEN type IN ('import', 'RETURN_SALE', 'DELETE_SALE') THEN quantity ELSE 0 END), 0) AS imp_qty,
                    COALESCE(SUM(CASE WHEN type IN ('import', 'RETURN_SALE', 'DELETE_SALE') THEN quantity * cost_price ELSE 0 END), 0) AS imp_val,
                    COALESCE(SUM(CASE WHEN type IN ('SALE', 'export', 'export_material', 'RETURN_IMPORT') THEN -quantity ELSE 0 END), 0) AS exp_qty,
                    COALESCE(SUM(CASE WHEN type IN ('SALE', 'export', 'export_material', 'RETURN_IMPORT') THEN -quantity * cost_price ELSE 0 END), 0) AS exp_val
                FROM stock_moves
                WHERE product_id = ? AND date >= ? AND date <= ?
            """, (pid, start_search, end_search))
            row = c.fetchone()
            imp_qty = float(row['imp_qty'] if row else 0)
            imp_val  = float(row['imp_val'] if row else 0)
            exp_qty  = float(row['exp_qty'] if row else 0)
            exp_val  = float(row['exp_val'] if row else 0)

            end_qty = beg_qty + imp_qty - exp_qty

            if beg_qty != 0 or imp_qty != 0 or exp_qty != 0 or end_qty != 0:
                report_items.append({
                    "product_name": prod['name'],
                    "unit_name": prod['unit'] or "Cái",
                    "beginning_quantity": round(beg_qty, 2),
                    "beginning_value": round(beg_qty * wac, 0),
                    "import_quantity": round(imp_qty, 2),
                    "import_value": round(imp_val, 0),
                    "export_quantity": round(exp_qty, 2),
                    "export_value": round(exp_val, 0),
                    "ending_quantity": round(end_qty, 2),
                    "ending_value": round(end_qty * wac, 0)
                })

        conn.close()

        return jsonify({
            "success": True,
            "doanhthu": round(doanhthu),  # [28] từ sale - đây là giá trị chính
            "expenses": {
                "ct24": round(ct24),
                "ct25": round(ct25),
                "ct26": round(ct26),
                "ct27": round(ct27),
                "ct28": round(ct28),
                "ct29": round(ct29),
                "ct30": round(ct30)
            },
            "items": report_items  # Chỉ liệt kê hàng hóa, không ảnh hưởng doanh thu
        })

    except Exception as e:
        print(f"LỖI THUẾ: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

#===API lấy dữ liệu từ tờ khai gốc để lập tờ khai bổ sung===#
@app.route('/api/tax_original', methods=['GET'])
@login_required
def api_tax_original():
    ky = request.args.get('ky')
    if not ky or '/' not in ky:
        return jsonify({"success": False, "error": "Kỳ không hợp lệ"}), 400

    try:
        q, y = map(int, ky.split('/'))
        start_month = (q - 1) * 3 + 1
        end_month = q * 3
        _, last_day = calendar.monthrange(y, end_month)

        start_search = f"{y}-{start_month:02d}-01 00:00:00"
        end_search = f"{y}-{end_month:02d}-{last_day:02d} 23:59:59"

        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        c = conn.cursor()

        # Giả sử bạn lưu tờ khai đã nộp vào bảng tax_declarations
        # (cột: ky_khai, loai_tkhai, ct24, ct25, ..., ct31, created_at, ...)
        c.execute("""
            SELECT ct24, ct25, ct26, ct27, ct28, ct29, ct30, ct31
            FROM tax_declarations
            WHERE ky_khai = ? AND loai_tkhai = 'C'
            ORDER BY created_at DESC
            LIMIT 1
        """, (ky,))
        
        row = c.fetchone()
        if row:
            data = {
                "ct24": row['ct24'] or 0,
                "ct25": row['ct25'] or 0,
                "ct26": row['ct26'] or 0,
                "ct27": row['ct27'] or 0,
                "ct28": row['ct28'] or 0,
                "ct29": row['ct29'] or 0,
                "ct30": row['ct30'] or 0,
                "ct31": row['ct31'] or 0,
            }
            return jsonify({"success": True, "data": data})
        else:
            return jsonify({
                "success": False,
                "message": f"Không tìm thấy tờ khai chính thức cho kỳ {ky}"
            })

    except Exception as e:
        print(f"Lỗi api_tax_original: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if 'conn' in locals():
            conn.close()

@app.route('/api/save_tax_declaration', methods=['POST'])
@login_required
def save_tax_declaration():
    try:
        data = request.form
        mst = data.get('mst')
        ky_khai = data.get('kyKKhai')
        loai_tkhai = data.get('loaiTKhai')

        if not mst:
            return jsonify({"success": False, "error": "Thiếu mã số thuế (mst)"}), 400

        conn = get_db_connection()
        c = conn.cursor()

        # ===== 1. XÓA TỜ KHAI CŨ NẾU TRÙNG =====
        c.execute("""
            DELETE FROM tax_declarations
            WHERE mst = ?
            AND ky_khai = ?
            AND loai_tkhai = ?
        """, (mst, ky_khai, loai_tkhai))

        # ===== 2. INSERT TỜ KHAI MỚI =====
        c.execute("""
            INSERT INTO tax_declarations (
                mst, mst_cu, tenNNT,
                ct05, ct06, ct09, ct09a, ct10, ct12b_soNha, ct12c_tenPhuong, ct12d_tenTinh, ct13a_soNha, ct13b_tenPhuong, ct13d_tenTinh, nguoiKy, ct24, ct25, ct26, ct27, ct28, ct29, ct30, ct31, tuGio, denGio, emailNNT, dthoaiNNT,
                ky_khai, loai_tkhai, so_lan,
                doanh_thu, thue_gtgt, thue_tncn,
                status
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            mst,
            data.get('mst_cu'),
            data.get('tenNNT'),
            data.get('ct05'),
            data.get('ct06'),
            float(data.get('ct09') or 0),
            float(data.get('ct09a') or 0),
            float(data.get('ct10') or 0),
            data.get('ct12b_soNha'),
            data.get('ct12c_tenPhuong'),
            data.get('ct12d_tenTinh'),
            data.get('ct13a_soNha'),
            data.get('ct13b_tenPhuong'),
            data.get('ct13d_tenTinh'),
            data.get('nguoiKy'),
            float(data.get('ct24') or 0),
            float(data.get('ct25') or 0),
            float(data.get('ct26') or 0),
            float(data.get('ct27') or 0),
            float(data.get('ct28') or 0),
            float(data.get('ct29') or 0),
            float(data.get('ct30') or 0),
            float(data.get('ct31') or 0),
            float(data.get('tuGio') or 0),
            float(data.get('denGio') or 0),
            data.get('emailNNT'),
            data.get('dthoaiNNT'),
            ky_khai,
            loai_tkhai,
            int(data.get('soLan') or 1),
            float(data.get('dt_gtgt_ct28') or 0),
            float(data.get('so_gtgt_ct28') or 0),
            float(data.get('so_tncn_ct28') or 0),
            'completed'
        ))

        # ===== 3. LƯU NGÀNH NGHỀ VÀO BẢNG hkd_nganh_nghe =====
        # Xóa ngành nghề cũ của MST này trước
        c.execute("DELETE FROM hkd_nganh_nghe WHERE mst = ?", (mst,))

        # Insert các ngành mới từ form (maNNghe_1, tenNNghe_1, ...)
        idx = 1
        while f"maNNghe_{idx}" in data:
            ma = data.get(f"maNNghe_{idx}", '').strip()
            ten = data.get(f"tenNNghe_{idx}", '').strip()
            if ma and ten:
                c.execute("""
                    INSERT OR REPLACE INTO hkd_nganh_nghe 
                    (mst, ma_nganh, ten_nganh, thu_tu)
                    VALUES (?, ?, ?, ?)
                """, (mst, ma, ten, idx))
            idx += 1

        conn.commit()
        conn.close()

        return jsonify({"success": True})

    except Exception as e:
        print("Lỗi save_tax_declaration:", e)
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/tax_report', methods=['GET', 'POST'])
def tax_report():
    if request.method == 'POST':
        data = request.form.to_dict()
        xml_content = generate_tax_xml(data)
        
        buffer = BytesIO()
        buffer.write(xml_content.encode('utf-8'))
        buffer.seek(0)
        
        mst = data.get('mst', 'UNKNOWN').replace('-', '')
        ky = data.get('kyKKhai', '').replace('/', '')
        
        # Quyết định đuôi file theo loại tờ khai
        loai_tkhai = data.get('loaiTKhai', 'C')  # 'C' = chính thức, 'B' = bổ sung thường gặp
        
        if loai_tkhai.upper() in ('C', 'CHINHTHUC', 'CHÍNH THỨC'):
            lan_bo_sung = '00'
        else:
            # là tờ khai bổ sung
            so_lan = data.get('soLan', '0')
            try:
                lan_bo_sung = str(int(so_lan)).zfill(2)  # 1 → 01, 2 → 02, ...
            except ValueError:
                lan_bo_sung = '01'  # fallback nếu dữ liệu lỗi
        
        filename = f"{mst}-01_CNKD_TT40-Q{ky}-L{lan_bo_sung}.xml"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/xml'
        )
    
    return render_template('tax_report.html')

import xml.etree.ElementTree as ET
from xml.dom import minidom
import sqlite3
import calendar
from datetime import datetime

def prettify(elem):
    """Pretty print XML"""
    rough_string = ET.tostring(elem, 'utf-8')
    reparsed = minidom.parseString(rough_string)
    return reparsed.toprettyxml(indent=" ")

def generate_tax_xml(data):
    NS = "http://kekhaithue.gdt.gov.vn/TKhaiThue"
    ET.register_namespace('', NS)
    xsi = "http://www.w3.org/2001/XMLSchema-instance"
    root = ET.Element("HSoThueDTu", {
        "xmlns": NS,
        "xmlns:xsi": xsi,
        "xsi:schemaLocation": f"{NS} ToKhaiThue.xsd"
    })
    hso = ET.SubElement(root, "HSoKhaiThue", {"id": "ID_1"})
    # 1. TTinChung
    ttinchung = ET.SubElement(hso, "TTinChung")
    ttindvu = ET.SubElement(ttinchung, "TTinDVu")
    ET.SubElement(ttindvu, "maDVu").text = "HTKK"
    ET.SubElement(ttindvu, "tenDVu").text = "HỖ TRỢ KÊ KHAI THUẾ"
    ET.SubElement(ttindvu, "pbanDVu").text = "5.5.6"
    ET.SubElement(ttindvu, "ttinNhaCCapDVu").text = "3D73CFFAB5DA6133D754BE7D6DB20D0B"
    ttintkhaithue = ET.SubElement(ttinchung, "TTinTKhaiThue")
    tkhaithue = ET.SubElement(ttintkhaithue, "TKhaiThue")
    ET.SubElement(tkhaithue, "maTKhai").text = "473"
    ET.SubElement(tkhaithue, "tenTKhai").text = "Tờ khai thuế đối với hộ kinh doanh, cá nhân kinh doanh"
    ET.SubElement(tkhaithue, "moTaBMau").text = "(Ban hành kèm theo Thông tư số 40/2021/TT-BTC ngày 01/6/2021 của Bộ trưởng Bộ Tài Chính)"
    ET.SubElement(tkhaithue, "pbanTKhaiXML").text = data.get('pbanTKhaiXML', '2.8.3')
    loai_tkhai = data.get('loaiTKhai', 'C')
    ET.SubElement(tkhaithue, "loaiTKhai").text = loai_tkhai
    ET.SubElement(tkhaithue, "soLan").text = data.get('soLan', '0')
    ky = ET.SubElement(tkhaithue, "KyKKhaiThue")
    ET.SubElement(ky, "kieuKy").text = "Q"
    ky_str = data.get('kyKKhai', '')
    ET.SubElement(ky, "kyKKhai").text = ky_str
    if '/' in ky_str:
        q, y = ky_str.split('/')
        q, y = int(q), int(y)
        start_month = (q - 1) * 3 + 1
        end_month = q * 3
        _, last_day = calendar.monthrange(y, end_month)
        ET.SubElement(ky, "kyKKhaiTuNgay").text = f"01/{start_month:02d}/{y}"
        ET.SubElement(ky, "kyKKhaiDenNgay").text = f"{last_day:02d}/{end_month:02d}/{y}"
        ET.SubElement(ky, "kyKKhaiTuThang").text = f"{start_month:02d}/{y}"
        ET.SubElement(ky, "kyKKhaiDenThang").text = f"{end_month:02d}/{y}"
    ET.SubElement(tkhaithue, "maCQTNoiNop").text = data.get('maCQTNoiNop', '')
    ET.SubElement(tkhaithue, "tenCQTNoiNop").text = data.get('tenCQTNoiNop', '')
    ET.SubElement(tkhaithue, "ngayLapTKhai").text = data.get('ngayLapTKhai', datetime.now().strftime('%Y-%m-%d'))
    ET.SubElement(tkhaithue, "nguoiKy").text = data.get('nguoiKy', '')
    ET.SubElement(tkhaithue, "ngayKy").text = data.get('ngayKy', datetime.now().strftime('%Y-%m-%d'))
    giahan = ET.SubElement(tkhaithue, "GiaHan")
    ET.SubElement(giahan, "maLyDoGiaHan").text = ""
    ET.SubElement(giahan, "lyDoGiaHan").text = ""
    # === NGÀNH NGHỀ: LẤY TỪ BẢNG hkd_nganh_nghe ===
    mst = str(data.get('mst', '')).strip()
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row # ✅ QUAN TRỌNG
    c = conn.cursor()
    c.execute("""
        SELECT TRIM(ma_nganh) as ma_nganh,
               TRIM(ten_nganh) as ten_nganh
        FROM hkd_nganh_nghe
        WHERE TRIM(mst) = ?
        ORDER BY thu_tu ASC
    """, (mst,))
    rows = c.fetchall()
    conn.close()
    nganh_list = []
    for row in rows:
        ma = str(row["ma_nganh"]).strip() if row["ma_nganh"] else ""
        ten = str(row["ten_nganh"]).strip() if row["ten_nganh"] else ""
        if ma and ten:
            nganh_list.append((ma, ten))
    print("DEBUG MST:", mst)
    print("DEBUG NGÀNH:", nganh_list)
    # 1️⃣ Thẻ chuỗi nối
    if nganh_list:
        nganh_text = ";".join([f"{ma}.-{ten}" for ma, ten in nganh_list])
    else:
        nganh_text = ""
    nganh_node = ET.SubElement(tkhaithue, "nganhNgheKD")
    nganh_node.text = nganh_text
    # NNT
    nnt = ET.SubElement(ttintkhaithue, "NNT")
    ET.SubElement(nnt, "mst").text = data.get('mst', '')
    ET.SubElement(nnt, "tenNNT").text = data.get('tenNNT', '')
    ET.SubElement(nnt, "dchiNNT").text = (
        (data.get('ct12b_soNha') or '') + ', ' +
        (data.get('ct12c_tenPhuong') or '') + ', ' +
        (data.get('ct12d_tenTinh') or '')
    ).strip(', ')
    ET.SubElement(nnt, "dthoaiNNT").text = data.get('dthoaiNNT', '')
    ET.SubElement(nnt, "emailNNT").text = data.get('emailNNT', '')
    # CTieuTKhaiChinh
    ctieu = ET.SubElement(hso, "CTieuTKhaiChinh")
    ET.SubElement(ctieu, "mst_cu").text = data.get('mst_cu', '')
    header = ET.SubElement(ctieu, "Header")
    ET.SubElement(header, "hkdcnkdnopthuekhoan").text = "0"
    ET.SubElement(header, "cnkdnopps").text = "0"
    ET.SubElement(header, "tccnkhainopthay").text = "0"
    ET.SubElement(header, "hkdcnkdnopkekhai").text = "1"
    ET.SubElement(header, "hkdcnkdnnxddoanhthu").text = "0"
    ET.SubElement(header, "hkdchuyendoipptinhthue").text = "0"
    ET.SubElement(header, "ct05").text = data.get('ct05', '')
    ET.SubElement(header, "ct06").text = data.get('ct06', '')
    # CT08 - Ngành nghề chi tiết (có mã và tên)
    ct08 = ET.SubElement(header, "CT08")
    for idx, (ma, ten) in enumerate(nganh_list, 1):
        nn = ET.SubElement(ct08, "NNgheKDoanh")
        nn.set("id", f"ID_{idx}")
        ET.SubElement(nn, "maNNgheKDoanh").text = ma
        ET.SubElement(nn, "tenNNgheKDoanh").text = f"{ma}.-{ten}"
    ET.SubElement(header, "ct08a").text = "0"
    ET.SubElement(header, "ct09").text = data.get('ct09', '1.00')
    ET.SubElement(header, "ct09a").text = data.get('ct09a', '0.50')
    ET.SubElement(header, "ct10").text = data.get('ct10', '0')
    # CT11 - Giờ mở cửa
    ct11 = ET.SubElement(header, "CT11")
    ET.SubElement(ct11, "tuGio").text = data.get('tuGio', '7')
    ET.SubElement(ct11, "tuPhut").text = data.get('tuPhut', '0')
    ET.SubElement(ct11, "denGio").text = data.get('denGio', '18')
    ET.SubElement(ct11, "denPhut").text = data.get('denPhut', '0')
    # CT12 - Địa chỉ kinh doanh
    ct12 = ET.SubElement(header, "CT12")
    ET.SubElement(ct12, "ct12a_tdtt").text = "0"
    ET.SubElement(ct12, "ct12b_soNha").text = data.get('ct12b_soNha', '')
    ET.SubElement(ct12, "ct12c_maPhuong").text = data.get('ct12c_maPhuong', '')
    ET.SubElement(ct12, "ct12c_tenPhuong").text = data.get('ct12c_tenPhuong', '')
    ET.SubElement(ct12, "ct12d_maQuan").text = ""
    ET.SubElement(ct12, "ct12d_tenQuan").text = ""
    ET.SubElement(ct12, "ct12d_maTinh").text = data.get('ct12d_maTinh', '701')
    ET.SubElement(ct12, "ct12d_tenTinh").text = data.get('ct12d_tenTinh', 'Thành phố Hồ Chí Minh')
    ET.SubElement(ct12, "ct12e_kdbiengioi").text = "0"
    # CT13 - Địa chỉ cư trú (nếu có)
    ct13 = ET.SubElement(header, "CT13")
    ET.SubElement(ct13, "ct13a_soNha").text = data.get('ct13a_soNha', '')
    ET.SubElement(ct13, "ct13b_maPhuong").text = data.get('ct13b_maPhuong', '')
    ET.SubElement(ct13, "ct13b_tenPhuong").text = data.get('ct13b_tenPhuong', '')
    ET.SubElement(ct13, "ct13c_maQuan").text = ""
    ET.SubElement(ct13, "ct13c_tenQuan").text = ""
    ET.SubElement(ct13, "ct13d_maTinh").text = data.get('ct13d_maTinh', '701')
    ET.SubElement(ct13, "ct13d_tenTinh").text = data.get('ct13d_tenTinh', 'Thành phố Hồ Chí Minh')
    # Tính tổng từ bảng kê hàng hóa
    tong_ct09 = tong_ct11 = tong_ct13 = tong_ct15 = 0.0
    pluc_root = ET.SubElement(hso, "PLuc")
    pl_01_2 = ET.SubElement(pluc_root, "PLuc_01_2_BK_HDKD")
    vlieu = ET.SubElement(pl_01_2, "VlieuDcuSPHH")
    bke_vlieu = ET.SubElement(vlieu, "BKeVLDCSPHH")
    j = 1
    while f"ct06_{j}" in data:
        item = ET.SubElement(bke_vlieu, "CTietHKDCNKD", {"id": f"ID_{j}"})
        for c in range(6, 16):
            tag = f"ct{c:02d}"
            val = data.get(f"{tag}_{j}", '0')
            ET.SubElement(item, tag).text = val
            if c == 9: tong_ct09 += float(val or 0)
            if c == 11: tong_ct11 += float(val or 0)
            if c == 13: tong_ct13 += float(val or 0)
            if c == 15: tong_ct15 += float(val or 0)
        j += 1
    ET.SubElement(bke_vlieu, "ct17").text = str(round(tong_ct09))
    ET.SubElement(bke_vlieu, "ct19").text = str(round(tong_ct11))
    ET.SubElement(bke_vlieu, "ct21").text = str(round(tong_ct13))
    ET.SubElement(bke_vlieu, "ct23").text = str(round(tong_ct15))
    # Chi phí quản lý
    chiphi = ET.SubElement(pl_01_2, "ChiPhiQL")
    for c in range(24, 32):
        ET.SubElement(chiphi, f"ct{c}").text = data.get(f"ct{c}", '0')
    # Kết quả tính thuế [28]
    doanh_thu = float(data.get('dt_gtgt_ct28', '0'))
    thue_gtgt = float(data.get('so_gtgt_ct28', '0'))
    thue_tncn = float(data.get('so_tncn_ct28', '0'))
    kk = ET.SubElement(ctieu, "KKThueGTGT_TNCN")
    for tag_name, val in [
        ("DoanhThuThueGTGT", doanh_thu),
        ("SoThueGTGT", thue_gtgt),
        ("DoanhThuThueTNCN", doanh_thu),
        ("SoThueTNCN", thue_tncn)
    ]:
        tag = ET.SubElement(kk, tag_name)
        ET.SubElement(tag, "ct28").text = str(round(val))
        for c in [29, 30, 31]:
            ET.SubElement(tag, f"ct{c}").text = "0"
        ET.SubElement(tag, "ct32").text = str(round(val))
    # Các phần rỗng khác
    kkhai_tt = ET.SubElement(ctieu, "KKhaiThueTTDB")
    ct_tt = ET.SubElement(kkhai_tt, "CTietKKhaiThueTTDB", {"id": "ID_1"})
    for tag in ["ct2_ma", "ct2_ten", "ct3", "ct4"]:
        ET.SubElement(ct_tt, tag).text = ""
    for tag in ["ct5", "ct6", "ct7"]:
        ET.SubElement(ct_tt, tag).text = "0"
    ET.SubElement(kkhai_tt, "tong_ct5").text = "0"
    ET.SubElement(kkhai_tt, "tong_ct7").text = "0"
    kkhai_tn = ET.SubElement(ctieu, "KKhaiTBVMT_TN")
    # ThueTaiNguyen
    thue_tn = ET.SubElement(kkhai_tn, "ThueTaiNguyen")
    ct_tn = ET.SubElement(thue_tn, "CTietThueTaiNguyen", {"id": "ID_1"})
    for tag in ["ct2_ma", "ct2_ten", "ct3", "ct4"]:
        ET.SubElement(ct_tn, tag).text = ""
    for tag in ["ct5", "ct6", "ct7", "ct8"]:
        ET.SubElement(ct_tn, tag).text = "0"
    ET.SubElement(thue_tn, "tongCong").text = "0"
    # ThueBVMT
    thue_bvmt = ET.SubElement(kkhai_tn, "ThueBVMT")
    ct_bvmt = ET.SubElement(thue_bvmt, "CTietThueBVMT", {"id": "ID_1"})
    for tag in ["ct2_ma", "ct2_ten", "ct3", "ct4"]:
        ET.SubElement(ct_bvmt, tag).text = ""
    for tag in ["ct5", "ct6", "ct8"]:
        ET.SubElement(ct_bvmt, tag).text = "0"
    ET.SubElement(thue_bvmt, "tongCong").text = "0"
    # PhiBVMT
    phi_bvmt = ET.SubElement(kkhai_tn, "PhiBVMT")
    ct_phi = ET.SubElement(phi_bvmt, "CTietPhiBVMT", {"id": "ID_1"})
    for tag in ["ct2_ma", "ct2_ten", "ct3", "ct4"]:
        ET.SubElement(ct_phi, tag).text = ""
    for tag in ["ct5", "ct6", "ct8"]:
        ET.SubElement(ct_phi, tag).text = "0"
    ET.SubElement(phi_bvmt, "tongCong").text = "0"
    # Phụ lục nếu là tờ khai bổ sung
    pluc_root = ET.SubElement(hso, "PLuc") if not pluc_root else pluc_root # Đảm bảo đã có
    if loai_tkhai == 'B':
        old_data = {}
        ky_str = data.get('kyKKhai')
        if ky_str and '/' in ky_str:
            try:
                conn = get_db_connection()
                conn.row_factory = sqlite3.Row
                c = conn.cursor()
                c.execute("""
                    SELECT ct24, ct25, ct26, ct27, ct28, ct29, ct30, ct31
                    FROM tax_declarations
                    WHERE ky_khai = ? AND loai_tkhai = 'C'
                    ORDER BY created_at DESC
                    LIMIT 1
                """, (ky_str,))
                row = c.fetchone()
                if row:
                    old_data = dict(row)
                conn.close()
            except Exception as e:
                print(f"Error fetching old data for bổ sung: {e}")
        ct_keys = ['ct24', 'ct25', 'ct26', 'ct27', 'ct28', 'ct29', 'ct30', 'ct31']
        for key in ct_keys:
            if key not in old_data:
                old_data[key] = 0
        labels = {
            'ct24': ('[24]', 'Lương nhân viên'),
            'ct25': ('[25]', 'Chi phí điện'),
            'ct26': ('[26]', 'Chi phí nước'),
            'ct27': ('[27]', 'Chi phí viễn thông'),
            'ct28': ('[28]', 'Chi phí thuê kho/mặt bằng'),
            'ct29': ('[29]', 'Chi phí quản lý khác'),
            'ct30': ('[30]', 'Chi phí khác'),
            'ct31': ('[31]', 'Tổng chi phí')
        }
        # PL01_KHBS (rỗng như mẫu)
        pl01_khbs = ET.SubElement(pluc_root, "PL01_KHBS")
        header_pl = ET.SubElement(pl01_khbs, "Header")
        ET.SubElement(header_pl, "maTKhai").text = "473"
        ET.SubElement(header_pl, "tenTKhai").text = "Tờ khai thuế đối với hộ kinh doanh, cá nhân kinh doanh"
        ET.SubElement(header_pl, "maGiaoDich").text = ""
        ky_pl = ET.SubElement(header_pl, "KyKKhaiThue")
        ET.SubElement(ky_pl, "kieuKy").text = "Q"
        ET.SubElement(ky_pl, "kyKKhai").text = ky_str
        ET.SubElement(ky_pl, "kyKKhaiTuNgay").text = f"01/{start_month:02d}/{y}"
        ET.SubElement(ky_pl, "kyKKhaiDenNgay").text = f"{last_day:02d}/{end_month:02d}/{y}"
        ET.SubElement(ky_pl, "kyKKhaiTuThang").text = f"{start_month:02d}/{y}"
        ET.SubElement(ky_pl, "kyKKhaiDenThang").text = f"{end_month:02d}/{y}"
        ET.SubElement(header_pl, "soLan").text = data.get('soLan', '0')
        ET.SubElement(header_pl, "mst").text = data.get('mst', '')
        ET.SubElement(header_pl, "tenNNT").text = data.get('tenNNT', '')
        ET.SubElement(header_pl, "mstDLyThue", {"xsi:nil": "true"})
        ET.SubElement(header_pl, "tenDLyThue").text = ""
        ET.SubElement(header_pl, "soHDongDLyThue").text = ""
        ET.SubElement(header_pl, "ngayKyHDDLyThue", {"xsi:nil": "true"})
        ET.SubElement(pl01_khbs, "ma_DonViTien").text = "VND"
        ET.SubElement(pl01_khbs, "ten_DonViTien").text = "Đồng Việt Nam"
        muc_a = ET.SubElement(pl01_khbs, "Muc_A")
        muc_i = ET.SubElement(muc_a, "Muc_I")
        muc_1 = ET.SubElement(muc_i, "Muc_1")
        chitiet_1 = ET.SubElement(muc_1, "ChiTiet", {"id": "ID_1"})
        ET.SubElement(chitiet_1, "ct2_ma").text = ""
        ET.SubElement(chitiet_1, "ct2_ten").text = ""
        ET.SubElement(chitiet_1, "ct3").text = "0"
        ET.SubElement(muc_1, "tongCong_ct10").text = "0"
        muc_2 = ET.SubElement(muc_i, "Muc_2")
        ds_pluc = ET.SubElement(muc_2, "DSachPLuc")
        bke_pluc = ET.SubElement(ds_pluc, "BKePLuc", {"id": "ID_1"})
        ET.SubElement(bke_pluc, "ma_PLuc").text = ""
        ET.SubElement(bke_pluc, "ten_PLuc").text = ""
        ctiet_pluc = ET.SubElement(bke_pluc, "CTietPLuc")
        chitiet_pl = ET.SubElement(ctiet_pluc, "ChiTiet", {"id": "ID_1"})
        for tag in ["ct2_ma", "ct2_ten", "ct3_ma", "ct3_ten", "ct04_ma", "ct04_ten", "ct05_ma", "ct05_ten", "ct06_ma", "ct06_ten"]:
            ET.SubElement(chitiet_pl, tag).text = ""
        ET.SubElement(chitiet_pl, "ct7").text = "0"
        ET.SubElement(ds_pluc, "tongCong_ct11").text = "0"
        muc_3 = ET.SubElement(muc_i, "Muc_3")
        ET.SubElement(muc_3, "ct3a_1", {"xsi:nil": "true"})
        ET.SubElement(muc_3, "ct3a_2").text = "0"
        ET.SubElement(muc_3, "ct3b").text = "0"
        muc_ii = ET.SubElement(muc_a, "Muc_II")
        chitiet_ii = ET.SubElement(muc_ii, "ChiTiet", {"id": "ID_1"})
        ET.SubElement(chitiet_ii, "ct2_ma").text = ""
        ET.SubElement(chitiet_ii, "ct2_ten").text = ""
        ET.SubElement(chitiet_ii, "ct3").text = "0"
        ET.SubElement(muc_ii, "tongCong_ct12").text = "0"
        muc_iii = ET.SubElement(muc_a, "Muc_III")
        chitiet_iii = ET.SubElement(muc_iii, "ChiTiet", {"id": "ID_1"})
        ET.SubElement(chitiet_iii, "ct2_ma").text = ""
        ET.SubElement(chitiet_iii, "ct2_ten").text = ""
        ET.SubElement(chitiet_iii, "ct3").text = "0"
        ET.SubElement(muc_iii, "tongCong_ct13").text = "0"
        muc_b = ET.SubElement(pl01_khbs, "Muc_B")
        muc_i_b = ET.SubElement(muc_b, "Muc_I")
        ET.SubElement(muc_i_b, "ct_1").text = "0"
        ET.SubElement(muc_i_b, "ct_2_so").text = ""
        ET.SubElement(muc_i_b, "ct_2_ngay", {"xsi:nil": "true"})
        ET.SubElement(muc_i_b, "ct_2_CQT_ma").text = ""
        ET.SubElement(muc_i_b, "ct_2_CQT_ten").text = ""
        ET.SubElement(muc_i_b, "ct_3_so").text = ""
        ET.SubElement(muc_i_b, "ct_3_ngay", {"xsi:nil": "true"})
        muc_ii_b = ET.SubElement(muc_b, "Muc_II")
        ET.SubElement(muc_ii_b, "ct_1").text = "0"
        ET.SubElement(muc_ii_b, "ct_2").text = "0"
        # PL01_1_KHBS - Bản giải trình khai bổ sung
        pl01_1_khbs = ET.SubElement(pluc_root, "PL01_1_KHBS")
        header_pl1 = ET.SubElement(pl01_1_khbs, "Header")
        ET.SubElement(header_pl1, "maTKhai").text = "473"
        ET.SubElement(header_pl1, "tenTKhai").text = "Tờ khai thuế đối với hộ kinh doanh, cá nhân kinh doanh"
        ET.SubElement(header_pl1, "maGiaoDich").text = ""
        ky_pl1 = ET.SubElement(header_pl1, "KyKKhaiThue")
        ET.SubElement(ky_pl1, "kieuKy").text = "Q"
        ET.SubElement(ky_pl1, "kyKKhai").text = ky_str
        ET.SubElement(ky_pl1, "kyKKhaiTuNgay").text = f"01/{start_month:02d}/{y}"
        ET.SubElement(ky_pl1, "kyKKhaiDenNgay").text = f"{last_day:02d}/{end_month:02d}/{y}"
        ET.SubElement(ky_pl1, "kyKKhaiTuThang").text = f"{start_month:02d}/{y}"
        ET.SubElement(ky_pl1, "kyKKhaiDenThang").text = f"{end_month:02d}/{y}"
        ET.SubElement(header_pl1, "soLan").text = data.get('soLan', '0')
        ET.SubElement(header_pl1, "mst").text = data.get('mst', '')
        ET.SubElement(header_pl1, "tenNNT").text = data.get('tenNNT', '')
        ET.SubElement(header_pl1, "mstDLyThue", {"xsi:nil": "true"})
        ET.SubElement(header_pl1, "tenDLyThue").text = ""
        ET.SubElement(header_pl1, "soHDongDLyThue").text = ""
        ET.SubElement(header_pl1, "ngayKyHDDLyThue", {"xsi:nil": "true"})
        ET.SubElement(pl01_1_khbs, "ma_DonViTien").text = "VND"
        ET.SubElement(pl01_1_khbs, "ten_DonViTien").text = "Đồng Việt Nam"
        muc_a1 = ET.SubElement(pl01_1_khbs, "Muc_A")
        dsach_hso = ET.SubElement(muc_a1, "DSachHSo")
        bke_hso = ET.SubElement(dsach_hso, "BKeHSo", {"id": "ID_1"})
        ET.SubElement(bke_hso, "ma_HSo").text = "010502"
        ET.SubElement(bke_hso, "ten_HSo").text = "01-2/BK-HĐKD"
        ctiet_hso = ET.SubElement(bke_hso, "CTietHSo")
        id_k = 1
        for key in ct_keys:
            old_val = old_data.get(key, 0)
            new_val = float(data.get(key, '0'))
            chenh = new_val - old_val
            chitiet = ET.SubElement(ctiet_hso, "ChiTiet", {"id": f"ID_{id_k}"})
            ET.SubElement(chitiet, "ct2_ma").text = ""
            ET.SubElement(chitiet, "ct2_ten").text = ""
            ma_label, ten_label = labels.get(key, (key.upper(), key.upper()))
            ET.SubElement(chitiet, "ct3_ma").text = ma_label
            ET.SubElement(chitiet, "ct3_ten").text = ten_label
            ET.SubElement(chitiet, "ct3_1_ma").text = ""
            ET.SubElement(chitiet, "ct3_1_ten").text = ""
            ET.SubElement(chitiet, "ct3_2_ma").text = ""
            ET.SubElement(chitiet, "ct3_2_ten").text = ""
            ET.SubElement(chitiet, "ct04").text = str(round(old_val))
            ET.SubElement(chitiet, "ct05").text = str(round(new_val))
            ET.SubElement(chitiet, "ct06").text = str(round(chenh))
            ET.SubElement(chitiet, "ct7").text = "0"
            ET.SubElement(chitiet, "ct8").text = data.get(f"ghiChu_{key}", "")
            id_k += 1
        tong_cong = ET.SubElement(dsach_hso, "TongCong")
        ET.SubElement(tong_cong, "tongCong_7").text = "0"
        ET.SubElement(tong_cong, "tongCong_8").text = "0"
        ET.SubElement(tong_cong, "tongCong_9").text = "0"
        muc_b1 = ET.SubElement(pl01_1_khbs, "Muc_B")
        ctiet_tl = ET.SubElement(muc_b1, "CTietTaiLieu", {"id": "ID_1"})
        ET.SubElement(ctiet_tl, "ma_TLieu").text = ""
        ET.SubElement(ctiet_tl, "ten_TLieu").text = ""
    # Ký số giả lập (thay bằng ký thật khi triển khai)
    signature = ET.SubElement(root, "Signature", {"xmlns": "http://www.w3.org/2000/09/xmldsig#"})
    ET.SubElement(signature, "SignedInfo")
    ET.SubElement(signature, "SignatureValue").text = "GIẢ LẬP CHỮ KÝ SỐ - THAY BẰNG KÝ THẬT"
    return prettify(root)

@app.route('/api/sign_xml', methods=['POST'])
@login_required
def sign_xml():
    try:
        data = request.form.to_dict()
        xml_content = generate_tax_xml(data)

        buffer = BytesIO(xml_content.encode('utf-8'))
        buffer.seek(0)

        return send_file(
            buffer,
            mimetype='text/xml',
            as_attachment=True,
            download_name=f'ToKhai_01_CNKD_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xml'
        )
    except Exception as e:
        print(f"Lỗi tạo XML: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

# === HÀM TÍNH DOANH THU THEO QUÝ === PHẦN HÀM DƯ THỪA XEM XÉT LOẠI BỎ.
def get_quarter_data(quarter_year):
    conn = get_db_connection()
    c = conn.cursor()
    quarter, year = map(int, quarter_year.split('/'))
    start_month = (quarter - 1) * 3 + 1
    end_month = quarter * 3
    start = f"{year}-{start_month:02d}-01"
    # tính ngày cuối cùng đúng của end_month
    last_day = calendar.monthrange(year, end_month)[1]
    end = f"{year}-{end_month:02d}-{last_day:02d}"
    c.execute("SELECT SUM(total) FROM sale WHERE date BETWEEN ? AND ?", (start, end))
    total = c.fetchone()[0] or 0
    tax_gtgt = total * 0.1
    c.execute("SELECT id, name, unit FROM products")
    products = []
    for row in c.fetchall():
        products.append({
            'id': row['id'],
            'name': row['name'],
            'unit': row['unit'],
            'begin_qty': 0, 'begin_value': 0,
            'import_qty': 0, 'import_value': 0,
            'sale_qty': 0, 'sale_value': 0,
            'end_qty': 0, 'end_value': 0
        })
    conn.close()
    return int(total), int(tax_gtgt), products

# === GỬI eTax API (ví dụ) ===
def submit_to_etax(signed_xml_path, mst):
    url = "https://api.etax.gdt.gov.vn/submit"
    try:
        with open(signed_xml_path, 'rb') as f:
            files = {'file': f}
            data = {'mst': mst, 'loaiTK': '01/CNKD'}
            response = requests.post(url, files=files, data=data, timeout=30)
        if response.status_code == 200:
            return {"status": "success", "data": response.json()}
        else:
            return {"status": "error", "message": response.text}
    except Exception as e:
        return {"status": "error", "message": str(e)}

#=============================================================================== End of Báo Thuế=============================================================================#

#======================================================================= Start of Báo Doanh Thu & Lợi Nhuận===================================================================#
def get_days_in_quarter(year, month):
    """Xác định quý, số ngày thực tế của quý và ngày bắt đầu/kết thúc quý đó."""
    quarter = (month - 1) // 3 + 1
    start_month = (quarter - 1) * 3 + 1
    end_month = start_month + 2
    
    q_start = datetime(year, start_month, 1)
    # Lấy ngày cuối cùng của tháng cuối quý
    last_day_of_q = calendar.monthrange(year, end_month)[1]
    q_end = datetime(year, end_month, last_day_of_q)
    
    total_days = (q_end - q_start).days + 1
    return total_days, q_start, q_end

@app.route('/api/reports/profit', methods=['GET'])
@login_required
def api_profit_report():
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    from_date_iso = request.args.get('from') # yyyy-mm-dd
    to_date_iso = request.args.get('to')     # yyyy-mm-dd
    
    if not from_date_iso or not to_date_iso:
        return jsonify({"error": "Thiếu thông tin ngày"}), 400
    
    try:
        start_dt = datetime.strptime(from_date_iso, '%Y-%m-%d')
        end_dt = datetime.strptime(to_date_iso, '%Y-%m-%d')
        
        # Khoảng tìm kiếm chuẩn cho Doanh thu & Giá vốn
        start_search = f"{from_date_iso} 00:00:00"
        end_search = f"{to_date_iso} 23:59:59"

        # --- LOGIC LÙI KỲ CHO CHI PHÍ VẬN HÀNH ---
        # 1. Lùi 1 Quý cho Thuế (Chi ở tương lai tính cho kỳ này)
        q_start_p = (start_dt + timedelta(days=60)).strftime('%Y-%m-%d') + " 00:00:00"
        q_end_p = (end_dt + timedelta(days=120)).strftime('%Y-%m-%d') + " 23:59:59"

        # 2. Lùi 1 Tháng cho Điện, Nước... (Chi ở tháng sau tính cho tháng này)
        m_start_p = (start_dt + timedelta(days=20)).strftime('%Y-%m-%d') + " 00:00:00"
        m_end_p = (end_dt + timedelta(days=45)).strftime('%Y-%m-%d') + " 23:59:59"

    except ValueError:
        return jsonify({"error": "Định dạng ngày không hợp lệ"}), 400

    try:
        # --- 1. DOANH THU ---
        c.execute("""
            SELECT COALESCE(SUM(total_amount), 0) FROM sale 
            WHERE status='completed' AND date BETWEEN ? AND ?
        """, (start_search, end_search))
        total_revenue = float(c.fetchone()[0])

        # --- 2. GIÁ VỐN HÀNG BÁN ---
        c.execute("""
            SELECT COALESCE(SUM(CASE 
                WHEN type IN ('SALE', 'export_material') THEN ABS(quantity) * cost_price 
                WHEN type = 'RETURN_SALE' THEN -quantity * cost_price 
                ELSE 0 END), 0) FROM stock_moves WHERE date BETWEEN ? AND ?
        """, (start_search, end_search))
        total_cogs = float(c.fetchone()[0])

        # --- 3. CHI PHÍ NHÂN CÔNG (Fix lỗi logic vòng lặp) ---
        months_years = []
        curr = datetime(start_dt.year, start_dt.month, 1)
        while curr <= datetime(end_dt.year, end_dt.month, 1):
            months_years.append((curr.month, curr.year))
            if curr.month == 12:
                curr = datetime(curr.year + 1, 1, 1)
            else:
                curr = datetime(curr.year, curr.month + 1, 1)

        cost_labor = 0.0
        if months_years:
            conditions = " OR ".join(["(month = ? AND year = ?)" for _ in months_years])
            params = [val for pair in months_years for val in pair]
            c.execute(f"SELECT COALESCE(SUM(salary_rate), 0) FROM Salary_Detail WHERE {conditions}", params)
            cost_labor = float(c.fetchone()[0])

        # --- 4. CHI PHÍ VẬN HÀNH (LOGIC LÙI KỲ) ---
        # 4.1. Nhóm Điện, Nước, VT, Lãi vay, Mặt bằng (Lùi 1 tháng)
        c.execute("""
            SELECT COALESCE(SUM(amount), 0) FROM phieu_chi 
            WHERE expense_type IN ('CP_DIEN', 'CP_NUOC', 'CP_VT', 'CP_MB', 'CP_TRALAIVAY')
            AND date BETWEEN ? AND ?
            AND (source_type IS NULL OR source_type != 'salary')
        """, (m_start_p, m_end_p))
        cost_utilities = float(c.fetchone()[0])

        # 4.2. Nhóm Thuế (Lùi 1 quý)
        c.execute("""
            SELECT COALESCE(SUM(amount), 0) FROM phieu_chi 
            WHERE expense_type = 'CP_THUE'
            AND date BETWEEN ? AND ?
            AND (source_type IS NULL OR source_type != 'salary')
        """, (q_start_p, q_end_p))
        cost_tax = float(c.fetchone()[0])

        # 4.3. Nhóm chi phí tức thì (VPP, Khác)
        c.execute("""
            SELECT COALESCE(SUM(amount), 0) FROM phieu_chi 
            WHERE expense_type IN ('CP_VPP', 'CP_KHAC')
            AND date BETWEEN ? AND ?
            AND (source_type IS NULL OR source_type != 'salary')
        """, (start_search, end_search))
        cost_immediate = float(c.fetchone()[0])
        cost_others = cost_tax + cost_immediate

        # --- 5. KHẤU HAO (GIỮ NGUYÊN LOGIC GỐC CỦA BẠN) ---
        cost_depreciation = 0.0
        c.execute("SELECT nguyen_gia_tinh_khau_hao, so_thang_khau_hao, ngay_bat_dau_su_dung FROM tai_san_co_dinh WHERE tinh_trang = 'Active'")
        assets = c.fetchall()
        
        days_in_q, q_start, q_end = get_days_in_quarter(start_dt.year, start_dt.month)

        for asset in assets:
            try:
                ng_gia = float(asset['nguyen_gia_tinh_khau_hao'] or 0)
                s_thang = int(asset['so_thang_khau_hao'] or 1)
                if ng_gia <= 0: continue

                raw_date = asset['ngay_bat_dau_su_dung']
                if isinstance(raw_date, str):
                    asset_start = datetime.strptime(raw_date.split(' ')[0], '%Y-%m-%d')
                elif hasattr(raw_date, 'year'):
                    asset_start = datetime(raw_date.year, raw_date.month, raw_date.day)
                else: continue

                # Tính ngày kết thúc (Ép kiểu int cho days để tránh lỗi trên VPS)
                asset_end = asset_start + timedelta(days=int(s_thang * 30.44))
                
                overlap_start = max(start_dt, asset_start)
                overlap_end = min(end_dt, asset_end)
                
                if overlap_start <= overlap_end:
                    days_selected = (overlap_end - overlap_start).days + 1
                    dep_per_month = ng_gia / s_thang
                    dep_per_quarter = dep_per_month * 3
                    cost_depreciation += (dep_per_quarter / days_in_q) * days_selected
            except:
                continue

        # --- 6. TỔNG HỢP ---
        total_op_exp = cost_labor + cost_depreciation + cost_utilities + cost_others
        gross_profit = total_revenue - total_cogs
        net_profit = gross_profit - total_op_exp

        return jsonify({
            "status": "success",
            "revenue": round(total_revenue),
            "cogs": round(total_cogs),
            "gross_profit": round(gross_profit),
            "operating_expenses": {
                "labor": round(cost_labor),
                "depreciation": round(cost_depreciation),
                "utilities": round(cost_utilities),
                "others": round(cost_others),
                "total": round(total_op_exp)
            },
            "net_profit": round(net_profit)
        })

    except Exception as e:
        print(f"Profit Report Error: {str(e)}")
        return jsonify({"error": "Lỗi hạch toán hệ thống"}), 500
    finally:
        conn.close()

# ⚡️ ENDPOINT BÁO CÁO DOANH THU ĐÃ CẬP NHẬT ⚡️
# ----------------------------------------------------------------------
@app.route('/api/reports/sale', methods=['GET'])
def get_sale_report():
    start_iso = request.args.get('start') # "2026-01-01"
    end_iso = request.args.get('end')     # "2026-01-30"

    # Mở rộng phạm vi giờ để bao quát trọn vẹn ngày được chọn
    # Ngày bắt đầu từ 0 giờ 0 phút 0 giây
    start_query = f"{start_iso} 00:00:00"
    # Ngày kết thúc đến 23 giờ 59 phút 59 giây
    end_query = f"{end_iso} 23:59:59"

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    # TRUY VẤN 1: DOANH THU HÀNG NGÀY
    # Sử dụng DATE(date) để nhóm tất cả các giờ trong ngày 30/01 vào 1 dòng duy nhất
    cursor.execute("""
        SELECT 
            DATE(date) as day, 
            SUM(total_amount) as revenue, 
            COUNT(id) as bills 
        FROM sale 
        WHERE date BETWEEN ? AND ? AND status = 'completed'
        GROUP BY DATE(date)
        ORDER BY day ASC
    """, (start_query, end_query))
    
    sale_data = [dict(row) for row in cursor.fetchall()]

    # TRUY VẤN 2: TOP SẢN PHẨM (Fix lỗi SUM nhầm total_amount của hóa đơn)
    cursor.execute("""
        SELECT 
            p.name, 
            SUM(si.quantity) as qty,
            SUM(si.quantity * si.price) as total -- Tính tiền từng dòng item
        FROM sale_items si
        JOIN sale s ON s.id = si.sale_id
        JOIN products p ON p.id = si.product_id
        WHERE s.date BETWEEN ? AND ? AND s.status = 'completed'
        GROUP BY p.id
        ORDER BY total DESC
        LIMIT 10
    """, (start_query, end_query))
    
    top_products = [dict(row) for row in cursor.fetchall()]

    conn.close()
    return jsonify({"sale": sale_data, "top_products": top_products})

@app.route('/reports/sale')
def sale_report_page():
    # Giả định tên file HTML là 'sale_report.html'
    return render_template('reports.html')

#======================================================================= End of Báo Cáo Doanh Thu & Lợi Nhuận===================================================================#


#===Hàm Tự Tắt Người Dùng khi Hết Hợp Đồng===#
def check_tenant_expirations():
    from datetime import datetime
    today = datetime.now().strftime('%Y-%m-%d')
    conn = get_db_connection()
    c = conn.cursor()
    # Tìm các tenant có expiry_date nhỏ hơn hôm nay và đang active
    c.execute("""
        UPDATE tenants 
        SET is_active = 0 
        WHERE expiry_date < ? AND is_active = 1
    """, (today,))
    conn.commit()
    conn.close()
    print(f"--- Đã kiểm tra và khóa các Tenant hết hạn ngày {today} ---")

#=== HÀM CHẠY LỊCH ĐỂ TỰ TẮT NGƯỜI DÙNG HẾT HẠN===#
from flask_apscheduler import APScheduler
scheduler = APScheduler()

@scheduler.task('cron', id='do_check_expiry', hour=0, minute=1)
def scheduled_task():
    with app.app_context():
        check_tenant_expirations()

scheduler.init_app(app)
scheduler.start()

#======================================================================= KẾT THÚC PHẦN MASTER SETTINGS ===============================================================================#


def backup_database():
    """Quét Registry và backup cho tất cả Tenant + Main DB"""
    try:
        # 1. Đảm bảo các thư mục gốc tồn tại
        if not os.path.exists(BACKUP_ROOT):
            os.makedirs(BACKUP_ROOT, exist_ok=True)

        # Danh sách công việc: (tenant_name, path_vật_lý)
        tasks = [('main', MAIN_DB_PATH)] 
        
        # 2. Lấy danh sách Tenant từ Registry
        try:
            conn_main = sqlite3.connect(MAIN_DB_PATH)
            # Chỉ lấy các shop đang hoạt động
            tenants = conn_main.execute("SELECT tenant_id, db_path FROM tenants WHERE is_active=1").fetchall()
            conn_main.close()
        except Exception as db_err:
            print(f"Lỗi truy cập Registry: {db_err}")
            return # Dừng nếu không đọc được file Main

        for t_id, t_path in tenants:
            if not t_id or not t_path: continue
            
            # Làm sạch ID để đặt tên thư mục an toàn
            t_id = str(t_id).strip()
            # Chuyển sang đường dẫn tuyệt đối
            abs_path = t_path if os.path.isabs(t_path) else os.path.join(BASE_DIR, t_path)
            tasks.append((t_id, abs_path))

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # 3. Vòng lặp sao lưu
        for tenant_id, db_path in tasks:
            try:
                if not os.path.exists(db_path):
                    print(f"Bỏ qua {tenant_id}: File không tồn tại tại {db_path}")
                    continue
                
                # Tạo thư mục con cho từng tenant
                tenant_backup_dir = os.path.join(BACKUP_ROOT, tenant_id)
                os.makedirs(tenant_backup_dir, exist_ok=True)

                filename = f"{tenant_id}_auto_{timestamp}.db"
                dest = os.path.join(tenant_backup_dir, filename)

                # Sử dụng phương pháp copy an toàn cho SQLite (Tránh lỗi file đang mở)
                # Nếu VPS của bạn dùng SQLite phiên bản mới, lệnh này là tốt nhất:
                # conn = sqlite3.connect(db_path)
                # conn.execute(f"VACUUM INTO '{dest}'")
                # conn.close()
                
                # Nếu không, dùng shutil.copy2 truyền thống nhưng cần try-catch kỹ:
                shutil.copy2(db_path, dest)

                # 4. Dọn dẹp bản cũ (Giữ 10 ngày)
                cutoff = (datetime.now() - timedelta(days=10)).timestamp()
                for f in os.listdir(tenant_backup_dir):
                    f_path = os.path.join(tenant_backup_dir, f)
                    # Chỉ xóa các file .db cũ, tránh xóa nhầm thư mục
                    if os.path.isfile(f_path) and f.endswith('.db'):
                        if os.path.getctime(f_path) < cutoff:
                            os.remove(f_path)
                
                print(f"[{datetime.now()}] Backup OK: {tenant_id}")

            except Exception as e:
                print(f"Lỗi khi backup tenant {tenant_id}: {e}")
                continue # Lỗi shop này vẫn tiếp tục shop sau

    except Exception as e:
        print(f"Lỗi hệ thống Backup (Tổng quát): {e}")

# --- KHỞI TẠO LẬP LỊCH (SCHEDULER) ---
# BackgroundScheduler chạy trong một thread riêng, không block Flask
scheduler = BackgroundScheduler()
scheduler.add_job(func=backup_database, trigger="cron", hour=20, minute=0)
scheduler.start()





#============================================================================ End of Tax and Accounting=========================================================================#


if __name__ == '__main__':
    from apscheduler.schedulers.background import BackgroundScheduler

    scheduler = BackgroundScheduler()
    scheduler.add_job(func=backup_database, trigger="cron", hour=20, minute=0)
    scheduler.start()

    print("POS System & Scheduler đã sẵn sàng.")
    print("Server running: http://127.0.0.1:5000")

    app.run(
        host='0.0.0.0',
        port=5000,
        debug=True,
        threaded=True
    )