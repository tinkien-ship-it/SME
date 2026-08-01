"""Xuất bộ BCTC SME (B01/B02/B03/B09) ra Excel."""
from __future__ import annotations

import io
import sqlite3
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

from Services.sme.b09_notes import notes_to_financial_statements
from Services.sme.bctc_report import balance_sheet, cash_flow_statement, income_statement


def _money(v) -> float:
    try:
        return round(float(v or 0), 0)
    except (TypeError, ValueError):
        return 0.0


def _write_rows(ws, rows: list[dict], amount_header: str = 'Số tiền') -> None:
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    header_font = Font(bold=True)
    ws.append(['Mã', 'Chỉ tiêu', amount_header])
    for cell in ws[1]:
        cell.font = header_font
        cell.border = thin
    for r in rows or []:
        kind = r.get('kind')
        code = '' if kind == 'header' else (r.get('code') or '')
        name = r.get('name') or ''
        amt = None if kind == 'header' else _money(r.get('amount'))
        ws.append([code, name, amt if amt is not None else ''])
        row = ws.max_row
        for col in range(1, 4):
            ws.cell(row, col).border = thin
        if r.get('bold') or kind == 'header':
            ws.cell(row, 2).font = Font(bold=True)
        ws.cell(row, 3).alignment = Alignment(horizontal='right')
        ws.cell(row, 3).number_format = '#,##0'
    for i, width in enumerate((12, 55, 18), 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def export_bctc_workbook(
    conn: sqlite3.Connection,
    *,
    fiscal_year: int,
    period_from: int = 1,
    period_to: int = 12,
    include_current_profit: bool = True,
) -> bytes:
    """Trả về bytes file .xlsx gồm 4 sheet B01/B02/B03/B09."""
    year = int(fiscal_year)
    p_from = max(1, min(12, int(period_from)))
    p_to = max(p_from, min(12, int(period_to)))

    b01 = balance_sheet(
        conn, fiscal_year=year, period_to=p_to,
        include_current_profit=include_current_profit,
    )
    b02 = income_statement(
        conn, fiscal_year=year, period_from=p_from, period_to=p_to,
    )
    b03 = cash_flow_statement(
        conn, fiscal_year=year, period_from=p_from, period_to=p_to,
    )
    try:
        b09 = notes_to_financial_statements(conn, fiscal_year=year, period_to=p_to)
    except Exception:
        b09 = {'narrative': {}, 'supplementary': {}, 'summary': {}}

    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'B01-CDKT'
    ws1['A1'] = f'Bảng cân đối kế toán B01-DN — Năm {year} đến T{p_to}'
    ws1['A1'].font = Font(bold=True, size=12)
    ws1.append([])
    _write_rows(ws1, b01.get('rows') or [], 'Số cuối kỳ')
    t = b01.get('totals') or {}
    ws1.append([])
    ws1.append(['', 'Tài sản', _money(t.get('total_assets'))])
    ws1.append(['', 'Nguồn vốn', _money(t.get('total_equity_and_liabilities'))])
    ws1.append(['', 'Cân đối', 'OK' if t.get('balanced') else f"Lệch {_money(t.get('difference'))}"])

    ws2 = wb.create_sheet('B02-KQKD')
    ws2['A1'] = f'Kết quả HĐKD B02-DN — T{p_from}–T{p_to}/{year}'
    ws2['A1'].font = Font(bold=True, size=12)
    ws2.append([])
    _write_rows(ws2, b02.get('rows') or [], 'Số kỳ này')

    ws3 = wb.create_sheet('B03-LCTT')
    ws3['A1'] = f'Lưu chuyển tiền tệ B03-DN — T{p_from}–T{p_to}/{year}'
    ws3['A1'].font = Font(bold=True, size=12)
    ws3.append([])
    _write_rows(ws3, b03.get('rows') or [], 'Số kỳ này')

    ws4 = wb.create_sheet('B09-TM')
    ws4['A1'] = f'Thuyết minh B09-DN — Năm {year} đến T{p_to}'
    ws4['A1'].font = Font(bold=True, size=12)
    ws4.append(['Mục', 'Mã', 'Nội dung / Chỉ tiêu', 'Giá trị'])
    for cell in ws4[2]:
        cell.font = Font(bold=True)
    narrative = b09.get('narrative') or {}
    for key in ('I', 'II', 'III', 'IV'):
        sec = narrative.get(key) or {}
        for it in sec.get('items') or []:
            ws4.append([key, it.get('code'), it.get('label'), it.get('value')])
    supp = b09.get('supplementary') or {}
    for key in ('V', 'VI'):
        sec = supp.get(key) or {}
        for note in sec.get('notes') or []:
            for r in note.get('rows') or []:
                val = r.get('amount')
                if val is None:
                    val = f"ĐN {_money(r.get('opening'))} / CK {_money(r.get('closing'))}"
                else:
                    val = _money(val)
                ws4.append([
                    key, note.get('code'),
                    f"{r.get('account_code')} — {r.get('name')}",
                    val,
                ])
    for i, width in enumerate((8, 12, 55, 40), 1):
        ws4.column_dimensions[get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_meta(
    *,
    fiscal_year: int,
    period_from: int,
    period_to: int,
) -> dict[str, Any]:
    return {
        'filename': f'BCTC_SME_{fiscal_year}_T{period_from}-{period_to}.xlsx',
        'fiscal_year': fiscal_year,
        'period_from': period_from,
        'period_to': period_to,
    }
