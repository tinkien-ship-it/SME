"""Số đầu năm BCTC — nhập tay trên báo cáo hoặc import Excel (phần mềm khác).

Hai nguồn:
- Chỉ tiêu B01/B02 (mã số) → lưu overlay, cột Số đầu năm / Năm trước.
- Số dư tài khoản (Nợ/Có) → ghi bút toán DK ngày 31/12 năm trước (vào sổ + BCTC).
"""
from __future__ import annotations

import io
import re
import sqlite3
import unicodedata
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

MONEY_Q = Decimal('0.01')
DK_DOC_TYPE = 'DK'
DK_BUSINESS = 'SO_DU_DAU_KY'


def _money(val) -> Decimal:
    if val is None or val == '':
        return Decimal('0.00')
    if isinstance(val, Decimal):
        return val.quantize(MONEY_Q, rounding=ROUND_HALF_UP)
    if isinstance(val, (int, float)):
        return Decimal(str(val)).quantize(MONEY_Q, rounding=ROUND_HALF_UP)
    s = str(val).strip().replace('\xa0', '').replace(' ', '')
    if not s or s in ('-', '—'):
        return Decimal('0.00')
    s = s.replace('₫', '').replace('đ', '').replace('VND', '').replace('vnd', '')
    if s.count(',') == 1 and s.count('.') >= 1:
        if s.rfind(',') > s.rfind('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    elif s.count('.') > 1:
        s = s.replace('.', '')
    elif s.count(',') > 1:
        s = s.replace(',', '')
    elif s.count(',') == 1 and s.count('.') == 0:
        left, right = s.split(',')
        if len(right) == 3 and left.replace('-', '').isdigit():
            s = s.replace(',', '')
        else:
            s = s.replace(',', '.')
    try:
        return Decimal(s).quantize(MONEY_Q, rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return Decimal('0.00')


def _now() -> str:
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')


def ensure_bctc_opening_schema(conn: sqlite3.Connection, *, commit: bool = False) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS sme_bctc_opening (
            fiscal_year INTEGER NOT NULL,
            report TEXT NOT NULL,
            line_code TEXT NOT NULL,
            amount REAL NOT NULL DEFAULT 0,
            source TEXT,
            updated_at TEXT,
            updated_by TEXT,
            PRIMARY KEY (fiscal_year, report, line_code)
        )
        """
    )
    if commit:
        conn.commit()


def _line_defs(conn: sqlite3.Connection, report: str) -> list[dict]:
    tt58 = False
    try:
        from Services.sme.regime_profile import get_ledger_profile
        tt58 = bool(get_ledger_profile(conn).get('is_tt58_micro'))
    except Exception:
        tt58 = False
    key = (report or 'B01').strip().upper()
    if tt58:
        from Services.sme.bctc_lines_tt58 import (
            B01_DNSN_BALANCE_SHEET,
            B02_DNSN_INCOME_STATEMENT,
        )
        return B02_DNSN_INCOME_STATEMENT if key == 'B02' else B01_DNSN_BALANCE_SHEET
    from Services.sme.bctc_lines import B01_BALANCE_SHEET, B02_INCOME_STATEMENT
    return B02_INCOME_STATEMENT if key == 'B02' else B01_BALANCE_SHEET


def leaf_codes(conn: sqlite3.Connection, report: str) -> set[str]:
    return {str(x['code']) for x in _line_defs(conn, report) if x.get('kind') == 'leaf'}


def list_opening_lines(
    conn: sqlite3.Connection,
    fiscal_year: int,
    report: str = 'B01',
) -> dict[str, Decimal]:
    ensure_bctc_opening_schema(conn)
    rows = conn.execute(
        """
        SELECT line_code, amount FROM sme_bctc_opening
        WHERE fiscal_year = ? AND report = ?
        """,
        (int(fiscal_year), (report or 'B01').strip().upper()),
    ).fetchall()
    out: dict[str, Decimal] = {}
    for r in rows:
        code = r[0] if not isinstance(r, sqlite3.Row) else r['line_code']
        amt = r[1] if not isinstance(r, sqlite3.Row) else r['amount']
        out[str(code)] = _money(amt)
    return out


def opening_meta(
    conn: sqlite3.Connection,
    fiscal_year: int,
    report: str = 'B01',
) -> dict[str, Any]:
    ensure_bctc_opening_schema(conn)
    row = conn.execute(
        """
        SELECT source, updated_at, updated_by, COUNT(*) AS n
        FROM sme_bctc_opening
        WHERE fiscal_year = ? AND report = ?
        GROUP BY source, updated_at, updated_by
        ORDER BY updated_at DESC
        LIMIT 1
        """,
        (int(fiscal_year), (report or 'B01').strip().upper()),
    ).fetchone()
    n = conn.execute(
        """
        SELECT COUNT(*) FROM sme_bctc_opening
        WHERE fiscal_year = ? AND report = ?
        """,
        (int(fiscal_year), (report or 'B01').strip().upper()),
    ).fetchone()[0]
    if not n:
        return {'has_override': False, 'count': 0, 'source': None, 'updated_at': None}
    src = None
    ts = None
    by = None
    if row:
        src = row[0] if not isinstance(row, sqlite3.Row) else row['source']
        ts = row[1] if not isinstance(row, sqlite3.Row) else row['updated_at']
        by = row[2] if not isinstance(row, sqlite3.Row) else row['updated_by']
    return {
        'has_override': True,
        'count': int(n or 0),
        'source': src,
        'updated_at': ts,
        'updated_by': by,
    }


def save_opening_lines(
    conn: sqlite3.Connection,
    *,
    fiscal_year: int,
    report: str,
    lines: list[dict] | dict[str, Any],
    source: str = 'manual',
    updated_by: str | None = None,
    replace: bool = True,
    commit: bool = True,
) -> dict[str, Any]:
    ensure_bctc_opening_schema(conn)
    year = int(fiscal_year)
    rep = (report or 'B01').strip().upper()
    if rep not in ('B01', 'B02'):
        raise ValueError('Chỉ lưu số đầu năm cho B01 hoặc B02')
    allowed = leaf_codes(conn, rep)
    items: dict[str, Decimal] = {}
    if isinstance(lines, dict):
        iterable = [{'code': k, 'amount': v} for k, v in lines.items()]
    else:
        iterable = list(lines or [])
    skipped: list[str] = []
    for raw in iterable:
        code = str(raw.get('code') or raw.get('line_code') or '').strip()
        if not code:
            continue
        if code not in allowed:
            skipped.append(code)
            continue
        items[code] = _money(raw.get('amount'))
    now = _now()
    if replace:
        conn.execute(
            'DELETE FROM sme_bctc_opening WHERE fiscal_year = ? AND report = ?',
            (year, rep),
        )
    for code, amt in items.items():
        conn.execute(
            """
            INSERT INTO sme_bctc_opening(
                fiscal_year, report, line_code, amount, source, updated_at, updated_by
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(fiscal_year, report, line_code) DO UPDATE SET
                amount = excluded.amount,
                source = excluded.source,
                updated_at = excluded.updated_at,
                updated_by = excluded.updated_by
            """,
            (year, rep, code, float(amt), source, now, updated_by),
        )
    if commit:
        conn.commit()
    return {
        'saved': len(items),
        'skipped': skipped,
        'fiscal_year': year,
        'report': rep,
        'source': source,
    }


def clear_opening_lines(
    conn: sqlite3.Connection,
    *,
    fiscal_year: int,
    report: str | None = None,
    commit: bool = True,
) -> int:
    ensure_bctc_opening_schema(conn)
    if report:
        cur = conn.execute(
            'DELETE FROM sme_bctc_opening WHERE fiscal_year = ? AND report = ?',
            (int(fiscal_year), report.strip().upper()),
        )
    else:
        cur = conn.execute(
            'DELETE FROM sme_bctc_opening WHERE fiscal_year = ?',
            (int(fiscal_year),),
        )
    if commit:
        conn.commit()
    return int(cur.rowcount or 0)


def merge_opening_leaf(
    journal_open: dict[str, Decimal],
    stored: dict[str, Decimal] | None,
) -> dict[str, Decimal]:
    """Ưu tiên số nhập tay/Excel; chỉ tiêu không nhập thì lấy sổ năm trước."""
    out = dict(journal_open or {})
    for code, amt in (stored or {}).items():
        out[code] = _money(amt)
    return out


def ending_from_opening_and_ytd(
    opening_leaf: dict[str, Decimal],
    journal_open: dict[str, Decimal],
    journal_close: dict[str, Decimal],
    *,
    leaf_keys: set[str],
) -> dict[str, Decimal]:
    """Số cuối = số đầu (đã merge) + phát sinh năm nay (cuối sổ − đầu sổ)."""
    out: dict[str, Decimal] = {}
    for code in leaf_keys:
        ytd = _money(journal_close.get(code, 0)) - _money(journal_open.get(code, 0))
        out[code] = _money(opening_leaf.get(code, 0)) + ytd
    return out


# ---------------------------------------------------------------------------
# Excel template + parse
# ---------------------------------------------------------------------------

def _norm_header(value: Any) -> str:
    s = str(value or '').replace('đ', 'd').replace('Đ', 'D')
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return re.sub(r'[^a-z0-9]+', '', s.lower())


_LINE_KEYS = {
    'ma', 'machitieu', 'machiso', 'code', 'line', 'linecode', 'chisao',
    'maso', 'mact', 'indicator',
}
_OPEN_KEYS = {
    'sodaunam', 'sodauky', 'dauky', 'opening', 'amountopening', 'columnd',
    'sotien', 'amount', 'giatri',
}
_PRIOR_KEYS = {
    'namtruoc', 'kytruoc', 'prioryear', 'amountprior', 'comparatives',
}
_ACCT_KEYS = {
    'matk', 'taikhoan', 'tk', 'account', 'accountcode', 'sotk', 'matkkt',
}
_DEBIT_KEYS = {'no', 'debit', 'duno', 'psno', 'openingdebit', 'nodauky'}
_CREDIT_KEYS = {'co', 'credit', 'duco', 'psco', 'openingcredit', 'codauky'}
_NAME_KEYS = {'chitieu', 'tentk', 'ten', 'name', 'diengiai', 'noidung'}


def _header_role(norm: str) -> str | None:
    if norm in _LINE_KEYS or norm.startswith('ma') and 'tk' not in norm:
        if norm in _ACCT_KEYS:
            return 'account'
        if 'tk' in norm:
            return 'account'
        return 'line'
    if norm in _ACCT_KEYS:
        return 'account'
    if norm in _DEBIT_KEYS:
        return 'debit'
    if norm in _CREDIT_KEYS:
        return 'credit'
    if norm in _PRIOR_KEYS:
        return 'prior'
    if norm in _OPEN_KEYS:
        return 'opening'
    if norm in _NAME_KEYS:
        return 'name'
    return None


def _detect_header_row(rows: list[tuple]) -> tuple[int, dict[str, int]] | None:
    for idx, row in enumerate(rows[:15]):
        mapping: dict[str, int] = {}
        for col, cell in enumerate(row):
            role = _header_role(_norm_header(cell))
            if role and role not in mapping:
                mapping[role] = col
        if 'line' in mapping or 'account' in mapping:
            if any(k in mapping for k in ('opening', 'prior', 'debit', 'credit')):
                return idx, mapping
    return None


def parse_opening_excel(file_obj) -> dict[str, Any]:
    """Đọc workbook: chỉ tiêu B01/B02 và/hoặc số dư TK."""
    wb = load_workbook(file_obj, data_only=True, read_only=True)
    b01: dict[str, Decimal] = {}
    b02: dict[str, Decimal] = {}
    trial: list[dict] = []
    errors: list[str] = []
    try:
        for ws in wb.worksheets:
            title = (ws.title or '').lower()
            rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
            if not rows:
                continue
            detected = _detect_header_row(rows)
            if not detected:
                continue
            header_idx, mapping = detected
            is_b02_sheet = 'b02' in title or 'kqkd' in title or 'kqhdkd' in title
            for ridx, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
                if not row or all(c is None or str(c).strip() == '' for c in row):
                    continue
                if 'account' in mapping and (
                    'debit' in mapping or 'credit' in mapping
                ):
                    acct = str(row[mapping['account']] or '').strip()
                    if not acct or _header_role(_norm_header(acct)):
                        continue
                    debit = _money(row[mapping['debit']]) if 'debit' in mapping else Decimal('0')
                    credit = _money(row[mapping['credit']]) if 'credit' in mapping else Decimal('0')
                    if debit == 0 and credit == 0:
                        continue
                    if debit > 0 and credit > 0:
                        errors.append(f'{ws.title} dòng {ridx}: TK {acct} vừa Nợ vừa Có')
                        continue
                    trial.append({
                        'account_code': acct,
                        'debit': float(debit),
                        'credit': float(credit),
                    })
                    continue
                code_col = mapping.get('line')
                if code_col is None:
                    continue
                code = str(row[code_col] or '').strip()
                if not code or not re.match(r'^[0-9A-Za-z]+$', code):
                    continue
                if is_b02_sheet or 'prior' in mapping:
                    amt = _money(row[mapping['prior']]) if 'prior' in mapping else _money(
                        row[mapping['opening']] if 'opening' in mapping else 0
                    )
                    if amt != 0:
                        b02[code] = amt
                else:
                    amt = _money(row[mapping['opening']]) if 'opening' in mapping else Decimal('0')
                    if amt != 0 or code in b01:
                        b01[code] = amt
    finally:
        wb.close()
    return {
        'b01': {k: float(v) for k, v in b01.items()},
        'b02': {k: float(v) for k, v in b02.items()},
        'trial_balance': trial,
        'errors': errors,
    }


def build_opening_template(conn: sqlite3.Connection) -> bytes:
    """Mẫu Excel: B01 số đầu năm, B02 năm trước, số dư TK."""
    from Services.sme.bctc_report import _coa_line_map

    b01_defs = _line_defs(conn, 'B01')
    b02_defs = _line_defs(conn, 'B02')
    wb = Workbook()
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    header_fill = PatternFill('solid', fgColor='1E3A8A')
    header_font = Font(bold=True, color='FFFFFF')
    leaf_fill = PatternFill('solid', fgColor='FFFBEB')

    def style_header(ws, cols: int) -> None:
        for col in range(1, cols + 1):
            cell = ws.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin

    ws0 = wb.active
    ws0.title = 'Huong_dan'
    ws0['A1'] = 'Nhập số đầu năm BCTC từ phần mềm khác'
    ws0['A1'].font = Font(bold=True, size=14)
    ws0['A3'] = (
        '1. Sheet B01_So_dau_nam: điền cột Số đầu năm theo mã chỉ tiêu (chỉ dòng màu vàng — chỉ tiêu gốc). '
        'Không sửa cột Mã. Dòng cộng (Tổng TS, Nguồn vốn…) hệ thống tự tính.'
    )
    ws0['A4'] = (
        '2. Sheet B02_Nam_truoc: điền số liệu năm trước (cột so sánh B02).'
    )
    ws0['A5'] = (
        '3. Sheet So_du_tai_khoan: nếu phần mềm khác xuất số dư TK (Nợ/Có), điền vào đây. '
        'Hệ thống ghi bút toán đầu kỳ (DK) ngày 31/12 năm trước — số dư vào sổ cái và BCTC.'
        ' Nợ phải bằng Có.'
    )
    ws0['A6'] = (
        '4. Có thể nhập file xuất sẵn (MISA/Fast/Bravo…): miễn có cột Mã chỉ tiêu + Số đầu năm, '
        'hoặc Mã TK + Nợ + Có. Header tiếng Việt hoặc tiếng Anh đều được.'
    )
    ws0['A7'] = '5. Tải file này, điền số, rồi dùng nút «Nhập Excel» trên trang Báo cáo tài chính.'
    ws0.column_dimensions['A'].width = 120
    ws0.row_dimensions[3].height = 36
    ws0.row_dimensions[5].height = 36

    ws1 = wb.create_sheet('B01_So_dau_nam')
    ws1.append(['Mã chỉ tiêu', 'Chỉ tiêu', 'Số đầu năm'])
    style_header(ws1, 3)
    for line in b01_defs:
        if line.get('kind') == 'header':
            ws1.append(['', line.get('name') or '', ''])
            ws1.cell(ws1.max_row, 2).font = Font(bold=True)
        elif line.get('kind') == 'leaf':
            ws1.append([line['code'], line.get('name') or '', 0])
            for col in range(1, 4):
                ws1.cell(ws1.max_row, col).fill = leaf_fill
                ws1.cell(ws1.max_row, col).border = thin
            ws1.cell(ws1.max_row, 3).number_format = '#,##0'
        else:
            ws1.append([line['code'], line.get('name') or '', ''])
            ws1.cell(ws1.max_row, 2).font = Font(bold=True, italic=True)
            ws1.cell(ws1.max_row, 3).fill = PatternFill('solid', fgColor='F1F5F9')
        for col in range(1, 4):
            ws1.cell(ws1.max_row, col).border = thin
    for i, w in enumerate((16, 70, 18), 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet('B02_Nam_truoc')
    ws2.append(['Mã chỉ tiêu', 'Chỉ tiêu', 'Năm trước'])
    style_header(ws2, 3)
    for line in b02_defs:
        if line.get('kind') == 'header':
            ws2.append(['', line.get('name') or '', ''])
            ws2.cell(ws2.max_row, 2).font = Font(bold=True)
        elif line.get('kind') == 'leaf':
            ws2.append([line['code'], line.get('name') or '', 0])
            for col in range(1, 4):
                ws2.cell(ws2.max_row, col).fill = leaf_fill
                ws2.cell(ws2.max_row, col).border = thin
            ws2.cell(ws2.max_row, 3).number_format = '#,##0'
        else:
            ws2.append([line['code'], line.get('name') or '', ''])
            ws2.cell(ws2.max_row, 2).font = Font(bold=True, italic=True)
        for col in range(1, 4):
            ws2.cell(ws2.max_row, col).border = thin
    for i, w in enumerate((16, 70, 18), 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws3 = wb.create_sheet('So_du_tai_khoan')
    ws3.append(['Mã TK', 'Tên tài khoản', 'Nợ', 'Có'])
    style_header(ws3, 4)
    try:
        accounts = [
            a for a in _coa_line_map(conn)
            if a.get('is_postable')
        ]
    except Exception:
        accounts = []
    if not accounts:
        try:
            accounts = [
                dict(r) for r in conn.execute(
                    """
                    SELECT code, name FROM sme_chart_of_accounts
                    WHERE is_active = 1 AND is_postable = 1
                    ORDER BY code
                    """
                ).fetchall()
            ]
        except sqlite3.Error:
            accounts = []
    for acc in accounts:
        ws3.append([acc.get('code'), acc.get('name') or '', 0, 0])
        ws3.cell(ws3.max_row, 3).number_format = '#,##0'
        ws3.cell(ws3.max_row, 4).number_format = '#,##0'
        for col in range(1, 5):
            ws3.cell(ws3.max_row, col).border = thin
            ws3.cell(ws3.max_row, col).fill = leaf_fill
    for i, w in enumerate((14, 55, 18, 18), 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _dk_document_no(fiscal_year: int) -> str:
    return f'DK{int(fiscal_year)}'


def _active_dk_ids(conn: sqlite3.Connection, fiscal_year: int) -> list[int]:
    rows = conn.execute(
        """
        SELECT id FROM sme_journal_entries
        WHERE document_type = ?
          AND document_no = ?
          AND status = 'posted'
          AND reverses_id IS NULL
        ORDER BY id
        """,
        (DK_DOC_TYPE, _dk_document_no(fiscal_year)),
    ).fetchall()
    return [int(r[0]) for r in rows]


def post_trial_balance_opening(
    conn: sqlite3.Connection,
    *,
    fiscal_year: int,
    lines: list[dict],
    created_by: str | None = None,
    replace_existing: bool = True,
) -> dict[str, Any]:
    """Ghi bút toán DK 31/12 năm trước từ số dư Nợ/Có từng TK."""
    from Services.sme.journal_engine import (
        post_journal_entry,
        resolve_postable_account,
        reverse_journal_entry,
    )

    year = int(fiscal_year)
    if year < 2000:
        raise ValueError('Năm tài chính không hợp lệ')
    posting_date = f'{year - 1}-12-31'
    prepared: list[dict] = []
    skipped: list[str] = []
    for raw in lines:
        code = str(raw.get('account_code') or raw.get('code') or '').strip()
        debit = _money(raw.get('debit'))
        credit = _money(raw.get('credit'))
        if not code or (debit == 0 and credit == 0):
            continue
        try:
            postable = resolve_postable_account(conn, code)
        except Exception:
            skipped.append(code)
            continue
        if debit > 0 and credit > 0:
            raise ValueError(f'TK {code}: không được vừa Nợ vừa Có')
        prepared.append({
            'account_code': postable,
            'debit': float(debit),
            'credit': float(credit),
            'description': f'Số dư đầu năm {year}',
        })
    if not prepared:
        raise ValueError('Không có dòng số dư tài khoản hợp lệ')

    reversed_ids: list[int] = []
    if replace_existing:
        for eid in _active_dk_ids(conn, year):
            info = reverse_journal_entry(
                conn, eid,
                posting_date=posting_date,
                created_by=created_by,
                reason=f'Thay số dư đầu năm {year}',
            )
            reversed_ids.append(int(info.get('id') or eid))

    posted = post_journal_entry(
        conn,
        posting_date=posting_date,
        document_date=posting_date,
        document_type=DK_DOC_TYPE,
        document_no=_dk_document_no(year),
        document_id=year,
        business_type=DK_BUSINESS,
        description=f'Số dư đầu năm {year} (nhập từ Excel / phần mềm khác)',
        created_by=created_by,
        lines=prepared,
        allow_locked_period=True,
        skip_cash_balance_check=True,
    )
    return {
        'posted': True,
        'entry_id': posted.get('id'),
        'document_no': _dk_document_no(year),
        'posting_date': posting_date,
        'lines': len(prepared),
        'skipped_accounts': skipped,
        'reversed_entry_ids': reversed_ids,
    }


def apply_excel_import(
    conn: sqlite3.Connection,
    *,
    fiscal_year: int,
    parsed: dict[str, Any],
    created_by: str | None = None,
    commit: bool = True,
) -> dict[str, Any]:
    """Lưu overlay B01/B02 và/hoặc ghi bút toán DK từ file đã parse."""
    year = int(fiscal_year)
    result: dict[str, Any] = {
        'fiscal_year': year,
        'b01_saved': 0,
        'b02_saved': 0,
        'dk': None,
        'errors': list(parsed.get('errors') or []),
    }
    b01 = parsed.get('b01') or {}
    b02 = parsed.get('b02') or {}
    trial = parsed.get('trial_balance') or []
    if not b01 and not b02 and not trial:
        raise ValueError(
            'Không nhận ra dữ liệu. Cần cột «Mã chỉ tiêu» + «Số đầu năm», '
            'hoặc «Mã TK» + «Nợ» + «Có».'
        )
    if b01:
        info = save_opening_lines(
            conn, fiscal_year=year, report='B01', lines=b01,
            source='excel', updated_by=created_by, commit=False,
        )
        result['b01_saved'] = info['saved']
        if info.get('skipped'):
            result['errors'].append(
                'B01 bỏ qua mã không thuộc mẫu: ' + ', '.join(info['skipped'][:20])
            )
    if b02:
        info = save_opening_lines(
            conn, fiscal_year=year, report='B02', lines=b02,
            source='excel', updated_by=created_by, commit=False,
        )
        result['b02_saved'] = info['saved']
        if info.get('skipped'):
            result['errors'].append(
                'B02 bỏ qua mã không thuộc mẫu: ' + ', '.join(info['skipped'][:20])
            )
    if trial:
        result['dk'] = post_trial_balance_opening(
            conn, fiscal_year=year, lines=trial, created_by=created_by,
        )
    if commit:
        conn.commit()
    return result
