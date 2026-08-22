# -*- coding: utf-8 -*-
"""Bảng kê thu mua hàng hóa, dịch vụ không có hóa đơn — mẫu 02/TNDN (TT 20/2026/TT-BTC).

Số căn cước người bán = mã số thuế (MST) khi lập phiếu nhập / danh mục NCC.
"""
from __future__ import annotations

import re
import sqlite3
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FORM_CODE = '02/TNDN'
FORM_TITLE = 'BẢNG KÊ THU MUA HÀNG HÓA, DỊCH VỤ KHÔNG CÓ HÓA ĐƠN'
FORM_LEGAL = 'Ban hành kèm theo Thông tư số 20/2026/TT-BTC của Bộ trưởng Bộ Tài chính'

# Cột Excel (dòng dữ liệu) — khớp mẫu giấy
EXCEL_HEADERS = [
    'STT',
    'Ngày mua hàng',
    'Tên người bán',
    'Địa chỉ',
    'Số căn cước (MST)',
    'Số điện thoại',
    'Tên hàng hóa, dịch vụ',
    'Số lượng',
    'ĐVT',
    'Đơn giá',
    'Tổng giá thanh toán',
    'Ghi chú',
]

_HEADER_ALIASES = {
    'stt': 'stt',
    'ngày': 'purchase_date',
    'ngay': 'purchase_date',
    'ngày mua': 'purchase_date',
    'ngày mua hàng': 'purchase_date',
    'tên người bán': 'seller_name',
    'ten nguoi ban': 'seller_name',
    'người bán': 'seller_name',
    'địa chỉ': 'seller_address',
    'dia chi': 'seller_address',
    'số căn cước': 'seller_id_no',
    'so can cuoc': 'seller_id_no',
    'căn cước': 'seller_id_no',
    'cccd': 'seller_id_no',
    'mst': 'seller_id_no',
    'mã số thuế': 'seller_id_no',
    'số căn cước (mst)': 'seller_id_no',
    'số điện thoại': 'seller_phone',
    'dien thoai': 'seller_phone',
    'điện thoại': 'seller_phone',
    'tên hàng hóa': 'item_name',
    'tên hàng hóa, dịch vụ': 'item_name',
    'hàng hóa': 'item_name',
    'dịch vụ': 'item_name',
    'số lượng': 'quantity',
    'số lượng, trọng lượng': 'quantity',
    'đvt': 'unit',
    'đơn vị': 'unit',
    'đơn giá': 'unit_price',
    'tổng giá': 'amount',
    'tổng giá thanh toán': 'amount',
    'thành tiền': 'amount',
    'ghi chú': 'note',
}


def ensure_purchase_02_tndn_tables(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS sme_purchase_02_tndn (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            period_month TEXT NOT NULL,
            purchase_date TEXT NOT NULL,
            seller_name TEXT NOT NULL DEFAULT '',
            seller_address TEXT NOT NULL DEFAULT '',
            seller_id_no TEXT NOT NULL DEFAULT '',
            seller_phone TEXT NOT NULL DEFAULT '',
            item_name TEXT NOT NULL DEFAULT '',
            quantity REAL NOT NULL DEFAULT 0,
            unit TEXT NOT NULL DEFAULT '',
            unit_price REAL NOT NULL DEFAULT 0,
            amount REAL NOT NULL DEFAULT 0,
            note TEXT NOT NULL DEFAULT '',
            purchase_place TEXT NOT NULL DEFAULT '',
            branch_code TEXT NOT NULL DEFAULT 'HQ',
            created_at TEXT,
            updated_at TEXT
        )
        """
    )
    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_sme_02_tndn_period
        ON sme_purchase_02_tndn(period_month, purchase_date)
        """
    )


def _f(v: Any) -> float:
    if v is None or v == '':
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(' ', '').replace('\u00a0', '')
    # VN: 1.234.567,89
    if re.search(r',\d{1,2}$', s) and s.count(',') == 1:
        s = s.replace('.', '').replace(',', '.')
    elif s.count('.') > 1 and ',' not in s:
        # VN nghìn: 1.234.567
        s = s.replace('.', '')
    else:
        s = s.replace(',', '')
    try:
        return float(Decimal(s))
    except (InvalidOperation, ValueError):
        return 0.0


def _norm_date(v: Any) -> str:
    """Trả về YYYY-MM-DD."""
    if v is None or v == '':
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    if hasattr(v, 'strftime'):
        try:
            return v.strftime('%Y-%m-%d')
        except Exception:
            pass
    s = str(v).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y'):
        try:
            return datetime.strptime(s[:10] if fmt == '%Y-%m-%d' else s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    # Excel serial
    try:
        from openpyxl.utils.datetime import from_excel
        n = float(s)
        if n > 20000:
            return from_excel(n).strftime('%Y-%m-%d')
    except Exception:
        pass
    return s[:10]


def _digits(s: str) -> str:
    return re.sub(r'\D', '', str(s or ''))


def load_tenant_business_info(conn: sqlite3.Connection | None = None) -> dict:
    """Đọc business_info của tenant DB hiện tại (ưu tiên hơn tenant_profile)."""
    own = False
    if conn is None:
        from db_utils import get_db_connection
        conn = get_db_connection()
        own = True
    try:
        try:
            row = conn.execute('SELECT * FROM business_info LIMIT 1').fetchone()
        except sqlite3.Error:
            return {}
        if not row:
            return {}
        return dict(row)
    finally:
        if own:
            try:
                conn.close()
            except Exception:
                pass


def default_purchase_place(biz: dict | None = None) -> str:
    """
    Mặc định «Địa chỉ nơi tổ chức thu mua» = thông tin DN trên tenant.
    Ưu tiên địa chỉ; nếu có tên DN thì ghép «Tên — Địa chỉ».
    """
    biz = biz or {}
    name = (
        biz.get('business_name')
        or biz.get('company_name')
        or biz.get('name')
        or ''
    ).strip()
    addr = (biz.get('address') or biz.get('business_address') or '').strip()
    if name and addr:
        return f'{name} — {addr}'
    return addr or name or ''


def biz_export_fields(biz: dict | None = None) -> dict:
    biz = biz or {}
    return {
        'business_name': (
            biz.get('business_name') or biz.get('company_name') or biz.get('name') or ''
        ).strip(),
        'tax_code': (biz.get('tax_code') or biz.get('mst') or '').strip(),
        'address': (biz.get('address') or biz.get('business_address') or '').strip(),
        'phone': (
            biz.get('phone') or biz.get('tel') or biz.get('mobile') or ''
        ).strip(),
        'purchase_place': default_purchase_place(biz),
    }


def resolve_branch_for_write(branch_code: str | None) -> str:
    """Lưu dòng: ALL → HQ."""
    code = (branch_code or '').strip().upper() or 'HQ'
    if code in ('ALL', '*'):
        return 'HQ'
    return code


def resolve_branch_for_read(branch_code: str | None) -> str | None:
    """Đọc: ALL → không lọc."""
    code = (branch_code or '').strip().upper()
    if not code or code in ('ALL', '*'):
        return None
    return code


def list_lines(
    conn: sqlite3.Connection,
    *,
    date_from: str,
    date_to: str,
    branch_code: str | None = None,
) -> list[dict]:
    ensure_purchase_02_tndn_tables(conn)
    sql = """
        SELECT * FROM sme_purchase_02_tndn
        WHERE purchase_date >= ? AND purchase_date <= ?
    """
    params: list[Any] = [date_from, date_to]
    if branch_code and branch_code not in ('ALL', '*', ''):
        sql += ' AND branch_code = ?'
        params.append(branch_code)
    sql += ' ORDER BY purchase_date, id'
    rows = conn.execute(sql, params).fetchall()
    out = []
    for r in rows:
        d = dict(r)
        out.append(d)
    return out


def replace_period_lines(
    conn: sqlite3.Connection,
    *,
    period_month: str,
    lines: list[dict],
    purchase_place: str = '',
    branch_code: str = 'HQ',
) -> int:
    """Thay toàn bộ dòng của một tháng (YYYY-MM)."""
    ensure_purchase_02_tndn_tables(conn)
    pm = (period_month or '')[:7]
    if not re.match(r'^\d{4}-\d{2}$', pm):
        raise ValueError('period_month phải dạng YYYY-MM')
    bc = resolve_branch_for_write(branch_code)
    place = (purchase_place or '').strip()
    if not place:
        place = default_purchase_place(load_tenant_business_info(conn))
    conn.execute(
        'DELETE FROM sme_purchase_02_tndn WHERE period_month = ? AND branch_code = ?',
        (pm, bc),
    )
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    n = 0
    for raw in lines or []:
        name = (raw.get('seller_name') or '').strip()
        item = (raw.get('item_name') or '').strip()
        if not name and not item:
            continue
        pdate = _norm_date(raw.get('purchase_date')) or f'{pm}-01'
        # Chuẩn hóa ngày sai định dạng
        if not re.match(r'^\d{4}-\d{2}-\d{2}$', pdate):
            pdate = f'{pm}-01'
        qty = _f(raw.get('quantity'))
        price = _f(raw.get('unit_price'))
        amount = _f(raw.get('amount'))
        if amount <= 0 and qty and price:
            amount = round(qty * price, 2)
        id_no = _digits(raw.get('seller_id_no') or '')
        row_place = (raw.get('purchase_place') or '').strip() or place
        conn.execute(
            """
            INSERT INTO sme_purchase_02_tndn (
                period_month, purchase_date, seller_name, seller_address,
                seller_id_no, seller_phone, item_name, quantity, unit,
                unit_price, amount, note, purchase_place, branch_code,
                created_at, updated_at
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                pm, pdate, name, (raw.get('seller_address') or '').strip(),
                id_no, (raw.get('seller_phone') or '').strip(), item,
                qty, (raw.get('unit') or '').strip(), price, amount,
                (raw.get('note') or '').strip(),
                row_place,
                bc, now, now,
            ),
        )
        n += 1
    return n


def parse_excel_rows(file_storage_or_bytes) -> list[dict]:
    """Đọc file Excel mẫu 02/TNDN → list dict."""
    if hasattr(file_storage_or_bytes, 'read'):
        data = file_storage_or_bytes.read()
        if hasattr(file_storage_or_bytes, 'seek'):
            try:
                file_storage_or_bytes.seek(0)
            except Exception:
                pass
        wb = load_workbook(BytesIO(data), data_only=True)
    else:
        wb = load_workbook(BytesIO(file_storage_or_bytes), data_only=True)
    ws = wb.active

    header_row = None
    col_map: dict[int, str] = {}
    for r in range(1, min(ws.max_row or 1, 25) + 1):
        vals = []
        for c in range(1, min(ws.max_column or 1, 20) + 1):
            v = ws.cell(r, c).value
            vals.append(str(v).strip().lower() if v is not None else '')
        joined = ' '.join(vals)
        if 'căn cước' in joined or 'người bán' in joined or (
            'ngày mua' in joined and 'đơn giá' in joined
        ) or ('stt' in vals and any('ngày' in x for x in vals)):
            header_row = r
            for c, raw in enumerate(vals, start=1):
                key = _HEADER_ALIASES.get(raw)
                if not key:
                    for ak, av in _HEADER_ALIASES.items():
                        if ak and ak in raw:
                            key = av
                            break
                if key and key != 'stt':
                    col_map[c] = key
            break

    if not header_row or not col_map:
        # Fallback: giả định hàng 6 = header chuẩn mẫu hệ thống
        header_row = 6
        for i, h in enumerate(EXCEL_HEADERS, start=1):
            key = _HEADER_ALIASES.get(h.lower())
            if key and key != 'stt':
                col_map[i] = key

    lines: list[dict] = []
    for r in range(header_row + 1, (ws.max_row or header_row) + 1):
        row: dict[str, Any] = {}
        empty = True
        for c, key in col_map.items():
            val = ws.cell(r, c).value
            if val is not None and str(val).strip() != '':
                empty = False
            row[key] = val
        if empty:
            continue
        # Bỏ dòng tổng / dòng trống tên
        seller = str(row.get('seller_name') or '').strip()
        item = str(row.get('item_name') or '').strip()
        name_probe = (seller or item).lower()
        if 'tổng' in name_probe:
            continue
        if not seller and not item:
            continue
        pdate = _norm_date(row.get('purchase_date'))
        qty = _f(row.get('quantity'))
        price = _f(row.get('unit_price'))
        amount = _f(row.get('amount'))
        if amount <= 0 and qty and price:
            amount = round(qty * price, 2)
        id_no = _digits(row.get('seller_id_no') or '')
        lines.append({
            'purchase_date': pdate,
            'seller_name': seller,
            'seller_address': str(row.get('seller_address') or '').strip(),
            'seller_id_no': id_no,
            'seller_phone': str(row.get('seller_phone') or '').strip(),
            'item_name': item,
            'quantity': qty,
            'unit': str(row.get('unit') or '').strip(),
            'unit_price': price,
            'amount': amount,
            'note': str(row.get('note') or '').strip(),
        })
    return lines


def build_excel(
    lines: list[dict],
    *,
    business_name: str = '',
    tax_code: str = '',
    address: str = '',
    phone: str = '',
    purchase_place: str = '',
    period_label: str = '',
) -> BytesIO:
    """Xuất Excel đẹp (kiểu Báo cáo tồn kho): tiêu đề, header xám, border, tổng."""
    wb = Workbook()
    ws = wb.active
    ws.title = '02-TNDN'

    thin = Border(
        top=Side(style='thin'),
        left=Side(style='thin'),
        bottom=Side(style='thin'),
        right=Side(style='thin'),
    )
    fill_header = PatternFill('solid', fgColor='FFE0E0E0')
    fill_title = PatternFill('solid', fgColor='FF1E40AF')
    font_title = Font(name='Calibri', bold=True, size=14, color='FFFFFFFF')
    font_bold = Font(name='Calibri', bold=True, size=11)
    font_norm = Font(name='Calibri', size=11)
    align_c = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_l = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_r = Alignment(horizontal='right', vertical='center')

    # Dòng 1–2: tiêu đề
    ws.merge_cells('A1:L1')
    c1 = ws['A1']
    c1.value = f'Mẫu số: {FORM_CODE}'
    c1.font = Font(name='Calibri', bold=True, size=12, color='FFFFFFFF')
    c1.fill = fill_title
    c1.alignment = align_c

    ws.merge_cells('A2:L2')
    c2 = ws['A2']
    c2.value = FORM_TITLE
    c2.font = font_title
    c2.fill = fill_title
    c2.alignment = align_c

    ws.merge_cells('A3:L3')
    ws['A3'] = FORM_LEGAL
    ws['A3'].font = Font(name='Calibri', italic=True, size=9, color='FF4B5563')
    ws['A3'].alignment = align_c

    ws.merge_cells('A4:L4')
    ws['A4'] = period_label or 'Kỳ: …'
    ws['A4'].font = font_bold
    ws['A4'].alignment = align_c

    # Thông tin DN
    info_rows = [
        (5, f'Tên doanh nghiệp: {business_name}'),
        (6, f'Mã số thuế: {tax_code}'),
        (7, f'Địa chỉ: {address}'),
        (8, f'Số điện thoại: {phone}'),
        (9, f'Địa chỉ nơi tổ chức thu mua: {purchase_place}'),
    ]
    for r, text in info_rows:
        ws.merge_cells(f'A{r}:L{r}')
        ws[f'A{r}'] = text
        ws[f'A{r}'].font = font_norm
        ws[f'A{r}'].alignment = align_l

    header_row = 11
    for col, h in enumerate(EXCEL_HEADERS, start=1):
        cell = ws.cell(header_row, col, h)
        cell.font = font_bold
        cell.fill = fill_header
        cell.border = thin
        cell.alignment = align_c

    data_start = header_row + 1
    total_amount = 0.0
    for i, line in enumerate(lines or [], start=1):
        r = data_start + i - 1
        pdate = line.get('purchase_date') or ''
        if pdate and len(pdate) >= 10:
            try:
                pdate = datetime.strptime(pdate[:10], '%Y-%m-%d').strftime('%d/%m/%Y')
            except ValueError:
                pass
        vals = [
            i,
            pdate,
            line.get('seller_name') or '',
            line.get('seller_address') or '',
            line.get('seller_id_no') or '',
            line.get('seller_phone') or '',
            line.get('item_name') or '',
            _f(line.get('quantity')),
            line.get('unit') or '',
            _f(line.get('unit_price')),
            _f(line.get('amount')),
            line.get('note') or '',
        ]
        total_amount += float(vals[10] or 0)
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(r, col, val)
            cell.font = font_norm
            cell.border = thin
            if col in (1, 2, 5, 6, 9):
                cell.alignment = align_c
            elif col in (10, 11):
                cell.alignment = align_r
                cell.number_format = '#,##0'
            elif col == 8:
                cell.alignment = align_r
                cell.number_format = '#,##0.###'
            else:
                cell.alignment = align_l

    total_row = data_start + len(lines or [])
    ws.merge_cells(f'A{total_row}:J{total_row}')
    cell_t = ws.cell(total_row, 1, 'TỔNG GIÁ TRỊ HÀNG HÓA, DỊCH VỤ MUA VÀO')
    cell_t.font = font_bold
    cell_t.alignment = align_c
    cell_t.fill = fill_header
    cell_t.border = thin
    for col in range(2, 11):
        try:
            c = ws.cell(total_row, col)
            c.border = thin
            c.fill = fill_header
        except Exception:
            pass
    cell_amt = ws.cell(total_row, 11, total_amount)
    cell_amt.font = font_bold
    cell_amt.alignment = align_r
    cell_amt.number_format = '#,##0'
    cell_amt.fill = fill_header
    cell_amt.border = thin
    try:
        ws.cell(total_row, 12).border = thin
        ws.cell(total_row, 12).fill = fill_header
    except Exception:
        pass

    # --- Phần dưới bảng (đúng mẫu 02/TNDN) ---
    try:
        from helpers import so_thanh_chu
        amount_words = so_thanh_chu(int(round(total_amount)))
    except Exception:
        amount_words = '........................................'

    r = total_row + 2
    ws.merge_cells(f'A{r}:L{r}')
    amount_fmt = f'{total_amount:,.0f}'.replace(',', '.')
    ws[f'A{r}'] = (
        f'- Tổng giá trị hàng hóa, dịch vụ mua vào: {amount_fmt}'
        f'  (Số tiền bằng chữ: {amount_words}).'
    )
    ws[f'A{r}'].font = font_bold
    ws[f'A{r}'].alignment = align_l

    # Ngày ký (từ period_label nếu có dd/mm/yyyy, không thì để trống)
    sign_day = sign_month = sign_year = '....'
    m = re.search(r'(\d{2})/(\d{2})/(\d{4})', period_label or '')
    if m:
        sign_day, sign_month, sign_year = m.group(1), m.group(2), m.group(3)

    r = total_row + 4
    ws.merge_cells(f'A{r}:D{r}')
    ws[f'A{r}'] = 'Người lập bảng kê'
    ws[f'A{r}'].font = font_bold
    ws[f'A{r}'].alignment = align_c

    ws.merge_cells(f'E{r}:H{r}')
    ws[f'E{r}'] = f'Ngày {sign_day} tháng {sign_month} năm {sign_year}'
    ws[f'E{r}'].font = font_norm
    ws[f'E{r}'].alignment = align_c

    ws.merge_cells(f'I{r}:L{r}')
    ws[f'I{r}'] = 'Người đại diện hoặc người được ủy quyền của doanh nghiệp'
    ws[f'I{r}'].font = font_bold
    ws[f'I{r}'].alignment = align_c

    r = total_row + 5
    ws.merge_cells(f'A{r}:D{r}')
    ws[f'A{r}'] = '(Ký, ghi rõ họ tên)'
    ws[f'A{r}'].font = Font(name='Calibri', italic=True, size=9)
    ws[f'A{r}'].alignment = align_c

    ws.merge_cells(f'I{r}:L{r}')
    ws[f'I{r}'] = '(Ký tên, đóng dấu)'
    ws[f'I{r}'].font = Font(name='Calibri', italic=True, size=9)
    ws[f'I{r}'].alignment = align_c

    # Khoảng trống chữ ký
    for blank in range(total_row + 6, total_row + 9):
        ws.row_dimensions[blank].height = 16

    r = total_row + 10
    ws.merge_cells(f'A{r}:L{r}')
    ws[f'A{r}'] = 'Ghi chú:'
    ws[f'A{r}'].font = font_bold

    notes = [
        '- Căn cứ vào số thực tế các hàng hóa, dịch vụ mà doanh nghiệp mua của người bán không có hóa đơn, '
        'lập bảng kê khai theo thứ tự thời gian mua, doanh nghiệp ghi đầy đủ các chỉ tiêu trên bảng kê, '
        'tổng hợp bảng kê hàng tháng.',
        '- Đối với doanh nghiệp có tổ chức các trạm thu mua ở nhiều nơi thì từng trạm thu mua phải lập từng '
        'bảng kê riêng. Doanh nghiệp lập bảng kê tổng hợp chung của các trạm.',
        '- Số căn cước người bán dùng làm mã số thuế (MST) khi lập phiếu nhập kho / danh mục nhà cung cấp.',
    ]
    for i, text in enumerate(notes):
        rr = r + 1 + i
        ws.merge_cells(f'A{rr}:L{rr}')
        ws[f'A{rr}'] = text
        ws[f'A{rr}'].font = Font(name='Calibri', italic=True, size=8, color='FF374151')
        ws[f'A{rr}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[rr].height = 28

    last_row = r + len(notes)

    # Độ rộng cột vừa A4 ngang (~277mm printable)
    widths = [5, 11, 14, 16, 12, 10, 16, 8, 6, 10, 11, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[header_row].height = 30

    # In: A4 ngang, fit bề rộng 1 trang
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f'A1:L{last_row}'
    ws.page_margins.left = 0.35
    ws.page_margins.right = 0.35
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2
    ws.print_options.horizontalCentered = True

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def build_template_excel(biz: dict | None = None) -> BytesIO:
    sample = [
        {
            'purchase_date': datetime.now().strftime('%Y-%m-%d'),
            'seller_name': 'Nguyễn Văn A',
            'seller_address': 'Ấp 1, xã …, huyện …',
            'seller_id_no': '079012345678',
            'seller_phone': '0901234567',
            'item_name': 'Ví dụ: Lúa tươi / Rau củ / Dịch vụ thu hoạch',
            'quantity': 100,
            'unit': 'Kg',
            'unit_price': 10000,
            'amount': 1000000,
            'note': 'Mẫu — xóa dòng này trước khi lưu',
        }
    ]
    fields = biz_export_fields(biz or {})
    return build_excel(
        sample,
        business_name=fields['business_name'],
        tax_code=fields['tax_code'],
        address=fields['address'],
        phone=fields['phone'],
        purchase_place=fields['purchase_place'],
        period_label='Mẫu nhập liệu — thay bằng dữ liệu thực tế',
    )


def group_lines_for_import(lines: list[dict]) -> list[dict]:
    """Gộp theo (CCCD/MST, tên, ngày) để nạp phiếu nhập."""
    groups: dict[tuple, dict] = {}
    for line in lines:
        id_no = _digits(line.get('seller_id_no') or '')
        name = (line.get('seller_name') or '').strip()
        pdate = _norm_date(line.get('purchase_date')) or ''
        key = (id_no or name.lower(), name.lower(), pdate)
        if key not in groups:
            groups[key] = {
                'seller_name': name,
                'seller_address': (line.get('seller_address') or '').strip(),
                'seller_id_no': id_no,
                'seller_phone': (line.get('seller_phone') or '').strip(),
                'purchase_date': pdate,
                'items': [],
                'total': 0.0,
            }
        g = groups[key]
        if not g['seller_address'] and line.get('seller_address'):
            g['seller_address'] = str(line.get('seller_address')).strip()
        amt = _f(line.get('amount'))
        g['items'].append({
            'invoice_name': (line.get('item_name') or '').strip(),
            'unit': (line.get('unit') or '').strip(),
            'qty': _f(line.get('quantity')) or 1,
            'price': _f(line.get('unit_price')),
            'tax_pct': 0,
            'note': (line.get('note') or '').strip(),
        })
        g['total'] += amt
    return list(groups.values())
